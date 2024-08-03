import pandas as pd
from tkinter import filedialog
import tkinter as tk
from tkinter import filedialog
import tkinter.messagebox as mb
from pathlib import Path
import openpyxl

import xlsxwriter


from openpyxl.styles import PatternFill


# def by_openpy():
#     wb = openpyxl.load_workbook('OZON-357.xlsx')
#     ws = wb.active
#
#
#     wb2 = openpyxl.load_workbook('Отказные_посылки.xlsx')
#     ws2 = wb2.active
#
#     def reverseCombiner(rowList):
#         # Don't do anything for empty list. Otherwise,
#         # make a copy and sort.
#
#         if len(rowList) == 0: return []
#         sortedList = rowList[:]
#         sortedList.sort()
#
#         # Init, empty tuple, use first item for previous and
#         # first in this run.
#
#         tupleList = []
#         firstItem = sortedList[0]
#         prevItem = sortedList[0]
#
#         # Process all other items in order.
#
#         for item in sortedList[1:]:
#             # If start of new run, add tuple and use new first-in-run.
#
#             if item != prevItem + 1:
#                 tupleList = [(firstItem, prevItem + 1 - firstItem)] + tupleList
#                 firstItem = item
#
#             # Regardless, current becomes previous for next loop.
#
#             prevItem = item
#
#         # Finish off the final run and return tuple list.
#
#         tupleList = [(firstItem, prevItem + 1 - firstItem)] + tupleList
#         return tupleList
#
#
#     i = 1
#     list_to_delete = []
#     for cell in ws['M']:
#         i += 1
#         print(i)
#         if cell.value != None and ws[f'N{i}'] != None and ws[f'O{i}'] != None and ws[f'R{i}'] != None:
#             print(ws[f'A{i}'].value)
#             list_to_delete.append(i)
#
#     tuples = reverseCombiner(list_to_delete)
#     print(f"Original: {list_to_delete}")
#     print(f"Tuples:   {tuples}\n")
#     for tuple in tuples:
#         if tuple[0]-1 != 2:
#             print(tuple[0]-1)
#             ws.delete_rows(tuple[0]-1, tuple[1])
#         else:
#             ws.delete_rows(tuple[0] - 1, tuple[1]-1)
#
#     wb.save('Шаблон_OZON-357_только пустые.xlsx')


def by_pandas():
    msg = "Выберите ШАБЛОН партии"
    mb.showinfo("Выгруженный из скарифа шаблон партии", msg)
    file_name = filedialog.askopenfilename()
    print(file_name)
    file_name_only = Path(file_name).stem
    df = pd.read_excel(file_name, sheet_name=0, engine='openpyxl', converters={1: str, 12: str, 13: str, 17: str})
    print(df)
    df.columns = ['Номер отправления ИМ', 'Номер пломбы', 'ФИО получателя', 'Фамилия',
                          'Имя', 'Отчество', 'Индекс', 'Область', 'Город', 'Адрес получателя', 'Телефон',
                          'Емейл', 'Серия паспорта', 'Номер паспорта', 'Дата выдачи', 'Орган выдачи', 'Дата рождения',
                             'Идентификационный налоговый номер', 'Ссылка на товар', 'Наименование товара', 'Код ТН ВЭД',
                             'Количество единиц товара', 'Стоимость ед. товарной позиции', 'Стоимость позиции',
                             'Общая стоимость накладной(посылки)', 'Валюта Объявленной стоимости товара', 'Вес брутто (Вес позиции)',
                             'Общий Вес места (накладной)', 'Длина коробки, см', 'Ширина, см', 'Высота коробки, см', 'Отправитель по AWB',
                              '№ AWB', 'Дата AWB', 'Страна отправления', 'Торгующая страна', 'Условия поставки', 'Код страны получателя',
                             'Краткое наименование страны получателя', 'Код документа (паспорта)', 'Номер накладной СДЭК',
                             'пол (1 - женский, 0 - мужской)', 'Признак платности (1 - платный, 0 - нет)', 'Примечание']

    msg = "Выберите файл 'Выгрузка по ПД'"
    mb.showinfo("Выгруженные из скарифа посылки с отказами", msg)
    file_name_empty_pd = filedialog.askopenfilename()

    df_empty_pd = pd.read_excel(file_name_empty_pd, usecols="B", sheet_name=0, engine='openpyxl')
    df_empty_pd = df_empty_pd.drop_duplicates()
    df_empty_pd["is_work"] = "True"
    df_empty_pd = pd.merge(df, df_empty_pd, how='left', left_on="Номер отправления ИМ", right_on="Посылки")
    df_empty_pd = df_empty_pd.loc[df_empty_pd["is_work"] == "True"]
    df_empty_pd = df_empty_pd.drop(['is_work'], axis='columns')
    print(df_empty_pd)
    msg = "Выберите файл 'Отказные по партии'"
    mb.showinfo("Выгруженные из скарифа посылки с отказами", msg)
    file_name_refuse = filedialog.askopenfilename()

    df_refuses = pd.read_excel(file_name_refuse, usecols="B, O", sheet_name=0, engine='openpyxl')
    df_refuses = df_refuses.loc[df_refuses["Код причины отказа"] == 420]
    df_merged = pd.merge(df_refuses, df, how="left", left_on="Трек-номер", right_on="Номер отправления ИМ")
    print(df_merged)
    df_merged = df_merged.drop(['Трек-номер', 'Код причины отказа'], axis='columns')
    df_concat = pd.concat([df_merged, df_empty_pd])
    df_concat["is_work"] = "True"
    df_concat = df_concat[['Номер отправления ИМ', "is_work"]].drop_duplicates('Номер отправления ИМ')
    print(df_concat)
    df_merged = pd.merge(df, df_concat, how="left", right_on="Номер отправления ИМ", left_on="Номер отправления ИМ")

    df_work = df_merged.loc[df_merged["is_work"] == "True"]
    df_work = df_work.drop(['is_work'], axis='columns')

    msg = "Выберите выгрузку ПД из Сайта"
    mb.showinfo("Собранные паспортные данные", msg)
    file_name_pd = filedialog.askopenfilename()
    df_pd = pd.read_excel(file_name_pd, sheet_name=0, engine='openpyxl', converters={7: str, 8: str, 11: str})

    df_pd = df_pd.rename(columns={'Фамилия': 'Фамилия_нов', 'Имя': 'Имя_нов', 'Отчество': 'Отчество_нов',
                                  'Серия паспорта': 'Серия паспорта_нов', 'Номер паспорта': 'Номер паспорта_нов',
                                  'Дата рождения': 'Дата рождения_нов'}).drop_duplicates('Номер телефона/Логин', keep='last')
    df_merged = pd.merge(df_work, df_pd, how="left", right_on="Номер телефона/Логин", left_on="Телефон")
    df_merged = df_merged[df_merged['ID'].notnull()]
    df_merged['Телефон'] = df_merged['Телефон'].astype(str).replace('nan', '')
    df_merged['Идентификационный налоговый номер'] = df_merged['Телефон'].astype(str).replace('nan', '')
    df_merged['ИНН'] = df_merged['ИНН'].astype(str).replace('nan', '')
    df_merged['Серия паспорта_нов'] = df_merged['Серия паспорта_нов'].astype(str).replace('nan', '')

    df_merged = df_merged[['Номер отправления ИМ', 'Номер пломбы', 'ФИО получателя', 'Фамилия',
                          'Имя', 'Отчество', 'Индекс', 'Область', 'Город', 'Адрес получателя', 'Телефон',
                          'Емейл', 'Серия паспорта', 'Номер паспорта', 'Дата выдачи', 'Орган выдачи', 'Дата рождения',
                             'Идентификационный налоговый номер', 'Ссылка на товар', 'Наименование товара', 'Код ТН ВЭД',
                             'Количество единиц товара', 'Стоимость ед. товарной позиции', 'Стоимость позиции',
                             'Общая стоимость накладной(посылки)', 'Валюта Объявленной стоимости товара', 'Вес брутто (Вес позиции)',
                             'Общий Вес места (накладной)', 'Длина коробки, см', 'Ширина, см', 'Высота коробки, см', 'Отправитель по AWB',
                              '№ AWB', 'Дата AWB', 'Страна отправления', 'Торгующая страна', 'Условия поставки', 'Код страны получателя',
                             'Краткое наименование страны получателя', 'Код документа (паспорта)', 'Номер накладной СДЭК',
                             'пол (1 - женский, 0 - мужской)', 'Признак платности (1 - платный, 0 - нет)', 'Примечание',
                           'Фамилия_нов', 'Имя_нов', 'Отчество_нов', 'Серия паспорта_нов', 'Номер паспорта_нов',
                           'Дата выдачи паспорта', 'ИНН', 'Дата рождения_нов']]


    writer = pd.ExcelWriter(f'{file_name_only} для сравнения со старыми пд.xlsx', engine='xlsxwriter')
    df_merged.to_excel(writer, sheet_name='Sheet1', index=False)
    writer.close()

    wb = openpyxl.load_workbook(f'{file_name_only} для сравнения со старыми пд.xlsx')
    ws = wb.active
    yelFill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')

    i = 1
    dict_latters = {'D': 'AS',
                    'E': 'AT',
                    'F': 'AU',
                    'M': 'AV',
                    'N': 'AW',
                    'O': 'AX',
                    'Q': 'AZ',
                    'R': 'AY'}
    for cell in ws['D']:
        i += 1
        print(i)
        for key, v in dict_latters.items():
            print(key, v)
            if ws[f'{key}{i}'].value == ws[f'{v}{i}'].value: #name
                ws[f'{key}{i}'].fill = yelFill
            else:
                ws[f'{key}{i}'].value = ws[f'{v}{i}'].value
        ws[f'C{i}'].value = (str(ws[f'AS{i}'].value) + ' ' + str(ws[f'AT{i}'].value) + ' ' + str(ws[f'AU{i}'].value)).replace(' None', '')
        print(ws[f'C{i}'].value)
    ws.delete_cols(45, 8)
    ws.delete_rows(ws.max_row, 1)
    wb.save(f'{file_name_only}_Собранные ПД.xlsx')

    msg = "ГОТОВО!'"
    mb.showinfo("Ура ура!", msg)

by_pandas()





