import datetime
import logging
import os
import sqlite3 as sl
from copy import copy
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from io import BytesIO
from smtplib import SMTP_SSL

import numpy as np
import openpyxl
import pandas as pd
import pytz
import requests
from apispec import APISpec
from apispec.ext.marshmallow import MarshmallowPlugin
from apscheduler.schedulers.background import BackgroundScheduler
from flask import Flask, jsonify, request, render_template, redirect, url_for, send_file
from flask import Response
from flask import abort
from flask import flash
from flask import make_response
from flask_apispec import use_kwargs, marshal_with
from flask_apispec.extension import FlaskApiSpec
from flask_jwt_extended import (
    JWTManager, jwt_required, get_jwt_identity, set_access_cookies,
    unset_jwt_cookies)
from flask_restful import Resource, Api
from flask_sqlalchemy import SQLAlchemy
from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet.page import PageMargins
from sqlalchemy import create_engine
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import sessionmaker, scoped_session

from schemas import UserSchema, AuthSchema

download_folder = 'C:/Users/User/Desktop/ДОКУМЕНТЫ/'
addition_folder = f'{download_folder}Места-Паллеты/'
if not os.path.isdir(download_folder):
    os.makedirs(download_folder, exist_ok=True)
if not os.path.isdir(addition_folder):
    os.makedirs(addition_folder, exist_ok=True)
pd.set_option("display.precision", 3)
pd.options.display.float_format = '{:.3f}'.format

app_svh = Flask(__name__)
app_svh.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024
app_svh.config['UPLOAD_EXTENSIONS'] = ['.xls', '.xlsx']
app_svh.config["USE_X_SENDFILE"] = True

app_svh.config['JWT_CSRF_CHECK_FORM'] = False
app_svh.config['JWT_TOKEN_LOCATION'] = ['cookies']
app_svh.config["SQLALCHEMY_DATABASE_URI"] =('sqlite:///db.sqlite')

db = SQLAlchemy(app_svh)

client = app_svh.test_client()
engine = create_engine('sqlite:///db.sqlite')
session = scoped_session(sessionmaker(autocommit=False, autoflush=False, bind=engine))

Base = declarative_base()
Base.query = session.query_property()

jwt = JWTManager(app_svh)
app_svh.secret_key = 'c9e779a3258b42338334daaed51bccf7'
app_svh.config['SESSION_TYPE'] = 'filesystem'
app_svh.config['JWT_SECRET_KEY'] = 'c9e779a3258b42338334daaed51bccf7'

api = Api(app_svh)

class UserLogin(Resource):
    def post(self):
        username = request.get_json()['username']
        password = request.get_json()['password']
        if username == 'admin' and password == 'habr':
            access_token = create_access_token(identity={
                'role': 'admin',
            }, expires_delta=False)
            result = {'token': access_token}
            return result
        return {'error': 'Invalid username and password'}

class ProtectArea(Resource):
    @jwt_required
    def get(self):
        return {'answer': 42}


api.add_resource(UserLogin, '/api/login/')
api.add_resource(ProtectArea, '/api/protect-area/')

docs = FlaskApiSpec()
docs.init_app(app_svh)
app_svh.config.update({
    'APICPEC_SPEC': APISpec(
        title='videoblog',
        version='v1',
        openapi_version='2.0',
        plugins=[MarshmallowPlugin()]
    ),
    'APISPEC_SWAGGER_URL': '/swagger/'
})

from models import *
Base.metadata.create_all(bind=engine)

@app_svh.route('/todo/api/v1.0/register', methods=['POST'])
@use_kwargs(UserSchema)
@marshal_with(AuthSchema)
def register(**kwargs):
    try:
        user = User(**kwargs)
        session.add(user)
        session.commit()
        token = user.get_token()
    except Exception as e:
        logger.warning(f'register error: {e}')
        return {'message': str(e)}, 400
    return {'access_token': token}

@app_svh.route('/home')
def home():
    return render_template("home.html")

@app_svh.route('/login', methods=['GET', 'POST'])
def login_start():
    return render_template("login.html")

@app_svh.route('/get_user_id')
@jwt_required()
def get_user_id():
    user_id = get_jwt_identity()
    resp = make_response(redirect(url_for('fechone_plomb')))
    resp.set_cookie('user_id', str(user_id))
    return resp

@app_svh.route('/todo/api/v1.0/login_insert', methods=['POST'])
def login_insert():
    try:
        email = request.form['email']
        password = request.form['password']
        log = client.post('/todo/api/v1.0/login', json={'email': email, 'password': password})
        access_token = log.get_json()['access_token']
        resp = make_response(redirect(url_for('get_user_id')))
        set_access_cookies(resp, access_token)
    except Exception as e:
        logger.warning(e)
        return {'message': str(e)}, 400
    return resp


@app_svh.route('/todo/api/v1.0/login', methods=['POST'])
@use_kwargs(UserSchema(only=('email', 'password')))
@marshal_with(AuthSchema)
def login(**kwargs):
    user = User.authenticate(**kwargs)
    token = user.get_token()
    return {'access_token': token}

@app_svh.route('/logout')
@jwt_required()
def logout():
    resp = make_response(redirect(url_for('login_start')))
    #resp.set_cookie('access_token', max_age=0)
    unset_jwt_cookies(resp)
    return resp
@app_svh.teardown_appcontext
def shutdown_session(exception=None):
    session.remove()

@app_svh.errorhandler(422)
def error_handlers(err):
    headers = err.data.get('headers', None)
    messages = err.data.get('messages', ['Invalid request'])
    if headers:
        return jsonify({'message': messages}), 400, headers
    else:
        return jsonify({'message': messages}), 400

@app_svh.route('/a')
def returnAudioFile():
    path_to_audio_file = "/statiс/SndNoPlomb.wav" #audio from project dir
    return send_file(
       path_to_audio_file,
       mimetype="audio/wav",
       as_attachment=True,
       attachment_filename="SndNoPlomb.wav")

def insert_user_action(object_name, comment):
    user_id = request.cookies.get('user_id')
    now_time = datetime.datetime.now().strftime("%d.%m.%Y %H:%M")
    con = sl.connect('BAZA-reports.db')
    with con:
        cur = con.cursor()
        action_date = now_time
        statement = "INSERT INTO report (user_id, object_name, comment, action_date) VALUES (?, ?, ?, ?)"
        cur.execute(statement, [user_id, object_name, comment, action_date])
        con.commit()

now = datetime.datetime.now().strftime("%d.%m.%Y")
now_time = datetime.datetime.now().strftime("%d.%m.%Y %H:%M")
def setup_logger(name, log_file, level=logging.INFO):
    logging.basicConfig(format=u'%(levelname)-8s [%(asctime)s] %(message)s')  # filename=u'mylog.log'
    handler = logging.FileHandler(log_file)
    logger = logging.getLogger(name)
    logger.setLevel(level)
    logger.addHandler(handler)
    return logger
logger = setup_logger('logger', 'mylog.log')

logger_change_plob = setup_logger('logger_change_plob', f'ИЗМЕНЕНИЯ ПЛОМБ.log')
logger_API = setup_logger('logger_API', f'API_insert.log')

#style = '<style>.dataframe th{background: rgb(0,94,73);background: linear-gradient(71deg, rgba(0,94,73,1) 0%, rgba(18,126,80,1) 100%);padding: 10px;font-family: lucida console;color: #343434;border:2px dotted;text-align:left !important;}</style>'
style = '<style>.dataframe th{background: rgb(255,255,255);background: radial-gradient(circle, rgba(255,255,255,1) 0%, rgba(236,236,236,1) 100%);padding: 5px;color: #343434;font-family: monospace;font-size: 110%;border:2px solid #e0e0e0;text-align:left !important;}.dataframe{border: 3px solid #ffebeb !important;}</style>'

map_eng_to_rus = {'registration_numb': 'Реестр', 'party_numb': 'Партия',
                                    'parcel_numb': 'Трек-номер', 'parcel_plomb_numb': 'Пломба', 'parcel_weight': 'вес',
                                    'custom_status': 'Статус ТО', 'custom_status_short': 'Статус ТО (кратк)',
                                    'decision_date': 'Дата решения',
                                    'refuse_reason': 'Причина отказа',
                                    'pallet': 'Паллет',
                                    'zone': 'Зона',
                                  'VH_status': 'Статус ВХ',
                                  'goods': 'Товары'}

con = sl.connect('BAZA-reports.db')
with con:
    baza = con.execute("select count(*) from sqlite_master where type='table' and name='report'")
    for row in baza:
        # если таких таблиц нет
        if row[0] == 0:
            # создаём таблицу
            with con:
                con.execute("""
                            CREATE TABLE report (
                            ID INTEGER PRIMARY KEY AUTOINCREMENT,
                            user_id INTEGER NOT NULL,
                            object_name VARCHAR(30),
                            comment VARCHAR(120),
                            action_date DATETIME
                            );
                        """)
con.commit()
con.close()

@app_svh.route('/api/get_decisions', methods=['POST'])
def get_parcel_info_API():
    parcel_details = request.get_json()
    print(parcel_details)
    df = pd.read_json(parcel_details)
    print(df)
    try:
        con = sl.connect('BAZA.db')
        with con:
            data = con.execute("select count(*) from sqlite_master where type='table' and name='api_parcels'")
            for row in data:
                # если таких таблиц нет
                if row[0] == 0:
                    # создаём таблицу
                    with con:
                        con.execute("""
                                                                            CREATE TABLE api_parcels (
                                                                            ID INTEGER PRIMARY KEY AUTOINCREMENT,
                                                                            parcel_numb VARCHAR(25) NOT NULL UNIQUE ON CONFLICT REPLACE
                                                                            );
                                                                            """)
            df.to_sql('api_parcels', con=con, if_exists='replace', index=False)
            query = """SELECT * FROM api_parcels"""
            data = con.execute(query).fetchall()
            for row in data:
                print(row)
        with con:
            query = """SELECT baza.registration_numb, baza.parcel_numb, baza.custom_status, baza.custom_status_short, baza.decision_date, baza.refuse_reason
                        FROM baza                          
                        JOIN api_parcels ON api_parcels.parcel_numb = baza.parcel_numb
                        """
            data = con.execute(query).fetchall()
            for row in data:
                print(row)
            df_decisions = pd.read_sql(query, con)
            print(df_decisions)
            df_decisions = df_decisions.loc[:, ~df_decisions.columns.duplicated()].copy()
            writer = pd.ExcelWriter('df_decisions.xlsx', engine='xlsxwriter')
            df_decisions.to_excel(writer, sheet_name='Sheet1', index=False)
            writer.save()
    except Exception as e:
        return {'message': str(e)}, 400

    return Response(df_decisions.to_json(orient="records", indent=2), mimetype='application/json')


@app_svh.route('/api/get_decisions_TSD', methods=['POST'])
def get_parcel_info_API_TSD():
    parcel_details = request.get_json()
    parcel_numb = parcel_details['parcel_numb']
    con = sl.connect('BAZA.db')
    try:
        with con:
            query_party_numb = f"SELECT party_numb from baza where parcel_numb = '{parcel_numb}'"
            data = con.execute(query_party_numb).fetchone()
            if data is not None:
                for row in data:
                    print(row)
                    party_numb = row
                query = f"SELECT party_numb, parcel_numb, parcel_plomb_numb, custom_status_short, custom_status, decision_date, refuse_reason from baza where party_numb = '{party_numb}'"
                data = con.execute(query).fetchall()
                for row in data:
                    print(row)
            else:
                query_party_numb = f"SELECT party_numb from baza where parcel_plomb_numb = '{parcel_numb}'"
                data = con.execute(query_party_numb).fetchone()
                for row in data:
                    print(row)
                    party_numb = row
                query = f"SELECT party_numb, parcel_numb, parcel_plomb_numb, custom_status_short, custom_status, decision_date, refuse_reason from baza where party_numb = '{party_numb}'"
                data = con.execute(query).fetchall()
                for row in data:
                    print(row)
            df_decisions = pd.read_sql(query, con)
            df_decisions = df_decisions.loc[:, ~df_decisions.columns.duplicated()].copy()
            writer = pd.ExcelWriter('df_decisions.xlsx', engine='xlsxwriter')
            df_decisions.to_excel(writer, sheet_name='Sheet1', index=False)
            writer.save()
    except Exception as e:
        print(e)
        return {'message': str(e)}, 400

    return Response(df_decisions.to_json(orient="records", indent=2), mimetype='application/json')

def insert_decision_API(registration_numb, parcel_numb, custom_status, custom_status_short, refuse_reason, decision_date, last_mile):
    con = sl.connect('BAZA.db')
    cur = con.cursor()
    registration_numb = last_mile + registration_numb
    row_isalready_in = pd.read_sql(f"Select * from baza where parcel_numb = '{parcel_numb}'", con)

    if row_isalready_in.empty:
        with con:
            statement = "INSERT INTO baza (registration_numb, parcel_numb, custom_status, custom_status_short, refuse_reason, decision_date) VALUES (?, ?, ?, ?, ?, ?)"
            cur.execute(statement, [registration_numb, parcel_numb, custom_status, custom_status_short, refuse_reason, decision_date])
    else:
        with con:
            con.execute(f"Update baza set"
                        f" registration_numb = '{registration_numb}',"
                        f" custom_status = '{custom_status}',"
                        f" custom_status_short = '{custom_status_short}',"
                        f" decision_date = '{decision_date}',"
                        f" refuse_reason = '{refuse_reason}' where parcel_numb = '{parcel_numb}'")
    return True

@app_svh.route('/api/add_decision', methods=['POST'])
def add_decision_API():
    event_details = request.get_json()
    registration_numb = event_details["registration_numb"]
    parcel_numb = event_details["parcel_numb"]
    custom_status = event_details["Event"]
    last_mile = event_details["Last_mile"]
    if 'выпуск' in str(custom_status).lower():
        custom_status_short = 'ВЫПУСК'
    else:
        custom_status_short = 'ИЗЪЯТИЕ'
    refuse_reason = event_details["Event_comment"]
    decision_date = datetime.datetime.strptime(event_details["Event_date"], "%Y-%m-%d %H:%M:%S").replace(tzinfo=pytz.UTC).astimezone(pytz.timezone("Europe/London"))
    result = insert_decision_API(registration_numb, parcel_numb, custom_status, custom_status_short, refuse_reason, decision_date, last_mile)
    return jsonify(result)


def insert_event_API_test(df, event_date):
    parcel_list = df['parcel_numb'].to_list()
    logger.warning(parcel_list)
    logger.warning(event_date)
    for parcel in parcel_list:
        body = {"parcel_numb": parcel, "Event": "Отгружен с таможенного склада",
                "Event_comment": "Уссурийск", "Event_date": event_date}
        headers = {'accept': 'application/json'}
        response = requests.post('http://164.132.182.145:5000/api/add/new_event', json=body,
                                 headers={'accept': 'application/json'})
        try:
            return response.json()
        except ValueError:
            pass

def send_mail(filepath, subject):
    basename = os.path.basename(filepath)
    address = "logistick.dv@yandex.ru"

    # Compose attachment
    part = MIMEBase('application', "octet-stream")
    part.set_payload(open(filepath, "rb").read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', 'attachment; filename="%s"' % basename)

    # Compose message
    msg = MIMEMultipart()
    msg['From'] = address
    msg['To'] = address
    msg['Subject'] = subject
    msg.attach(part)

    # Send mail
    smtp = SMTP_SSL('smtp.yandex.ru')
    smtp.connect('smtp.yandex.ru')
    smtp.login(address, 'rwlefgatbfpewlmt')
    smtp.sendmail(address, address, msg.as_string())
    smtp.quit()

@app_svh.route('/TEST', methods=['GET', 'POST'])
def test(df):
    logger.warning('start')
    # create an output stream
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='xlsxwriter')
    # taken from the original question
    df.to_excel(writer)
    # the writer has done its job
    writer.close()
    # go back to the beginning of the stream
    output.seek(0)
    # finally return the file
    logger.warning(output)
    return send_file(output, download_name="testing.xlsx", as_attachment=True)

@app_svh.route('/')
def fechone_plomb():
    con = sl.connect('BAZA.db')
    con.row_factory = sl.Row
    try:
        with con:
            data = pd.read_sql("Select DISTINCT ID, party_numb from baza", con)
            data = data.sort_values(by='ID', ascending=False)
            parties = data.drop_duplicates(subset='party_numb')['party_numb']
            parties = parties.fillna(value=np.nan)
        con.close()
    except Exception as e:
        logger.warning(e)
    return render_template('index.html', parties=parties)

@app_svh.route('/info', methods=['GET', 'POST'])
def object_info():
    numb = request.form['numb']
    con = sl.connect('BAZA.db')
    with con:
        df = pd.read_sql(f"Select * from baza where party_numb = '{numb}'", con)
        df['decision_date'] = df['decision_date'].str.slice(0, 17)
        df = df.drop_duplicates(subset='parcel_plomb_numb')
        df = df.rename(columns=map_eng_to_rus)
        df = df[['Партия', 'Пломба', 'Паллет', 'Зона']]
        object_name = 'Партия'
        index = False
        if df.empty:
            df = pd.read_sql(f"Select * from baza where registration_numb = '{numb}'", con)
            df['decision_date'] = df['decision_date'].str.slice(0, 17)
            df = df.rename(columns=map_eng_to_rus)
            df['Товары'] = df['Товары'].str.slice(0, 50)
            object_name = 'Реестр ПТДЭГ'
            index = False
            if df.empty:
                df = pd.read_sql(f"Select * from baza where pallet = '{numb}'", con)
                df['decision_date'] = df['decision_date'].str.slice(0, 17)
                df = df.drop_duplicates(subset='parcel_plomb_numb')
                df = df.rename(columns=map_eng_to_rus)
                df['№'] = np.arange(len(df))[::+1] + 1
                df = df[['№', 'Пломба', 'Зона']]
                object_name = 'Паллет'
                index = False
                if df.empty:
                    df = pd.read_sql(f"Select * from baza where parcel_plomb_numb = '{numb}'", con)
                    df['decision_date'] = df['decision_date'].str.slice(0, 17)
                    df = df.rename(columns=map_eng_to_rus)
                    df['Товары'] = df['Товары'].str.slice(0, 50)
                    object_name = 'Пломба'
                    index = False
                    if df.empty:
                        df = pd.read_sql(f"Select * from baza where parcel_numb = '{numb}'", con)
                        df['decision_date'] = df['decision_date'].str.slice(0, 17)
                        df = df.rename(columns=map_eng_to_rus)
                        df = df.transpose()
                        index = True
                        df = df.rename(columns={0: ''})
                        object_name = 'Посылка'
    # create an output stream
    return render_template('info.html', tables=[style + df.to_html(classes='mystyle', index=index,
                                                                              float_format='{:2,.2f}'.format)],
                           titles=['Информация'],
                           numb=numb, object_name=object_name)

@app_svh.route('/party_info/<string:row>')
def party_info(row):
    con = sl.connect('BAZA.db')
    data = pd.read_sql(f"Select * from baza where party_numb = '{row}'", con)
    a = data.drop_duplicates(subset='parcel_plomb_numb')
    logger.warning(a)
    quonty_plomb = len(a)
    quonty_parcels = len(data)
    b = data.loc[data['custom_status_short'] == 'ИЗЪЯТИЕ']
    quonty_parcels_refuse = len(b)
    quonty_plomb_refuse = len(b.drop_duplicates(subset='parcel_plomb_numb'))
    return render_template('party_info.html', row=row,
                           quonty_plomb=quonty_plomb,
                           quonty_plomb_refuse=quonty_plomb_refuse,
                           quonty_parcels=quonty_parcels,
                           quonty_parcels_refuse=quonty_parcels_refuse)


@app_svh.route('/add/load_tracks', methods=['GET', 'POST'])
def load_tracks():
    if request.method == 'POST':
        uploaded_file = request.files['file']
        filename = uploaded_file.filename
        logger.warning(filename)
        if filename != '':
            file_ext = os.path.splitext(filename)[1]
            if file_ext not in app_svh.config['UPLOAD_EXTENSIONS']:
                abort(400)
            uploaded_file.save(uploaded_file.filename)
            df_track = pd.read_excel(filename, sheet_name=0, engine='openpyxl', usecols='A, AO')
            df_track = df_track[['Номер отправления ИМ', 'Номер накладной СДЭК']]
            df_track = df_track.rename(columns={'Номер отправления ИМ': 'parcel_numb',
                                    'Номер накладной СДЭК': 'track_numb'})
            if df_track['track_numb'].str.contains('#н/д').any() or df_track['track_numb'].isnull().any():
                flash(f'Ошибка загрузки треков: В колонке Треков есть пустые значения или #н/д, поправьте и загрузите заново', category='error')
            else:
                con_track = sl.connect('TRACKS.db')
                with con_track:
                    data = con_track.execute("select count(*) from sqlite_master where type='table' and name='tracks'")
                    for row in data:
                        # если таких таблиц нет
                        if row[0] == 0:
                            # создаём таблицу
                            with con_track:
                                con_track.execute("""
                                                        CREATE TABLE tracks (
                                                        ID INTEGER PRIMARY KEY AUTOINCREMENT,
                                                        parcel_numb VARCHAR(25) NOT NULL UNIQUE ON CONFLICT REPLACE,
                                                        track_numb VARCHAR(25)
                                                        );
                                                        """)
                    df_track.to_sql('tracks', con=con_track, if_exists='append', index=False)
                    con_track.commit()
                flash(f'Треки загружены')

    return render_template('add_manifest_sample.html')


@app_svh.route('/add/load_sample_manifest', methods=['GET', 'POST'])
def load_sample_manifest():
    if request.method == 'POST':
        uploaded_file = request.files['file']
        filename = uploaded_file.filename
        logger.warning(filename)
        if filename != '':
            file_ext = os.path.splitext(filename)[1]
            if file_ext not in app_svh.config['UPLOAD_EXTENSIONS']:
                abort(400)
            uploaded_file.save(uploaded_file.filename)
            df = pd.read_excel(filename, sheet_name=0, engine='openpyxl')
            df = df[['Номер отправления ИМ', 'Номер пломбы', 'Наименование товара', '№ AWB', 'Вес брутто (Вес позиции)']]
            df = df.rename(columns={'Номер отправления ИМ': 'parcel_numb',
                                   'Номер пломбы': 'parcel_plomb_numb',
                                   'Наименование товара': 'goods',
                                   '№ AWB': 'party_numb', 'Вес брутто (Вес позиции)': 'parcel_weight'})
            logger.warning(df)
            df_group = df.groupby('parcel_numb', sort=False)['goods'].agg(','.join)
            logger.warning(df_group)
            df = pd.merge(df, df_group, how='left', left_on='parcel_numb', right_on='parcel_numb')
            df = df.drop_duplicates(subset=['parcel_numb'], keep='first')
            df['goods'] = df['goods_y']
            df = df[['parcel_numb', 'parcel_plomb_numb', 'goods', 'party_numb', 'parcel_weight']]
            df['custom_status'] = 'Unknown'
            df['custom_status_short'] = 'ИЗЪЯТИЕ'
            con = sl.connect('BAZA.db')
            # открываем базу

            with con:
                baza = con.execute("select count(*) from sqlite_master where type='table' and name='baza'")
                for row in baza:
                    # если таких таблиц нет
                    if row[0] == 0:
                        # создаём таблицу
                        with con:
                            con.execute("""
                                                    CREATE TABLE baza (
                                                    ID INTEGER PRIMARY KEY AUTOINCREMENT,
                                                    registration_numb VARCHAR(25),
                                                    party_numb VARCHAR(20),
                                                    parcel_numb VARCHAR(20) NOT NULL UNIQUE ON CONFLICT REPLACE,
                                                    parcel_plomb_numb VARCHAR(20),
                                                    parcel_weight FLOAT,
                                                    custom_status VARCHAR(400),
                                                    custom_status_short VARCHAR(8),
                                                    decision_date VARCHAR(20),
                                                    refuse_reason VARCHAR(400),
                                                    pallet VARCHAR(10),
                                                    zone VARCHAR(5),
                                                    VH_status VARCHAR(10),
                                                    goods
                                                    );
                                                """)
                df_to_append = pd.DataFrame()
                party_numb = df['party_numb'].values[0]
                party_numb_isalready_in = pd.read_sql(f"Select party_numb from baza where party_numb = '{party_numb}'",
                                                      con)
                if party_numb_isalready_in.empty:
                    df.to_sql('baza', con=con, if_exists='append', index=False)
                else:
                    for parcel_numb in df['parcel_numb']:
                        row_isalready_in = pd.read_sql(f"Select * from baza where parcel_numb = '{parcel_numb}'", con)
                        row = df.loc[df['parcel_numb'] == parcel_numb]
                        if row_isalready_in.empty:
                            df_to_append = df_to_append.append(row)
                            logger.warning(row)
                        else:
                            goods = df.loc[df['parcel_numb'] == parcel_numb]['goods'].values[0]
                            logger.warning(goods)
                            party_numb_in_base = row_isalready_in['party_numb'].values[0]
                            if party_numb_in_base is None:
                                party_numb = df.loc[df['parcel_numb'] == parcel_numb]['party_numb'].values[0]
                                parcel_plomb_numb = df.loc[df['parcel_numb'] == parcel_numb]['parcel_plomb_numb'].values[0]
                                parcel_weight = df.loc[df['parcel_numb'] == parcel_numb]['parcel_weight'].values[0]
                                con.execute("Update baza set goods = ?, "
                                            "party_numb = ?, "
                                            "parcel_plomb_numb = ?,"
                                            "parcel_weight = ?"
                                            "where parcel_numb = ?",
                                            (goods, party_numb, parcel_plomb_numb, parcel_weight, parcel_numb))
                            else:
                                con.execute("Update baza set party_numb = ?, goods = ? where parcel_numb = ?",
                                            (party_numb, goods, parcel_numb))
                    logger.warning(df_to_append)
                    df_to_append.to_sql('baza', con=con, if_exists='append', index=False)
                flash(f'Шаблон загружен')
    return render_template('add_manifest_sample.html')

@app_svh.route('/add/decisions', methods=['GET', 'POST'])
def load_decisions():
    logger.warning('OK')
    if request.method == 'POST':
        uploaded_file = request.files['file']
        filename = uploaded_file.filename
        logger.warning(filename)
        if filename != '':
            file_ext = os.path.splitext(filename)[1]
            if file_ext not in app_svh.config['UPLOAD_EXTENSIONS']:
                abort(400)
            uploaded_file.save(uploaded_file.filename)
            df = pd.read_excel(filename, sheet_name=0, engine='openpyxl')
            df['Вес брутто'] = df['Вес брутто'].replace(',', '.', regex=True).astype(float)
            df['Статус_ТО'] = df['Статус ТО']
            df['Статус_ТО'] = df['Статус_ТО'].replace(to_replace='10-в',
                                                      value='В', regex=True)
            df['Статус_ТО'] = df['Статус_ТО'].replace(to_replace='32-в',
                                                      value='В', regex=True)
            df['Статус_ТО'] = df['Статус_ТО'].replace(to_replace='Выпуск товаров без уплаты таможенных платежей', value='ВЫПУСК', regex=True)
            df['Статус_ТО'] = df['Статус_ТО'].replace(to_replace='Выпуск товаров разрешен, таможенные платежи уплачены', value='ВЫПУСК', regex=True)
            for cel in df['Статус_ТО']:
                if cel != 'ВЫПУСК':
                    df.loc[df['Статус_ТО'] == cel, 'Статус_ТО'] = 'ИЗЪЯТИЕ'
            df['Пломба'] = df['Пломба'].astype(str)

            df = df.rename(columns={'Рег. номер': 'registration_numb', 'Общая накладная': 'party_numb',
                                    'Трек-номер': 'parcel_numb', 'Пломба': 'parcel_plomb_numb',
                                    'Вес брутто': 'parcel_weight',
                                    'Статус ТО': 'custom_status', 'Статус_ТО': 'custom_status_short',
                                    'Дата решения': 'decision_date',
                                    'Причина отказа ТО': 'refuse_reason'})
            con = sl.connect('BAZA.db')
            # открываем базу
            with con:
                baza = con.execute("select count(*) from sqlite_master where type='table' and name='baza'")
                for row in baza:
                    # если таких таблиц нет
                    if row[0] == 0:
                        # создаём таблицу
                        with con:
                            con.execute("""
                                        CREATE TABLE baza (
                                        ID INTEGER PRIMARY KEY AUTOINCREMENT,
                                        registration_numb VARCHAR(25) NOT NULL,
                                        party_numb VARCHAR(20),
                                        parcel_numb VARCHAR(20) NOT NULL UNIQUE ON CONFLICT REPLACE,
                                        parcel_plomb_numb VARCHAR(20),
                                        parcel_weight FLOAT,
                                        custom_status VARCHAR(400),
                                        custom_status_short VARCHAR(8),
                                        decision_date VARCHAR(20),
                                        refuse_reason VARCHAR(400),
                                        pallet VARCHAR(10),
                                        zone VARCHAR(5),
                                        VH_status VARCHAR(10),
                                        goods
                                        );
                                    """)
                df_to_append = pd.DataFrame()
                for parcel_numb in df['parcel_numb']:
                    row_isalready_in = pd.read_sql(f"Select * from baza where parcel_numb = '{parcel_numb}'", con)
                    row = df.loc[df['parcel_numb'] == parcel_numb]
                    if row_isalready_in.empty:
                        # row.to_sql('baza', con=con, if_exists='append', index=False)
                        df_to_append = df_to_append.append(row)
                        logger.warning(row)
                    else:
                        logger.warning(row)
                        custom_status_short = df.loc[df['parcel_numb'] == parcel_numb]['custom_status_short'].values[0]
                        logger.warning(custom_status_short)
                        custom_status = df.loc[df['parcel_numb'] == parcel_numb]['custom_status'].values[0]
                        logger.warning(custom_status)
                        registration_numb = df.loc[df['parcel_numb'] == parcel_numb]['registration_numb'].values[0]
                        logger.warning(registration_numb)
                        parcel_weight = df.loc[df['parcel_numb'] == parcel_numb]['parcel_weight'].values[0]
                        decision_date = df.loc[df['parcel_numb'] == parcel_numb]['decision_date'].values[0]
                        refuse_reason = df.loc[df['parcel_numb'] == parcel_numb]['refuse_reason'].values[0]

                        con.execute(f"Update baza set registration_numb = '{registration_numb}',"
                                    f" custom_status = '{custom_status}',"
                                    f" custom_status_short = '{custom_status_short}',"
                                    f" parcel_weight = '{parcel_weight}',"
                                    f" decision_date = '{decision_date}',"
                                    f" refuse_reason = '{refuse_reason}' where parcel_numb = '{parcel_numb}'")
                        row_isalready_in = pd.read_sql(f"Select * from baza where parcel_numb = '{parcel_numb}'", con)
                        logger.warning(row_isalready_in)
                    # df_isalready_in = df_isalready_in.append(row_isalready_in)
                logger.warning(df_to_append)
                df_to_append.to_sql('baza', con=con, if_exists='append', index=False)
                flash(f'Решения загружены')
    return render_template('add_decisions.html')

@app_svh.route('/plomb', methods=['GET'])
@jwt_required()
def plomb():
    return render_template('plomb.html')

@app_svh.route('/pallet', methods=['GET'])
def pallet():
    return render_template('pallet.html')

@app_svh.route('/search/plomb', methods=['GET'])
@jwt_required()
def plomb_searh():
    try:
        parcel_plomb_numb = request.args.get('parcel_plomb_numb')
        return render_template('plomb_search.html', search=parcel_plomb_numb)
    except Exception as e:
        return {'message': str(e)}, 400

@app_svh.route('/search/parcel', methods=['GET'])
def parcel_searh():
    try:
        parcel_numb = request.args.get('parcel_numb')
        return render_template('parcel_search.html', search=parcel_numb)
    except Exception as e:
        return {'message': str(e)}, 400

@app_svh.route('/search/parcel_goods_info', methods=['GET'])
def parcel_goods_info():
    try:
        parcel_numb = request.args.get('parcel_numb')
        return render_template('parcel_goods.html', search=parcel_numb)
    except Exception as e:
        return {'message': str(e)}, 400

@app_svh.route('/search/plomb_info', methods=['POST', 'GET'])
def get_plomb_info():
    parcel_plomb_numb = request.form['parcel_plomb_numb']
    title_status = ''
    try:
        con = sl.connect('BAZA.db')
        with con:
            df_search_plomb = pd.read_sql(f"SELECT * FROM baza where parcel_plomb_numb = '{parcel_plomb_numb}'", con)
            df_search_plomb['№'] = np.arange(len(df_search_plomb))[::+1] + 1
            try:
                df_search_plomb['parcel_weight'] = df_search_plomb['parcel_weight'].round(3)
            except:
                pass
            df_search_plomb = df_search_plomb[['№', 'parcel_numb', 'custom_status_short', 'VH_status', 'parcel_weight', 'goods']]
            df_search_plomb = df_search_plomb.rename(columns={'parcel_numb': 'Трек-номер',
                                                              'custom_status_short': 'Статус',
                                                              'parcel_weight': 'Вес',
                                                              'VH_status': 'ВХ', 'goods': 'Товары'})
            df_search_plomb['Товары'] = df_search_plomb['Товары'].str.slice(0, 50)
            df_parc_quont = len(df_search_plomb)
            df_parc_refuse_quont = len(df_search_plomb.loc[df_search_plomb['Статус'] == 'ИЗЪЯТИЕ'])
            if df_search_plomb.empty:
                flash(f'Пломба не найдена!', category='error')
            elif df_search_plomb['Статус'].str.contains("ИЗЪЯТИЕ").any():
                flash(f'Открываем место!', category='error')
                title_status = 'ИЗЪЯТИЕ'
            else:
                title_status = 'ВЫПУСК'
            df_search_plomb.fillna("", inplace=True)
            def highlight_last_row(df_search_plomb):
                    return ['background-color: #FF0000' if 'ИЗЪЯТИЕ' in str(i) else '' for i in df_search_plomb]
            def highlight_last_row_2(df_search_plomb):
                    return ['background-color: #7CFC00' if 'На ВХ' in str(i) else '' for i in df_search_plomb]
            if 'На ВХ' in df_search_plomb['ВХ'].values:
                df_search_plomb = df_search_plomb.style.apply(highlight_last_row_2).hide_index()
                logger.warning('VH')
            else:
                df_search_plomb = df_search_plomb.style.apply(highlight_last_row).hide_index()
            df_search_plomb = df_search_plomb.format(precision=2)
        object_name = parcel_plomb_numb
        logger.warning(object_name)
        comment = 'Отбор пломб: Просмотрена пломба'
        insert_user_action(object_name, comment)
    except Exception as e:
        return {'message': str(e)}, 400


    content = render_template('plomb_info.html', tables=[style + df_search_plomb.to_html(index=False,
                                                                              float_format='{:2,.2f}'.format)],
                           titles=['Информация о посылке'],
                           parcel_plomb_numb=parcel_plomb_numb, df_parc_quont=df_parc_quont,
                           df_parc_refuse_quont=df_parc_refuse_quont, title_status=title_status)
    resp = make_response(content)
    resp.headers['csrf_token'] = 'kjllljlj'
    return resp

done_parcels = pd.DataFrame()

@app_svh.route('/search/parcel_info', methods=['POST', 'GET'])
def get_parcel_info():
    global done_parcels
    global done_parcel
    parcel_numb = request.form['parcel_numb']
    try:
        con = sl.connect('BAZA.db')
        with con:
            df_parc_events = pd.read_sql(f"SELECT * FROM baza where parcel_numb = '{parcel_numb}'", con)
            parcel_plomb_numb = df_parc_events['parcel_plomb_numb'].values[0]
            df_parcel_plomb_info = pd.read_sql(f"SELECT * FROM baza where parcel_plomb_numb = '{parcel_plomb_numb}'", con)
            df_parc_quont = len(df_parcel_plomb_info)
            df_parcel_plomb_refuse_info = df_parcel_plomb_info.loc[df_parcel_plomb_info['custom_status_short'] == 'ИЗЪЯТИЕ']
            df_parc_refuse_quont = len(df_parcel_plomb_refuse_info)
            df_parcel_plomb_refuse_info['№1'] = np.arange(len(df_parcel_plomb_refuse_info))[::+1]+1
            df_parcel_plomb_refuse_info = df_parcel_plomb_refuse_info[['№1', 'parcel_numb', 'custom_status_short', 'parcel_plomb_numb',
                                                                       'VH_status', 'parcel_weight', 'goods']]
            df_parcel_plomb_refuse_info = df_parcel_plomb_refuse_info.rename(columns={'№1': '№', 'parcel_numb': 'Трек-номер', 'custom_status_short': 'Статус',
                         'parcel_plomb_numb': 'Пломба', 'VH_status': 'ВХ', 'parcel_weight': 'вес', 'goods': 'Товары'})
            df_parcel_plomb_refuse_info['Товары'] = df_parcel_plomb_refuse_info['Товары'].str.slice(0, 200)
            if parcel_plomb_numb == '':
                df_parcel_plomb_refuse_info = pd.DataFrame()
            df_parc_events['№'] = np.arange(len(df_parc_events))[::+1]+1
            df_parc_events = df_parc_events[['parcel_numb', 'custom_status_short', 'parcel_plomb_numb', 'VH_status']]
            df_parc_events = df_parc_events.rename(
                columns={'parcel_numb': 'Трек-номер', 'custom_status_short': 'Статус',
                         'parcel_plomb_numb': 'Пломба', 'VH_status': 'ВХ'})
            status = df_parc_events['Статус'][0]
            if df_parc_events.empty:
                flash(f'Посылка не найдена!', category='error')
            elif 'На ВХ' in df_parc_events['ВХ'].values:
                flash(f'Уже размещено', category='error')
            elif df_parc_events.loc[df_parc_events['Статус'] == 'ИЗЪЯТИЕ'].empty:
                pass
            else:
                flash(f'ИЗЪЯТИЕ на склад', category='error')
            done_parcels = done_parcels.append(df_parc_events).drop_duplicates(subset=['Трек-номер'], keep='first')
            done_parcels['№'] = np.arange(len(done_parcels))[::+1] + 1
            done_parcels_plomb_info = done_parcels.loc[done_parcels['Пломба'] == parcel_plomb_numb]
            if len(done_parcels_plomb_info) == df_parc_quont:
                flash(f'Место завершено', category='success')
            done_parcels_styl = done_parcels.reset_index()
            done_parcels_styl = done_parcels_styl.drop('index', axis=1)
            done_parcels_styl = done_parcels_styl[
                ['№', 'Трек-номер', 'Статус', 'Пломба', 'ВХ']].drop_duplicates(subset=['Трек-номер'], keep='first')
            trigger_color1 = df_parc_events['Статус']
            trigger_color2 = df_parc_events['ВХ']
            done_parcels_styl.fillna("", inplace=True)
            df_parc_events.fillna("", inplace=True)
            df_parcel_plomb_refuse_info.fillna("", inplace=True)
            if len(df_parcel_plomb_refuse_info) == len(done_parcels_styl.loc[((done_parcels_styl.Статус == 'ИЗЪЯТИЕ') & (done_parcels_styl.Пломба == parcel_plomb_numb))]):
                flash(f'Все отказы найдены!', category='success')

            def highlight_RED(df_parc_events):
                return ['background-color: #FF0000' if 'ИЗЪЯТИЕ' in str(df_parc_events) else '']# for i in df_parc_events
            def highlight_GREEN(df_parc_events):
                return ['background-color: #7CFC00' if 'На ВХ' in str(df_parc_events) else '']# for i in df_parc_events
            def highlight_last_row_2(done_parcels_styl):
                    return ['background-color: #7CFC00' if 'На ВХ' in str(i) else '' for i in done_parcels_styl]
            if 'На ВХ' in done_parcels_styl['ВХ'].values:
                done_parcels_styl = done_parcels_styl.style.apply(highlight_last_row_2).hide_index()
            if 'На ВХ' in trigger_color2.values:
                df_parc_events = df_parc_events.style.apply(highlight_GREEN).hide_index()
            elif 'ИЗЪЯТИЕ' in trigger_color1.values:
                df_parc_events = df_parc_events.style.apply(highlight_RED).hide_index()
            else:
                df_parc_events = df_parc_events.style.apply(highlight_GREEN).hide_index()

        object_name = parcel_numb
        comment = 'Отбор посылок: Просмотрена посылка'
        insert_user_action(object_name, comment)
    except Exception as e:
        flash(f'Посылка не найдена!', category='error')
        return render_template('parcel_info.html')
        #return {'message': str(f'{e}  "PARCEL NOT FOUND"')}, 400
        pass
    return render_template('parcel_info.html', tables=[df_parc_events.to_html(classes='mystyle', index=False,
                                                                              float_format='{:2,.2f}'.format),
                                                       df_parcel_plomb_refuse_info.to_html(classes='mystyle', index=False,
                                                                                 float_format='{:2,.2f}'.format),
                                                       done_parcels_styl.to_html(classes='mystyle', index=False,
                                                                                 float_format='{:2,.2f}'.format)],
                           titles=['na', 'Отбор посылок', 'Нужно найти в мешке:', '\n\nОтработанные:'],
                           parcel_numb=parcel_numb, status=status, df_parc_quont=df_parc_quont,
                           df_parc_refuse_quont=df_parc_refuse_quont, parcel_plomb_numb=parcel_plomb_numb)

@app_svh.route('/search/clean', methods=['POST'])
def clean_working_place():
    global done_parcels
    global done_parcel
    done_parcels_VH = done_parcels.loc[done_parcels['Статус'] == 'ИЗЪЯТИЕ']
    con = sl.connect('BAZA.db')
    with con:
        for parcel_numb in done_parcels_VH['Трек-номер']:
            con.execute(f"Update baza set VH_status = 'На ВХ', parcel_plomb_numb = ''  where parcel_numb = '{parcel_numb}'")
    con.commit()
    con.close()
    parcel_plomb_numb = done_parcels['Пломба'].values[0]
    done_parcels = pd.DataFrame()
    done_parcel = None
    object_name = parcel_plomb_numb
    comment = 'Отбор посылок: Завершено место'
    insert_user_action(object_name, comment)
    return render_template('parcel_info.html')

@app_svh.route('/search/manifest', methods=['GET'])
def make_manifest():
    try:
        global parcel_plomb_numb
        parcel_plomb_numb = request.args.get('parcel_plomb_numb')
        object_name = parcel_plomb_numb
        comment = 'Манифест огрузки: Пломба добавленна для отгрузки'
        insert_user_action(object_name, comment)
        return render_template('plomb_search_manifest.html', search=parcel_plomb_numb)
    except Exception as e:
        return {'message': str(e)}, 400

df_plomb_to_manifest = pd.DataFrame()
df_plomb_to_manifest_total = pd.DataFrame()
df_to_manifest_HTML = pd.DataFrame()
df_refuse_plombs = pd.DataFrame()
@app_svh.route('/search/plomb_to_manifest', methods=['POST', 'GET'])
def plomb_to_manifest():
    global df_plomb_to_manifest
    global df_plomb_to_manifest_total
    global parcel_plomb_numb
    global df_to_manifest_HTML
    global df_refuse_plombs
    parcel_plomb_numb = request.form['parcel_plomb_numb']
    if df_plomb_to_manifest.empty:
        df_plomb_to_manifest = pd.DataFrame({'parcel_plomb_numb': [parcel_plomb_numb]})
    try:
        con = sl.connect('BAZA.db')
        with con:
            df_plomb_to_manifest = pd.read_sql(f"SELECT * FROM baza where parcel_plomb_numb = '{parcel_plomb_numb}'",
                                               con)
            logger.warning(df_plomb_to_manifest['custom_status_short'])
            # if that is parcel:
            if df_plomb_to_manifest.empty:
                df_plomb_to_manifest = pd.read_sql(f"SELECT * FROM baza where parcel_numb = '{parcel_plomb_numb}'", con)
                df_all_pallet_plombs = df_plomb_to_manifest.drop_duplicates(subset=['parcel_plomb_numb', 'parcel_numb'],
                                                                                                                     keep='first')
                df_all_pallet_plombs['Тип'] = "Посылка-место"
                df_plomb_to_manifest_refuse_parcel = df_all_pallet_plombs['parcel_numb'].values[0]
                if not df_plomb_to_manifest.loc[df_plomb_to_manifest['custom_status_short'] == 'ИЗЪЯТИЕ'].empty:
                    flash(f'ПОСЫЛКА {df_plomb_to_manifest_refuse_parcel} СО СТАТУСОМ ИЗЪЯТИЕ (НЕ ВЫПУЩЕНА!!!)', category='error')
            # if that is plomb:
            else:
                df_plomb_to_manifest = df_plomb_to_manifest.drop_duplicates(subset='parcel_plomb_numb', keep='first')
                df_plomb_to_manifest['quont_plomb'] = None
                df_plomb_to_manifest['Тип'] = "Пломба"
                df_plomb_to_manifest['parcel_numb'] = ''
                pallet = df_plomb_to_manifest['pallet'].values[0]
                df_all_pallet_plombs = pd.read_sql(f"SELECT * FROM baza where pallet = '{pallet}'", con)
                df_all_pallet_plombs_refuse = df_all_pallet_plombs.loc[df_all_pallet_plombs['custom_status_short'] == 'ИЗЪЯТИЕ']
                if not df_all_pallet_plombs_refuse.empty:
                    df_all_pallet_plombs_refuse.drop_duplicates(subset='parcel_plomb_numb', keep='first')
                    logger.warning(df_all_pallet_plombs_refuse)
                    df_all_pallet_plombs_refuse = df_all_pallet_plombs_refuse['parcel_plomb_numb'].to_string(index=False)
                    flash(f'НА ПАЛЛЕТЕ {pallet} ЕСТЬ ОТКАЗНЫЕ (НЕ ОТРАБОТАННЫЕ) ПЛОМБЫ: {df_all_pallet_plombs_refuse}', category='error')
                df_all_pallet_plombs = df_all_pallet_plombs.drop_duplicates(subset='parcel_plomb_numb', keep='first')
                df_all_pallet_plombs['Тип'] = "Пломба"
                df_all_pallet_plombs['parcel_numb'] = ''
            df_all_pallet_plombs['quont_plomb'] = len(df_all_pallet_plombs)
            if df_all_pallet_plombs.empty:
                df_plomb_to_manifest_total = df_plomb_to_manifest_total.append(df_plomb_to_manifest)
                object_name = parcel_plomb_numb
                comment = 'Манифест огрузки: Пломба добавленна для отгрузки'
                insert_user_action(object_name, comment)
            else:
                df_plomb_to_manifest_total = df_plomb_to_manifest_total.append(df_all_pallet_plombs)
                object_name = pallet
                comment = 'Манифест огрузки: Паллет добавлен для отгрузки'
                insert_user_action(object_name, comment)
            df_plomb_to_manifest_total = df_plomb_to_manifest_total.drop_duplicates(subset=['parcel_plomb_numb', 'parcel_numb'], keep='first')
            df_plomb_to_manifest_total['№1'] = np.arange(len(df_plomb_to_manifest_total))[::+1] + 1

            df_to_manifest_HTML = df_plomb_to_manifest_total[['№1', 'parcel_plomb_numb', 'pallet', 'party_numb', 'quont_plomb', 'Тип', 'parcel_numb']]

            df_to_manifest_HTML = df_to_manifest_HTML.rename(columns={'№1': '№', 'parcel_plomb_numb': 'Пломба',
                                                                      'pallet': 'Паллет', 'party_numb': 'Партия',
                                                                      'quont_plomb': 'кол. пломб', 'parcel_numb': 'Трек'})
            df_to_manifest_HTML.fillna("", inplace=True)

    except Exception as e:
        flash(f'Пломба не найдена!', category='error')
        #return {'message': str(e)}, 400
    return render_template('plomb_info_manifest.html', tables=[style + df_to_manifest_HTML.to_html(classes='mystyle', index=False)],
                           titles=['n/a', 'Добавленные'],
                           parcel_plomb_numb=parcel_plomb_numb)

@app_svh.route('/search/clean_working_place_manifest', methods=['POST'])
def clean_working_place_manifest():
    global df_plomb_to_manifest_total
    global df_to_manifest_HTML
    object_name = None
    comment = 'Манифест огрузки: Таблица отгрузки очищена'
    insert_user_action(object_name, comment)
    df_plomb_to_manifest_total = pd.DataFrame()
    df_plomb_to_manifest_total = pd.DataFrame()
    return render_template('plomb_info_manifest.html')

def manifest_to_xls(df_manifest_total):
    now = datetime.datetime.now().strftime("%d.%m.%Y")
    now_time = datetime.datetime.now().strftime("%d.%m.%Y %H-%M")
    df = df_manifest_total
    delta = datetime.timedelta(hours=-10, minutes=0)
    event_date = datetime.datetime.now() + delta
    event_date = event_date.strftime("%Y-%m-%d %H:%M:%S")
    #insert_event_API_test(df, event_date)
    party_numb = df_manifest_total['party_numb'].values[0]
    df_manifest_total['№ п.п.'] = np.arange(len(df_manifest_total))[::+1] + 1
    df_manifest_total['Номер индивидуальной     накладной'] = df_manifest_total['parcel_numb']
    df_manifest_total['Трекинг'] = df_manifest_total['parcel_numb']
    df_manifest_total['Номер накладной'] = df_manifest_total['party_numb']
    df_manifest_total['Вес посылки'] = df_manifest_total['parcel_weight']
    quont_of_plomb = len(df_manifest_total.drop_duplicates(subset='parcel_plomb_numb'))
    group_Weight_df = df_manifest_total.groupby('parcel_plomb_numb')['parcel_weight'].sum()
    group_Weight_df = group_Weight_df.rename('Вес_мешка', inplace=True)
    df_total = pd.merge(df_manifest_total, group_Weight_df, how='left', left_on='parcel_plomb_numb',
                        right_on='parcel_plomb_numb')

    df_total['Мест'] = df_total.Вес_мешка.eq(df_total.Вес_мешка.shift()).astype('str')
    df_total['Вес_мешка'] = np.where(df_total['Мест'] == 'True', '0', df_total['Вес_мешка'])

    df_total['Мест'] = df_total.parcel_plomb_numb.eq(df_total.parcel_plomb_numb.shift()).astype('str')
    df_total['parcel_plomb_numb'] = np.where(df_total['Мест'] == 'True', '', df_total['parcel_plomb_numb'])

    df_total['Вес мешка'] = df_total['Вес_мешка'].astype(float)

    df_total = df_total.reindex(columns=['№ п.п.', 'Номер индивидуальной     накладной',
                                         'Трекинг', 'Номер накладной', 'Вес посылки',
                                         'parcel_plomb_numb', 'Вес мешка'])

    df_total = df_total.drop_duplicates(subset='Трекинг', keep='first')



    writer = pd.ExcelWriter('system.xlsx', engine='xlsxwriter')
    df_total.to_excel(writer, sheet_name='Sheet1', index=False)
    for column in df_total:
        column_width = max(df_total[column].astype(str).map(len).max(), len(column))
        col_idx = df_total.columns.get_loc(column)
        writer.sheets['Sheet1'].set_column(col_idx, col_idx, column_width)
        writer.sheets['Sheet1'].set_column(0, 3, 10)
        writer.sheets['Sheet1'].set_column(1, 3, 20)
        writer.sheets['Sheet1'].set_column(2, 3, 20)
        writer.sheets['Sheet1'].set_column(3, 3, 20)
        writer.sheets['Sheet1'].set_column(4, 3, 30)
        writer.sheets['Sheet1'].set_column(5, 3, 20)

    writer.save()
    wb2 = openpyxl.load_workbook('system.xlsx')
    ws2 = wb2.active
    ws2.insert_rows(1, 2)

    wb = openpyxl.load_workbook('Akt.xlsx')
    ws = wb.active
    ws['B1'].value = f'{now} а/м М246ВР пломба № 52388385'

    head = "A1:G3"  # Заголовок таблицы, в котором есть объединенные ячейки
    for _range in ws.merged_cells.ranges:
        boundaries = range_boundaries(str(_range))
        ws2.merge_cells(start_column=boundaries[0], start_row=boundaries[1],
                        end_column=boundaries[2], end_row=boundaries[3])
    for row_number, row in enumerate(ws[head]):
        for col_number, cell in enumerate(row):
            ws2.cell(row_number + 1, col_number + 1, cell.value)
            if cell.has_style:
                ws2.cell(row_number + 1, col_number + 1).font = copy(cell.font)
                ws2.cell(row_number + 1, col_number + 1).fill = copy(cell.fill)
                ws2.cell(row_number + 1, col_number + 1).border = copy(cell.border)
                ws2.cell(row_number + 1, col_number + 1).number_format = copy(cell.number_format)
                ws2.cell(row_number + 1, col_number + 1).protection = copy(cell.protection)
                ws2.cell(row_number + 1, col_number + 1).alignment = copy(cell.alignment)
                ws2.cell(row_number + 1, col_number + 1).quotePrefix = copy(cell.quotePrefix)
                ws2.cell(row_number + 1, col_number + 1).pivotButton = copy(cell.pivotButton)

    len_A = len(ws2['A']) + 1
    ws2[f"D{len_A}"] = 'ИТОГО (вес / пломб):'
    ws2[f"E{len_A}"] = ws2[f"E{len_A}"].number_format = '0.000'
    ws2[f"E{len_A}"] = f"=SUM(E4:E{len_A - 1})"
    ws2[f"F{len_A}"] = quont_of_plomb
    Manifest_name = f'{download_folder}Manifest {now_time} ({party_numb}).xlsx'
    Manifest_name_short_name = f'Manifest {now_time} ({party_numb})'
    wb2.save(Manifest_name)

    send_mail(Manifest_name, Manifest_name_short_name)

    df_manifest_for_driver = df_total.drop_duplicates(subset='parcel_plomb_numb')
    writer = pd.ExcelWriter('system.xlsx', engine='xlsxwriter')
    df_manifest_for_driver.to_excel(writer, sheet_name='Sheet1', index=False)
    for column in df_manifest_for_driver:
        column_width = max(df_manifest_for_driver[column].astype(str).map(len).max(), len(column))
        col_idx = df_manifest_for_driver.columns.get_loc(column)
        writer.sheets['Sheet1'].set_column(col_idx, col_idx, column_width)
        writer.sheets['Sheet1'].set_column(0, 3, 10)
        writer.sheets['Sheet1'].set_column(1, 3, 20)
        writer.sheets['Sheet1'].set_column(2, 3, 20)
        writer.sheets['Sheet1'].set_column(3, 3, 20)
        writer.sheets['Sheet1'].set_column(4, 3, 30)
        writer.sheets['Sheet1'].set_column(5, 3, 20)

    writer.save()
    wb2 = openpyxl.load_workbook('system.xlsx')
    ws2 = wb2.active
    ws2.insert_rows(1, 2)

    wb = openpyxl.load_workbook('Akt.xlsx')
    ws = wb.active
    ws['B1'].value = f'{now} а/м М246ВР пломба № 52388385'

    head = "A1:G3"  # Заголовок таблицы, в котором есть объединенные ячейки
    for _range in ws.merged_cells.ranges:
        boundaries = range_boundaries(str(_range))
        ws2.merge_cells(start_column=boundaries[0], start_row=boundaries[1],
                        end_column=boundaries[2], end_row=boundaries[3])
    for row_number, row in enumerate(ws[head]):
        for col_number, cell in enumerate(row):
            ws2.cell(row_number + 1, col_number + 1, cell.value)
            if cell.has_style:
                ws2.cell(row_number + 1, col_number + 1).font = copy(cell.font)
                ws2.cell(row_number + 1, col_number + 1).fill = copy(cell.fill)
                ws2.cell(row_number + 1, col_number + 1).border = copy(cell.border)
                ws2.cell(row_number + 1, col_number + 1).number_format = copy(cell.number_format)
                ws2.cell(row_number + 1, col_number + 1).protection = copy(cell.protection)
                ws2.cell(row_number + 1, col_number + 1).alignment = copy(cell.alignment)
                ws2.cell(row_number + 1, col_number + 1).quotePrefix = copy(cell.quotePrefix)
                ws2.cell(row_number + 1, col_number + 1).pivotButton = copy(cell.pivotButton)

    len_A = len(ws2['A']) + 1
    ws2[f"E{len_A}"] = 'ИТОГО (пломб / вес):'
    ws2[f"G{len_A}"] = ws2[f"G{len_A}"].number_format = '0.000'
    ws2[f"G{len_A}"] = f"=SUM(G4:G{len_A - 1})"
    ws2[f"F{len_A}"] = quont_of_plomb
    ws2[f"D{len_A + 4}"] = 'Груз сдал_________'
    ws2[f"F{len_A + 4}"] = 'Груз принял_________'

    ws2.sheet_properties.pageSetUpPr.fitToPage = True
    ws2.page_setup.fitToHeight = False
    cm = int(1 / 4)
    ws2.page_margins = PageMargins(left=cm, right=cm, top=cm, bottom=cm)
    Manifest_name2 = f'{download_folder}Short Manifest {now_time} ({party_numb}).xlsx'
    Manifest_name_short_name2 = f'Для водителя Manifest {now_time} ({party_numb})'
    wb2.save(Manifest_name2)
    send_mail(Manifest_name2, Manifest_name_short_name2)
    object_name = Manifest_name_short_name
    comment = 'Манифест огрузки: Манифест сформирован'
    insert_user_action(object_name, comment)
    df_plomb_to_manifest = pd.DataFrame()
    df_plomb_to_manifest_total = pd.DataFrame()
    flash(f'Манифест {Manifest_name_short_name} сформирован', category='success')

@app_svh.route('/save_manifest', methods=['POST', 'GET'])
def save_manifest():
    global df_plomb_to_manifest
    global df_plomb_to_manifest_total
    logger.warning(df_plomb_to_manifest_total)
    con = sl.connect('BAZA.db')
    df_manifest_total = pd.DataFrame()
    df_manifest_total_refuses = df_plomb_to_manifest_total.loc[df_plomb_to_manifest_total['custom_status_short'] == 'ИЗЪЯТИЕ']
    df_manifest_total_refuses_parcels = df_manifest_total_refuses.drop_duplicates(subset='parcel_numb')
    df_manifest_total_refuses_parcels = df_manifest_total_refuses_parcels['parcel_numb'].to_list()
    if df_manifest_total_refuses.empty:
        df_parcelc_to_manifest_total = df_plomb_to_manifest_total.loc[df_plomb_to_manifest_total['Тип'] == 'Посылка-место']
        df_plomb_to_manifest_total = df_plomb_to_manifest_total.loc[df_plomb_to_manifest_total['Тип'] == 'Пломба']
        for parcel_plomb_numb in df_plomb_to_manifest_total['parcel_plomb_numb']:
            with con:
                df_manifest = pd.read_sql(f"SELECT * FROM baza where parcel_plomb_numb = '{parcel_plomb_numb}'", con)
                con.execute("Update baza set VH_status = 'ОТГРУЖЕН' where parcel_plomb_numb = ?", (parcel_plomb_numb, ))
                df_manifest_total = df_manifest_total.append(df_manifest)
            logger.warning(df_manifest_total)
        for parcel_numb in df_parcelc_to_manifest_total['parcel_numb']:
            with con:
                df_manifest = pd.read_sql(f"SELECT * FROM baza where parcel_numb = '{parcel_numb}'", con)
                con.execute("Update baza set VH_status = 'ОТГРУЖЕН' where parcel_numb = ?", (parcel_numb, ))
                df_manifest_total = df_manifest_total.append(df_manifest)
            logger.warning(df_manifest_total)
        manifest_to_xls(df_manifest_total)
    else:
        flash(f'{df_manifest_total_refuses_parcels} СО СТАТУСОМ ИЗЪЯТИЕ (НЕ ВЫПУЩЕНА!!!)', category='error')
    return render_template('plomb_info_manifest.html')

@app_svh.route('/party_to_manifest', methods=['POST', 'GET'])
def party_to_manifest():
    con = sl.connect('BAZA.db')
    with con:
        df_parties = pd.read_sql(f"SELECT * FROM baza", con).drop_duplicates(subset='party_numb', keep='first')
        df_parties = df_parties['party_numb'].to_list()
    return render_template('party_to_manifest.html', df_parties=df_parties)


@app_svh.route('/party_to_manifest_button', methods=['POST', 'GET'])
def party_to_manifest_button():
    party_numb = request.form['party_numb']
    con = sl.connect('BAZA.db')
    with con:
        df_manifest_total = pd.read_sql(f"SELECT * FROM baza where party_numb = '{party_numb}' "
                                  f"and custom_status_short = 'ВЫПУСК'", con)
        con.execute("Update baza set VH_status = 'ОТГРУЖЕН' where party_numb = ? AND custom_status_short = ?", (party_numb, 'ВЫПУСК'))
    df_manifest_total = df_manifest_total.sort_values(by='parcel_plomb_numb', ascending=False)
    object_name = party_numb
    comment = 'Манифест огрузки: Партия отгружена'
    insert_user_action(object_name, comment)
    manifest_to_xls(df_manifest_total)
    return render_template('plomb_search_manifest.html')

df_changed_plomb = pd.DataFrame()
@app_svh.route('/search1/old_plomb/add', methods=['GET', 'POST'])
def old_plomb_add():
    now_time = datetime.datetime.now().strftime("%d.%m.%Y %H:%M")
    global df_changed_plomb
    try:
        if request.method == 'POST':
            old_plomb_numb = request.form['old_plomb_numb']
            new_plomb_numb = request.form['new_plomb_numb']
            con = sl.connect('BAZA.db')
            with con:
                df = pd.read_sql(f"SELECT * FROM baza where parcel_plomb_numb = '{old_plomb_numb}' ", con)
                if df.empty:
                    flash(f'Пломба {old_plomb_numb} не найдена!', category='error')
                else:

                    cursor = con.cursor()
                    query = "Update baza set parcel_plomb_numb = ? where parcel_plomb_numb = ?"
                    data = (new_plomb_numb, old_plomb_numb)
                    logger.warning(data)
                    cursor.execute(query, data)
                    con.commit()
                    cursor.close()
                    flash(f'Пломба {old_plomb_numb} обновлена!', category='success')
                    df_changed_plomb_to_append = pd.DataFrame({'старая': [old_plomb_numb], 'новая': [new_plomb_numb]})
                    df_changed_plomb = df_changed_plomb.append(df_changed_plomb_to_append)
                    df_changed_plomb['№'] = np.arange(len(df_changed_plomb))[::+1] + 1
                    df_changed_plomb = df_changed_plomb[['№', 'старая', 'новая']]
                    df_changed_plomb = df_changed_plomb.iloc[::-1]
                    logger_change_plob.info(f'Время: {now_time} Старая пломба: {old_plomb_numb} Новая пломба: {new_plomb_numb}')
                    object_name = old_plomb_numb
                    comment = f'Новая пломба: Изменена на {new_plomb_numb}'
                    insert_user_action(object_name, comment)
        return render_template('old_plomb.html', tables=[style + df_changed_plomb.to_html(classes='mystyle', index=False,
                                                                                        float_format='{:2,.2f}'.format)],
                           titles=['na', 'Изменены:'])

    except Exception as e:
        logger.warning(f'error {e}')
        return {'message': str(e)}, 400

@app_svh.route('/new_place', methods=['POST', 'GET'])
def new_place():
    try:
        parcel_numb = request.args.get('parcel_numb')
        object_name = parcel_numb
        comment = f'Сформировать место: Посылка отмечена для формирования нового места'
        insert_user_action(object_name, comment)
        return render_template('parcel_search_new_place.html', search=parcel_numb)
    except Exception as e:
        return {'message': str(e)}, 400

done_parcels_np = pd.DataFrame()
@app_svh.route('/save_new_place', methods=['POST', 'GET'])
def making_new_place():
    global done_parcels_np
    global parcel_numb_np
    global df_parc_events_np
    global done_parcels_styl_np
    parcel_numb_np = request.form['parcel_numb']
    try:
        con = sl.connect('BAZA.db')
        with con:
            df_parc_events_np = pd.read_sql(f"SELECT * FROM baza where parcel_numb = '{parcel_numb_np}'", con)
            df_parc_events_np['№'] = np.arange(len(df_parc_events_np))[::+1] + 1
            df_parc_events_np = df_parc_events_np[['parcel_numb', 'custom_status_short', 'parcel_plomb_numb', 'VH_status']]
            df_parc_events_np = df_parc_events_np.rename(
                columns={'parcel_numb': 'Трек-номер', 'custom_status_short': 'Статус',
                         'parcel_plomb_numb': 'Пломба', 'VH_status': 'ВХ'})
            status = df_parc_events_np['Статус'][0]
            if df_parc_events_np.empty:
                flash(f'Посылка не найдена!', category='error')
            else:
                pass
            done_parcels_np = done_parcels_np.append(df_parc_events_np).drop_duplicates(subset=['Трек-номер'], keep='first')
            done_parcels_np['№'] = np.arange(len(done_parcels_np))[::+1] + 1
            done_parcels_styl_np = done_parcels_np.reset_index()
            done_parcels_styl_np = done_parcels_styl_np.drop('index', axis=1)
            done_parcels_styl_np = done_parcels_styl_np[
                ['№', 'Трек-номер', 'Статус', 'Пломба', 'ВХ']].drop_duplicates(subset=['Трек-номер'], keep='first')
            trigger_color1 = df_parc_events_np['Статус']
            trigger_color2 = df_parc_events_np['ВХ']
            def highlight_GREEN(df_parc_events):
                return ['background-color: #7CFC00' if 'На ВХ' in str(i) else '' for i in df_parc_events]
            def highlight_RED(df_parc_events_np):
                return ['background-color: #FF0000' if 'ИЗЪЯТИЕ' in str(i) else '' for i in df_parc_events_np]
            if 'На ВХ' in trigger_color2.values:
                df_parc_events_np = df_parc_events_np.style.apply(highlight_GREEN).hide_index()
            elif 'ИЗЪЯТИЕ' in trigger_color1.values:
                df_parc_events_np = df_parc_events_np.style.apply(highlight_RED).hide_index()
            if 'На ВХ' in done_parcels_styl_np['ВХ'].values:
                done_parcels_styl_np = done_parcels_styl_np.style.apply(highlight_GREEN).hide_index()
            elif 'ИЗЪЯТИЕ' in done_parcels_styl_np['Статус'].values:
                done_parcels_styl_np = done_parcels_styl_np.style.apply(highlight_RED).hide_index()
            object_name = parcel_numb_np
            comment = f'Сформировать место: Посылка отмечена для формирования нового места'
            insert_user_action(object_name, comment)
    except Exception as e:
        flash(f'Посылка не найдена!', category='error')
        return render_template('parcel_info_new_place.html')
        #return {'message': str(e)}, 400
        pass
    return render_template('parcel_info_new_place.html', tables=[df_parc_events_np.to_html(classes='mystyle', index=False,
                                                                              float_format='{:2,.2f}'.format),
                                                       done_parcels_styl_np.to_html(classes='mystyle', index=False,
                                                                                 float_format='{:2,.2f}'.format)],
                           titles=['na', 'Посылка:', '\n\nОтработанные:'],
                           parcel_numb=parcel_numb_np)

@app_svh.route('/delete_last_parcel', methods=['POST', 'GET'])
def delete_last_parcel():
    global done_parcels_np
    global df_parc_events_np
    global done_parcels_styl_np
    logger.warning(done_parcels_np)
    done_parcels_np = done_parcels_np[:-1]
    done_parcels_styl_np.data = done_parcels_styl_np.data[:-1]
    logger.warning(done_parcels_np)
    return render_template('parcel_info_new_place.html', tables=[df_parc_events_np.to_html(classes='mystyle', index=False,
                                                                                        float_format='{:2,.2f}'.format),
                                                                 style + done_parcels_styl_np.to_html(classes='mystyle',
                                                                                           index=False,
                                                                                           float_format='{:2,.2f}'.format)],
                           titles=['na', 'Посылка:', '\n\nОтработанные:'],
                           parcel_numb=parcel_numb_np)

@app_svh.route('/save_new_place/place_numb', methods=['POST', 'GET'])
def make_place_numb():
    global done_parcels_np
    plomb_toreplace_new = request.form['plomb_toreplace_new']
    con = sl.connect('BAZA.db')
    # открываем базу
    with con:
        for parcel_numb in done_parcels_np['Трек-номер']:
            #plomb_toreplace = done_parcels_np.loc[done_parcels_np['Трек-номер'] == parcel_numb]['Пломба'].values[0]
            #logger.warning(plomb_toreplace)
            con.execute(f"Update baza set parcel_plomb_numb = '{plomb_toreplace_new}' where parcel_numb = '{parcel_numb}'")
    df_new_place = pd.read_sql(f"Select * from baza where parcel_plomb_numb = '{plomb_toreplace_new}'", con)
    writer = pd.ExcelWriter(f'{addition_folder}Место {plomb_toreplace_new}.xlsx', engine='xlsxwriter')
    df_new_place.to_excel(writer, sheet_name='Sheet1', index=False)
    for column in df_new_place:
        column_width = max(df_new_place[column].astype(str).map(len).max(), len(column))
        col_idx = df_new_place.columns.get_loc(column)
        writer.sheets['Sheet1'].set_column(col_idx, col_idx, column_width)
        writer.sheets['Sheet1'].set_column(0, 3, 10)
        writer.sheets['Sheet1'].set_column(1, 3, 20)
        writer.sheets['Sheet1'].set_column(2, 3, 20)
        writer.sheets['Sheet1'].set_column(3, 3, 20)
        writer.sheets['Sheet1'].set_column(4, 3, 30)
        writer.sheets['Sheet1'].set_column(5, 3, 20)
    object_name = plomb_toreplace_new
    comment = f'Сформировать место: Место сформированно'
    insert_user_action(object_name, comment)
    writer.save()
    con.commit()
    con.close()
    done_parcels_np = pd.DataFrame()
    flash(f'Место сформировано', category='success')
    return render_template('parcel_info_new_place.html')

parcel_plomb_numb_np = None
df_plombs_np = pd.DataFrame()
@app_svh.route('/create_new_pallet', methods=['POST', 'GET'])
def create_pallet():
    global df_plombs_html
    global df_plombs_np
    global parcel_plomb_numb_np
    try:
        parcel_plomb_numb_np = request.form['parcel_plomb_numb_np']
    except:
        pass
    try:
        con = sl.connect('BAZA.db')
        with con:
            df_all_pallets = pd.read_sql(f"SELECT pallet FROM baza", con).drop_duplicates(subset='pallet', keep='first')
            try:
                df_all_pallets['pallet'] = df_all_pallets['pallet'].fillna(value=np.nan).fillna(0).astype(int)
                df_all_pallets = df_all_pallets.sort_values(by='pallet')
                logger.warning(df_all_pallets)

                last_pall_numb = int(df_all_pallets.values[-1].tolist()[0])
                i = last_pall_numb + 1
            except:
                last_pall_numb = df_all_pallets.values[0].tolist()[0]
                i = 1
            logger.warning(last_pall_numb)
            df_plomb = pd.read_sql(f"SELECT * FROM baza where parcel_plomb_numb = '{parcel_plomb_numb_np}'", con).drop_duplicates(subset=['parcel_plomb_numb'], keep='first')
            df_plomb['Тип'] = 'Пломба'
            if df_plomb.empty and parcel_plomb_numb_np != None:
                df_plomb = pd.read_sql(f"SELECT * FROM baza where parcel_numb = '{parcel_plomb_numb_np}'", con)
                df_plomb['Тип'] = "Посылка-место"
                logger.warning(df_plomb)
                if df_plomb.empty and parcel_plomb_numb_np != None:
                    flash(f'Пломба не найдена!', category='error')
                else:
                    pass
            else:
                pass
            try:
                df_plombs_np = df_plombs_np.append(df_plomb)
                df_plombs_np['№1'] = np.arange(len(df_plombs_np))[::+1] + 1
                df_plombs_html = df_plombs_np.rename(columns={
                    '№1': '№',
                    'parcel_plomb_numb': 'Пломба',
                    'parcel_numb': 'Трек',
                    'pallet': '№ Паллет',
                    'zone': 'Зона',
                    'party_numb': 'Партия'
                })
                df_plombs_html = df_plombs_html[['№', 'Пломба', '№ Паллет',
                                               'Зона', 'Партия', 'Трек', 'Тип']]
                df_plombs_html.fillna("", inplace=True)
                df_plombs_html = df_plombs_html.reset_index()
                df_plombs_html = df_plombs_html.drop('index', axis=1)
            except:
                pass
        object_name = parcel_plomb_numb_np
        comment = f'Сформировать новый паллет: пломба отмечена для добавления на паллет'
        insert_user_action(object_name, comment)
    except Exception as e:
        flash(f'Пломба не найдена!', category='error')
        #return render_template('parcel_info_new_place.html')
        return {'message': str(e)}, 400
        pass
    return render_template('New_pallet.html', tables=[style + df_plombs_html.to_html(classes='mystyle', index=False,
                                                                              float_format='{:2,.2f}'.format)],
                           titles=['na', '\n\nМешки\места:'],
                           parcel_plomb_numb_np=parcel_plomb_numb_np, i=i)

@app_svh.route('/delete_last_plomb', methods=['POST', 'GET'])
def delete_last_plomb():
    global parcel_plomb_numb_np
    global df_plombs_np
    global df_plombs_html
    logger.warning(df_plombs_html)
    df_plombs_html = df_plombs_html[:-1]
    df_plombs_np = df_plombs_np[:-1]
    logger.warning(df_plombs_html)
    return render_template('New_pallet.html', tables=[style + df_plombs_html.to_html(classes='mystyle', index=False,
                                                                             float_format='{:2,.2f}'.format)],
                           titles=['na', '\n\nМешки\места:'],
                          parcel_plomb_numb_np=parcel_plomb_numb_np)

@app_svh.route('/delete_last_plomb_addpallet', methods=['POST', 'GET'])
def delete_last_plomb_addpallet():
    global parcel_plomb_numb_np
    global df_plombs_np
    global df_plombs_html
    logger.warning(df_plombs_html)
    df_plombs_html = df_plombs_html[:-1]
    df_plombs_np = df_plombs_np[:-1]
    logger.warning(df_plombs_html)
    return render_template('add_to_pallet.html', tables=[style + df_plombs_html.to_html(classes='mystyle', index=False,
                                                                             float_format='{:2,.2f}'.format)],
                           titles=['na', '\n\nМешки\места:'],
                          parcel_plomb_numb_np=parcel_plomb_numb_np)

@app_svh.route('/search/clean_working_place_clean_working_place_pallet', methods=['POST'])
def clean_working_place_pallet():
    global parcel_plomb_numb_np
    global df_plombs_np
    global df_plombs_html
    df_plombs_html = pd.DataFrame()
    df_plombs_np = pd.DataFrame()
    object_name = None
    comment = f'Сформировать новый паллет: Таблица очищена'
    insert_user_action(object_name, comment)
    return render_template('New_pallet.html')

@app_svh.route('/search/clean_working_place_clean_working_place_addpallet', methods=['POST'])
def clean_working_place_addpallet():
    global parcel_plomb_numb_np
    global df_plombs_np
    global df_plombs_html
    df_plombs_html = pd.DataFrame()
    df_plombs_np = pd.DataFrame()
    object_name = None
    comment = f'Добавить на паллет: Таблица очищена'
    insert_user_action(object_name, comment)
    return render_template('add_to_pallet.html')

@app_svh.route('/insert_new_pallet', methods=['POST', 'GET'])
def insert_pallet():
    global df_plombs_np
    global parcel_plomb_numb_np
    pallet_new = request.form['pallet_new']
    # открываем базу
    con = sl.connect('BAZA.db')
    with con:
        df_plombs_np_typeplomb = df_plombs_np.loc[df_plombs_np['Тип'] == 'Пломба']
        logger.warning(df_plombs_np_typeplomb)
        df_plombs_np_typeparcel = df_plombs_np.loc[df_plombs_np['Тип'] == 'Посылка-место']
        logger.warning(df_plombs_np_typeparcel)
        for parcel_plomb_numb in df_plombs_np_typeplomb['parcel_plomb_numb']:
                con.execute(
                    f"Update baza set pallet = '{pallet_new}' where parcel_plomb_numb = '{parcel_plomb_numb}'")
        for parcel_numb in df_plombs_np_typeparcel['parcel_numb']:
                con.execute(
                    f"Update baza set pallet = '{pallet_new}' where parcel_numb = '{parcel_numb}'")
        df_new_pallet = pd.read_sql(f"Select * from baza where pallet = '{pallet_new}'", con)
        writer = pd.ExcelWriter(f'{addition_folder}Паллет {pallet_new}.xlsx', engine='xlsxwriter')
        df_new_pallet.to_excel(writer, sheet_name='Sheet1', index=False)
        for column in df_new_pallet:
            column_width = max(df_new_pallet[column].astype(str).map(len).max(), len(column))
            col_idx = df_new_pallet.columns.get_loc(column)
            writer.sheets['Sheet1'].set_column(col_idx, col_idx, column_width)
            writer.sheets['Sheet1'].set_column(0, 3, 10)
            writer.sheets['Sheet1'].set_column(1, 3, 20)
            writer.sheets['Sheet1'].set_column(2, 3, 20)
            writer.sheets['Sheet1'].set_column(3, 3, 20)
            writer.sheets['Sheet1'].set_column(4, 3, 30)
            writer.sheets['Sheet1'].set_column(5, 3, 20)

        writer.save()
        con.commit()
        df_plombs_np = pd.DataFrame()
        parcel_plomb_numb_np = None
        object_name = pallet_new
        comment = f'Сформировать новый паллет: Паллет сформирован'
        insert_user_action(object_name, comment)
        flash(f'Паллет сформирован!', category='success')
    return render_template('New_pallet.html')


@app_svh.route('/add_to_pallet', methods=['POST', 'GET'])
def add_to_pallet():
    global df_plombs_html
    global df_plombs_np
    global parcel_plomb_numb_np
    try:
        parcel_plomb_numb_np = request.form['parcel_plomb_numb_np']
    except:
        pass
    try:
        con = sl.connect('BAZA.db')
        with con:
            #to show all pallets from system for choice
            df_all_pallets = pd.read_sql(f"SELECT pallet FROM baza", con).drop_duplicates(subset='pallet', keep='first')
            df_all_pallets['pallet'] = df_all_pallets['pallet'].fillna(value=np.nan).fillna(0).astype(int)
            df_all_pallets = df_all_pallets.sort_values(by='pallet', na_position='last', ascending=False)
            df_all_pallets = df_all_pallets['pallet'].to_list()
            logger.warning(df_all_pallets)
            #select row with current plomb (parcel) number
            df_plomb = pd.read_sql(f"SELECT * FROM baza where parcel_plomb_numb = '{parcel_plomb_numb_np}'", con)
            df_plomb['Тип'] = 'Пломба'
            #if that is parcel
            if df_plomb.empty and parcel_plomb_numb_np != None:
                df_plomb = pd.read_sql(f"SELECT * FROM baza where parcel_numb = '{parcel_plomb_numb_np}'", con)
                df_plomb['Тип'] = "Посылка-место"
                logger.warning(df_plomb)
                if df_plomb.empty and parcel_plomb_numb_np != None:
                    flash(f'Пломба не найдена!', category='error')
            else:
                pass
            try:
                df_plombs_np = df_plombs_np.append(df_plomb).drop_duplicates(subset=['parcel_plomb_numb'], keep='first')
                df_plombs_np['№1'] = np.arange(len(df_plombs_np))[::+1] + 1
                df_plombs_html = df_plombs_np.rename(columns={
                    '№1': '№',
                    'parcel_plomb_numb': 'Пломба',
                    'parcel_numb': 'Трек',
                    'pallet': '№ Паллет',
                    'zone': 'Зона',
                    'party_numb': 'Партия'
                })
                df_plombs_html = df_plombs_html[['№', 'Пломба', '№ Паллет',
                                                 'Зона', 'Партия', 'Трек', 'Тип']]
                df_plombs_html.fillna("", inplace=True)
                df_plombs_html = df_plombs_html.reset_index()
                df_plombs_html = df_plombs_html.drop('index', axis=1)
                object_name = parcel_plomb_numb_np
                comment = f'Добавить на паллет: Пломба отмечена для добавления'
                insert_user_action(object_name, comment)
            except:
                pass

    except Exception as e:
        flash(f'Пломба не найдена!', category='error')
        #return render_template('parcel_info_new_place.html')
        return {'message': str(e)}, 400
        pass
    return render_template('add_to_pallet.html', tables=[style + df_plombs_html.to_html(classes='mystyle', index=False,
                                                                                     float_format='{:2,.2f}'.format)],
                           titles=['na', '\n\nМешки\места:'],
                           parcel_plomb_numb_np=parcel_plomb_numb_np, df_all_pallets=df_all_pallets)

@app_svh.route('/add_to_pallet_button', methods=['POST', 'GET'])
def add_to_pallet_button():
    global df_plombs_np
    global parcel_plomb_numb_np
    pallet_new = request.form['pallet_new']
    # открываем базу
    con = sl.connect('BAZA.db')
    with con:
        df_new_pallet = pd.read_sql(f"Select * from baza where pallet = '{pallet_new}'", con)
        logger.warning(df_new_pallet)
        df_plombs_np_typeplomb = df_plombs_np.loc[df_plombs_np['Тип'] == 'Пломба']
        logger.warning(df_plombs_np_typeplomb)
        df_plombs_np_typeparcel = df_plombs_np.loc[df_plombs_np['Тип'] == 'Посылка-место']
        logger.warning(df_plombs_np_typeparcel)
        if not df_new_pallet.empty:
            for parcel_plomb_numb in df_plombs_np_typeplomb['parcel_plomb_numb']:
                    # plomb_toreplace = done_parcels_np.loc[done_parcels_np['Трек-номер'] == parcel_numb]['Пломба'].values[0]
                    # logger.warning(plomb_toreplace)
                    con.execute(
                        f"Update baza set pallet = '{pallet_new}' where parcel_plomb_numb = '{parcel_plomb_numb}'")
            for parcel_numb in df_plombs_np_typeparcel['parcel_numb']:
                con.execute(
                    f"Update baza set pallet = '{pallet_new}' where parcel_numb = '{parcel_numb}'")
            df_new_pallet = pd.read_sql(f"Select * from baza where pallet = '{pallet_new}'", con)
            writer = pd.ExcelWriter(f'{addition_folder}Паллет {pallet_new}.xlsx', engine='xlsxwriter')
            df_new_pallet.to_excel(writer, sheet_name='Sheet1', index=False)
            for column in df_new_pallet:
                column_width = max(df_new_pallet[column].astype(str).map(len).max(), len(column))
                col_idx = df_new_pallet.columns.get_loc(column)
                writer.sheets['Sheet1'].set_column(col_idx, col_idx, column_width)
                writer.sheets['Sheet1'].set_column(0, 3, 10)
                writer.sheets['Sheet1'].set_column(1, 3, 20)
                writer.sheets['Sheet1'].set_column(2, 3, 20)
                writer.sheets['Sheet1'].set_column(3, 3, 20)
                writer.sheets['Sheet1'].set_column(4, 3, 30)
                writer.sheets['Sheet1'].set_column(5, 3, 20)

            writer.save()
            con.commit()
            df_plombs_np = pd.DataFrame()
            parcel_plomb_numb_np = None
            object_name = pallet_new
            comment = f'Добавить на паллет: Пломбы добавленны'
            insert_user_action(object_name, comment)
            flash(f'Добавлено!', category='success')
        else:
            flash(f'Паллет не найден!', category='error')
    return render_template('add_to_pallet.html')

parc_quont_pallet_info = 0
plomb_quont_pallet_info = 0
@app_svh.route('/pallet_info', methods=['POST', 'GET'])
def pallet_info():
    global parc_quont_pallet_info
    global plomb_quont_pallet_info
    global pallet
    df_refuses = []
    numb = 0
    try:
        numb = request.form['numb']
        logger.warning(numb)
    except:
        pass
    con = sl.connect('BAZA.db')
    if numb == 0:
        df = pd.DataFrame()
    else:
        with con:
            df = pd.read_sql(f"Select * from baza where pallet = '{numb}'", con)
            df = df.rename(columns=map_eng_to_rus)
            df['№'] = np.arange(len(df))[::+1] + 1
            df = df[['№', 'Паллет', 'Пломба', 'Трек-номер', 'Статус ВХ', 'Статус ТО (кратк)',  'Статус ТО', 'Партия']]
            if df.empty:
                df = pd.read_sql(f"Select * from baza where parcel_plomb_numb = '{numb}'", con)
                try:
                    pallet = df['pallet'].values[0]
                    df = pd.read_sql(f"Select * from baza where pallet = '{pallet}'", con)
                    df = df.rename(columns=map_eng_to_rus)
                    df['№'] = np.arange(len(df))[::+1] + 1
                    df = df[['№', 'Паллет', 'Пломба', 'Трек-номер', 'Статус ВХ', 'Статус ТО (кратк)', 'Статус ТО', 'Партия']]
                except:
                    pass
                if df.empty:
                    df = pd.read_sql(f"Select * from baza where parcel_numb = '{numb}'", con)
                    try:
                        pallet = df['pallet'].values[0]
                        df = pd.read_sql(f"Select * from baza where pallet = '{pallet}'", con)
                        df = df.rename(columns=map_eng_to_rus)
                        df['№'] = np.arange(len(df))[::+1] + 1
                        df = df[['№', 'Паллет', 'Пломба', 'Трек-номер', 'Статус ВХ', 'Статус ТО (кратк)', 'Статус ТО', 'Партия']]
                    except:
                        pass
            df_refuses = df.loc[df['Статус ТО (кратк)'] == 'ИЗЪЯТИЕ']
            df_refuses = df_refuses['Трек-номер'].to_list()
            logger.warning(df_refuses)
            parc_quont_pallet_info = len(df)
            plomb_quont_pallet_info = len(df.drop_duplicates(subset='Пломба', keep='first'))
            pallet = df['Паллет'].values[0]
            def highlight_RED(df):
                    return ['background-color: #FF0000' if 'ИЗЪЯТИЕ' in str(i) else '' for i in df]

            df = df.style.apply(highlight_RED).hide_index()
    return render_template('pallet_info.html', tables=[style + df.to_html(classes='mystyle', index=False,
                                                                   float_format='{:2,.2f}'.format)],
                           titles=['Информация'], df_refuses=df_refuses,
                           pallet=pallet, numb=numb, parc_quont_pallet_info=parc_quont_pallet_info, plomb_quont_pallet_info=plomb_quont_pallet_info)


def chunks(lst, n):
    """Yield successive n-sized chunks from lst."""
    for i in range(0, len(lst), n):
        yield lst[i:i + n]


def GBS_request_events():
    con = sl.connect('BAZA.db')
    len_id = con.execute('SELECT max(id) from baza').fetchone()[0]
    print(len_id)
    id_for_job = len_id - 1000000
    print(id_for_job)
    df = pd.read_sql(f"Select parcel_numb from baza "
                     f"where party_numb LIKE '%GBS%' "
                     f"AND ID > {id_for_job} "  # where ID > (len(ID) - 200 000)
                     f"AND custom_status_short = 'ИЗЪЯТИЕ' "
                     f"AND custom_status != 'Return in process'", con).drop_duplicates(subset='parcel_numb')
    print(df)
    list_chanks = list(chunks(df['parcel_numb'], 25))
    #print(list_chanks)
    n = 0
    for chank in list_chanks:
        print(chank)
        list_of_dicts = []
        for i in chank:
            json_parc = {"HWBRefNumber": i}
            list_of_dicts.append(json_parc)
        print(list_of_dicts)
        print(len(list_of_dicts))
        url = "https://api.gbs-broker.ru/"
        body = {
                 "jsonrpc": "2.0",
                 "method": "get_events_public",
                 "params": {
                     "Filter": {
                     "HWB": list_of_dicts

                     },
                     "TextLang": "ru"
                     },

                 "id": "c52f1b33-dg"
                }
        headers = {'Content-Type': 'application/json'}
        response = requests.post(url=url, json=body,  # http://164.132.182.145:5001
                                 headers=headers)

        print(response)
        json_events = response.json()
        print(json_events)
        result = json_events['result']
        """Event_triger = ['CI',
                        'AI',
                        'CR',
                        'CR2',
                        'CR3',
                        'PCD',
                        'ASF',
                        'HBA',
                        'HBA41',
                        'HBA42',
                        'HBA43',
                        'HBA44',
                        'HBA45',
                        'HBA46',
                        'NOID',
                        'DOCOK',
                        'RIC'
                        ]"""
        with con:
            for parcel_slot in result:
                try:
                    parcel_numb = parcel_slot['HWBRefNumber']
                    #event_code = parcel_slot['events'][0]['event_code']
                    custom_status = parcel_slot['events'][0]['event_text']
                    events = parcel_slot['events']
                    if 'clearance complete' in custom_status or 'Released by customs' in custom_status:
                        custom_status_short = 'ВЫПУСК'
                        decision_date = parcel_slot['events'][0]['event_time']
                        refuse_reason = parcel_slot['events'][0]['event_comment']
                    else:
                        for event in events:
                            if 'CR' in event['event_code'] or 'CR2' in event['event_code'] or 'CR3' in event['event_code']:
                                custom_status_short = 'ВЫПУСК'
                                decision_date = event['event_time']
                                refuse_reason = event['event_comment']
                                break
                            else:

                                custom_status_short = 'ИЗЪЯТИЕ'
                                decision_date = event['event_time']
                                refuse_reason = event['event_comment']
                                print(refuse_reason)


                        # custom_status_short = 'ИЗЪЯТИЕ'

                    con.execute(f"Update baza set "
                                f" custom_status = '{custom_status}',"
                                f" custom_status_short = '{custom_status_short}',"
                                f" decision_date = '{decision_date}',"
                                f" refuse_reason = '{refuse_reason}' where parcel_numb = '{parcel_numb}'")
                    print('updated')
                except Exception as e:
                    logger.warning(f'parcel_slot: {parcel_slot} - ERROR: {e}')
        n += 25
        print(n)

def Scarif_request_events():
    map_scarif_status = {'exds10': 'выпуск товаров без уплаты таможенных платежей',
                         'exds30': 'выпуск возвращаемых товаров разрешен',
                         'exds31': 'требуется уплата таможенных платежей',
                         'exds32': 'выпуск товаров разрешен, таможенные платежи уплачены',
                         'exds33': 'выпуск разрешён, ожидание по временному ввозу',
                         'exds40': 'разрешение на отзыв',
                         'exds70': 'продление срока выпуска',
                         'exds90': 'отказ в выпуске товаров',
                         'exds0': 'статус не определен'}
    con = sl.connect('BAZA.db')
    #con.execute('pragma journal_mode=wal')
    df = pd.read_sql(f"Select parcel_numb from baza "
                     f"where party_numb NOT LIKE '%GBS%' and party_numb NOT LIKE '%TEST%' and party_numb NOT LIKE '%тест%'"  # where ID > (len(ID) - 200 000)
                     f"AND custom_status_short = 'ИЗЪЯТИЕ' AND custom_status != 'РЕЭКСПОРТ ЗАПРЕТ ВЫВОЗА'", con).drop_duplicates(subset='parcel_numb')
    for parcel_numb in df['parcel_numb']:
        url = f"https://cellog.deklarant.ru/api/external/parcel-status/{parcel_numb}"
        headers = {"api-token": "40e2f498-450c-4b9f-a509-7f4c8877a6ff"}
        try:
            response = requests.get(url=url,  # http://164.132.182.145:5001
                                    headers=headers)

            json_events = response.json()
            registration_numb = json_events['registryNumber']
            custom_status = json_events["externalStatus"]
            for key in map_scarif_status.keys():
                custom_status = custom_status.replace(key, map_scarif_status[key])
            print(custom_status)
            if 'Выпуск' in str(custom_status) or 'выпуск ' in str(custom_status):
                custom_status_short = 'ВЫПУСК'
            else:
                custom_status_short = 'ИЗЪЯТИЕ'
            decision_date = json_events["decisionAt"]
            refuse_reason = json_events["reasonMessage"]
            with con:

                con.execute(f"Update baza set"
                            f" registration_numb = '{registration_numb}',"
                            f" custom_status = '{custom_status}',"
                            f" custom_status_short = '{custom_status_short}',"
                            f" decision_date = '{decision_date}',"
                            f" refuse_reason = '{refuse_reason}'"
                            f"where parcel_numb = '{parcel_numb}'")
        except Exception as e:
            print(e)


@app_svh.route('/search', methods=['GET'])
def parc_searh():
    return render_template('parc_search.html')


@app_svh.route('/get_info', methods=['POST', 'GET'])
def get_parcel_info_list():
    parcel_numbs = request.form['parcel_numbs'].replace(' ', ',')
    parcels_list = parcel_numbs.split(",")
    df_all_parcels = pd.DataFrame()
    for parcel_numb in parcels_list:
        try:
            con = sl.connect('BAZA.db')
            with con:
                df_parc_events = pd.read_sql(
                    f"SELECT * FROM baza where parcel_numb = '{parcel_numb}' ORDER BY ID", con)
                print(df_parc_events)
                df_all_parcels = df_all_parcels.append(df_parc_events)
        except Exception as e:
            logger.warning(f'parcel {parcel_numb}  - read action faild with error: {e}')
            return {'message': str(e)}, 400
    return render_template('parc_info.html', tables=[df_all_parcels.to_html(classes='mystyle', index=False)],
                           titles=['na', 'ALL'],
                           parcel_numb=parcel_numb)


@app_svh.route('/api_insert_decisions/', methods=['POST', 'GET'])
def api_insert_decisions():
    event_details = request.get_json()
    parcel_numb = event_details["parcel_numb"]
    try:
        registration_numb = event_details["registration_numb"]
    except:
        registration_numb = 'unknown'
    custom_status = event_details["custom_status"]
    custom_status_short = event_details["custom_status_short"]
    refuse_reason = event_details["refuse_reason"]
    decision_date = event_details["decision_date"]
    try:
        con = sl.connect('BAZA.db')
        with con:
            row_isalready_in = pd.read_sql(f"Select * from baza where parcel_numb = '{parcel_numb}'", con)
            print(row_isalready_in)
            if row_isalready_in.empty:
                statement = "INSERT INTO baza (registration_numb, parcel_numb, custom_status, custom_status_short, refuse_reason, decision_date) VALUES (?, ?, ?, ?, ?, ?)"
                con.execute(statement, [registration_numb, parcel_numb, custom_status, custom_status_short, refuse_reason, decision_date])
                print('INSERT OK')
            else:
                con.execute(f"Update baza set"
                            f" registration_numb = '{registration_numb}',"
                            f" custom_status = '{custom_status}',"
                            f" custom_status_short = '{custom_status_short}',"
                            f" decision_date = '{decision_date}',"
                            f" refuse_reason = '{refuse_reason}' where parcel_numb = '{parcel_numb}'")

                print('Update OK')
        return jsonify(True)
    except Exception as e:
        print(e)


def progress(status, remaining, total):
    print(f'Copied {total-remaining} of {total} pages...')


def backup():
    con = sl.connect("BAZA.db")
    bck = sl.connect('BAZA_backup.db')
    with bck:
        con.backup(bck, pages=1, progress=progress)
    bck.close()
    con.close()


def check_and_backup():
    con = sl.connect("BAZA.db")
    cur = con.cursor()
    try:
        cur.execute("PRAGMA integrity_check")
        print('Baza is OK')
        backup()
    except sl.DatabaseError:
        con.close()


check_and_backup()

#scheduler = BackgroundScheduler(daemon=True)


# Create the job
#scheduler.add_job(func=back_up, trigger='interval', seconds=30) #trigger='cron', hour='22', minute='30'
#scheduler.start()

if __name__ == '__main__':
    app_svh.secret_key = 'c9e779a3258b42338334daaed51bccf7'
    app_svh.config['SESSION_TYPE'] = 'filesystem'
    app_svh.run(host='0.0.0.0', port=5001, debug=True)  #http://127.0.0.1:5000
