from flask import Flask, jsonify, request, render_template, redirect, url_for, send_file
from flask import abort
from flask import flash
import pandas as pd
import sqlite3 as sl
import numpy as np
import logging
import openpyxl
from copy import copy
from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet.page import PageMargins
from openpyxl.styles import Font
from openpyxl.styles import Alignment
import datetime
import json
from io import BytesIO
from smtplib import SMTP_SSL
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import os
import requests
import pytz
from flask_sqlalchemy import SQLAlchemy
from flask_restful import Resource,  Api
from flask import Response
from sqlalchemy.orm import sessionmaker, scoped_session
from sqlalchemy import create_engine
from sqlalchemy.ext.declarative import declarative_base
from apispec.ext.marshmallow import MarshmallowPlugin
from apispec import APISpec
from flask_apispec.extension import FlaskApiSpec
from flask_apispec import use_kwargs, marshal_with
from flask import make_response
from flask_jwt_extended import (
    JWTManager, jwt_required, create_access_token,
    create_refresh_token,
    get_jwt_identity, set_access_cookies,
    set_refresh_cookies, unset_jwt_cookies)
from schemas import VideoSchema, UserSchema, AuthSchema
import winsound
from apscheduler.schedulers.background import BackgroundScheduler


download_folder = 'C:/Users/User/Desktop/ДОКУМЕНТЫ/'
download_folder_allmanif = 'C:/Users/User/Desktop/ДОКУМЕНТЫ/ОТГРУЖЕННОЕ'
addition_folder = f'{download_folder}Места-Паллеты/'
if not os.path.isdir(download_folder):
    os.makedirs(download_folder, exist_ok=True)
if not os.path.isdir(addition_folder):
    os.makedirs(addition_folder, exist_ok=True)
if not os.path.isdir(download_folder_allmanif):
    os.makedirs(download_folder_allmanif, exist_ok=True)
pd.set_option("display.precision", 3)
pd.options.display.float_format = '{:.3f}'.format


app_svh = Flask(__name__)
app_svh.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024
app_svh.config['UPLOAD_EXTENSIONS'] = ['.xls', '.xlsx']
#app_svh.config["USE_X_SENDFILE"] = True


app_svh.config['JWT_CSRF_CHECK_FORM'] = False
app_svh.config['JWT_TOKEN_LOCATION'] = ['cookies']
app_svh.config["SQLALCHEMY_DATABASE_URI"] =('sqlite:///db.sqlite')


db = SQLAlchemy(app_svh)

client = app_svh.test_client()
engine = create_engine('sqlite:///db.sqlite')
session = scoped_session(sessionmaker(autocommit=False, autoflush=False, bind=engine))

Base = declarative_base()
Base.query = session.query_property()

jwt = JWTManager(app_svh)
app_svh.secret_key = 'c9e779a3258b42338334daaed51bccf7'
app_svh.config['SESSION_TYPE'] = 'filesystem'
app_svh.config['JWT_SECRET_KEY'] = 'c9e779a3258b42338334daaed51bccf7'

api = Api(app_svh)

class UserLogin(Resource):
    def post(self):
        username = request.get_json()['username']
        password = request.get_json()['password']
        if username == 'admin' and password == 'habr':
            access_token = create_access_token(identity={
                'role': 'admin',
            }, expires_delta=False)
            result = {'token': access_token}
            return result
        return {'error': 'Invalid username and password'}

class ProtectArea(Resource):
    @jwt_required
    def get(self):
        return {'answer': 42}


api.add_resource(UserLogin, '/api/login/')
api.add_resource(ProtectArea, '/api/protect-area/')

docs = FlaskApiSpec()
docs.init_app(app_svh)
app_svh.config.update({
    'APICPEC_SPEC': APISpec(
        title='videoblog',
        version='v1',
        openapi_version='2.0',
        plugins=[MarshmallowPlugin()]
    ),
    'APISPEC_SWAGGER_URL': '/swagger/'
})

from models import *
Base.metadata.create_all(bind=engine)

@app_svh.route('/a', methods=['GET', 'POST'])
def playAudioFile():
    return render_template('audio.html')

@app_svh.route('/todo/api/v1.0/register', methods=['POST'])
@use_kwargs(UserSchema)
@marshal_with(AuthSchema)
def register(**kwargs):
    try:
        user = User(**kwargs)
        session.add(user)
        session.commit()
        token = user.get_token()
    except Exception as e:
        logger.warning(f'register error: {e}')
        return {'message': str(e)}, 400
    return {'access_token': token}

@app_svh.route('/home')
def home():
    return render_template("home.html")

@app_svh.route('/login', methods=['GET', 'POST'])
def login_start():
    return render_template("login.html")

@app_svh.route('/get_user_id')
@jwt_required()
def get_user_id():
    user_id = get_jwt_identity()
    resp = make_response(redirect(url_for('fetchmany_party')))
    resp.set_cookie('user_id', str(user_id))
    return resp

@app_svh.route('/todo/api/v1.0/login_insert', methods=['POST'])
def login_insert():
    try:
        email = request.form['email']
        password = request.form['password']
        log = client.post('/todo/api/v1.0/login', json={'email': email, 'password': password})
        access_token = log.get_json()['access_token']
        resp = make_response(redirect(url_for('get_user_id')))
        set_access_cookies(resp, access_token)
    except Exception as e:
        logger.warning(e)
        return {'message': str(e)}, 400
    return resp


@app_svh.route('/todo/api/v1.0/login', methods=['POST'])
@use_kwargs(UserSchema(only=('email', 'password')))
@marshal_with(AuthSchema)
def login(**kwargs):
    user = User.authenticate(**kwargs)
    token = user.get_token()
    return {'access_token': token}

@app_svh.route('/logout')
@jwt_required()
def logout():
    resp = make_response(redirect(url_for('login_start')))
    #resp.set_cookie('access_token', max_age=0)
    unset_jwt_cookies(resp)
    return resp
@app_svh.teardown_appcontext
def shutdown_session(exception=None):
    session.remove()

@app_svh.errorhandler(422)
def error_handlers(err):
    headers = err.data.get('headers', None)
    messages = err.data.get('messages', ['Invalid request'])
    if headers:
        return jsonify({'message': messages}), 400, headers
    else:
        return jsonify({'message': messages}), 400

@app_svh.route('/a')
def returnAudioFile():
    path_to_audio_file = "/statiс/SndNoPlomb.wav" #audio from project dir
    return send_file(
       path_to_audio_file,
       mimetype="audio/wav",
       as_attachment=True,
       attachment_filename="SndNoPlomb.wav")

def insert_user_action(object_name, comment):
    try:
        user_id = request.cookies.get('user_id')
    except:
        user_id = "API"
    if user_id is None:
        user_id = "API"
    print(user_id)
    now_time = datetime.datetime.now().strftime("%d.%m.%Y %H:%M")
    con = sl.connect('BAZA-reports.db')
    with con:
        cur = con.cursor()
        action_date = now_time
        statement = "INSERT INTO report (user_id, object_name, comment, action_date) VALUES (?, ?, ?, ?)"
        cur.execute(statement, [user_id, object_name, comment, action_date])
        con.commit()


now = datetime.datetime.now().strftime("%d.%m.%Y")
now_time = datetime.datetime.now().strftime("%d.%m.%Y %H:%M")
def setup_logger(name, log_file, level=logging.INFO):
    logging.basicConfig(format=u'%(levelname)-8s [%(asctime)s] %(message)s')  # filename=u'mylog.log'
    handler = logging.FileHandler(log_file)
    logger = logging.getLogger(name)
    logger.setLevel(level)
    logger.addHandler(handler)
    return logger
logger = setup_logger('logger', 'mylog.log')

logger_change_plob = setup_logger('logger_change_plob', f'ИЗМЕНЕНИЯ ПЛОМБ.log')
logger_API = setup_logger('logger_API', f'API_insert.log')

#style = '<style>.dataframe th{background: rgb(0,94,73);background: linear-gradient(71deg, rgba(0,94,73,1) 0%, rgba(18,126,80,1) 100%);padding: 10px;font-family: lucida console;color: #343434;border:2px dotted;text-align:left !important;}</style>'
style = '<style>.dataframe th{background: rgb(255,255,255);background: radial-gradient(circle, rgba(255,255,255,1) 0%, rgba(236,236,236,1) 100%);padding: 5px;color: #343434;font-family: monospace;font-size: 110%;border:2px solid #e0e0e0;text-align:left !important;}.dataframe{border: 3px solid #ffebeb !important;}</style>'

map_eng_to_rus = {'registration_numb': 'Реестр', 'party_numb': 'Партия',
                                    'parcel_numb': 'Трек-номер', 'parcel_plomb_numb': 'Пломба', 'parcel_weight': 'вес',
                                    'custom_status': 'Статус ТО', 'custom_status_short': 'Статус ТО (кратк)',
                                    'decision_date': 'Дата решения',
                                    'refuse_reason': 'Причина отказа',
                                    'pallet': 'Паллет',
                                    'zone': 'Зона',
                                  'VH_status': 'Статус ВХ',
                                  'goods': 'Товары'}

con = sl.connect('BAZA-reports.db')
with con:
    baza = con.execute("select count(*) from sqlite_master where type='table' and name='report'")
    for row in baza:
        # если таких таблиц нет
        if row[0] == 0:
            # создаём таблицу
            with con:
                con.execute("""
                            CREATE TABLE report (
                            ID INTEGER PRIMARY KEY AUTOINCREMENT,
                            user_id INTEGER NOT NULL,
                            object_name VARCHAR(30),
                            comment VARCHAR(120),
                            action_date DATETIME
                            );
                        """)

con = sl.connect('BAZA.db')
with con:
    baza = con.execute("select count(*) from sqlite_master where type='table' and name='df_plomb_to_manifest'")
    for row in baza:
        # если таких таблиц нет
        if row[0] == 0:
            # создаём таблицу
            with con:
                con.execute("""
                                                        CREATE TABLE df_plomb_to_manifest (
                                                        ID INTEGER PRIMARY KEY AUTOINCREMENT,
                                                        parcel_plomb_numb VARCHAR(20)
                                                        );
                                                    """)
con.commit()
con.close()

con = sl.connect('BAZA.db')
with con:
    baza = con.execute("select count(*) from sqlite_master where type='table' and name='baza'")
    for row in baza:
        # если таких таблиц нет
        if row[0] == 0:
            # создаём таблицу
            with con:
                con.execute("""
                                        CREATE TABLE baza (
                                        ID INTEGER PRIMARY KEY AUTOINCREMENT,
                                        registration_numb VARCHAR(25),
                                        party_numb VARCHAR(20),
                                        parcel_numb VARCHAR(20) NOT NULL UNIQUE ON CONFLICT REPLACE,
                                        parcel_plomb_numb VARCHAR(20),
                                        parcel_weight FLOAT,
                                        custom_status VARCHAR(400),
                                        custom_status_short VARCHAR(8),
                                        decision_date VARCHAR(20),
                                        refuse_reason VARCHAR(400),
                                        pallet VARCHAR(10),
                                        zone VARCHAR(5),
                                        VH_status VARCHAR(10),
                                        goods
                                        );
                                    """)
con.commit()
con.close()

def insert_decision_API(parcel_numb, custom_status, custom_status_short, refuse_reason, decision_date):
    conn = sl.connect('BAZA.db')
    cur = conn.cursor()
    statement = "INSERT INTO events2 (parcel_numb, custom_status, custom_status_short, refuse_reason, decision_date) VALUES (?, ?, ?, ?, ?)"
    cur.execute(statement, [parcel_numb, custom_status, custom_status_short, refuse_reason, decision_date])
    conn.commit()
    conn.close()
    return True

@app_svh.route('/api/add_decision', methods=['POST'])
def add_decision_API():
    event_details = request.get_json()
    parcel_numb = event_details["parcel_numb"]
    custom_status = event_details["Event"]
    if 'Выпуск' in str(custom_status):
        custom_status_short = 'ВЫПУСК'
    else:
        custom_status_short = 'ИЗЪЯТИЕ'
    refuse_reason = event_details["Event_comment"]
    decision_date = datetime.datetime.strptime(event_details["Event_date"], "%Y-%m-%d %H:%M:%S").replace(tzinfo=pytz.UTC).astimezone(pytz.timezone("Europe/London"))
    result = insert_decision_API(parcel_numb, custom_status, custom_status_short, refuse_reason, decision_date)
    return jsonify(result)

def insert_event_API_test(df, event_date):
    parcel_list = df['parcel_numb'].to_list()
    logger.warning(parcel_list)
    logger.warning(event_date)
    for parcel in parcel_list:
        body = {"parcel_numb": parcel, "Event": "Отгружен с таможенного склада",
                "Event_comment": "Уссурийск", "Event_date": event_date}
        headers = {'accept': 'application/json'}
        response = requests.post('http://164.132.182.145:5000/api/add/new_event', json=body,
                                 headers={'accept': 'application/json'})
        try:
            return response.json()
        except ValueError:
            pass

@app_svh.route('/api/del_plomb_TSD', methods=['POST'])
def api_del_plomb_TSD():
    parcel_details = request.get_json()
    parcel_plomb_numb = parcel_details['parcel_plomb_numb']
    con = sl.connect('BAZA.db')
    print(parcel_plomb_numb)
    try:
        with con:
            con.execute(f"Update baza set VH_status = 'На ВХ', parcel_plomb_numb = ''  where parcel_plomb_numb = '{parcel_plomb_numb}' and custom_status_short = 'ИЗЪЯТИЕ'")
            print('ok')
        #object_name = parcel_numb
        #logger.warning(object_name)
        #comment = 'TSD: Завершено место с посылкой на изъятие'
        #insert_user_action(object_name, comment)
    except Exception as e:
        print(e)
        return {'message': str(e)}, 400
    return jsonify('True')

@app_svh.route('/api/get_plomb_info_API_TSD', methods=['POST'])
def get_plomb_info_API_TSD():
    plomb_details = request.get_json()
    parcel_plomb_numb = plomb_details['parcel_plomb_numb']
    print(parcel_plomb_numb)
    con = sl.connect('BAZA.db')
    try:
        with con:
            query_plomb_status = f"SELECT parcel_numb, custom_status_short from baza where parcel_plomb_numb = '{parcel_plomb_numb}'"
            df_plomb_status = pd.read_sql(query_plomb_status, con)
            print(df_plomb_status)
    except Exception as e:
        print(e)
        return {'message': str(e)}, 400

    return Response(df_plomb_status.to_json(orient="records", indent=2), mimetype='application/json')

@app_svh.route('/api/create_pallet_API_TSD', methods=['POST'])
def create_pallet_API_TSD():
    plomb_details = request.get_json()
    print(plomb_details)
    df_plombs = pd.json_normalize(json.loads(plomb_details))
    print(df_plombs)
    try:
        con = sl.connect('BAZA.db')
        with con:
            df_all_pallets = pd.read_sql(f"SELECT DISTINCT pallet FROM baza", con).drop_duplicates(subset='pallet', keep='first')
            try:
                df_all_pallets['pallet'] = df_all_pallets['pallet'].fillna(0).astype(int)
                df_all_pallets = df_all_pallets.sort_values(by='pallet')
                last_pall_numb = int(df_all_pallets.values[-1].tolist()[0])
                pallet_new = last_pall_numb + 1
            except:
                pallet_new = 1
            for parcel_plomb_numb in df_plombs['parcel_plomb_numb']:
                print(parcel_plomb_numb)
                con.execute(
                    f"Update baza set pallet = '{pallet_new}' where parcel_plomb_numb = '{parcel_plomb_numb}'")
            df_new_pallet = pd.read_sql(f"Select * from baza where pallet = '{pallet_new}'", con)
            writer = pd.ExcelWriter(f'{addition_folder}Паллет {pallet_new}.xlsx', engine='xlsxwriter')
            df_new_pallet.to_excel(writer, sheet_name='Sheet1', index=False)
            for column in df_new_pallet:
                column_width = max(df_new_pallet[column].astype(str).map(len).max(), len(column))
                col_idx = df_new_pallet.columns.get_loc(column)
                writer.sheets['Sheet1'].set_column(col_idx, col_idx, column_width)
                writer.sheets['Sheet1'].set_column(0, 3, 10)
                writer.sheets['Sheet1'].set_column(1, 3, 20)
                writer.sheets['Sheet1'].set_column(2, 3, 20)
                writer.sheets['Sheet1'].set_column(3, 3, 20)
                writer.sheets['Sheet1'].set_column(4, 3, 30)
                writer.sheets['Sheet1'].set_column(5, 3, 20)

            writer.save()
            con.commit()
            object_name = pallet_new
            comment = f'Сформировать новый паллет: Паллет сформирован'
            insert_user_action(object_name, comment)
            flash(f'Паллет сформирован!', category='success')
            winsound.PlaySound('Snd\Pallet_made.wav', winsound.SND_FILENAME)
        result = jsonify({"pallet": pallet_new})
    except Exception as e:
        result = jsonify({"Error": e})
    return result

@app_svh.route('/api/update_decisions/<party_numb>', methods=['POST', 'GET'])
def API_update_decisions(party_numb):
    con = sl.connect('BAZA.db')
    with con:
        df_request_decisions = pd.read_sql(f"SELECT parcel_numb FROM baza WHERE party_numb = '{party_numb}'", con)
    body = df_request_decisions.to_json(orient="records", indent=2)
    # body = {'parcel_numb': 'CEL7000012753CD'}
    print(body)
    headers = {'accept': 'application/json'}
    response = requests.post('http://164.132.182.145:5001/api/get_decisions',
                             json=body)  # http://127.0.0.1:5000  # 'http://164.132.182.145:5001/api/get_decisions'
    try:
        json_decisions = response.json()
        print(type(json_decisions))
        print(json_decisions)
        df_loaded_decisions = pd.DataFrame.from_records(json_decisions)
        print(df_loaded_decisions)
        with con:
            df_to_append = pd.DataFrame()
            for parcel_numb in df_loaded_decisions['parcel_numb']:
                row_isalready_in = pd.read_sql(f"Select * from baza where parcel_numb = '{parcel_numb}'", con)
                row = df_loaded_decisions.loc[df_loaded_decisions['parcel_numb'] == parcel_numb]
                if row_isalready_in.empty:
                    # row.to_sql('baza', con=con, if_exists='append', index=False)
                    df_to_append = df_to_append.append(row)
                    print('appended')
                else:
                    custom_status_short = df_loaded_decisions.loc[df_loaded_decisions['parcel_numb'] == parcel_numb][
                        'custom_status_short'].values[0]
                    custom_status = df_loaded_decisions.loc[df_loaded_decisions['parcel_numb'] == parcel_numb]['custom_status'].values[
                        0]
                    decision_date = df_loaded_decisions.loc[df_loaded_decisions['parcel_numb'] == parcel_numb]['decision_date'].values[
                        0]
                    refuse_reason = df_loaded_decisions.loc[df_loaded_decisions['parcel_numb'] == parcel_numb]['refuse_reason'].values[
                        0]

                    con.execute(f"Update baza set "
                                f" custom_status = '{custom_status}',"
                                f" custom_status_short = '{custom_status_short}',"
                                f" decision_date = '{decision_date}',"
                                f" refuse_reason = '{refuse_reason}' where parcel_numb = '{parcel_numb}'")
                    print('updated')
                #    row_isalready_in = pd.read_sql(f"Select * from baza where parcel_numb = '{parcel_numb}'", con)
                # df_isalready_in = df_isalready_in.append(row_isalready_in)
            df_to_append.to_sql('baza', con=con, if_exists='append', index=False)
            print(df_to_append)
            flash(f'Решения по партии {party_numb} загружены')
            object_name = party_numb
            comment = 'Решения по партии загружены вручную'
            insert_user_action(object_name, comment)
    except:
        print(response)
        flash(f'Ошибка загрузки решений {response}')
    return render_template('index.html')

def send_mail(filepath, subject):
    basename = os.path.basename(filepath)
    address = "logistick.dv@yandex.ru"

    # Compose attachment
    part = MIMEBase('application', "octet-stream")
    part.set_payload(open(filepath, "rb").read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', 'attachment; filename="%s"' % basename)

    # Compose message
    msg = MIMEMultipart()
    msg['From'] = address
    msg['To'] = address
    msg['Subject'] = subject
    msg.attach(part)

    # Send mail
    smtp = SMTP_SSL('smtp.yandex.ru')
    smtp.connect('smtp.yandex.ru')
    smtp.login(address, 'rwlefgatbfpewlmt')
    smtp.sendmail(address, address, msg.as_string())
    smtp.quit()
df = pd.DataFrame()
@app_svh.route('/TEST', methods=['GET', 'POST'])
def test(df):
    logger.warning('start')
    # create an output stream
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='xlsxwriter')
    # taken from the original question
    df.to_excel(writer)
    # the writer has done its job
    writer.close()
    # go back to the beginning of the stream
    output.seek(0)
    # finally return the file
    logger.warning(output)
    return send_file(output, download_name="testing.xlsx", as_attachment=True)

@app_svh.route('/parties_analitic')
def parties_analitic():
    con = sl.connect('BAZA.db')
    con.row_factory = sl.Row
    with con:
        len_id = con.execute('SELECT max(id) from baza').fetchone()[0]
        id_for_job = len_id - 100000
        data = f"Select DISTINCT ID, party_numb from baza where ID > {id_for_job}"
        #data_df = pd.DataFrame(data)
        data_df = pd.read_sql(data, con).sort_values(by='ID', ascending=False)

        parties = data_df.drop_duplicates(subset='party_numb')['party_numb']
        parties = parties.fillna(value=np.nan)
        list_of_parties = parties.to_list()
        tuple_of_parties = tuple(list_of_parties)
        #print(tuple_of_parties)
        data2 = f'Select party_numb, parcel_plomb_numb, parcel_numb, custom_status_short, custom_status, VH_status, parcel_weight  from baza where party_numb in {tuple_of_parties}'
        parts_analit_table = pd.read_sql(data2, con)
        print(parts_analit_table)

@app_svh.route('/')
def fetchmany_party():
    con = sl.connect('BAZA.db')
    con.row_factory = sl.Row
    with con:
        len_id = con.execute('SELECT max(id) from baza').fetchone()[0]
        id_for_job = len_id - 100000
        data_start = f"Select DISTINCT ID, party_numb from baza where ID > {id_for_job}"
        #data_df = pd.DataFrame(data)
        data_df = pd.read_sql(data_start, con).sort_values(by='ID', ascending=False)
        print(data_df)
        parties = data_df.drop_duplicates(subset='party_numb')['party_numb'].dropna()

        print(parties)
        list_of_parties = parties.to_list()
        tuple_of_parties = tuple(list_of_parties)
        #print(tuple_of_parties)
        data_query = (f'Select party_numb, parcel_plomb_numb, parcel_numb, custom_status_short, '
                      f'custom_status, VH_status, parcel_weight from baza where party_numb in {tuple_of_parties}')
        data_all = pd.read_sql(data_query, con)
        print(data_all)
        analit_table = pd.DataFrame(index=None)
        for party_numb in list_of_parties:
            data = data_all.loc[(data_all['party_numb'] == party_numb)]
            a = data.drop_duplicates(subset='parcel_plomb_numb').loc[data['parcel_plomb_numb'] != '']
            df_not_shipped = a.loc[a['VH_status'] != 'ОТГРУЖЕН']
            logger.warning(a)
            quonty_plomb = len(a)
            quonty_parcels = len(data)
            not_shipt_quont = len(df_not_shipped)
            df_refuses = data.loc[(data['custom_status_short'] == 'ИЗЪЯТИЕ')]
            quonty_parcels_refuse = len(df_refuses)
            quonty_plomb_refuse = len(
                df_refuses.loc[data['parcel_plomb_numb'] != ''].drop_duplicates(subset='parcel_plomb_numb'))
            custom_control = data.loc[data['custom_status'].str.contains('родление')]
            custom_control_quont = len(custom_control)
            dont_declarate = data.loc[data['custom_status'].str.contains('Unknown')]
            dont_declarate = dont_declarate[['parcel_numb', 'parcel_plomb_numb']]
            dont_declarate_quont = len(dont_declarate)
            weight = round(data['parcel_weight'].sum(), 3)
            df_to_append = pd.DataFrame(data={"Партия": [party_numb], "Кол-во пломб": [quonty_plomb], "Кол-во посылок": [quonty_parcels],
                                              "Отказных пломб": [quonty_plomb_refuse], "Отказных посылок": [quonty_parcels_refuse],
                                              "вес партии": [weight], "Не отгруженные места": [not_shipt_quont]}, index=None)
            analit_table = analit_table.append(df_to_append)
        analit_table = analit_table.reset_index(drop=True)
        analit_table = analit_table.transpose()
        print(analit_table)
    con.close()
    return render_template('index.html', parties=parties,
                           analit_table=analit_table)

@app_svh.route('/all_parties')
def all_parties():
    con = sl.connect('BAZA.db')
    con.row_factory = sl.Row
    with con:
        data = pd.read_sql("Select DISTINCT ID, party_numb from baza", con)
        data = data.sort_values(by='ID', ascending=False)
        parties = data.drop_duplicates(subset='party_numb')['party_numb']
        parties = parties.fillna(value=np.nan)
        logger.warning(parties)
    con.close()
    return render_template('index2.html', parties=parties)

@app_svh.route('/info/not_shiped', methods=['GET', 'POST'])
def all_not_shipped():
    global df_not_shipped
    con = sl.connect('BAZA.db')
    with con:
        df_not_shipped = pd.read_sql("Select * from baza where VH_status != 'ОТГРУЖЕН' and custom_status != 'Unknown'", con)
        df_not_shipped['decision_date'] = df_not_shipped['decision_date'].str.slice(0, 17)
        df_not_shipped = df_not_shipped.rename(columns=map_eng_to_rus)
        df_not_shipped['Товары'] = df_not_shipped['Товары'].str.slice(0, 50)

    return render_template('info_not_shipped.html', tables=[style + df_not_shipped.to_html(index=False,
                                                                  float_format='{:2,.2f}'.format)],
                           titles=[''])
@app_svh.route('/info/not_shiped_to_xl', methods=['GET', 'POST'])
def not_shiped_to_xl():
    global df_not_shipped
    now_time = datetime.datetime.now().strftime("%d.%m.%Y %H-%M")
    writer = pd.ExcelWriter(f'{download_folder}Не отгруженное на {now_time}.xlsx', engine='xlsxwriter')
    df_not_shipped.to_excel(writer, sheet_name='Sheet1', index=False)
    for column in df:
        column_width = max(df[column].astype(str).map(len).max(), len(column))
        col_idx = df.columns.get_loc(column)
        writer.sheets['Sheet1'].set_column(col_idx, col_idx, column_width)
        writer.sheets['Sheet1'].set_column(0, 3, 10)
        writer.sheets['Sheet1'].set_column(1, 3, 20)
        writer.sheets['Sheet1'].set_column(2, 3, 20)
        writer.sheets['Sheet1'].set_column(3, 3, 20)
        writer.sheets['Sheet1'].set_column(4, 3, 30)
        writer.sheets['Sheet1'].set_column(5, 3, 20)
    writer.save()
    flash(f'Не отгруженные список выгружен в excel!', category='success')
    object_name = ''
    comment = f'Не отгруженные список выгружен в excel!'
    insert_user_action(object_name, comment)
    return render_template('info_not_shipped.html', tables=[style + df_not_shipped.to_html(index=False,
                                                          float_format='{:2,.2f}'.format)],
                   titles=[''])

@app_svh.route('/info', methods=['GET', 'POST'])
def object_info():
    numb = request.form['numb']
    con = sl.connect('BAZA.db')
    with con:
        df = pd.read_sql(f"Select * from baza where parcel_numb = '{numb}'", con)
        df['decision_date'] = df['decision_date'].str.slice(0, 17)
        df = df.rename(columns=map_eng_to_rus)
        if df['Статус ТО (кратк)'].str.contains("ИЗЪЯТИЕ").any():
            flash(f'ИЗЪЯТИЕ на склад', category='error')
        elif df['Статус ТО (кратк)'].str.contains("ВЫПУСК").any():
            flash(f'ВЫПУСК!', category='success')
        else:
            pass
        df = df.transpose()
        index = True
        df = df.rename(columns={0: ''})
        object_name = 'Посылка'
        try:
            con_vect = sl.connect('VECTORS.db')
            with con_vect:
                vector = con_vect.execute(f"SELECT vector FROM VECTORS where parcel_numb = '{numb}'").fetchone()
                flash(f'{vector[0]}!', category='success')
        except Exception as e:
            print(e)
            pass
        if df.empty:
            df = pd.read_sql(f"Select * from baza where party_numb = '{numb}'", con)
            df['decision_date'] = df['decision_date'].str.slice(0, 17)
            df = df.drop_duplicates(subset='parcel_plomb_numb')
            df = df.rename(columns=map_eng_to_rus)
            df = df[['Партия', 'Пломба', 'Паллет', 'Зона']]
            object_name = 'Партия'
            index = False
            if df.empty:
                df = pd.read_sql(f"Select * from baza where registration_numb = '{numb}'", con)
                df['decision_date'] = df['decision_date'].str.slice(0, 17)
                df = df.rename(columns=map_eng_to_rus)
                df['Товары'] = df['Товары'].str.slice(0, 50)
                object_name = 'Реестр ПТДЭГ'
                index = False
                if df.empty:
                    df = pd.read_sql(f"Select * from baza where pallet = '{numb}'", con)
                    df['decision_date'] = df['decision_date'].str.slice(0, 17)
                    df = df.drop_duplicates(subset='parcel_plomb_numb')
                    df = df.rename(columns=map_eng_to_rus)
                    df['№'] = np.arange(len(df))[::+1] + 1
                    df = df[['№', 'Пломба', 'Зона']]
                    object_name = 'Паллет'
                    index = False
                    if df.empty:
                        df = pd.read_sql(f"Select * from baza where parcel_plomb_numb = '{numb}' COLLATE NOCASE", con)
                        df['decision_date'] = df['decision_date'].str.slice(0, 17)
                        df = df.rename(columns=map_eng_to_rus)
                        df['Товары'] = df['Товары'].str.slice(0, 50)
                        object_name = 'Пломба'
                        index = False
                        try:
                            con_vect = sl.connect('VECTORS.db')
                            with con_vect:
                                vector = con_vect.execute(
                                    f"SELECT vector FROM VECTORS where parcel_plomb_numb = '{numb}'").fetchone()
                                flash(f'{vector[0]}!', category='success')
                        except Exception as e:
                            print(e)
                            pass


    # create an output stream
    return render_template('info.html', tables=[style + df.to_html(classes='mystyle', index=index,
                                                                              float_format='{:2,.2f}'.format)],
                           titles=['Информация'],
                           numb=numb, object_name=object_name)

@app_svh.route('/party_info/<string:row>', methods=['GET', 'POST'])
def party_info(row):
    con = sl.connect('BAZA.db')
    data = pd.read_sql(f"Select * from baza where party_numb = '{row}'", con)
    a = data.drop_duplicates(subset='parcel_plomb_numb').loc[data['parcel_plomb_numb'] != '']
    logger.warning(a)
    quonty_plomb = len(a)
    quonty_parcels = len(data)
    df_refuses = data.loc[data['custom_status_short'] == 'ИЗЪЯТИЕ']
    quonty_parcels_refuse = len(df_refuses)
    quonty_plomb_refuse = len(df_refuses.loc[data['parcel_plomb_numb'] != ''].drop_duplicates(subset='parcel_plomb_numb'))
    print(quonty_plomb_refuse)
    custom_control = data.loc[data['custom_status'].str.contains('родление')]
    custom_control = custom_control[['registration_numb', 'parcel_numb', 'goods']]
    custom_control = custom_control.rename(columns={'registration_numb': 'Рег. номер', 'parcel_numb': 'Трек', 'goods': 'Товары'})
    custom_control_quont = len(custom_control)
    dont_declarate = data.loc[data['custom_status'].str.contains('Unknown')]
    dont_declarate = dont_declarate[['registration_numb', 'parcel_numb', 'parcel_plomb_numb', 'goods']]
    dont_declarate = dont_declarate.rename(
        columns={'registration_numb': 'Рег. номер', 'parcel_numb': 'Трек', 'parcel_plomb_numb': 'Пломба', 'goods': 'Товары'})
    dont_declarate_quont = len(dont_declarate)
    return render_template('party_info.html', row=row,
                           quonty_plomb=quonty_plomb,
                           quonty_plomb_refuse=quonty_plomb_refuse,
                           quonty_parcels=quonty_parcels,
                           quonty_parcels_refuse=quonty_parcels_refuse,
                           custom_control=custom_control,
                           custom_control_quont=custom_control_quont,
                           dont_declarate_quont=dont_declarate_quont,
                           df_refuses=df_refuses,
                           tables=[style + custom_control.to_html(index=False,
                                                                   float_format='{:2,.2f}'.format),
                                   style + dont_declarate.to_html(index=False,
                                                                   float_format='{:2,.2f}'.format)
                                   ],
                           titles=['Информация о посылке', f'Продление срока выпуска {custom_control_quont}шт:',
                                   f'Неподанные: {dont_declarate_quont}шт']
                           )

@app_svh.route('/party_info_create_pallet/<string:row>', methods=['GET', 'POST'])
def party_info_create_pallet(row):

    return render_template('party_info_create_pallet.html', row=row,)


@app_svh.route('/party_info_allnotshipped_to_pallet/<string:row>', methods=['GET', 'POST'])
def party_info_allnotshipped_to_pallet(row):
    con = sl.connect('BAZA.db')
    try:
        with con:
            df_all_pallets = pd.read_sql(f"SELECT pallet FROM baza", con).drop_duplicates(subset='pallet', keep='first')
            try:
                df_all_pallets['pallet'] = df_all_pallets['pallet'].fillna(value=np.nan).fillna(0).astype(int)
                df_all_pallets = df_all_pallets.sort_values(by='pallet')
                last_pall_numb = int(df_all_pallets.values[-1].tolist()[0])
                new_pallet_numb = last_pall_numb + 1
            except:
                new_pallet_numb = 0
            print(row)
            print(new_pallet_numb)
            con.execute(
                f"Update baza set pallet = '{new_pallet_numb}' where party_numb = '{row}' and VH_status is ?", (None,))
            df_new_pallet = pd.read_sql(f"Select * from baza where pallet = '{new_pallet_numb}'", con)
            writer = pd.ExcelWriter(f'{addition_folder}Паллет {new_pallet_numb}.xlsx', engine='xlsxwriter')
            df_new_pallet.to_excel(writer, sheet_name='Sheet1', index=False)
            for column in df_new_pallet:
                column_width = max(df_new_pallet[column].astype(str).map(len).max(), len(column))
                col_idx = df_new_pallet.columns.get_loc(column)
                writer.sheets['Sheet1'].set_column(col_idx, col_idx, column_width)
                writer.sheets['Sheet1'].set_column(0, 3, 10)
                writer.sheets['Sheet1'].set_column(1, 3, 20)
                writer.sheets['Sheet1'].set_column(2, 3, 20)
                writer.sheets['Sheet1'].set_column(3, 3, 20)
                writer.sheets['Sheet1'].set_column(4, 3, 30)
                writer.sheets['Sheet1'].set_column(5, 3, 20)

            writer.save()
    except Exception as e:
        return {'message': str(e)}, 400

    flash(f'Паллет {new_pallet_numb} c не отгруженными из {row} сформирован!', category='success')
    object_name = new_pallet_numb
    comment = f'Паллет c не отгруженными из {row} сформирован!'
    insert_user_action(object_name, comment)
    return render_template('party_info.html', row=row
                           )

@app_svh.route('/party_info_vectors_to_pallet/<string:row>', methods=['GET', 'POST'])
def party_info_vectors_to_pallet(row):
    con = sl.connect('BAZA.db')
    print(row)
    try:
        df_party = pd.read_sql(f"SELECT * FROM baza where party_numb = '{row}'", con)
        con_vect = sl.connect('VECTORS.db')
        with con_vect:
            df_data = pd.read_sql(f"SELECT * FROM vectors where party_numb = '{row}'", con_vect)
            df_merge_vector_party = pd.merge(df_party, df_data,
                                             how='left',
                                             left_on='parcel_numb',
                                             right_on='parcel_numb').drop_duplicates(subset='parcel_plomb_numb_x')
            print(df_merge_vector_party)
            df_vectors = df_merge_vector_party.drop_duplicates(subset='vector')
            dict_of_pallets = {}
            for vector in df_vectors['vector']:
                with con:
                    df_all_pallets = pd.read_sql(f"SELECT pallet FROM baza", con).drop_duplicates(subset='pallet',
                                                                                                  keep='first')
                    try:
                        df_all_pallets['pallet'] = df_all_pallets['pallet'].fillna(value=np.nan).fillna(0).astype(int)
                        df_all_pallets = df_all_pallets.sort_values(by='pallet')
                        last_pall_numb = int(df_all_pallets.values[-1].tolist()[0])
                        new_pallet_numb = last_pall_numb + 1
                    except:
                        new_pallet_numb = 0
                    print(new_pallet_numb)
                    print(vector)
                    dict_of_pallets[vector] = new_pallet_numb
                    df_create_pallet_vector = df_merge_vector_party.loc[df_merge_vector_party['vector'] == vector]
                    for parcel_plomb_numb in df_create_pallet_vector['parcel_plomb_numb_x']:
                        con.execute(f"Update baza set pallet = '{new_pallet_numb}' where parcel_plomb_numb = '{parcel_plomb_numb}' and VH_status is ?", (None,))
                df_new_pallet = pd.read_sql(f"Select * from baza where pallet = '{new_pallet_numb}'", con)
                writer = pd.ExcelWriter(f'{addition_folder}Паллет {new_pallet_numb} - {vector}.xlsx', engine='xlsxwriter')
                df_new_pallet.to_excel(writer, sheet_name='Sheet1', index=False)
                for column in df_new_pallet:
                    column_width = max(df_new_pallet[column].astype(str).map(len).max(), len(column))
                    col_idx = df_new_pallet.columns.get_loc(column)
                    writer.sheets['Sheet1'].set_column(col_idx, col_idx, column_width)
                    writer.sheets['Sheet1'].set_column(0, 3, 10)
                    writer.sheets['Sheet1'].set_column(1, 3, 20)
                    writer.sheets['Sheet1'].set_column(2, 3, 20)
                    writer.sheets['Sheet1'].set_column(3, 3, 20)
                    writer.sheets['Sheet1'].set_column(4, 3, 30)
                    writer.sheets['Sheet1'].set_column(5, 3, 20)

                writer.save()
    except Exception as e:
        return {'message': str(e)}, 400

    flash(f'Паллеты {dict_of_pallets} c не отгруженными из {row} сформирован!', category='success')
    object_name = new_pallet_numb
    comment = f'Паллет c не отгруженными из {row} сформирован!'
    insert_user_action(object_name, comment)
    return render_template('party_info.html', row=row)



@app_svh.route('/party_info_issues_to_pallet/<string:row>', methods=['GET', 'POST'])
def party_info_issues_to_pallet(row):
    con = sl.connect('BAZA.db')
    try:
        with con:
            df_all_pallets = pd.read_sql(f"SELECT pallet FROM baza", con).drop_duplicates(subset='pallet', keep='first')
            try:
                df_all_pallets['pallet'] = df_all_pallets['pallet'].fillna(value=np.nan).fillna(0).astype(int)
                df_all_pallets = df_all_pallets.sort_values(by='pallet')
                last_pall_numb = int(df_all_pallets.values[-1].tolist()[0])
                new_pallet_numb = last_pall_numb + 1
            except:
                new_pallet_numb = 0
            print(new_pallet_numb)
            data = pd.read_sql(f"Select parcel_plomb_numb, custom_status_short from baza where party_numb = '{row}'", con)
            data_refuse_plombs = data.loc[data['custom_status_short'] == 'ИЗЪЯТИЕ'].drop_duplicates(subset='parcel_plomb_numb')
            all_plombs = data.drop_duplicates(subset='parcel_plomb_numb').loc[data['parcel_plomb_numb'] != '']
            df_plomb_not_refuse = all_plombs[~all_plombs.parcel_plomb_numb.isin(data_refuse_plombs.parcel_plomb_numb)]
            print(df_plomb_not_refuse)
            for parcel_plomb_numb in df_plomb_not_refuse['parcel_plomb_numb']:
                con.execute(
                    f"Update baza set pallet = '{new_pallet_numb}' where parcel_plomb_numb = '{parcel_plomb_numb}' and VH_status is ?", (None,))
                print(f'{parcel_plomb_numb} - pallet updated')
            df_new_pallet = pd.read_sql(f"Select * from baza where pallet = '{new_pallet_numb}'", con)
            writer = pd.ExcelWriter(f'{addition_folder}Паллет {new_pallet_numb}.xlsx', engine='xlsxwriter')
            df_new_pallet.to_excel(writer, sheet_name='Sheet1', index=False)
            for column in df_new_pallet:
                column_width = max(df_new_pallet[column].astype(str).map(len).max(), len(column))
                col_idx = df_new_pallet.columns.get_loc(column)
                writer.sheets['Sheet1'].set_column(col_idx, col_idx, column_width)
                writer.sheets['Sheet1'].set_column(0, 3, 10)
                writer.sheets['Sheet1'].set_column(1, 3, 20)
                writer.sheets['Sheet1'].set_column(2, 3, 20)
                writer.sheets['Sheet1'].set_column(3, 3, 20)
                writer.sheets['Sheet1'].set_column(4, 3, 30)
                writer.sheets['Sheet1'].set_column(5, 3, 20)

            writer.save()
    except Exception as e:
        return {'message': str(e)}, 400

    flash(f'Паллет {new_pallet_numb} c 0-ми из {row} сформирован!', category='success')
    object_name = new_pallet_numb
    comment = f'Паллет c 0-ми из {row} сформирован!'
    insert_user_action(object_name, comment)
    return render_template('party_info.html', row=row
                           )
@app_svh.route('/party_info_refuses/<string:row>', methods=['GET', 'POST'])
def party_info_refuses(row):
    con = sl.connect('BAZA.db')
    data = pd.read_sql(f"Select * from baza where party_numb = '{row}'", con)
    a = data.drop_duplicates(subset='parcel_plomb_numb').loc[data['parcel_plomb_numb'] != '']
    logger.warning(a)
    quonty_plomb = len(a)
    quonty_parcels = len(data)
    df_refuses = data.loc[data['custom_status_short'] == 'ИЗЪЯТИЕ']
    quonty_parcels_refuse = len(df_refuses)
    quonty_plomb_refuse = len(df_refuses.loc[data['parcel_plomb_numb'] != ''].drop_duplicates(subset='parcel_plomb_numb'))
    df_refuses = df_refuses.rename(columns=map_eng_to_rus)
    df_refuses['№'] = np.arange(len(df_refuses))[::+1] + 1
    df_refuses = df_refuses[['№', 'Партия', 'Трек-номер', 'Пломба', 'Статус ТО', 'Статус ТО (кратк)', 'Дата решения', 'Причина отказа', 'вес',
                             'Товары', 'Паллет']]
    return render_template('party_info_refuses.html', row=row,
                           quonty_plomb=quonty_plomb,
                           quonty_plomb_refuse=quonty_plomb_refuse,
                           quonty_parcels=quonty_parcels,
                           quonty_parcels_refuse=quonty_parcels_refuse,
                           tables=[style + df_refuses.to_html(index=False,
                                                                  float_format='{:2,.2f}'.format)],
                           titles=['', f'Список изъятий по партии {row}']
                           )

@app_svh.route('/party_info_refuses_excel/<string:row>', methods=['GET', 'POST'])
def party_info_refuses_excel(row):
    con = sl.connect('BAZA.db')
    data = pd.read_sql(f"Select * from baza where party_numb = '{row}'", con)
    a = data.drop_duplicates(subset='parcel_plomb_numb').loc[data['parcel_plomb_numb'] != '']
    quonty_plomb = len(a)
    quonty_parcels = len(data)
    df_refuses = data.loc[data['custom_status_short'] == 'ИЗЪЯТИЕ']
    quonty_parcels_refuse = len(df_refuses)
    quonty_plomb_refuse = len(
        df_refuses.loc[data['parcel_plomb_numb'] != ''].drop_duplicates(subset='parcel_plomb_numb'))
    df_refuses = df_refuses.rename(columns=map_eng_to_rus)
    writer = pd.ExcelWriter(f'{download_folder}ИЗЪЯТИЯ {row}.xlsx', engine='xlsxwriter')
    df_refuses.to_excel(writer, sheet_name='Sheet1', index=False)
    for column in df_refuses:
        column_width = max(df_refuses[column].astype(str).map(len).max(), len(column))
        col_idx = df_refuses.columns.get_loc(column)
        writer.sheets['Sheet1'].set_column(col_idx, col_idx, column_width)
        writer.sheets['Sheet1'].set_column(0, 3, 10)
        writer.sheets['Sheet1'].set_column(1, 3, 20)
        writer.sheets['Sheet1'].set_column(2, 3, 20)
        writer.sheets['Sheet1'].set_column(3, 3, 20)
        writer.sheets['Sheet1'].set_column(4, 3, 30)
        writer.sheets['Sheet1'].set_column(5, 3, 20)
    writer.save()
    flash(f'Изъятия по партии {row} выгружены в excel!', category='success')
    object_name = row
    comment = f'Изъятия по партии {row} выгружены в excel!'
    insert_user_action(object_name, comment)
    return render_template('party_info_refuses.html', row=row,
    quonty_plomb=quonty_plomb,
    quonty_plomb_refuse = quonty_plomb_refuse,
    quonty_parcels = quonty_parcels,
    quonty_parcels_refuse = quonty_parcels_refuse)

@app_svh.route('/party_info_vectors/<string:row>', methods=['GET', 'POST'])
def party_info_vectors(row):
    con =  sl.connect('BAZA.db')
    con_vect = sl.connect('VECTORS.db')
    print(row)
    with con_vect:
        df_vectors = pd.read_sql(f"Select * from vectors where party_numb = '{row}'", con_vect)
    print(df_vectors)
    a = df_vectors.drop_duplicates(subset='parcel_plomb_numb').loc[df_vectors['parcel_plomb_numb'] != '']
    logger.warning(a)
    quonty_plomb = len(a)
    quonty_parcels = len(df_vectors)
    df_vectors['№'] = np.arange(len(df_vectors))[::+1] + 1
    df_vectors['quont'] = 1
    print(df_vectors)
    group_df_vectors = df_vectors.drop_duplicates(subset='parcel_plomb_numb').loc[df_vectors['parcel_plomb_numb'] != '']
    with con:
        query = f"Select DISTINCT parcel_plomb_numb from baza where party_numb = '{row}' and VH_status is ?"
        df_not_shiped = pd.read_sql(sql=query, con=con, params=(None,))
        print(df_not_shiped)
        df_not_shiped['quont2'] = 1
        df_not_shiped['quont2'] = df_not_shiped['quont2'].astype(int)
        pd.options.display.float_format = '{:,.0f}'.format
    group_df_vectors_all = pd.merge(group_df_vectors, df_not_shiped, how='left', left_on='parcel_plomb_numb', right_on='parcel_plomb_numb')

    group_df_vectors = group_df_vectors_all.groupby('vector')['quont', 'quont2'].sum().reset_index() #.to_frame()
    print(group_df_vectors)
    group_df_vectors = group_df_vectors.rename(columns={'vector': 'Направление', 'quont': 'Кол-во пломб', 'quont2': 'Не отгруженны'})
    group_df_vectors_all = group_df_vectors_all.loc[group_df_vectors_all['quont2'] == 1]
    group_df_vectors_all['№п/п'] = np.arange(len(group_df_vectors_all))[::+1] + 1
    group_df_vectors_all = group_df_vectors_all.rename(columns={'parcel_plomb_numb': 'Пломба',
                                                                'vector': 'Направление'})
    group_df_vectors_all = group_df_vectors_all.reindex(columns=['№п/п',
                                                                'Пломба',
                                                                'Направление'])
    return render_template('party_info_vectors.html', row=row,
                           quonty_plomb=quonty_plomb,
                           quonty_parcels=quonty_parcels,
                           group_df_vectors_all=group_df_vectors_all,
                           tables=[style + group_df_vectors.to_html(index=False,
                                                                  float_format='{:,.0f}'.format),
                                   style + group_df_vectors_all.to_html(index=False,
                                                                        float_format='{:,.0f}'.format)
                                   ],
                           titles=['', f'Направления по партии {row}', 'Неотгруженные пломбы:']
                           )

@app_svh.route('/check_refuses/<string:row>', methods=['GET', 'POST'])
def check_refuses(row):
    party_numb = row
    con = sl.connect('BAZA.db')
    with con:
        try:
            con = sl.connect('BAZA.db')
            with con:
                baza = con.execute("select count(*) from sqlite_master where type='table' and name='party_refuses'")
                for row in baza:
                    # если таких таблиц нет
                    if row[0] == 0:
                        # создаём таблицу
                        with con:
                            con.execute("""
                                                    CREATE TABLE party_refuses (
                                                    ID INTEGER PRIMARY KEY AUTOINCREMENT,
                                                    party_numb VARCHAR(20),
                                                    parcel_numb VARCHAR(20),
                                                    custom_status_short VARCHAR(8),
                                                    parcel_find_status VARCHAR(8)
                                                    );
                                                """)

                df = pd.read_sql(f"SELECT party_numb, parcel_numb, custom_status_short FROM baza where party_numb = '{party_numb}' and custom_status_short = 'ИЗЪЯТИЕ' COLLATE NOCASE", con)
                df['parcel_find_status'] = " "
                df['№'] = np.arange(len(df))[::+1] + 1
                df = df[['№', 'party_numb', 'parcel_numb', 'parcel_find_status']]
                index = False
                print(df)
                df.to_sql('party_refuses', con=con, if_exists='replace', index=False)
        except Exception as e:
            return {'message': str(e)}, 400
    return render_template('party_refuses_work.html', tables=[style + df.to_html(classes='mystyle', index=index,
                                                                                float_format='{:2,.2f}'.format)],
                           titles=['Информация'],
                           party_numb=party_numb)

@app_svh.route('/check_refuses_work/<string:party_numb>', methods=['GET', 'POST'])
def check_refuses_work(party_numb):
    parcel_numb = request.form['parcel_numb']
    con = sl.connect('BAZA.db')
    with con:
        df_check_parcel = pd.read_sql(
            f"SELECT * FROM party_refuses where parcel_numb = '{parcel_numb}' COLLATE NOCASE", con)
        if df_check_parcel.empty:
            flash(f'Посылка {parcel_numb} не входит в список изъятий или не найдена!', category='error')
            winsound.PlaySound('Snd\Snd_CancelIssue.wav', winsound.SND_FILENAME)
        else:
            con.execute(
                f"Update party_refuses set parcel_find_status = 'НАЙДЕНА' where parcel_numb = '{parcel_numb}'")
        df = pd.read_sql(f"SELECT * FROM party_refuses where party_numb = '{party_numb}' COLLATE NOCASE", con)
        df['№'] = np.arange(len(df))[::+1] + 1
        df = df[['№', 'party_numb', 'parcel_numb', 'parcel_find_status']]
        if df.loc[df['parcel_find_status'] == ' '].empty:
            flash(f'Все изъятия найдены', category='success')
            winsound.PlaySound('Snd\se_mesta_naid.wav', winsound.SND_FILENAME)
        index = False
        quont_refuses_all = len(df)
        quont_refuses_done = len(df.loc[df['parcel_find_status'] == 'НАЙДЕНА'])
        quon_not_done = quont_refuses_all - quont_refuses_done
        object_name = parcel_numb
        comment = f'Проверка изъятий по партии {party_numb}: Посылка просмотренна'
        insert_user_action(object_name, comment)

    return render_template('party_refuses_work2.html', tables=[style + df.to_html(classes='mystyle', index=index,
                                                                                float_format='{:2,.2f}'.format)],
                           titles=['Информация'],
                           party_numb=party_numb, quont_refuses_all=quont_refuses_all,
                           quont_refuses_done=quont_refuses_done, quon_not_done=quon_not_done)

@app_svh.route('/add/load_tracks', methods=['GET', 'POST'])
def load_tracks():
    if request.method == 'POST':
        uploaded_file = request.files['file']
        filename = uploaded_file.filename
        logger.warning(filename)
        if filename != '':
            file_ext = os.path.splitext(filename)[1]
            if file_ext not in app_svh.config['UPLOAD_EXTENSIONS']:
                abort(400)
            uploaded_file.save(uploaded_file.filename)
            df_track = pd.read_excel(filename, sheet_name=0, engine='openpyxl', usecols='A, AO')
            df_track = df_track[['Номер отправления ИМ', 'Номер накладной СДЭК']]
            df_track = df_track.rename(columns={'Номер отправления ИМ': 'parcel_numb',
                                    'Номер накладной СДЭК': 'track_numb'})
            if df_track['track_numb'].str.contains('#н/д').any() or df_track['track_numb'].isnull().any():
                flash(f'Ошибка загрузки треков: В колонке Треков есть пустые значения или #н/д, поправьте и загрузите заново', category='error')
            else:
                con_track = sl.connect('TRACKS.db')
                with con_track:
                    data = con_track.execute("select count(*) from sqlite_master where type='table' and name='tracks'")
                    for row in data:
                        # если таких таблиц нет
                        if row[0] == 0:
                            # создаём таблицу
                            with con_track:
                                con_track.execute("""
                                                        CREATE TABLE tracks (
                                                        ID INTEGER PRIMARY KEY AUTOINCREMENT,
                                                        parcel_numb VARCHAR(25) NOT NULL UNIQUE ON CONFLICT REPLACE,
                                                        track_numb VARCHAR(25)
                                                        );
                                                        """)
                    df_track.to_sql('tracks', con=con_track, if_exists='append', index=False)
                    con_track.commit()
                flash(f'Треки загружены')
                winsound.PlaySound('Snd\sample_load.wav', winsound.SND_FILENAME)

    return render_template('add_manifest_sample.html')

@app_svh.route('/add/vectors', methods=['GET', 'POST'])
def load_vectors():
    if request.method == 'POST':
        uploaded_file = request.files['file']
        filename = uploaded_file.filename
        logger.warning(filename)
        if filename != '':
            file_ext = os.path.splitext(filename)[1]
            if file_ext not in app_svh.config['UPLOAD_EXTENSIONS']:
                abort(400)
            uploaded_file.save(uploaded_file.filename)
            df_vector = pd.read_excel(filename, sheet_name=0, engine='openpyxl')
            df_vector = df_vector[['Партия', 'Номер пломбы', 'Номер отправления ИМ', 'Направление']]
            df_vector = df_vector.rename(columns={'Партия':'party_numb', 'Номер пломбы': 'parcel_plomb_numb', 'Номер отправления ИМ': 'parcel_numb',
                                    'Направление': 'vector'})
            if df_vector['vector'].str.contains('#н/д').any() or df_vector['parcel_numb'].isnull().any():
                flash(f'Ошибка загрузки Направлений: В колонке направлений есть пустые значения или #н/д, поправьте и загрузите заново', category='error')
            else:
                con_vector = sl.connect('VECTORS.db')
                with con_vector:
                    data = con_vector.execute("select count(*) from sqlite_master where type='table' and name='vectors'")
                    for row in data:
                        # если таких таблиц нет
                        if row[0] == 0:
                            # создаём таблицу
                            with con:
                                con.execute("""
                                                            CREATE TABLE vectors (
                                                            ID INTEGER PRIMARY KEY AUTOINCREMENT,
                                                            party_numb VARCHAR(20),
                                                            parcel_plomb_numb VARCHAR(20),
                                                            parcel_numb VARCHAR(20) NOT NULL UNIQUE ON CONFLICT REPLACE,
                                                            vector
                                                            );
                                                        """)
                    df_vector.to_sql('vectors', con=con_vector, if_exists='append', index=False)
                    con_vector.commit()
                flash(f'Направления загружены')
                winsound.PlaySound('Snd\sample_load.wav', winsound.SND_FILENAME)

    return render_template('add_manifest_sample.html')

@app_svh.route('/add/load_sample_manifest', methods=['GET', 'POST'])
def load_sample_manifest():
    if request.method == 'POST':
        uploaded_file = request.files['file']
        filename = uploaded_file.filename
        logger.warning(filename)
        if filename != '':
            file_ext = os.path.splitext(filename)[1]
            if file_ext not in app_svh.config['UPLOAD_EXTENSIONS']:
                abort(400)
            uploaded_file.save(uploaded_file.filename)
            df_track = pd.read_excel(filename, sheet_name=0, engine='openpyxl', usecols='A, AO')
            df_track = df_track[['Номер отправления ИМ', 'Номер накладной СДЭК']]
            df_track = df_track.rename(columns={'Номер отправления ИМ': 'parcel_numb',
                                    'Номер накладной СДЭК': 'track_numb'})
            print(df_track)
            if df_track['track_numb'].astype(str).str.contains('#н/д').any() or df_track['track_numb'].isnull().any():
                flash(f'Ошибка загрузки треков: В колонке Треков есть пустые значения или #н/д, поправьте и загрузите заново', category='error')
            else:
                con_track = sl.connect('TRACKS.db')
                with con_track:
                    data = con_track.execute("select count(*) from sqlite_master where type='table' and name='tracks'")
                    for row in data:
                        # если таких таблиц нет
                        if row[0] == 0:
                            # создаём таблицу
                            with con_track:
                                con_track.execute("""
                                                        CREATE TABLE tracks (
                                                        ID INTEGER PRIMARY KEY AUTOINCREMENT,
                                                        parcel_numb VARCHAR(25) NOT NULL UNIQUE ON CONFLICT REPLACE,
                                                        track_numb VARCHAR(25)
                                                        );
                                                        """)
                    df_track.to_sql('tracks', con=con_track, if_exists='append', index=False)
                    con_track.commit()
                    print(df_track)
            df = pd.read_excel(filename, sheet_name=0, engine='openpyxl')
            df = df[['Номер отправления ИМ', 'Номер пломбы', 'Наименование товара', '№ AWB', 'Общий Вес места (накладной)']]
            df = df.rename(columns={'Номер отправления ИМ': 'parcel_numb',
                                   'Номер пломбы': 'parcel_plomb_numb',
                                   'Наименование товара': 'goods',
                                   '№ AWB': 'party_numb', 'Общий Вес места (накладной)': 'parcel_weight'})
            logger.warning(df)
            df_group = df.groupby('parcel_numb', sort=False)['goods'].agg(','.join)
            logger.warning(df_group)
            df = pd.merge(df, df_group, how='left', left_on='parcel_numb', right_on='parcel_numb')
            df = df.drop_duplicates(subset=['parcel_numb'], keep='first')
            df['goods'] = df['goods_y']
            df = df[['parcel_numb', 'parcel_plomb_numb', 'goods', 'party_numb', 'parcel_weight']]
            df['custom_status'] = 'Unknown'
            df['custom_status_short'] = 'ИЗЪЯТИЕ'
            con = sl.connect('BAZA.db')
            # открываем базу
            with con:
                baza = con.execute("select count(*) from sqlite_master where type='table' and name='baza'")
                for row in baza:
                    # если таких таблиц нет
                    if row[0] == 0:
                        # создаём таблицу
                        with con:
                            con.execute("""
                                                    CREATE TABLE baza (
                                                    ID INTEGER PRIMARY KEY AUTOINCREMENT,
                                                    registration_numb VARCHAR(25),
                                                    party_numb VARCHAR(20),
                                                    parcel_numb VARCHAR(20) NOT NULL UNIQUE ON CONFLICT REPLACE,
                                                    parcel_plomb_numb VARCHAR(20),
                                                    parcel_weight FLOAT,
                                                    custom_status VARCHAR(400),
                                                    custom_status_short VARCHAR(8),
                                                    decision_date VARCHAR(20),
                                                    refuse_reason VARCHAR(400),
                                                    pallet VARCHAR(10),
                                                    zone VARCHAR(5),
                                                    VH_status VARCHAR(10),
                                                    goods
                                                    );
                                                """)
                df_to_append = pd.DataFrame()
                party_numb = df['party_numb'].values[0]
                party_numb_isalready_in = pd.read_sql(f"Select party_numb from baza where party_numb = '{party_numb}'", con)
                if party_numb_isalready_in.empty:
                    df.to_sql('baza', con=con, if_exists='append', index=False)
                else:
                    for parcel_numb in df['parcel_numb']:
                        row_isalready_in = pd.read_sql(f"Select * from baza where parcel_numb = '{parcel_numb}'", con)
                        row = df.loc[df['parcel_numb'] == parcel_numb]
                        if row_isalready_in.empty:
                            df_to_append = df_to_append.append(row)
                            logger.warning(row)
                        else:
                            goods = df.loc[df['parcel_numb'] == parcel_numb]['goods'].values[0]
                            parcel_weight = df.loc[df['parcel_numb'] == parcel_numb]['parcel_weight'].values[0]
                            logger.warning(goods)

                            party_numb_in_base = row_isalready_in['party_numb'].values[0]
                            if party_numb_in_base is None:
                                party_numb = df.loc[df['parcel_numb'] == parcel_numb]['party_numb'].values[0]
                                parcel_plomb_numb = str(df.loc[df['parcel_numb'] == parcel_numb]['parcel_plomb_numb'].values[0])
                                parcel_weight = df.loc[df['parcel_numb'] == parcel_numb]['parcel_weight'].values[0]
                                con.execute("Update baza set goods = ?, "
                                            "party_numb = ?, "
                                            "parcel_plomb_numb = ?,"
                                            "parcel_weight = ?"
                                            "where parcel_numb = ?",
                                            (goods, party_numb, parcel_plomb_numb, parcel_weight, parcel_numb))
                            else:
                                con.execute("Update baza set party_numb = ?, goods = ?, parcel_weight = ? where parcel_numb = ?",
                                            (party_numb, goods, parcel_weight, parcel_numb))
                    logger.warning(df_to_append)
                    df_to_append.to_sql('baza', con=con, if_exists='append', index=False)
                try:
                    df_vector = pd.read_excel(filename, sheet_name=0, engine='openpyxl')
                    df_vector = df_vector[['№ AWB', 'Номер пломбы', 'Номер отправления ИМ', 'Примечание']]
                    df_vector = df_vector.rename(columns={'№ AWB': 'party_numb', 'Номер пломбы': 'parcel_plomb_numb', 'Номер отправления ИМ': 'parcel_numb',
                                                          'Примечание': 'vector'})
                    con = sl.connect('VECTORS.db')
                    with con:
                        baza = con.execute("select count(*) from sqlite_master where type='table' and name='vectors'")
                        for row in baza:
                            # если таких таблиц нет
                            if row[0] == 0:
                                # создаём таблицу
                                with con:
                                    con.execute("""
                                                                CREATE TABLE vectors (
                                                                ID INTEGER PRIMARY KEY AUTOINCREMENT,
                                                                party_numb VARCHAR(20),
                                                                parcel_plomb_numb VARCHAR(20),
                                                                parcel_numb VARCHAR(20) NOT NULL UNIQUE ON CONFLICT REPLACE,
                                                                vector
                                                                );
                                                            """)
                        df_vector.to_sql('vectors', con=con, if_exists='append', index=False)
                    flash(f'Инфо по последней миле загружена')
                except Exception as e:
                    flash(f'Нет инфо по последней миле {e}')

                flash(f'Шаблон загружен')
                winsound.PlaySound('Snd\sample_load.wav', winsound.SND_FILENAME)
    return render_template('add_manifest_sample.html')

@app_svh.route('/add/decisions', methods=['GET', 'POST'])
def load_decisions():
    logger.warning('OK')
    if request.method == 'POST':
        uploaded_file = request.files['file']
        filename = uploaded_file.filename
        logger.warning(filename)
        if filename != '':
            file_ext = os.path.splitext(filename)[1]
            if file_ext not in app_svh.config['UPLOAD_EXTENSIONS']:
                abort(400)
            uploaded_file.save(uploaded_file.filename)
            df = pd.read_excel(filename, sheet_name=0, engine='openpyxl')
            df['Вес брутто'] = df['Вес брутто'].replace(',', '.', regex=True).astype(float)
            df['Статус_ТО'] = df['Статус ТО']
            df['Статус_ТО'] = df['Статус_ТО'].replace(to_replace='10-в',
                                                      value='В', regex=True)
            df['Статус_ТО'] = df['Статус_ТО'].replace(to_replace='32-в',
                                                      value='В', regex=True)
            df['Статус_ТО'] = df['Статус_ТО'].replace(to_replace='Выпуск товаров без уплаты таможенных платежей', value='ВЫПУСК', regex=True)
            df['Статус_ТО'] = df['Статус_ТО'].replace(to_replace='Выпуск товаров разрешен, таможенные платежи уплачены', value='ВЫПУСК', regex=True)
            for cel in df['Статус_ТО']:
                if cel != 'ВЫПУСК':
                    df.loc[df['Статус_ТО'] == cel, 'Статус_ТО'] = 'ИЗЪЯТИЕ'
            df['Пломба'] = df['Пломба'].astype(str)

            df = df.rename(columns={'Рег. номер': 'registration_numb', 'Общая накладная': 'party_numb',
                                    'Трек-номер': 'parcel_numb', 'Пломба': 'parcel_plomb_numb', 'Вес брутто': 'parcel_weight',
                                    'Статус ТО': 'custom_status', 'Статус_ТО': 'custom_status_short', 'Дата решения': 'decision_date',
                                    'Причина отказа ТО': 'refuse_reason'})
            con = sl.connect('BAZA.db')
            # открываем базу
            with con:
                baza = con.execute("select count(*) from sqlite_master where type='table' and name='baza'")
                for row in baza:
                    # если таких таблиц нет
                    if row[0] == 0:
                        # создаём таблицу
                        with con:
                            con.execute("""
                                        CREATE TABLE baza (
                                        ID INTEGER PRIMARY KEY AUTOINCREMENT,
                                        registration_numb VARCHAR(25) NOT NULL,
                                        party_numb VARCHAR(20),
                                        parcel_numb VARCHAR(20) NOT NULL UNIQUE ON CONFLICT REPLACE,
                                        parcel_plomb_numb VARCHAR(20),
                                        parcel_weight FLOAT,
                                        custom_status VARCHAR(400),
                                        custom_status_short VARCHAR(8),
                                        decision_date VARCHAR(20),
                                        refuse_reason VARCHAR(400),
                                        pallet VARCHAR(10),
                                        zone VARCHAR(5),
                                        VH_status VARCHAR(10),
                                        goods
                                        );
                                    """)
                df_to_append = pd.DataFrame()
                for parcel_numb in df['parcel_numb']:
                    row_isalready_in = pd.read_sql(f"Select * from baza where parcel_numb = '{parcel_numb}'", con)
                    row = df.loc[df['parcel_numb'] == parcel_numb]
                    if row_isalready_in.empty:
                        # row.to_sql('baza', con=con, if_exists='append', index=False)
                        df_to_append = df_to_append.append(row)
                        logger.warning(row)
                    else:
                        logger.warning(row)
                        custom_status_short = df.loc[df['parcel_numb'] == parcel_numb]['custom_status_short'].values[0]
                        logger.warning(custom_status_short)
                        custom_status = df.loc[df['parcel_numb'] == parcel_numb]['custom_status'].values[0]
                        logger.warning(custom_status)
                        registration_numb = df.loc[df['parcel_numb'] == parcel_numb]['registration_numb'].values[0]
                        logger.warning(registration_numb)
                        parcel_weight = df.loc[df['parcel_numb'] == parcel_numb]['parcel_weight'].values[0]
                        decision_date = df.loc[df['parcel_numb'] == parcel_numb]['decision_date'].values[0]
                        refuse_reason = df.loc[df['parcel_numb'] == parcel_numb]['refuse_reason'].values[0]

                        con.execute(f"Update baza set registration_numb = '{registration_numb}',"
                                    f" custom_status = '{custom_status}',"
                                    f" custom_status_short = '{custom_status_short}',"
                                    f" parcel_weight = '{parcel_weight}',"
                                    f" decision_date = '{decision_date}',"
                                    f" refuse_reason = '{refuse_reason}' where parcel_numb = '{parcel_numb}'")
                        row_isalready_in = pd.read_sql(f"Select * from baza where parcel_numb = '{parcel_numb}'", con)
                        logger.warning(row_isalready_in)
                    # df_isalready_in = df_isalready_in.append(row_isalready_in)
                logger.warning(df_to_append)
                df_to_append.to_sql('baza', con=con, if_exists='append', index=False)
                flash(f'Решения загружены')
                winsound.PlaySound('Snd\esheniya_zagruzhenu.wav', winsound.SND_FILENAME)

    return render_template('add_decisions.html')

@app_svh.route('/plomb', methods=['GET'])
@jwt_required()
def plomb():
    return render_template('plomb.html')

@app_svh.route('/pallet', methods=['GET'])
def pallet():
    return render_template('pallet.html')

@app_svh.route('/search/plomb', methods=['GET'])
@jwt_required()
def plomb_searh():
    try:
        parcel_plomb_numb = request.args.get('parcel_plomb_numb')
        return render_template('plomb_search.html', search=parcel_plomb_numb)
    except Exception as e:
        return {'message': str(e)}, 400

@app_svh.route('/search/parcel', methods=['GET'])
def parcel_searh():
    try:
        parcel_numb = request.args.get('parcel_numb')
        return render_template('parcel_search.html', search=parcel_numb)
    except Exception as e:
        return {'message': str(e)}, 400

@app_svh.route('/search/parcel_goods_info', methods=['GET'])
def parcel_goods_info():
    try:
        parcel_numb = request.args.get('parcel_numb')
        return render_template('parcel_goods.html', search=parcel_numb)
    except Exception as e:
        return {'message': str(e)}, 400

@app_svh.route('/search/plomb_info', methods=['POST', 'GET'])
def get_plomb_info():
    parcel_plomb_numb = request.form['parcel_plomb_numb']
    title_status = ''
    vector = None
    try:
        con = sl.connect('BAZA.db')
        with con:
            data = con.execute('SELECT * FROM plombs WHERE parcel_plomb_numb=?', (parcel_plomb_numb,)).fetchone()
            try:
                if data is None:
                    pass
                else:
                    con.execute(f"Update plombs set parcel_plomb_status = 'Принят' where parcel_plomb_numb = '{parcel_plomb_numb}'")
            except:
                pass
            df_search_plomb = pd.read_sql(f"SELECT * FROM baza where parcel_plomb_numb = '{parcel_plomb_numb}' COLLATE NOCASE", con)
            df_search_plomb['№'] = np.arange(len(df_search_plomb))[::+1] + 1
            first_parcel = df_search_plomb['parcel_numb'].values[0]
            print(first_parcel)
            con_vect = sl.connect('VECTORS.db')
            with con_vect:
                try:
                    vector = con_vect.execute(
                        f"SELECT vector FROM VECTORS where parcel_plomb_numb = '{parcel_plomb_numb}'").fetchone()[0]
                    flash(f'{vector}!', category='success')
                    print(vector)
                except:
                    try:
                        vector = con_vect.execute(
                            f"SELECT vector FROM VECTORS where parcel_numb = '{first_parcel}'").fetchone()[
                            0]
                        flash(f'{vector}!', category='success')
                        print(vector)
                    except Exception as e:
                        print(e)
            try:
                df_search_plomb['parcel_weight'] = df_search_plomb['parcel_weight'].round(3)
            except:
                pass
            df_search_plomb = df_search_plomb[['№', 'parcel_numb', 'custom_status_short', 'VH_status', 'parcel_weight', 'goods']]
            df_search_plomb = df_search_plomb.rename(columns={'parcel_numb': 'Трек-номер',
                                                              'custom_status_short': 'Статус',
                                                              'parcel_weight': 'Вес',
                                                              'VH_status': 'ВХ', 'goods': 'Товары'})
            df_search_plomb['Товары'] = df_search_plomb['Товары'].str.slice(0, 50)
            df_parc_quont = len(df_search_plomb)
            df_parc_refuse_quont = len(df_search_plomb.loc[df_search_plomb['Статус'] == 'ИЗЪЯТИЕ'])
            if df_search_plomb.empty:
                flash(f'Пломба не найдена!', category='error')
                winsound.PlaySound('Snd\Snd_NoPlomb.wav', winsound.SND_FILENAME)
            elif df_search_plomb['Статус'].str.contains("ИЗЪЯТИЕ").any():
                flash(f'Открываем место!', category='error')
                winsound.PlaySound('Snd\Snd_Open_Bag.wav', winsound.SND_FILENAME)
                title_status = 'ИЗЪЯТИЕ'
            else:
                title_status = 'ВЫПУСК'
            df_search_plomb.fillna("", inplace=True)
            def highlight_last_row(df_search_plomb):
                    return ['background-color: #FF0000' if 'ИЗЪЯТИЕ' in str(i) else '' for i in df_search_plomb]
            def highlight_last_row_2(df_search_plomb):
                    return ['background-color: #7CFC00' if 'На ВХ' in str(i) else '' for i in df_search_plomb]
            if 'На ВХ' in df_search_plomb['ВХ'].values:
                df_search_plomb = df_search_plomb.style.apply(highlight_last_row_2).hide_index()
                logger.warning('VH')
            else:
                df_search_plomb = df_search_plomb.style.apply(highlight_last_row).hide_index()
            df_search_plomb = df_search_plomb.format(precision=2)
        object_name = parcel_plomb_numb
        logger.warning(object_name)
        comment = 'Отбор пломб: Просмотрена пломба'
        insert_user_action(object_name, comment)
    except Exception as e:
        winsound.PlaySound('Snd\Snd_NoPlomb.wav', winsound.SND_FILENAME)
        flash(f'Пломба не найдена! {e}', category='error')
        return render_template('plomb_info.html')


    return render_template('plomb_info.html', tables=[style + df_search_plomb.to_html(index=False,
                                                                              float_format='{:2,.2f}'.format)],
                           titles=['Информация о посылке'],
                           parcel_plomb_numb=parcel_plomb_numb, df_parc_quont=df_parc_quont,
                           df_parc_refuse_quont=df_parc_refuse_quont, title_status=title_status,
                              vector=vector)


@app_svh.route('/party/plomb_come/<party_numb>', methods=['POST', 'GET'])
def get_plomb_come(party_numb):
    try:
        con = sl.connect('BAZA.db')
        with con:
            baza = con.execute("select count(*) from sqlite_master where type='table' and name='plombs'")
            for row in baza:
                # если таких таблиц нет
                if row[0] == 0:
                    # создаём таблицу
                    with con:
                        con.execute("""
                                                CREATE TABLE plombs (
                                                ID INTEGER PRIMARY KEY AUTOINCREMENT,
                                                party_numb VARCHAR(20),
                                                parcel_plomb_numb VARCHAR(20),
                                                parcel_plomb_status VARCHAR(8)
                                                );
                                            """)
            row_isalready_in_plombs = pd.read_sql(f"Select * from plombs where party_numb = '{party_numb}'", con)
            if row_isalready_in_plombs.empty:
                df = pd.read_sql(f"SELECT * FROM baza where party_numb = '{party_numb}' COLLATE NOCASE", con)
                df = df.drop_duplicates(subset='parcel_plomb_numb')
                df['parcel_plomb_status'] = 'Ожидаем'
                df = df[['parcel_plomb_numb', 'party_numb', 'parcel_plomb_status']]
                index = False
                print(df)
                df.to_sql('plombs', con=con, if_exists='append', index=False)
            else:
                df = pd.read_sql(f"SELECT * FROM plombs where party_numb = '{party_numb}' COLLATE NOCASE", con)
                df = df.drop_duplicates(subset='parcel_plomb_numb').loc[df['parcel_plomb_numb'] != '']
                df['№'] = np.arange(len(df))[::+1] + 1
                df = df[['№', 'parcel_plomb_numb', 'party_numb', 'parcel_plomb_status']]
                index = False
                print(df, 'j')
            object_name = party_numb
            comment = f'Приемка по местам: Открыта приемка по партии {party_numb})'
            insert_user_action(object_name, comment)
    except Exception as e:
        return {'message': str(f'{e}')}, 400
    return render_template('party_plombs.html', tables=[style + df.to_html(classes='mystyle', index=index,
                                                                   float_format='{:2,.2f}'.format)],
                           titles=['Информация'],
                           party_numb=party_numb)

@app_svh.route('/party/plomb_come_work/<party_numb>', methods=['POST', 'GET'])
def get_plomb_come_work(party_numb):
    parcel_plomb_numb = request.form['parcel_plomb_numb']
    con_vect = sl.connect('VECTORS.db')
    vector = None
    with con_vect:
        try:
            vector = con_vect.execute(
                f"SELECT vector FROM VECTORS where parcel_plomb_numb = '{parcel_plomb_numb}'").fetchone()[0]
            flash(f'{vector}!', category='success')
            print(vector)
        except Exception as e:
            print(e)
    con = sl.connect('BAZA.db')
    df_not_custom = pd.read_excel('неподанные.xlsx', sheet_name=0, engine='openpyxl', header=None)
    if df_not_custom[0].astype(str).str.contains(parcel_plomb_numb).any():
        flash(f'НЕПОДАНА Пломба {parcel_plomb_numb} !!', category='error')
        winsound.PlaySound('Snd\Snd_CancelIssue.wav', winsound.SND_FILENAME)
    with con:
        df_check_plomb = pd.read_sql(f"SELECT * FROM plombs where parcel_plomb_numb = '{parcel_plomb_numb}' COLLATE NOCASE", con)
        if df_check_plomb.empty:
            flash(f'Пломба {parcel_plomb_numb} не найдена в партии!', category='error')
            winsound.PlaySound('Snd\Snd_CancelIssue.wav', winsound.SND_FILENAME)
        else:
            df_check_refuse = pd.read_sql(
                f"SELECT parcel_plomb_numb FROM baza where parcel_plomb_numb = '{parcel_plomb_numb}' and custom_status_short = 'ИЗЪЯТИЕ'",
                con)
            if df_check_refuse.empty:
                flash(f'ВЫПУСК МЕСТА', category='success')
                winsound.PlaySound('Snd\zvuk-vezeniya.wav', winsound.SND_FILENAME)
            con.execute(f"Update plombs set parcel_plomb_status = 'Принят' where parcel_plomb_numb = '{parcel_plomb_numb}'")
        df = pd.read_sql(f"SELECT * FROM plombs where party_numb = '{party_numb}' COLLATE NOCASE", con)
        df = df.drop_duplicates(subset='parcel_plomb_numb').loc[df['parcel_plomb_numb'] != '']
        df['№'] = np.arange(len(df))[::+1] + 1
        df = df[['№', 'parcel_plomb_numb', 'party_numb', 'parcel_plomb_status']]
        if df.loc[df['parcel_plomb_status'] == 'Ожидаем'].empty:
            flash(f'Все места найдены', category='success')
            winsound.PlaySound('Snd\se_mesta_naid.wav', winsound.SND_FILENAME)
        index = False
        quont_all_plombs = len(df)
        quont_plomb_done = len(df.loc[df['parcel_plomb_status'] == 'Принят'])
        quon_not_done = quont_all_plombs - quont_plomb_done
        object_name = parcel_plomb_numb
        comment = f'Приемка по местам: Пломба принята (Партия: {party_numb})'
        insert_user_action(object_name, comment)
    return render_template('party_plombs_work.html', tables=[style + df.to_html(classes='mystyle', index=index,
                                                                           float_format='{:2,.2f}'.format)],
                           titles=['Информация'],
                           party_numb=party_numb, quont_all_plombs=quont_all_plombs,
                           quont_plomb_done=quont_plomb_done, quon_not_done=quon_not_done, parcel_plomb_numb=parcel_plomb_numb,
                           vector=vector)
done_parcels = pd.DataFrame()

@app_svh.route('/search/parcel_info', methods=['POST', 'GET'])
def get_parcel_info():
    global done_parcels
    global done_parcel
    audiofile = 'None'
    parcel_numb = request.form['parcel_numb']
    try:
        con = sl.connect('BAZA.db')
        with con:
            df_parc_events = pd.read_sql(f"SELECT * FROM baza where parcel_numb = '{parcel_numb}'", con)
            parcel_plomb_numb = df_parc_events['parcel_plomb_numb'].values[0]
            df_parcel_plomb_info = pd.read_sql(f"SELECT * FROM baza where parcel_plomb_numb = '{parcel_plomb_numb}'",
                                               con)
            df_parc_quont = len(df_parcel_plomb_info)
            df_parcel_plomb_refuse_info = df_parcel_plomb_info.loc[
                df_parcel_plomb_info['custom_status_short'] == 'ИЗЪЯТИЕ']
            df_parc_refuse_quont = len(df_parcel_plomb_refuse_info)
            df_parcel_plomb_refuse_info['№1'] = np.arange(len(df_parcel_plomb_refuse_info))[::+1] + 1
            df_parcel_plomb_refuse_info = df_parcel_plomb_refuse_info[
                ['№1', 'parcel_numb', 'custom_status_short', 'parcel_plomb_numb',
                 'VH_status', 'parcel_weight', 'goods']]
            df_parcel_plomb_refuse_info = df_parcel_plomb_refuse_info.rename(
                columns={'№1': '№', 'parcel_numb': 'Трек-номер', 'custom_status_short': 'Статус',
                         'parcel_plomb_numb': 'Пломба', 'VH_status': 'ВХ', 'parcel_weight': 'вес', 'goods': 'Товары'})
            df_parcel_plomb_refuse_info['Товары'] = df_parcel_plomb_refuse_info['Товары'].str.slice(0, 200)
            if parcel_plomb_numb == '':
                df_parcel_plomb_refuse_info = pd.DataFrame()
            df_parc_events['№'] = np.arange(len(df_parc_events))[::+1] + 1
            df_parc_events = df_parc_events[['parcel_numb', 'custom_status_short', 'parcel_plomb_numb', 'VH_status']]
            df_parc_events = df_parc_events.rename(
                columns={'parcel_numb': 'Трек-номер', 'custom_status_short': 'Статус',
                         'parcel_plomb_numb': 'Пломба', 'VH_status': 'ВХ'})
            if df_parc_events.empty:
                flash(f'Посылка не найдена!', category='error')
                #winsound.PlaySound('Snd\Snd_Parcel_Not_Found.wav', winsound.SND_FILENAME)
                audiofile = 'Snd_Parcel_Not_Found.wav'
            elif 'На ВХ' in df_parc_events['ВХ'].values:
                flash(f'Уже размещено', category='success')
            if df_parc_events.loc[df_parc_events['Статус'] == 'ИЗЪЯТИЕ'].empty:
                flash(f'ВЫПУСК')
            else:
                flash(f'ИЗЪЯТИЕ на склад', category='error')
                #winsound.PlaySound('Snd\Snd_CancelIssue.wav', winsound.SND_FILENAME)
                audiofile = 'Snd_CancelIssue.wav'
            done_parcels = done_parcels.append(df_parc_events).drop_duplicates(subset=['Трек-номер'], keep='first')
            done_parcels['№'] = np.arange(len(done_parcels))[::+1] + 1
            done_parcels_plomb_info = done_parcels.loc[done_parcels['Пломба'] == parcel_plomb_numb]
            if len(done_parcels_plomb_info) == df_parc_quont:
                flash(f'Место завершено', category='success')
                winsound.PlaySound('Snd\se_mesta_naid.wav', winsound.SND_FILENAME)
            done_parcels_styl = done_parcels.reset_index()
            done_parcels_styl = done_parcels_styl.drop('index', axis=1)
            done_parcels_styl = done_parcels_styl[
                ['№', 'Трек-номер', 'Статус', 'Пломба', 'ВХ']].drop_duplicates(subset=['Трек-номер'], keep='first')
            #trigger_color1 = df_parc_events['Статус']
            #trigger_color2 = df_parc_events['ВХ']
            done_parcels_styl.fillna("", inplace=True)
            df_parc_events.fillna("", inplace=True)
            df_parcel_plomb_refuse_info.fillna("", inplace=True)
            if len(df_parcel_plomb_refuse_info) == len(done_parcels_styl.loc[(
                    (done_parcels_styl.Статус == 'ИЗЪЯТИЕ') & (done_parcels_styl.Пломба == parcel_plomb_numb))]):
                flash(f'Все отказы найдены!', category='success')
            #def highlight_RED(df_parc_events):
            #    return [
            #        'background-color: #FF0000' if 'ИЗЪЯТИЕ' in str(df_parc_events) else '']  # for i in df_parc_events

            #def highlight_GREEN(df_parc_events):
            #    return [
            #        'background-color: #7CFC00' if 'На ВХ' in str(df_parc_events) else '']  # for i in df_parc_events

            def highlight_last_row_2(done_parcels_styl):
                return ['background-color: #FF0000' if 'ИЗЪЯТИЕ' in str(i) else '' for i in done_parcels_styl]

            if 'ИЗЪЯТИЕ' in done_parcels_styl['Статус'].values:
                done_parcels_styl = done_parcels_styl.style.apply(highlight_last_row_2).hide_index()

        object_name = parcel_numb
        comment = 'Отбор посылок: Просмотрена посылка'
        insert_user_action(object_name, comment)
    except Exception as e:
        flash(f'Посылка не найдена! {e}', category='error')
        #winsound.PlaySound('Snd\Snd_Parcel_Not_Found.wav', winsound.SND_FILENAME)
        audiofile = 'Snd_CancelIssue.wav'
        return render_template('parcel_info.html', audiofile=audiofile)
        #return {'message': str(f'{e}  "PARCEL NOT FOUND"')}, 400
        pass
    return render_template('parcel_info.html', tables=[
                                                       df_parcel_plomb_refuse_info.to_html(classes='mystyle',
                                                                                           index=False,
                                                                                           float_format='{:2,.2f}'.format),
                                                       done_parcels_styl.to_html(classes='mystyle', index=False,
                                                                                 float_format='{:2,.2f}'.format)],
                           titles=['na', 'Нужно найти в мешке:', '\n\nОтработанные:'],
                           parcel_numb=parcel_numb, df_parc_quont=df_parc_quont,
                           df_parc_refuse_quont=df_parc_refuse_quont, parcel_plomb_numb=parcel_plomb_numb, audiofile=audiofile)


@app_svh.route('/search/clean', methods=['POST'])
def clean_working_place():
    global done_parcels
    global done_parcel
    done_parcels_VH = done_parcels.loc[done_parcels['Статус'] == 'ИЗЪЯТИЕ']
    con = sl.connect('BAZA.db')
    with con:
        for parcel_numb in done_parcels_VH['Трек-номер']:
            con.execute(f"Update baza set VH_status = 'На ВХ', parcel_plomb_numb = ''  where parcel_numb = '{parcel_numb}'")
    con.commit()
    con.close()
    parcel_plomb_numb = done_parcels_VH['Пломба'].values[0]  #done_parcels
    done_parcels = pd.DataFrame()
    done_parcel = None
    object_name = parcel_plomb_numb
    comment = 'Отбор посылок: Завершено место'
    insert_user_action(object_name, comment)
    return render_template('parcel_info.html')
parcel_plomb_numb = None
@app_svh.route('/search/manifest', methods=['GET'])
def make_manifest():
    try:
        global parcel_plomb_numb
        return render_template('plomb_search_manifest.html', search=parcel_plomb_numb)
    except Exception as e:
        return {'message': str(e)}, 400

df_plomb_to_manifest = pd.DataFrame()
df_plomb_to_manifest_total = pd.DataFrame()
df_to_manifest_HTML = pd.DataFrame()
df_refuse_plombs = pd.DataFrame()
@app_svh.route('/search/plomb_to_manifest', methods=['POST', 'GET'])
def plomb_to_manifest():
    global df_plomb_to_manifest
    global df_plomb_to_manifest_total
    global parcel_plomb_numb
    global df_to_manifest_HTML
    global df_refuse_plombs
    parcel_plomb_numb = request.form['parcel_plomb_numb']
    print(parcel_plomb_numb)
    if df_plomb_to_manifest.empty:
        df_plomb_to_manifest = pd.DataFrame({'parcel_plomb_numb': [parcel_plomb_numb]})
    try:
        con = sl.connect('BAZA.db')
        with con:
            df_plomb_to_manifest = pd.read_sql(f"SELECT * FROM baza where parcel_plomb_numb = '{parcel_plomb_numb}'",
                                               con)
            logger.warning(df_plomb_to_manifest['custom_status_short'])
            # if that is parcel:
            if df_plomb_to_manifest.empty:
                df_plomb_to_manifest = pd.read_sql(f"SELECT * FROM baza where parcel_numb = '{parcel_plomb_numb}'", con)
                df_all_pallet_plombs = df_plomb_to_manifest.drop_duplicates(subset=['parcel_plomb_numb', 'parcel_numb'],
                                                                                                                     keep='first')
                df_all_pallet_plombs['Тип'] = "Посылка-место"
                df_plomb_to_manifest_refuse_parcel = df_all_pallet_plombs['parcel_numb'].values[0]
                if not df_plomb_to_manifest.loc[df_plomb_to_manifest['custom_status_short'] == 'ИЗЪЯТИЕ'].empty:
                    flash(f'ПОСЫЛКА {df_plomb_to_manifest_refuse_parcel} СО СТАТУСОМ ИЗЪЯТИЕ (НЕ ВЫПУЩЕНА!!!)', category='error')
                    winsound.PlaySound('Snd\imanie izyyatie.wav', winsound.SND_FILENAME)
            # if that is plomb:
            else:
                df_plomb_to_manifest = df_plomb_to_manifest.drop_duplicates(subset='parcel_plomb_numb', keep='first')
                df_plomb_to_manifest['quont_plomb'] = None
                df_plomb_to_manifest['Тип'] = "Пломба"
                df_plomb_to_manifest['parcel_numb'] = ''
                pallet = df_plomb_to_manifest['pallet'].values[0]
                df_all_pallet_plombs = pd.read_sql(f"SELECT * FROM baza where pallet = '{pallet}'", con)
                df_all_pallet_plombs_refuse = df_all_pallet_plombs.loc[df_all_pallet_plombs['custom_status_short'] == 'ИЗЪЯТИЕ']
                if not df_all_pallet_plombs_refuse.empty:
                    df_all_pallet_plombs_refuse.drop_duplicates(subset='parcel_plomb_numb', keep='first')
                    df_all_pallet_plombs_refuse = df_all_pallet_plombs_refuse['parcel_plomb_numb'].drop_duplicates().to_string(index=False)
                    flash(f'НА ПАЛЛЕТЕ {pallet} ЕСТЬ ОТКАЗНЫЕ (НЕ ОТРАБОТАННЫЕ) ПЛОМБЫ: {df_all_pallet_plombs_refuse}', category='error')
                    winsound.PlaySound('Snd\imanie izyyatie.wav', winsound.SND_FILENAME)
                df_all_pallet_plombs = df_all_pallet_plombs.drop_duplicates(subset='parcel_plomb_numb', keep='first')
                df_all_pallet_plombs['Тип'] = "Пломба"
                df_all_pallet_plombs['parcel_numb'] = ''
            df_all_pallet_plombs['quont_plomb'] = len(df_all_pallet_plombs)
            if df_all_pallet_plombs.empty:
                df_plomb_to_manifest_total = df_plomb_to_manifest_total.append(df_plomb_to_manifest)
                object_name = parcel_plomb_numb
                comment = 'Манифест огрузки: Пломба добавленна для отгрузки'
                insert_user_action(object_name, comment)
            else:
                df_plomb_to_manifest_total = df_plomb_to_manifest_total.append(df_all_pallet_plombs)
                object_name = pallet
                comment = 'Манифест огрузки: Паллет добавлен для отгрузки'
                insert_user_action(object_name, comment)
            df_plomb_to_manifest_total = df_plomb_to_manifest_total.drop_duplicates(subset=['parcel_plomb_numb', 'parcel_numb'], keep='first')
            df_plomb_to_manifest_total['№1'] = np.arange(len(df_plomb_to_manifest_total))[::+1] + 1

            df_to_manifest_HTML = df_plomb_to_manifest_total[['№1', 'parcel_plomb_numb', 'pallet', 'party_numb', 'quont_plomb', 'Тип', 'parcel_numb']]

            df_to_manifest_HTML = df_to_manifest_HTML.rename(columns={'№1': '№', 'parcel_plomb_numb': 'Пломба',
                                                                      'pallet': 'Паллет', 'party_numb': 'Партия',
                                                                      'quont_plomb': 'кол. пломб', 'parcel_numb': 'Трек'})
            df_to_manifest_HTML.fillna("", inplace=True)
        object_name = parcel_plomb_numb
        comment = 'Манифест огрузки: Пломба добавленна для отгрузки'
        insert_user_action(object_name, comment)
    except Exception as e:
        flash(f'Пломба не найдена!', category='error')
        winsound.PlaySound('Snd\Snd_NoPlomb.wav', winsound.SND_FILENAME)
        return {'message': str(e)}, 400
    return render_template('plomb_info_manifest.html', tables=[style + df_to_manifest_HTML.to_html(classes='mystyle', index=False)],
                           titles=['n/a', 'Добавленные'],
                           parcel_plomb_numb=parcel_plomb_numb)

@app_svh.route('/search/clean_working_place_manifest', methods=['POST'])
def clean_working_place_manifest():
    global df_plomb_to_manifest_total
    global df_to_manifest_HTML
    object_name = None
    comment = 'Манифест огрузки: Таблица отгрузки очищена'
    insert_user_action(object_name, comment)
    df_plomb_to_manifest_total = pd.DataFrame()
    return render_template('plomb_info_manifest.html')

def manifest_to_xls(df_manifest_total):
    now = datetime.datetime.now().strftime("%d.%m.%Y")
    now_month = datetime.datetime.now().strftime("%m.%Y")
    now_time = datetime.datetime.now().strftime("%d.%m.%Y %H-%M")
    delta = datetime.timedelta(hours=-10, minutes=0)
    event_date = datetime.datetime.now() + delta
    event_date = event_date.strftime("%Y-%m-%d %H:%M:%S")
    #insert_event_API_test(df, event_date)
    party_numb = df_manifest_total['party_numb'].values[0]
    with open("manif_number.txt") as manif_number_file:
        manif_number = int(manif_number_file.read())+1
    with open("manif_number.txt", 'w') as manif_number_file:
        manif_number_file.write(str(manif_number))
    if not df_manifest_total['custom_status_short'].astype(str).str.contains('ИЗЪЯТИЕ').any():
        df_manifest_total = df_manifest_total.sort_values(by='parcel_plomb_numb')
        df_manifest_total['№ п.п.'] = np.arange(len(df_manifest_total))[::+1] + 1
        df_manifest_total['Номер индивидуальной     накладной'] = df_manifest_total['parcel_numb']
        df_manifest_total['Трекинг'] = df_manifest_total['parcel_numb']
        df_manifest_total['Номер накладной'] = df_manifest_total['party_numb']
        df_manifest_total['Вес посылки'] = df_manifest_total['parcel_weight']
        quont_of_plomb = len(df_manifest_total.drop_duplicates(subset='parcel_plomb_numb'))
        group_Weight_df = df_manifest_total.groupby('parcel_plomb_numb')['parcel_weight'].sum()
        group_Weight_df = group_Weight_df.rename('Вес_мешка', inplace=True)
        df_total = pd.merge(df_manifest_total, group_Weight_df, how='left', left_on='parcel_plomb_numb',
                            right_on='parcel_plomb_numb')

        list_of_parties1 = df_manifest_total['Номер накладной'].drop_duplicates().str[-8:].to_list()
        list_of_parties = ' '.join(list_of_parties1)[:180]

        writer = pd.ExcelWriter('test.xlsx', engine='xlsxwriter')
        df_total.to_excel(writer, sheet_name='Sheet1', index=False)
        writer.save()

        df_parcels = df_total[['parcel_numb']].drop_duplicates()
        print(df_parcels)
        con_vect = sl.connect('VECTORS.db')
        data = con_vect.execute("select count(*) from sqlite_master where type='table' and name='parcels_to_manifest'")
        for row in data:
            # если таких таблиц нет
            if row[0] == 0:
                # создаём таблицу
                with con_vect:
                    con_vect.execute("""
                                                            CREATE TABLE parcels_to_manifest (
                                                            ID INTEGER PRIMARY KEY AUTOINCREMENT,
                                                            parcel_numb VARCHAR(20)
                                                            );
                                                        """)
        df_parcels.to_sql('parcels_to_manifest', con=con_vect, index=False, if_exists='replace')
        query = """SELECT vectors.parcel_numb, vectors.vector
                                FROM vectors                          
                                JOIN parcels_to_manifest ON parcels_to_manifest.parcel_numb = vectors.parcel_numb
                                """
        data = con_vect.execute(query).fetchall()
        for row in data:
            print(row)
        df_manif_vectors = pd.read_sql(query, con_vect)
        print(df_manif_vectors)
        df_total = pd.merge(df_total, df_manif_vectors, how='left', left_on='parcel_numb' ,right_on='parcel_numb')
        #df_total['Мест'] = df_total.Вес_мешка.eq(df_total.Вес_мешка.shift()).astype('str')
        df_total['Мест'] = df_total.parcel_plomb_numb.eq(df_total.parcel_plomb_numb.shift()).astype('str')
        df_total['Вес_мешка'] = np.where(df_total['Мест'] == 'True', '0', df_total['Вес_мешка'])
        df_total['parcel_plomb_numb'] = np.where(df_total['Мест'] == 'True', '', df_total['parcel_plomb_numb'])

        df_total['Вес мешка'] = df_total['Вес_мешка'].astype(float)
        df_total['Направление'] = df_total['vector']
        df_total = df_total.reindex(columns=['№ п.п.', 'Номер индивидуальной     накладной',
                                             'Трекинг', 'Номер накладной', 'Вес посылки',
                                             'parcel_plomb_numb', 'Вес мешка', 'Направление'])


        df_total = df_total.drop_duplicates(subset='Трекинг', keep='first')
        total_weight = df_manifest_total['parcel_weight'].sum().round(3)


        writer = pd.ExcelWriter('system.xlsx', engine='xlsxwriter')
        df_total.to_excel(writer, sheet_name='Sheet1', index=False)
        for column in df_total:
            column_width = max(df_total[column].astype(str).map(len).max(), len(column))
            col_idx = df_total.columns.get_loc(column)
            writer.sheets['Sheet1'].set_column(col_idx, col_idx, column_width)
            writer.sheets['Sheet1'].set_column(0, 3, 10)
            writer.sheets['Sheet1'].set_column(1, 3, 20)
            writer.sheets['Sheet1'].set_column(2, 3, 20)
            writer.sheets['Sheet1'].set_column(3, 3, 20)
            writer.sheets['Sheet1'].set_column(4, 3, 30)
            writer.sheets['Sheet1'].set_column(5, 3, 20)

        writer.save()
        wb2 = openpyxl.load_workbook('system.xlsx')
        ws2 = wb2.active
        ws2.insert_rows(1, 2)

        wb = openpyxl.load_workbook('Akt.xlsx')
        ws = wb.active
        ws['B1'].value = f'{now} а/м М246ВР пломба № 52388385'

        head = "A1:H4"  # Заголовок таблицы, в котором есть объединенные ячейки
        for _range in ws.merged_cells.ranges:
            boundaries = range_boundaries(str(_range))
            ws2.merge_cells(start_column=boundaries[0], start_row=boundaries[1],
                            end_column=boundaries[2], end_row=boundaries[3])
        for row_number, row in enumerate(ws[head]):
            for col_number, cell in enumerate(row):
                ws2.cell(row_number + 1, col_number + 1, cell.value)
                if cell.has_style:
                    ws2.cell(row_number + 1, col_number + 1).font = copy(cell.font)
                    ws2.cell(row_number + 1, col_number + 1).fill = copy(cell.fill)
                    ws2.cell(row_number + 1, col_number + 1).border = copy(cell.border)
                    ws2.cell(row_number + 1, col_number + 1).number_format = copy(cell.number_format)
                    ws2.cell(row_number + 1, col_number + 1).protection = copy(cell.protection)
                    ws2.cell(row_number + 1, col_number + 1).alignment = copy(cell.alignment)
                    ws2.cell(row_number + 1, col_number + 1).quotePrefix = copy(cell.quotePrefix)
                    ws2.cell(row_number + 1, col_number + 1).pivotButton = copy(cell.pivotButton)

        len_A = len(ws2['A']) + 1
        ws2[f"D{len_A}"] = 'ИТОГО (вес / пломб):'
        ws2[f"E{len_A}"] = ws2[f"E{len_A}"].number_format = '0.000'
        ws2[f"E{len_A}"] = f"=SUM(E4:E{len_A - 1})"
        ws2[f"F{len_A}"] = quont_of_plomb
        Manifest_name = f'{download_folder}Manifest {manif_number} - pacs-{quont_of_plomb} ({total_weight} kg) от {now_time} - {list_of_parties}.xlsx'
        Manifest_name_short_name = f'МАНИФЕСТ ОТГРУЗКИ {manif_number} - мест {quont_of_plomb} ({total_weight} кг) от {now_time} - {list_of_parties}'
        wb2.save(Manifest_name)

        #send_mail(Manifest_name, Manifest_name_short_name)

        df_total['Имя манифеста'] = f'{Manifest_name}'
        df_total['Дата манифеста'] = now_time

        try:
            df_all_sent = pd.read_excel(f'C:/Users/User/Desktop/ДОКУМЕНТЫ/ОТГРУЖЕННОЕ/{now_month}-ALL_Manifests.xlsx')
        except:
            df_all_sent = pd.DataFrame()

        df_all_sent_new = df_all_sent.append(df_total)

        writer = pd.ExcelWriter(f'C:/Users/User/Desktop/ДОКУМЕНТЫ/ОТГРУЖЕННОЕ/{now_month}-ALL_Manifests.xlsx', engine='xlsxwriter')
        df_all_sent_new.to_excel(writer, sheet_name='Sheet1', index=False)
        for column in df_all_sent_new:
            column_width = max(df_all_sent_new[column].astype(str).map(len).max(), len(column))
            col_idx = df_all_sent_new.columns.get_loc(column)
            writer.sheets['Sheet1'].set_column(col_idx, col_idx, column_width)
            writer.sheets['Sheet1'].set_column(0, 3, 10)
            writer.sheets['Sheet1'].set_column(1, 3, 20)
            writer.sheets['Sheet1'].set_column(2, 3, 20)
            writer.sheets['Sheet1'].set_column(3, 3, 20)
            writer.sheets['Sheet1'].set_column(4, 3, 30)
            writer.sheets['Sheet1'].set_column(5, 3, 20)

        writer.save()

        df_total = df_total.drop(['Имя манифеста', 'Дата манифеста'], axis=1)
        print(df_total)
        df_manifest_for_driver = df_total.drop_duplicates(subset='parcel_plomb_numb')
        df_manifest_for_driver = df_manifest_for_driver[df_manifest_for_driver['parcel_plomb_numb'].astype(bool)]
        writer = pd.ExcelWriter('system.xlsx', engine='xlsxwriter')
        df_manifest_for_driver.to_excel(writer, sheet_name='Sheet1', index=False)
        for column in df_manifest_for_driver:
            column_width = max(df_manifest_for_driver[column].astype(str).map(len).max(), len(column))
            col_idx = df_manifest_for_driver.   columns.get_loc(column)
            writer.sheets['Sheet1'].set_column(col_idx, col_idx, column_width)
            writer.sheets['Sheet1'].set_column(0, 3, 10)
            writer.sheets['Sheet1'].set_column(1, 3, 20)
            writer.sheets['Sheet1'].set_column(2, 3, 20)
            writer.sheets['Sheet1'].set_column(3, 3, 20)
            writer.sheets['Sheet1'].set_column(4, 3, 30)
            writer.sheets['Sheet1'].set_column(5, 3, 20)

        writer.save()
        wb2 = openpyxl.load_workbook('system.xlsx')
        ws2 = wb2.active
        ws2.insert_rows(1, 2)

        wb = openpyxl.load_workbook('Akt.xlsx')
        ws = wb.active
        ws['B1'].value = f'№{manif_number} от {now} а/м М246ВР пломба № 52388385'

        head = "A1:G3"  # Заголовок таблицы, в котором есть объединенные ячейки
        for _range in ws.merged_cells.ranges:
            boundaries = range_boundaries(str(_range))
            ws2.merge_cells(start_column=boundaries[0], start_row=boundaries[1],
                            end_column=boundaries[2], end_row=boundaries[3])
        for row_number, row in enumerate(ws[head]):
            for col_number, cell in enumerate(row):
                ws2.cell(row_number + 1, col_number + 1, cell.value)
                if cell.has_style:
                    ws2.cell(row_number + 1, col_number + 1).font = copy(cell.font)
                    ws2.cell(row_number + 1, col_number + 1).fill = copy(cell.fill)
                    ws2.cell(row_number + 1, col_number + 1).border = copy(cell.border)
                    ws2.cell(row_number + 1, col_number + 1).number_format = copy(cell.number_format)
                    ws2.cell(row_number + 1, col_number + 1).protection = copy(cell.protection)
                    ws2.cell(row_number + 1, col_number + 1).alignment = copy(cell.alignment)
                    ws2.cell(row_number + 1, col_number + 1).quotePrefix = copy(cell.quotePrefix)
                    ws2.cell(row_number + 1, col_number + 1).pivotButton = copy(cell.pivotButton)

        len_A = len(ws2['A']) + 1
        ws2[f"E{len_A}"] = 'ИТОГО (пломб / вес):'
        ws2[f"G{len_A}"] = ws2[f"G{len_A}"].number_format = '0.000'
        ws2[f"G{len_A}"] = f"=SUM(G4:G{len_A - 1})"
        ws2[f"F{len_A}"] = quont_of_plomb
        ws2[f"D{len_A + 4}"] = 'Груз сдал_________'
        ws2[f"F{len_A + 4}"] = 'Груз принял_________'

        ws2.sheet_properties.pageSetUpPr.fitToPage = True
        ws2.page_setup.fitToHeight = False
        cm = int(1 / 4)
        ws2.page_margins = PageMargins(left=cm, right=cm, top=cm, bottom=cm)
        Manifest_name2 = f'{download_folder}TTN {manif_number} - мест {quont_of_plomb} - от {now_time}  - {list_of_parties}.xlsx'
        Manifest_name_short_name2 = f'ТТН {manif_number} - мест {quont_of_plomb} - от {now_time} - {list_of_parties})'
        wb2.save(Manifest_name2)
        #send_mail(Manifest_name2, Manifest_name_short_name2)
        object_name = Manifest_name_short_name
        comment = 'Манифест огрузки: Манифест сформирован'
        insert_user_action(object_name, comment)


        flash(f'Манифест {Manifest_name_short_name} сформирован', category='success')
        winsound.PlaySound('Snd\priezjayte-k-nam-esche.wav', winsound.SND_FILENAME)
    else:
        flash(f'ЕСТЬ ОТКАЗНЫЕ (НЕ ОТРАБОТАННЫЕ) ПЛОМБЫ!', category='error')

def manifest_to_xls_GBS(df_manifest_total):
    now = datetime.datetime.now().strftime("%d.%m.%Y")
    now_nextday = (datetime.datetime.now() + datetime.timedelta(days=1)).strftime("%d.%m.%Y")
    now_time = datetime.datetime.now().strftime("%d.%m.%Y %H-%M")
    df = df_manifest_total
    delta = datetime.timedelta(hours=-10, minutes=0)
    event_date = datetime.datetime.now() + delta
    event_date = event_date.strftime("%Y-%m-%d %H:%M:%S")
    #insert_event_API_test(df, event_date)
    party_numb = df_manifest_total['party_numb'].values[0]

    group_Weight_df = df_manifest_total.groupby('parcel_plomb_numb')['parcel_weight'].sum()
    group_Weight_df = group_Weight_df.rename('Вес_мешка', inplace=True)
    df_total = pd.merge(df_manifest_total, group_Weight_df, how='left', left_on='parcel_plomb_numb',
                        right_on='parcel_plomb_numb')
    df_total['№ п.п.'] = np.arange(len(df_total))[::+1] + 1
    df_total['Номер индивидуальной     накладной'] = df_manifest_total['parcel_numb']
    df_total['Трекинг'] = df_total['parcel_numb']
    df_total['Номер накладной'] = df_total['party_numb']
    df_total['Вес посылки'] = df_total['parcel_weight']
    df_total['Мест'] = df_total.Вес_мешка.eq(df_total.Вес_мешка.shift()).astype('str')
    df_total['Вес_мешка'] = np.where(df_total['Мест'] == 'True', '0', df_total['Вес_мешка'])

    df_total['Мест'] = df_total.parcel_plomb_numb.eq(df_total.parcel_plomb_numb.shift()).astype('str')
    df_total['parcel_plomb_numb'] = np.where(df_total['Мест'] == 'True', '', df_total['parcel_plomb_numb'])

    df_total['Вес мешка'] = df_total['Вес_мешка'].astype(float)

    df_total = df_total.reindex(columns=['№ п.п.', 'Номер индивидуальной     накладной',
                                         'Трекинг', 'Номер накладной', 'Вес посылки',
                                         'parcel_plomb_numb', 'Вес мешка'])

    df_total = df_total.drop_duplicates(subset='Трекинг', keep='first')
    if not df_manifest_total['custom_status_short'].astype(str).str.contains('ИЗЪЯТИЕ').any():

        df_parc_numb = df_manifest_total[['parcel_numb', 'parcel_weight', 'parcel_plomb_numb']]
        print(df_parc_numb)
        con_track = sl.connect('TRACKS.db')
        with con_track:
            data = con_track.execute("select count(*) from sqlite_master where type='table' and name='parcels_search_tracks'")
            for row in data:
                # если таких таблиц нет
                if row[0] == 0:
                    # создаём таблицу
                    with con_track:
                        con_track.execute("""
                                                                CREATE TABLE parcels_search_tracks (
                                                                ID INTEGER PRIMARY KEY AUTOINCREMENT,
                                                                parcel_numb VARCHAR(25) NOT NULL UNIQUE ON CONFLICT REPLACE,
                                                                parcel_weight FLOAT,
                                                                parcel_plomb_numb VARCHAR(25)
                                                                );
                                                                """)
            df_parc_numb.to_sql('parcels_search_tracks', con=con_track, if_exists='replace', index=False)
        with con_track:
            query = """SELECT * FROM parcels_search_tracks                          
                                LEFT JOIN tracks ON tracks.parcel_numb = parcels_search_tracks.parcel_numb
                                """
            df_manifest_total = pd.read_sql(query, con_track)

        quont_of_plomb = len(df_manifest_total.drop_duplicates(subset='parcel_plomb_numb'))
        quont_of_parc = len(df_manifest_total.drop_duplicates(subset='parcel_numb'))
        weight_total = df_manifest_total['parcel_weight'].sum().round(3)
        df_manifest_total = df_manifest_total.loc[:, ~df_manifest_total.columns.duplicated()].copy()
        df_manifest_total['№ п.п.'] = np.arange(len(df_manifest_total))[::+1] + 1
        df_manifest_total = df_manifest_total[['№ п.п.', 'track_numb', 'parcel_numb', 'parcel_weight', 'parcel_plomb_numb']]

        df_manifest_total = df_manifest_total.rename(columns={'№ п.п.': '№ п/п',
                                                              'track_numb': 'Номер индивидуальной накладной',
                                                              'parcel_numb': 'Трекинг',
                                                              'parcel_weight': 'Вес посылки',
                                                              'parcel_plomb_numb': 'Пломба'
                                                              })


        Manifest_name = f'{download_folder}GBS Manifest {party_numb} pacs-{quont_of_plomb} ({weight_total} kg) to {now_nextday}.xlsx'
        Manifest_name_short_name = f'МАНИФЕСТ ОТГРУЗКИ GBS {party_numb} мест {quont_of_plomb} ({weight_total} кг) к {now_nextday}'

        writer = pd.ExcelWriter(Manifest_name, engine='xlsxwriter')
        df_manifest_total.to_excel(writer, sheet_name='Sheet1', index=False)

        writer.save()

        wb = openpyxl.load_workbook(Manifest_name)
        ws = wb.active
        ws.insert_rows(1, 1)
        ws['F2'].value = 'Номер реестра выпуска'
        ws['G2'].value = 'Дата уведомления о выпуске'
        ws['H2'].value = 'Подготовлено к отгрузке'
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length + 5

        len_table = len(ws['A'])
        ws['A1'].value = f'готово к выдаче {quont_of_parc} посылка'
        ws['A1'].font = Font(size=12, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:H1')

        ws[f'B{len_table + 1}'].value = f'ИТОГО: {quont_of_parc} ПОСЫЛКА'
        ws[f'B{len_table + 1}'].font = Font(size=12, bold=True)
        ws[f'B{len_table + 1}'].alignment = Alignment(horizontal='center')
        ws.merge_cells(f'B{len_table + 1}:H{len_table + 1}')

        """ws.column_dimensions['A'].width = 9
        ws.column_dimensions['B'].width = 31
        ws.column_dimensions['C'].width = 31
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 31"""


        wb.save(Manifest_name)

        df_manifest_for_driver = df_total.drop_duplicates(subset='parcel_plomb_numb')
        df_manifest_for_driver = df_manifest_for_driver[df_manifest_for_driver['parcel_plomb_numb'].astype(bool)]
        writer = pd.ExcelWriter('system.xlsx', engine='xlsxwriter')
        df_manifest_for_driver.to_excel(writer, sheet_name='Sheet1', index=False)
        for column in df_manifest_for_driver:
            column_width = max(df_manifest_for_driver[column].astype(str).map(len).max(), len(column))
            col_idx = df_manifest_for_driver.columns.get_loc(column)
            writer.sheets['Sheet1'].set_column(col_idx, col_idx, column_width)
            writer.sheets['Sheet1'].set_column(0, 3, 10)
            writer.sheets['Sheet1'].set_column(1, 3, 20)
            writer.sheets['Sheet1'].set_column(2, 3, 20)
            writer.sheets['Sheet1'].set_column(3, 3, 20)
            writer.sheets['Sheet1'].set_column(4, 3, 30)
            writer.sheets['Sheet1'].set_column(5, 3, 20)

        writer.save()
        wb2 = openpyxl.load_workbook('system.xlsx')
        ws2 = wb2.active
        ws2.insert_rows(1, 2)

        wb = openpyxl.load_workbook('Akt.xlsx')
        ws = wb.active
        ws['B1'].value = f'{now} а/м М246ВР пломба № 52388385'

        head = "A1:G3"  # Заголовок таблицы, в котором есть объединенные ячейки
        for _range in ws.merged_cells.ranges:
            boundaries = range_boundaries(str(_range))
            ws2.merge_cells(start_column=boundaries[0], start_row=boundaries[1],
                            end_column=boundaries[2], end_row=boundaries[3])
        for row_number, row in enumerate(ws[head]):
            for col_number, cell in enumerate(row):
                ws2.cell(row_number + 1, col_number + 1, cell.value)
                if cell.has_style:
                    ws2.cell(row_number + 1, col_number + 1).font = copy(cell.font)
                    ws2.cell(row_number + 1, col_number + 1).fill = copy(cell.fill)
                    ws2.cell(row_number + 1, col_number + 1).border = copy(cell.border)
                    ws2.cell(row_number + 1, col_number + 1).number_format = copy(cell.number_format)
                    ws2.cell(row_number + 1, col_number + 1).protection = copy(cell.protection)
                    ws2.cell(row_number + 1, col_number + 1).alignment = copy(cell.alignment)
                    ws2.cell(row_number + 1, col_number + 1).quotePrefix = copy(cell.quotePrefix)
                    ws2.cell(row_number + 1, col_number + 1).pivotButton = copy(cell.pivotButton)

        len_A = len(ws2['A']) + 1
        ws2[f"E{len_A}"] = 'ИТОГО (пломб / вес):'
        ws2[f"G{len_A}"] = ws2[f"G{len_A}"].number_format = '0.000'
        ws2[f"G{len_A}"] = f"=SUM(G4:G{len_A - 1})"
        ws2[f"F{len_A}"] = quont_of_plomb
        ws2[f"D{len_A + 4}"] = 'Груз сдал_________'
        ws2[f"F{len_A + 4}"] = 'Груз принял_________'

        ws2.sheet_properties.pageSetUpPr.fitToPage = True
        ws2.page_setup.fitToHeight = False
        cm = int(1 / 4)
        ws2.page_margins = PageMargins(left=cm, right=cm, top=cm, bottom=cm)
        Manifest_name2 = f'{download_folder}TTN {party_numb} мест {quont_of_plomb} - {now_time}.xlsx'
        Manifest_name_short_name2 = f'ТТН {party_numb} мест {quont_of_plomb} - {now_time})'
        wb2.save(Manifest_name2)

        #send_mail(Manifest_name, Manifest_name_short_name)
        object_name = Manifest_name_short_name
        comment = 'Манифест огрузки: Манифест GBS сформирован'
        insert_user_action(object_name, comment)
        df_plomb_to_manifest = pd.DataFrame()
        df_plomb_to_manifest_total = pd.DataFrame()
        flash(f'Манифест {Manifest_name_short_name} сформирован', category='success')
        winsound.PlaySound('Snd\priezjayte-k-nam-esche.wav', winsound.SND_FILENAME)
    else:
        flash(f'ЕСТЬ ОТКАЗНЫЕ (НЕ ОТРАБОТАННЫЕ) ПЛОМБЫ!', category='error')

@app_svh.route('/save_manifest<string:partner>', methods=['POST', 'GET'])
def save_manifest(partner):
    #global df_plomb_to_manifest
    global df_plomb_to_manifest_total
    con = sl.connect('BAZA.db', timeout=30)
    df_manifest_total_refuses = df_plomb_to_manifest_total.loc[df_plomb_to_manifest_total['custom_status_short'] == 'ИЗЪЯТИЕ']
    df_manifest_total_refuses_parcels = df_manifest_total_refuses.drop_duplicates(subset='parcel_numb')
    df_manifest_total_refuses_parcels = df_manifest_total_refuses_parcels['parcel_numb'].to_list()
    if df_manifest_total_refuses.empty:
        df_parcelc_to_manifest_total = df_plomb_to_manifest_total.loc[df_plomb_to_manifest_total['Тип'] == 'Посылка-место']
        df_plomb_to_manifest_total = df_plomb_to_manifest_total.loc[df_plomb_to_manifest_total['Тип'] == 'Пломба']
        with con:
            df_plombs_to_sql = df_plomb_to_manifest_total['parcel_plomb_numb']
            df_plombs_to_sql.to_sql('df_plomb_to_manifest', con=con, if_exists='replace')
        with con:
            query = """Select * FROM baza
                    JOIN df_plomb_to_manifest ON df_plomb_to_manifest.parcel_plomb_numb = baza.parcel_plomb_numb"""

            df_manifest_total = pd.read_sql(query, con)
            i = 0
            ploms_to_manifest = df_plomb_to_manifest_total['parcel_plomb_numb']

            for parcel_numb in df_parcelc_to_manifest_total['parcel_numb']:

                df_manifest = pd.read_sql(f"SELECT * FROM baza where parcel_numb = '{parcel_numb}'", con)
                con.execute("Update baza set VH_status = 'ОТГРУЖЕН' where parcel_numb = ?", (parcel_numb, ))

                df_manifest_total = df_manifest_total.append(df_manifest)
            logger.warning(df_manifest_total)
        df_manifest_total = df_manifest_total.loc[:, ~df_manifest_total.columns.duplicated()].copy()
        writer = pd.ExcelWriter('test.xlsx', engine='xlsxwriter')
        df_manifest_total.to_excel(writer, sheet_name='Sheet1', index=False)
        writer.save()
        print(partner)
        if partner == 'CEL':
            manifest_to_xls(df_manifest_total)
        elif partner == 'GBS':
            manifest_to_xls_GBS(df_manifest_total)
        for parcel_plomb_numb in ploms_to_manifest:
            i += 1
            con.execute(
                "Update baza set VH_status = 'ОТГРУЖЕН' where parcel_plomb_numb = ? and custom_status_short = ?",
                (parcel_plomb_numb, 'ВЫПУСК'))
            print(f'{i} - {parcel_plomb_numb} updated')
        logger.warning(df_manifest_total)
    else:
        flash(f'{df_manifest_total_refuses_parcels} СО СТАТУСОМ ИЗЪЯТИЕ (НЕ ВЫПУЩЕНА!!!)', category='error')
        winsound.PlaySound('Snd\imanie izyyatie.wav', winsound.SND_FILENAME)
    return render_template('plomb_info_manifest.html')

@app_svh.route('/party_to_manifest', methods=['POST', 'GET'])
def party_to_manifest():
    con = sl.connect('BAZA.db')
    with con:
        df_parties = pd.read_sql(f"SELECT * FROM baza", con).drop_duplicates(subset='party_numb', keep='first')
        df_parties = df_parties['party_numb'].to_list()
    return render_template('party_to_manifest.html', df_parties=df_parties)


@app_svh.route('/party_to_manifest_button', methods=['POST', 'GET'])
def party_to_manifest_button():
    party_numb = request.form['party_numb']
    con = sl.connect('BAZA.db')
    with con:
        df_manifest_total = pd.read_sql(f"SELECT * FROM baza where party_numb = '{party_numb}' "
                                  f"and custom_status_short = 'ВЫПУСК'", con)
        con.execute("Update baza set VH_status = 'ОТГРУЖЕН' where party_numb = ? AND custom_status_short = ?", (party_numb, 'ВЫПУСК'))
    df_manifest_total = df_manifest_total.sort_values(by='parcel_plomb_numb', ascending=False)
    object_name = party_numb
    comment = 'Манифест огрузки: Партия отгружена'
    insert_user_action(object_name, comment)
    manifest_to_xls(df_manifest_total)
    return render_template('plomb_search_manifest.html')

df_changed_plomb = pd.DataFrame()
@app_svh.route('/search1/old_plomb/add', methods=['GET', 'POST'])
def old_plomb_add():
    now_time = datetime.datetime.now().strftime("%d.%m.%Y %H:%M")
    global df_changed_plomb
    try:
        if request.method == 'POST':
            old_plomb_numb = request.form['old_plomb_numb']
            new_plomb_numb = request.form['new_plomb_numb']
            con = sl.connect('BAZA.db')
            with con:
                df = pd.read_sql(f"SELECT * FROM baza where parcel_plomb_numb = '{old_plomb_numb}' ", con)
                if df.empty:
                    flash(f'Пломба {old_plomb_numb} не найдена!', category='error')
                    winsound.PlaySound('Snd\Snd_NoPlomb.wav', winsound.SND_FILENAME)
                else:

                    cursor = con.cursor()
                    query = "Update baza set parcel_plomb_numb = ? where parcel_plomb_numb = ?"
                    data = (new_plomb_numb, old_plomb_numb)
                    logger.warning(data)
                    cursor.execute(query, data)
                    con.commit()
                    cursor.close()
                    flash(f'Пломба {old_plomb_numb} обновлена!', category='success')
                    winsound.PlaySound('Snd\plomba_obnovlena.wav', winsound.SND_FILENAME)
                    df_changed_plomb_to_append = pd.DataFrame({'старая': [old_plomb_numb], 'новая': [new_plomb_numb]})
                    df_changed_plomb = df_changed_plomb.append(df_changed_plomb_to_append)
                    df_changed_plomb['№'] = np.arange(len(df_changed_plomb))[::+1] + 1
                    df_changed_plomb = df_changed_plomb[['№', 'старая', 'новая']]
                    df_changed_plomb = df_changed_plomb.iloc[::-1]
                    logger_change_plob.info(f'Время: {now_time} Старая пломба: {old_plomb_numb} Новая пломба: {new_plomb_numb}')
                    object_name = old_plomb_numb
                    comment = f'Новая пломба: Изменена на {new_plomb_numb}'
                    insert_user_action(object_name, comment)
        return render_template('old_plomb.html', tables=[style + df_changed_plomb.to_html(classes='mystyle', index=False,
                                                                                        float_format='{:2,.2f}'.format)],
                           titles=['na', 'Изменены:'])

    except Exception as e:
        logger.warning(f'error {e}')
        return {'message': str(e)}, 400

@app_svh.route('/new_place', methods=['POST', 'GET'])
def new_place():
    try:
        parcel_numb = request.args.get('parcel_numb')
        object_name = parcel_numb
        comment = f'Сформировать место: Посылка отмечена для формирования нового места'
        insert_user_action(object_name, comment)
        return render_template('parcel_search_new_place.html', search=parcel_numb)
    except Exception as e:
        return {'message': str(e)}, 400

done_parcels_np = pd.DataFrame()
@app_svh.route('/save_new_place', methods=['POST', 'GET'])
def making_new_place():
    global done_parcels_np
    global parcel_numb_np
    global df_parc_events_np
    global done_parcels_styl_np
    parcel_numb_np = request.form['parcel_numb']
    con_vec = sl.connect('VECTORS.db')
    try:
        vector = con_vec.execute(
            f"SELECT vector FROM VECTORS where parcel_numb = '{parcel_numb_np}'").fetchone()[
            0]
        flash(f'{vector}!', category='success')
    except Exception as e:
        vector = None
        print(e)

    try:
        con = sl.connect('BAZA.db')
        with con:
            df_parc_events_np = pd.read_sql(f"SELECT * FROM baza where parcel_numb = '{parcel_numb_np}'", con)
            df_parc_events_np['№1'] = np.arange(len(df_parc_events_np))[::+1] + 1
            df_parc_events_np = df_parc_events_np[['parcel_numb', 'custom_status_short', 'parcel_plomb_numb', 'VH_status']]
            df_parc_events_np = df_parc_events_np.rename(
                columns={'parcel_numb': 'Трек-номер', 'custom_status_short': 'Статус',
                         'parcel_plomb_numb': 'Пломба', 'VH_status': 'ВХ'})
            status = df_parc_events_np['Статус'][0]
            if df_parc_events_np.empty:
                flash(f'Посылка не найдена!', category='error')
                winsound.PlaySound('Snd\Snd_Parcel_Not_Found.wav', winsound.SND_FILENAME)
            else:
                pass
            done_parcels_np = done_parcels_np.append(df_parc_events_np).drop_duplicates(subset=['Трек-номер'], keep='first')
            done_parcels_np.index = done_parcels_np.index + 1  # shifting index
            done_parcels_np.sort_index(inplace=True)
            done_parcels_np['№'] = np.arange(len(done_parcels_np))[::-1] + 1
            done_parcels_styl_np = done_parcels_np.reset_index()
            done_parcels_styl_np = done_parcels_styl_np.drop('index', axis=1)
            done_parcels_styl_np = done_parcels_styl_np[
                ['№', 'Трек-номер', 'Статус', 'Пломба', 'ВХ']].drop_duplicates(subset=['Трек-номер'], keep='first')
            trigger_color1 = df_parc_events_np['Статус']
            #trigger_color2 = df_parc_events_np['ВХ']
            #def highlight_GREEN(df_parc_events):
            #    return ['background-color: #7CFC00' if 'На ВХ' in str(i) else '' for i in df_parc_events]
            def highlight_RED(df_parc_events_np):
                return ['background-color: #FF0000' if 'ИЗЪЯТИЕ' in str(i) else '' for i in df_parc_events_np]
            #if 'На ВХ' in trigger_color2.values:
            #    df_parc_events_np = df_parc_events_np.style.apply(highlight_GREEN).hide_index()
            if 'ИЗЪЯТИЕ' in trigger_color1.values:
                flash(f'ИЗЪЯТИЕ!', category='error')
                df_parc_events_np = df_parc_events_np.style.apply(highlight_RED).hide_index()
            #if 'На ВХ' in done_parcels_styl_np['ВХ'].values:
            #    done_parcels_styl_np = done_parcels_styl_np.style.apply(highlight_GREEN).hide_index()
            if 'ИЗЪЯТИЕ' in done_parcels_styl_np['Статус'].values:
                done_parcels_styl_np = done_parcels_styl_np.style.apply(highlight_RED).hide_index()
            object_name = parcel_numb_np
            comment = f'Сформировать место: Посылка отмечена для формирования нового места'
            insert_user_action(object_name, comment)
    except Exception as e:
        flash(f'Посылка не найдена!', category='error')
        winsound.PlaySound('Snd\Snd_Parcel_Not_Found.wav', winsound.SND_FILENAME)
        print(e)
        return render_template('parcel_info_new_place.html')
        #return {'message': str(e)}, 400
        pass
    return render_template('parcel_info_new_place.html', tables=[done_parcels_styl_np.to_html(classes='mystyle', index=False,
                                                                                 float_format='{:2,.2f}'.format)],
                           titles=['na', '\n\nОтработанные:'],
                           parcel_numb=parcel_numb_np, vector=vector)

@app_svh.route('/delete_last_parcel', methods=['POST', 'GET'])
def delete_last_parcel():
    global done_parcels_np
    global df_parc_events_np
    global done_parcels_styl_np
    logger.warning(done_parcels_np)
    done_parcels_np = done_parcels_np[1:]
    done_parcels_styl_np = done_parcels_styl_np[1:]
    logger.warning(done_parcels_np)
    return render_template('parcel_info_new_place.html', tables=[df_parc_events_np.to_html(classes='mystyle', index=False,
                                                                                        float_format='{:2,.2f}'.format),
                                                                 style + done_parcels_styl_np.to_html(classes='mystyle',
                                                                                           index=False,
                                                                                           float_format='{:2,.2f}'.format)],
                           titles=['na', 'Посылка:', '\n\nОтработанные:'],
                           parcel_numb=parcel_numb_np)

@app_svh.route('/save_new_place/place_numb', methods=['POST', 'GET'])
def make_place_numb():
    global done_parcels_np
    plomb_toreplace_new = request.form['plomb_toreplace_new']
    con = sl.connect('BAZA.db')
    # открываем базу
    with con:
        for parcel_numb in done_parcels_np['Трек-номер']:
            #plomb_toreplace = done_parcels_np.loc[done_parcels_np['Трек-номер'] == parcel_numb]['Пломба'].values[0]
            #logger.warning(plomb_toreplace)
            con.execute(f"Update baza set parcel_plomb_numb = '{plomb_toreplace_new}' where parcel_numb = '{parcel_numb}'")
    df_new_place = pd.read_sql(f"Select * from baza where parcel_plomb_numb = '{plomb_toreplace_new}'", con)
    writer = pd.ExcelWriter(f'{addition_folder}Место {plomb_toreplace_new}.xlsx', engine='xlsxwriter')
    df_new_place.to_excel(writer, sheet_name='Sheet1', index=False)
    for column in df_new_place:
        column_width = max(df_new_place[column].astype(str).map(len).max(), len(column))
        col_idx = df_new_place.columns.get_loc(column)
        writer.sheets['Sheet1'].set_column(col_idx, col_idx, column_width)
        writer.sheets['Sheet1'].set_column(0, 3, 10)
        writer.sheets['Sheet1'].set_column(1, 3, 20)
        writer.sheets['Sheet1'].set_column(2, 3, 20)
        writer.sheets['Sheet1'].set_column(3, 3, 20)
        writer.sheets['Sheet1'].set_column(4, 3, 30)
        writer.sheets['Sheet1'].set_column(5, 3, 20)
    object_name = plomb_toreplace_new
    comment = f'Сформировать место: Место сформированно'
    insert_user_action(object_name, comment)
    writer.save()
    con.commit()
    con.close()
    done_parcels_np = pd.DataFrame()
    flash(f'Место сформировано', category='success')
    winsound.PlaySound('Snd\mesto_sformirovano.wav', winsound.SND_FILENAME)
    return render_template('parcel_info_new_place.html')

parcel_plomb_numb_np = None
df_plombs_np = pd.DataFrame()
vector = None

@app_svh.route('/modal')
def modal():
    return render_template('modal.html')


@app_svh.route('/create_new_pallet', methods=['POST', 'GET'])
def create_pallet():
    global df_plombs_html
    global df_plombs_np
    global parcel_plomb_numb_np
    global vector
    modal = 0
    try:
        parcel_plomb_numb_np = request.form['parcel_plomb_numb_np']
    except:
        pass
    try:
        con = sl.connect('BAZA.db')
        with con:
            df_all_pallets = pd.read_sql(f"SELECT DISTINCT pallet FROM baza", con).drop_duplicates(subset='pallet', keep='first')
            try:
                df_all_pallets['pallet'] = df_all_pallets['pallet'].fillna(0).astype(int)
                df_all_pallets = df_all_pallets.sort_values(by='pallet')
                last_pall_numb = int(df_all_pallets.values[-1].tolist()[0])
                i = last_pall_numb + 1
            except:
                last_pall_numb = df_all_pallets.values[0].tolist()[0]
                i = 1
            logger.warning(last_pall_numb)
            df_plomb = pd.read_sql(f"SELECT * FROM baza where parcel_plomb_numb = '{parcel_plomb_numb_np}'", con)
            try:
                first_parcel = df_plomb['parcel_numb'].values[0]
                print(first_parcel)
            except:
                pass
            if df_plomb['custom_status_short'].astype(str).str.contains('ИЗЪЯТИЕ').any():
                flash(f'ВНИМАНИЕ, пломба с неотвязанными посылками на изъятие!', category='error')
                winsound.PlaySound('Snd\Snd_CancelIssue.wav', winsound.SND_FILENAME)
            df_plomb = df_plomb.drop_duplicates(subset=['parcel_plomb_numb'], keep='first')
            df_plomb['Тип'] = 'Пломба'
            if df_plomb.empty and parcel_plomb_numb_np is not None:
                df_plomb = pd.read_sql(f"SELECT * FROM baza where parcel_numb = '{parcel_plomb_numb_np}'", con)
                df_plomb['Тип'] = "Посылка-место"
                logger.warning(df_plomb)
                if df_plomb.empty and parcel_plomb_numb_np is not None:
                    flash(f'Пломба не найдена!', category='error')
                    winsound.PlaySound('Snd\Snd_NoPlomb.wav', winsound.SND_FILENAME)
                else:
                    pass
            else:
                pass

            try:
                df_plombs_np = df_plombs_np.append(df_plomb).drop_duplicates(subset=['parcel_plomb_numb'], keep='first')
                quont_plombs = len(df_plombs_np)
                df_plombs_np.index = df_plombs_np.index + 1  # shifting index
                df_plombs_np.sort_index(inplace=True)
                df_plombs_np['№1'] = np.arange(len(df_plombs_np))[::-1] + 1
                df_plombs_html = df_plombs_np.rename(columns={
                    '№1': '№',
                    'parcel_plomb_numb': 'Пломба',
                    'parcel_numb': 'Трек',
                    'pallet': '№ Паллет',
                    'zone': 'Зона',
                    'party_numb': 'Партия'
                })
                df_plombs_html = df_plombs_html[['№', 'Пломба', '№ Паллет',
                                               'Зона', 'Партия', 'Трек', 'Тип']]
                df_plombs_html.fillna("", inplace=True)
                df_plombs_html = df_plombs_html.reset_index()
                df_plombs_html = df_plombs_html.drop('index', axis=1)

                object_name = parcel_plomb_numb_np
                comment = f'Сформировать новый паллет: пломба отмечена для добавления на паллет №{i}'
                insert_user_action(object_name, comment)
                winsound.PlaySound('Snd\zvuk-vezeniya.wav', winsound.SND_FILENAME)
            except:
                pass
        try:
            con = sl.connect('VECTORS.db')
            with con:
                try:
                    try:
                        vector1 = con.execute(f"SELECT vector FROM VECTORS where parcel_plomb_numb = '{parcel_plomb_numb_np}'").fetchone()[0]
                        flash(f'{vector1}!', category='success')
                        print(vector1)
                    except:
                        try:
                            vector1 = con.execute(
                                f"SELECT vector FROM VECTORS where parcel_numb = '{first_parcel}'").fetchone()[
                                0]
                            flash(f'{vector1}!', category='success')
                            print(vector1)
                        except Exception as e:
                            vector1 = None
                            print(e)

                    if vector != vector1 and vector is not None:
                        flash(f'Направления не совпадают!! Было ({vector})!', category='error')
                        winsound.PlaySound('Snd\Snd_CancelIssue.wav', winsound.SND_FILENAME)
                        modal = 1
                    vector = vector1
                except Exception as e:
                    print(e)

        except Exception as e:
            flash(f'{e}!', category='error')


    except Exception as e:
        flash(f'Пломба не найдена! {e}', category='error')
        winsound.PlaySound('Snd\Snd_NoPlomb.wav', winsound.SND_FILENAME)
        #return render_template('parcel_info_new_place.html')
        return {'message': str(e)}, 400

    return render_template('New_pallet.html', tables=[style + df_plombs_html.to_html(classes='mystyle', index=False,
                                                                              float_format='{:2,.2f}'.format)],
                           titles=['na', '\n\nМешки\места:'],
                           parcel_plomb_numb_np=parcel_plomb_numb_np, i=i, quont_plombs=quont_plombs,
                           modal=modal, vector=vector)

@app_svh.route('/delete_last_plomb', methods=['POST', 'GET'])
def delete_last_plomb():
    global parcel_plomb_numb_np
    global df_plombs_np
    global df_plombs_html
    logger.warning(df_plombs_html)
    df_plombs_html = df_plombs_html[1:]
    df_plombs_np = df_plombs_np[1:]
    object_name = parcel_plomb_numb_np
    comment = f'Удалена пломба из создания паллета'
    insert_user_action(object_name, comment)
    return render_template('New_pallet.html', tables=[style + df_plombs_html.to_html(classes='mystyle', index=False,
                                                                             float_format='{:2,.2f}'.format)],
                           titles=['na', '\n\nМешки\места:'],
                          parcel_plomb_numb_np=parcel_plomb_numb_np)

@app_svh.route('/delete_last_plomb_addpallet', methods=['POST', 'GET'])
def delete_last_plomb_addpallet():
    global parcel_plomb_numb_np
    global df_plombs_np
    global df_plombs_html
    logger.warning(df_plombs_html)
    df_plombs_html = df_plombs_html[:-1]
    df_plombs_np = df_plombs_np[:-1]
    object_name = parcel_plomb_numb_np
    comment = f'Удалена пломба из добавления на паллет'
    insert_user_action(object_name, comment)
    return render_template('add_to_pallet.html', tables=[style + df_plombs_html.to_html(classes='mystyle', index=False,
                                                                             float_format='{:2,.2f}'.format)],
                           titles=['na', '\n\nМешки\места:'],
                          parcel_plomb_numb_np=parcel_plomb_numb_np)

@app_svh.route('/search/clean_working_place_clean_working_place_pallet', methods=['POST'])
def clean_working_place_pallet():
    global parcel_plomb_numb_np
    global df_plombs_np
    global df_plombs_html
    df_plombs_html = pd.DataFrame()
    df_plombs_np = pd.DataFrame()
    parcel_plomb_numb_np = None
    object_name = None
    comment = f'Сформировать новый паллет: Таблица очищена'
    insert_user_action(object_name, comment)
    return render_template('New_pallet.html')

@app_svh.route('/search/clean_working_place_clean_working_place_addpallet', methods=['POST'])
def clean_working_place_addpallet():
    global parcel_plomb_numb_np
    global df_plombs_np
    global df_plombs_html
    df_plombs_html = pd.DataFrame()
    df_plombs_np = pd.DataFrame()
    object_name = None
    comment = f'Добавить на паллет: Таблица очищена'
    insert_user_action(object_name, comment)
    return render_template('add_to_pallet.html')

@app_svh.route('/insert_new_pallet', methods=['POST', 'GET'])
def insert_pallet():
    global df_plombs_np
    global parcel_plomb_numb_np
    pallet_new = request.form['pallet_new']
    con = sl.connect('BAZA.db')
    if pallet_new is None or pallet_new == '':
        with con:
            df_all_pallets = pd.read_sql(f"SELECT pallet FROM baza", con).drop_duplicates(subset='pallet', keep='first')
            try:
                print(df_all_pallets['pallet'])
                df_all_pallets['pallet'] = df_all_pallets['pallet'].fillna(0).astype(int)
                df_all_pallets = df_all_pallets.sort_values(by='pallet')
                logger.warning(df_all_pallets)
                last_pall_numb = int(df_all_pallets.values[-1].tolist()[0])
                pallet_new = last_pall_numb + 1
            except Exception as e:
                print(e)
    else:
        pass
    # открываем базу
    con = sl.connect('BAZA.db')
    with con:
        df_plombs_np_typeplomb = df_plombs_np.loc[df_plombs_np['Тип'] == 'Пломба']
        logger.warning(df_plombs_np_typeplomb)
        df_plombs_np_typeparcel = df_plombs_np.loc[df_plombs_np['Тип'] == 'Посылка-место']
        logger.warning(df_plombs_np_typeparcel)
        for parcel_plomb_numb in df_plombs_np_typeplomb['parcel_plomb_numb']:
                con.execute(
                    f"Update baza set pallet = '{pallet_new}' where parcel_plomb_numb = '{parcel_plomb_numb}'")
        for parcel_numb in df_plombs_np_typeparcel['parcel_numb']:
                con.execute(
                    f"Update baza set pallet = '{pallet_new}' where parcel_numb = '{parcel_numb}'")
        df_new_pallet = pd.read_sql(f"Select * from baza where pallet = '{pallet_new}'", con)
        writer = pd.ExcelWriter(f'{addition_folder}Паллет {pallet_new}.xlsx', engine='xlsxwriter')
        df_new_pallet.to_excel(writer, sheet_name='Sheet1', index=False)
        for column in df_new_pallet:
            column_width = max(df_new_pallet[column].astype(str).map(len).max(), len(column))
            col_idx = df_new_pallet.columns.get_loc(column)
            writer.sheets['Sheet1'].set_column(col_idx, col_idx, column_width)
            writer.sheets['Sheet1'].set_column(0, 3, 10)
            writer.sheets['Sheet1'].set_column(1, 3, 20)
            writer.sheets['Sheet1'].set_column(2, 3, 20)
            writer.sheets['Sheet1'].set_column(3, 3, 20)
            writer.sheets['Sheet1'].set_column(4, 3, 30)
            writer.sheets['Sheet1'].set_column(5, 3, 20)

        writer.save()
        con.commit()
        df_plombs_np = pd.DataFrame()
        parcel_plomb_numb_np = None
        object_name = pallet_new
        comment = f'Сформировать новый паллет: Паллет сформирован'
        insert_user_action(object_name, comment)
        flash(f'Паллет сформирован!', category='success')
        winsound.PlaySound('Snd\Pallet_made.wav', winsound.SND_FILENAME)
    return render_template('New_pallet.html')


@app_svh.route('/add_to_pallet', methods=['POST', 'GET'])
def add_to_pallet():
    global df_plombs_html
    global df_plombs_np
    global parcel_plomb_numb_np
    try:
        parcel_plomb_numb_np = request.form['parcel_plomb_numb_np']
    except:
        pass
    try:
        con = sl.connect('BAZA.db')
        with con:
            #to show all pallets from system for choice
            df_all_pallets = pd.read_sql(f"SELECT pallet FROM baza", con).drop_duplicates(subset='pallet', keep='first')
            df_all_pallets['pallet'] = df_all_pallets['pallet'].fillna(value=np.nan).fillna(0).astype(int)
            df_all_pallets = df_all_pallets.sort_values(by='pallet', na_position='last', ascending=False)
            df_all_pallets = df_all_pallets['pallet'].to_list()
            logger.warning(df_all_pallets)
            #select row with current plomb (parcel) number
            df_plomb = pd.read_sql(f"SELECT * FROM baza where parcel_plomb_numb = '{parcel_plomb_numb_np}'", con)
            df_plomb['Тип'] = 'Пломба'
            #if that is parcel
            if df_plomb.empty and parcel_plomb_numb_np != None:
                df_plomb = pd.read_sql(f"SELECT * FROM baza where parcel_numb = '{parcel_plomb_numb_np}'", con)
                df_plomb['Тип'] = "Посылка-место"
                logger.warning(df_plomb)
                if df_plomb.empty and parcel_plomb_numb_np != None:
                    flash(f'Пломба не найдена!', category='error')
                    winsound.PlaySound('Snd\Snd_NoPlomb.wav', winsound.SND_FILENAME)
            else:
                pass
            try:
                df_plombs_np = df_plombs_np.append(df_plomb).drop_duplicates(subset=['parcel_plomb_numb'], keep='first')
                df_plombs_np['№1'] = np.arange(len(df_plombs_np))[::+1] + 1
                df_plombs_html = df_plombs_np.rename(columns={
                    '№1': '№',
                    'parcel_plomb_numb': 'Пломба',
                    'parcel_numb': 'Трек',
                    'pallet': '№ Паллет',
                    'zone': 'Зона',
                    'party_numb': 'Партия'
                })
                df_plombs_html = df_plombs_html[['№', 'Пломба', '№ Паллет',
                                                 'Зона', 'Партия', 'Трек', 'Тип']]
                df_plombs_html.fillna("", inplace=True)
                df_plombs_html = df_plombs_html.reset_index()
                df_plombs_html = df_plombs_html.drop('index', axis=1)
                object_name = parcel_plomb_numb_np
                comment = f'Добавить на паллет: Пломба отмечена для добавления'
                insert_user_action(object_name, comment)
            except:
                pass

    except Exception as e:
        flash(f'Пломба не найдена!', category='error')
        winsound.PlaySound('Snd\Snd_NoPlomb.wav', winsound.SND_FILENAME)
        #return render_template('parcel_info_new_place.html')
        return {'message': str(e)}, 400
        pass
    return render_template('add_to_pallet.html', tables=[style + df_plombs_html.to_html(classes='mystyle', index=False,
                                                                                     float_format='{:2,.2f}'.format)],
                           titles=['na', '\n\nМешки\места:'],
                           parcel_plomb_numb_np=parcel_plomb_numb_np, df_all_pallets=df_all_pallets)

@app_svh.route('/add_to_pallet_button', methods=['POST', 'GET'])
def add_to_pallet_button():
    global df_plombs_np
    global parcel_plomb_numb_np
    pallet_new = request.form['pallet_new']
    # открываем базу
    con = sl.connect('BAZA.db')
    with con:
        df_new_pallet = pd.read_sql(f"Select * from baza where pallet = '{pallet_new}'", con)
        logger.warning(df_new_pallet)
        df_plombs_np_typeplomb = df_plombs_np.loc[df_plombs_np['Тип'] == 'Пломба']
        logger.warning(df_plombs_np_typeplomb)
        df_plombs_np_typeparcel = df_plombs_np.loc[df_plombs_np['Тип'] == 'Посылка-место']
        logger.warning(df_plombs_np_typeparcel)
        if not df_new_pallet.empty:
            for parcel_plomb_numb in df_plombs_np_typeplomb['parcel_plomb_numb']:
                    # plomb_toreplace = done_parcels_np.loc[done_parcels_np['Трек-номер'] == parcel_numb]['Пломба'].values[0]
                    # logger.warning(plomb_toreplace)
                    con.execute(
                        f"Update baza set pallet = '{pallet_new}' where parcel_plomb_numb = '{parcel_plomb_numb}'")
            for parcel_numb in df_plombs_np_typeparcel['parcel_numb']:
                con.execute(
                    f"Update baza set pallet = '{pallet_new}' where parcel_numb = '{parcel_numb}'")
            df_new_pallet = pd.read_sql(f"Select * from baza where pallet = '{pallet_new}'", con)
            writer = pd.ExcelWriter(f'{addition_folder}Паллет {pallet_new}.xlsx', engine='xlsxwriter')
            df_new_pallet.to_excel(writer, sheet_name='Sheet1', index=False)
            for column in df_new_pallet:
                column_width = max(df_new_pallet[column].astype(str).map(len).max(), len(column))
                col_idx = df_new_pallet.columns.get_loc(column)
                writer.sheets['Sheet1'].set_column(col_idx, col_idx, column_width)
                writer.sheets['Sheet1'].set_column(0, 3, 10)
                writer.sheets['Sheet1'].set_column(1, 3, 20)
                writer.sheets['Sheet1'].set_column(2, 3, 20)
                writer.sheets['Sheet1'].set_column(3, 3, 20)
                writer.sheets['Sheet1'].set_column(4, 3, 30)
                writer.sheets['Sheet1'].set_column(5, 3, 20)

            writer.save()
            con.commit()
            df_plombs_np = pd.DataFrame()
            parcel_plomb_numb_np = None
            object_name = pallet_new
            comment = f'Добавить на паллет: Пломбы добавленны'
            insert_user_action(object_name, comment)
            flash(f'Добавлено!', category='success')
            winsound.PlaySound('Snd\dd_.wav', winsound.SND_FILENAME)
        else:
            flash(f'Паллет не найден!', category='error')
            winsound.PlaySound('Snd\Pallet_not_found.wav', winsound.SND_FILENAME)
    return render_template('add_to_pallet.html')

parc_quont_pallet_info = 0
plomb_quont_pallet_info = 0
@app_svh.route('/pallet_info', methods=['POST', 'GET'])
def pallet_info():
    global parc_quont_pallet_info
    global plomb_quont_pallet_info
    global pallet
    df_refuses = []
    numb = 0
    try:
        numb = request.form['numb']
        logger.warning(numb)
    except:
        pass
    con = sl.connect('BAZA.db')
    if numb == 0:
        df = pd.DataFrame()
    else:
        with con:
            df = pd.read_sql(f"Select * from baza where pallet = '{numb}'", con)
            df = df.rename(columns=map_eng_to_rus)
            df['№'] = np.arange(len(df))[::+1] + 1
            df = df[['№', 'Паллет', 'Пломба', 'Трек-номер', 'Статус ВХ', 'Статус ТО (кратк)',  'Статус ТО', 'Партия']]
            if df.empty:
                df = pd.read_sql(f"Select * from baza where parcel_plomb_numb = '{numb}'", con)
                try:
                    pallet = df['pallet'].values[0]
                    df = pd.read_sql(f"Select * from baza where pallet = '{pallet}'", con)
                    df = df.rename(columns=map_eng_to_rus)
                    df['№'] = np.arange(len(df))[::+1] + 1
                    df = df[['№', 'Паллет', 'Пломба', 'Трек-номер', 'Статус ВХ', 'Статус ТО (кратк)', 'Статус ТО', 'Партия']]
                except:
                    pass
                if df.empty:
                    df = pd.read_sql(f"Select * from baza where parcel_numb = '{numb}'", con)
                    try:
                        pallet = df['pallet'].values[0]
                        df = pd.read_sql(f"Select * from baza where pallet = '{pallet}'", con)
                        df = df.rename(columns=map_eng_to_rus)
                        df['№'] = np.arange(len(df))[::+1] + 1
                        df = df[['№', 'Паллет', 'Пломба', 'Трек-номер', 'Статус ВХ', 'Статус ТО (кратк)', 'Статус ТО', 'Партия']]
                    except:
                        pass
            df_refuses = df.loc[df['Статус ТО (кратк)'] == 'ИЗЪЯТИЕ']
            df_refuses = df_refuses['Трек-номер'].to_list()
            logger.warning(df_refuses)
            parc_quont_pallet_info = len(df)
            plomb_quont_pallet_info = len(df.drop_duplicates(subset='Пломба', keep='first'))
            pallet = df['Паллет'].values[0]
            def highlight_RED(df):
                    return ['background-color: #FF0000' if 'ИЗЪЯТИЕ' in str(i) else '' for i in df]

            df = df.style.apply(highlight_RED).hide_index()
    return render_template('pallet_info.html', tables=[style + df.to_html(classes='mystyle', index=False,
                                                                   float_format='{:2,.2f}'.format)],
                           titles=['Информация'], df_refuses=df_refuses,
                           pallet=pallet, numb=numb, parc_quont_pallet_info=parc_quont_pallet_info, plomb_quont_pallet_info=plomb_quont_pallet_info)

@app_svh.route('/pallet_info_callback_refuses', methods=['POST', 'GET'])
def pallet_info_callback_refuses():
    global pallet
    con = sl.connect('BAZA.db')
    with con:
        con.execute(f"Update baza set pallet = '0', parcel_plomb_numb = '' where pallet = '{pallet}' AND custom_status_short = 'ИЗЪЯТИЕ'")
        print('updated')
        object_name = pallet
        comment = f'Отвязанны посылки с изъятием от паллета'
        insert_user_action(object_name, comment)
    flash(f'Посылки со статусом ИЗЪЯТИЕ успешно отвязаны от паллета {pallet}! Убедитесь, что их реально изъяли!', category='success')
    return render_template('pallet_info.html', pallet=pallet)

def chunks(lst, n):
    """Yield successive n-sized chunks from lst."""
    for i in range(0, len(lst), n):
        yield lst[i:i + n]

def server_request_events():
    print("start updating")
    con = sl.connect('BAZA.db')
    len_id = con.execute('SELECT max(id) from baza').fetchone()[0]
    id_for_job = len_id - 1900000
    df = pd.read_sql(f"Select parcel_numb from baza "
                     f"where ID > {id_for_job} "
                     f"AND custom_status_short = 'ИЗЪЯТИЕ' ", con).drop_duplicates(subset='parcel_numb')
    list_chanks = list(chunks(df, 100))
    #print(list_chanks)
    i = 0
    for chank in list_chanks:
        i += 1
        with con:
            body = chank.to_json(orient="records", indent=2)
            # body = {'parcel_numb': 'CEL7000012753CD'}
            headers = {'accept': 'application/json'}
            response = requests.post('http://164.132.182.145:5001/api/get_decisions',
                                     json=body)  # http://127.0.0.1:5000  # 'http://164.132.182.145:5001/api/get_decisions'
            try:
                json_decisions = response.json()
                df_loaded_decisions = pd.DataFrame.from_records(json_decisions)
                with con:
                    df_to_append = pd.DataFrame()

                    for parcel_numb in df_loaded_decisions['parcel_numb']:
                        row_isalready_in = pd.read_sql(f"Select * from baza where parcel_numb = '{parcel_numb}'", con)
                        row = df_loaded_decisions.loc[df_loaded_decisions['parcel_numb'] == parcel_numb]

                        custom_status_short = df_loaded_decisions.loc[df_loaded_decisions['parcel_numb'] == parcel_numb][
                            'custom_status_short'].values[0]
                        custom_status = df_loaded_decisions.loc[df_loaded_decisions['parcel_numb'] == parcel_numb][
                            'custom_status'].values[0]
                        decision_date = df_loaded_decisions.loc[df_loaded_decisions['parcel_numb'] == parcel_numb][
                            'decision_date'].values[0]
                        refuse_reason = df_loaded_decisions.loc[df_loaded_decisions['parcel_numb'] == parcel_numb][
                            'refuse_reason'].values[0]

                        con.execute(f"Update baza set "
                                    f" custom_status = '{custom_status}',"
                                    f" custom_status_short = '{custom_status_short}',"
                                    f" decision_date = '{decision_date}',"
                                    f" refuse_reason = '{refuse_reason}' where parcel_numb = '{parcel_numb}'")
                        #    row_isalready_in = pd.read_sql(f"Select * from baza where parcel_numb = '{parcel_numb}'", con)
                        # df_isalready_in = df_isalready_in.append(row_isalready_in)
                    df_to_append.to_sql('baza', con=con, if_exists='append', index=False)

            except:
                print(response)
        print(f"chunk{i} updated")


scheduler = BackgroundScheduler(daemon=True, job_defaults={'max_instances': 5})


#Create the job
scheduler.add_job(func=server_request_events, trigger='interval', seconds=500) #trigger='cron', hour='22', minute='30'
scheduler.start()

if __name__ == '__main__':
    app_svh.secret_key = 'c9e779a3258b42338334daaed51bccf7'
    app_svh.config['SESSION_TYPE'] = 'filesystem'
    app_svh.run(host='0.0.0.0', port=5000, threaded=True)  # cancel debug=True  and make threaded=True