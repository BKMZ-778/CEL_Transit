from flask import flash
import pandas as pd
import sqlite3 as sl
import numpy as np
import logging
import openpyxl
from copy import copy
from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet.page import PageMargins
from openpyxl.styles import Font
from openpyxl.styles import Alignment
import datetime
import winsound
from openpyxl.styles import PatternFill

from .services import insert_user_action, download_folder


def manifest_to_xls(df_manifest_total):
    now = datetime.datetime.now().strftime("%d.%m.%Y")
    now_month = datetime.datetime.now().strftime("%m.%Y")
    now_time = datetime.datetime.now().strftime("%d.%m.%Y %H-%M")
    delta = datetime.timedelta(hours=-10, minutes=0)
    event_date = datetime.datetime.now() + delta
    event_date = event_date.strftime("%Y-%m-%d %H:%M:%S")
    #insert_event_API_test(df, event_date)
    party_numb = df_manifest_total['party_numb'].values[0]
    with open("manif_number.txt") as manif_number_file:
        manif_number = int(manif_number_file.read())+1
    with open("manif_number.txt", 'w') as manif_number_file:
        manif_number_file.write(str(manif_number))
    # number for particular SVH warehouse
    manif_number = 'US-' + str(manif_number)


    if not df_manifest_total['custom_status_short'].astype(str).str.contains('ИЗЪЯТИЕ').any():
        df_manifest_total = df_manifest_total.sort_values(by='parcel_plomb_numb')
        df_manifest_total['№ п.п.'] = np.arange(len(df_manifest_total))[::+1] + 1
        df_manifest_total['Номер индивидуальной     накладной'] = df_manifest_total['parcel_numb']
        df_manifest_total['Трекинг'] = df_manifest_total['parcel_numb']
        df_manifest_total['Номер накладной'] = df_manifest_total['party_numb']
        df_manifest_total['Вес посылки'] = df_manifest_total['parcel_weight']
        quont_of_plomb = len(df_manifest_total.drop_duplicates(subset='parcel_plomb_numb'))
        group_Weight_df = df_manifest_total.groupby('parcel_plomb_numb')['parcel_weight'].sum()
        group_Weight_df = group_Weight_df.rename('Вес_мешка', inplace=True)
        df_total = pd.merge(df_manifest_total, group_Weight_df, how='left', left_on='parcel_plomb_numb',
                            right_on='parcel_plomb_numb')

        list_of_parties1 = df_manifest_total['Номер накладной'].drop_duplicates().str[-8:].to_list()
        list_of_parties = ' '.join(list_of_parties1)[:100]

        writer = pd.ExcelWriter('test.xlsx', engine='xlsxwriter')
        df_total.to_excel(writer, sheet_name='Sheet1', index=False)
        writer.save()

        df_parcels = df_total[['parcel_numb']].drop_duplicates()
        df_plombs = df_total[['parcel_plomb_numb']].drop_duplicates()
        print(df_plombs)
        con_vect = sl.connect('VECTORS.db')
        data = con_vect.execute("select count(*) from sqlite_master where type='table' and name='parcels_to_manifest'")
        for row in data:
            # если таких таблиц нет
            if row[0] == 0:
                # создаём таблицу
                with con_vect:
                    con_vect.execute("""
                                                            CREATE TABLE parcels_to_manifest (
                                                            ID INTEGER PRIMARY KEY AUTOINCREMENT,
                                                            parcel_numb VARCHAR(20)
                                                            );
                                                        """)
                    con_vect.execute("""
                                                            CREATE TABLE plombs_to_manifest (
                                                            ID INTEGER PRIMARY KEY AUTOINCREMENT,
                                                            parcel_plomb_numb VARCHAR(20)
                                                            );
                                                        """)

        df_parcels.to_sql('parcels_to_manifest', con=con_vect, index=False, if_exists='replace')
        df_plombs.to_sql('plombs_to_manifest', con=con_vect, index=False, if_exists='replace')
        query = """SELECT vectors.parcel_numb, vectors.vector
                                FROM vectors                          
                                JOIN parcels_to_manifest 
                                ON parcels_to_manifest.parcel_numb = vectors.parcel_numb
                                """

        con = sl.connect('BAZA.db')
        with con:
            df_with_pullings = pd.read_sql("Select parcel_plomb_numb from plombs_with_pullings", con)
        with con_vect:
            df_manif_vectors = pd.read_sql(query, con_vect)
        df_with_pullings['sign'] = '1'
        print(df_with_pullings)
        df_plomb_merge = pd.merge(df_plombs, df_with_pullings, how='left', left_on="parcel_plomb_numb", right_on="parcel_plomb_numb")
        print(df_plomb_merge)
        print(df_plomb_merge)
        writer = pd.ExcelWriter('system.xlsx', engine='xlsxwriter')
        df_with_pullings.to_excel(writer, sheet_name='Sheet1', index=False)
        writer.save()

        df_total = pd.merge(df_total, df_manif_vectors, how='left', left_on='parcel_numb' ,right_on='parcel_numb')

        #df_total['Мест'] = df_total.Вес_мешка.eq(df_total.Вес_мешка.shift()).astype('str')
        df_total = pd.merge(df_total, df_plomb_merge, how='left', left_on='parcel_plomb_numb',
                            right_on='parcel_plomb_numb')
        df_total['Мест'] = df_total.parcel_plomb_numb.eq(df_total.parcel_plomb_numb.shift()).astype('str')
        df_total['Вес_мешка'] = np.where(df_total['Мест'] == 'True', '0', df_total['Вес_мешка'])
        df_total['parcel_plomb_numb'] = np.where(df_total['Мест'] == 'True', '', df_total['parcel_plomb_numb'])

        df_total['Вес мешка'] = df_total['Вес_мешка'].astype(float)
        df_total['Направление'] = df_total['vector']


        df_total = df_total.reindex(columns=['№ п.п.', 'Номер индивидуальной     накладной',
                                             'Трекинг', 'Номер накладной', 'Вес посылки',
                                             'parcel_plomb_numb', 'Вес мешка', 'Направление', 'sign'])


        df_total = df_total.drop_duplicates(subset='Трекинг', keep='first')
        total_weight = df_manifest_total['parcel_weight'].sum().round(3)
        writer = pd.ExcelWriter('system.xlsx', engine='xlsxwriter')
        df_total.to_excel(writer, sheet_name='Sheet1', index=False)
        for column in df_total:
            column_width = max(df_total[column].astype(str).map(len).max(), len(column))
            col_idx = df_total.columns.get_loc(column)
            writer.sheets['Sheet1'].set_column(col_idx, col_idx, column_width)
            writer.sheets['Sheet1'].set_column(0, 3, 10)
            writer.sheets['Sheet1'].set_column(1, 3, 20)
            writer.sheets['Sheet1'].set_column(2, 3, 20)
            writer.sheets['Sheet1'].set_column(3, 3, 20)
            writer.sheets['Sheet1'].set_column(4, 3, 30)
            writer.sheets['Sheet1'].set_column(5, 3, 20)

        writer.save()
        wb2 = openpyxl.load_workbook('system.xlsx')
        ws2 = wb2.active
        ws2.insert_rows(1, 2)

        wb = openpyxl.load_workbook('Akt.xlsx')
        ws = wb.active
        ws['B1'].value = f'{now} а/м М246ВР пломба № 52388385'

        head = "A1:H4"  # Заголовок таблицы, в котором есть объединенные ячейки
        for _range in ws.merged_cells.ranges:
            boundaries = range_boundaries(str(_range))
            ws2.merge_cells(start_column=boundaries[0], start_row=boundaries[1],
                            end_column=boundaries[2], end_row=boundaries[3])
        for row_number, row in enumerate(ws[head]):
            for col_number, cell in enumerate(row):
                ws2.cell(row_number + 1, col_number + 1, cell.value)
                if cell.has_style:
                    ws2.cell(row_number + 1, col_number + 1).font = copy(cell.font)
                    ws2.cell(row_number + 1, col_number + 1).fill = copy(cell.fill)
                    ws2.cell(row_number + 1, col_number + 1).border = copy(cell.border)
                    ws2.cell(row_number + 1, col_number + 1).number_format = copy(cell.number_format)
                    ws2.cell(row_number + 1, col_number + 1).protection = copy(cell.protection)
                    ws2.cell(row_number + 1, col_number + 1).alignment = copy(cell.alignment)
                    ws2.cell(row_number + 1, col_number + 1).quotePrefix = copy(cell.quotePrefix)
                    ws2.cell(row_number + 1, col_number + 1).pivotButton = copy(cell.pivotButton)

        len_A = len(ws2['A']) + 1
        ws2[f"D{len_A}"] = 'ИТОГО (вес / пломб):'
        ws2[f"E{len_A}"] = ws2[f"E{len_A}"].number_format = '0.000'
        ws2[f"E{len_A}"] = f"=SUM(E4:E{len_A - 1})"
        ws2[f"F{len_A}"] = quont_of_plomb
        ws2["A1"] = manif_number
        len_F = len(ws2['F'])
        yelFill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')

        for i in range(len_F):
            try:
                if ws2[f'I{i}'].value == '1':
                    ws2[f'B{i}'].fill = yelFill
                    ws2[f'C{i}'].fill = yelFill
                    ws2[f'F{i}'].fill = yelFill
                    print('yellow')
            except:
                pass
        ws2.delete_cols(9, 1)
        Manifest_name = f'{download_folder}Manifest {manif_number} - pacs-{quont_of_plomb} ({total_weight} kg) от {now_time} - {list_of_parties}.xlsx'
        Manifest_name_short_name = f'МАНИФЕСТ ОТГРУЗКИ {manif_number} - мест {quont_of_plomb} ({total_weight} кг) от {now_time} - {list_of_parties}'
        wb2.save(Manifest_name)

        #send_mail(Manifest_name, Manifest_name_short_name)

        df_total['Имя манифеста'] = f'{Manifest_name}'
        df_total['Дата манифеста'] = now_time

        try:
            df_all_sent = pd.read_excel(f'C:/Users/User/Desktop/ДОКУМЕНТЫ/ОТГРУЖЕННОЕ/{now_month}-ALL_Manifests.xlsx')
        except:
            df_all_sent = pd.DataFrame()

        df_all_sent_new = df_all_sent.append(df_total)

        writer = pd.ExcelWriter(f'C:/Users/User/Desktop/ДОКУМЕНТЫ/ОТГРУЖЕННОЕ/{now_month}-ALL_Manifests.xlsx', engine='xlsxwriter')
        df_all_sent_new.to_excel(writer, sheet_name='Sheet1', index=False)
        for column in df_all_sent_new:
            column_width = max(df_all_sent_new[column].astype(str).map(len).max(), len(column))
            col_idx = df_all_sent_new.columns.get_loc(column)
            writer.sheets['Sheet1'].set_column(col_idx, col_idx, column_width)
            writer.sheets['Sheet1'].set_column(0, 3, 10)
            writer.sheets['Sheet1'].set_column(1, 3, 20)
            writer.sheets['Sheet1'].set_column(2, 3, 20)
            writer.sheets['Sheet1'].set_column(3, 3, 20)
            writer.sheets['Sheet1'].set_column(4, 3, 30)
            writer.sheets['Sheet1'].set_column(5, 3, 20)

        writer.save()

        df_total = df_total.drop(['Имя манифеста', 'Дата манифеста'], axis=1)
        print(df_total)
        df_manifest_for_driver = df_total.drop_duplicates(subset='parcel_plomb_numb')
        df_manifest_for_driver = df_manifest_for_driver[df_manifest_for_driver['parcel_plomb_numb'].astype(bool)]
        writer = pd.ExcelWriter('system.xlsx', engine='xlsxwriter')
        df_manifest_for_driver.to_excel(writer, sheet_name='Sheet1', index=False)
        for column in df_manifest_for_driver:
            column_width = max(df_manifest_for_driver[column].astype(str).map(len).max(), len(column))
            col_idx = df_manifest_for_driver.   columns.get_loc(column)
            writer.sheets['Sheet1'].set_column(col_idx, col_idx, column_width)
            writer.sheets['Sheet1'].set_column(0, 3, 10)
            writer.sheets['Sheet1'].set_column(1, 3, 20)
            writer.sheets['Sheet1'].set_column(2, 3, 20)
            writer.sheets['Sheet1'].set_column(3, 3, 20)
            writer.sheets['Sheet1'].set_column(4, 3, 30)
            writer.sheets['Sheet1'].set_column(5, 3, 20)

        writer.save()
        wb2 = openpyxl.load_workbook('system.xlsx')
        ws2 = wb2.active
        ws2.insert_rows(1, 2)

        wb = openpyxl.load_workbook('Akt.xlsx')
        ws = wb.active
        ws['B1'].value = f'№{manif_number} от {now} а/м М246ВР пломба № 52388385'

        head = "A1:G3"  # Заголовок таблицы, в котором есть объединенные ячейки
        for _range in ws.merged_cells.ranges:
            boundaries = range_boundaries(str(_range))
            ws2.merge_cells(start_column=boundaries[0], start_row=boundaries[1],
                            end_column=boundaries[2], end_row=boundaries[3])
        for row_number, row in enumerate(ws[head]):
            for col_number, cell in enumerate(row):
                ws2.cell(row_number + 1, col_number + 1, cell.value)
                if cell.has_style:
                    ws2.cell(row_number + 1, col_number + 1).font = copy(cell.font)
                    ws2.cell(row_number + 1, col_number + 1).fill = copy(cell.fill)
                    ws2.cell(row_number + 1, col_number + 1).border = copy(cell.border)
                    ws2.cell(row_number + 1, col_number + 1).number_format = copy(cell.number_format)
                    ws2.cell(row_number + 1, col_number + 1).protection = copy(cell.protection)
                    ws2.cell(row_number + 1, col_number + 1).alignment = copy(cell.alignment)
                    ws2.cell(row_number + 1, col_number + 1).quotePrefix = copy(cell.quotePrefix)
                    ws2.cell(row_number + 1, col_number + 1).pivotButton = copy(cell.pivotButton)

        len_A = len(ws2['A']) + 1
        ws2[f"E{len_A}"] = 'ИТОГО (пломб / вес):'
        ws2[f"G{len_A}"] = ws2[f"G{len_A}"].number_format = '0.000'
        ws2[f"G{len_A}"] = f"=SUM(G4:G{len_A - 1})"
        ws2[f"F{len_A}"] = quont_of_plomb
        ws2[f"D{len_A + 4}"] = 'Груз сдал_________'
        ws2[f"F{len_A + 4}"] = 'Груз принял_________'

        ws2.sheet_properties.pageSetUpPr.fitToPage = True
        ws2.page_setup.fitToHeight = False
        cm = int(1 / 4)
        ws2.page_margins = PageMargins(left=cm, right=cm, top=cm, bottom=cm)
        Manifest_name2 = f'{download_folder}TTN {manif_number} - мест {quont_of_plomb} - от {now_time}  - {list_of_parties}.xlsx'
        Manifest_name_short_name2 = f'ТТН {manif_number} - мест {quont_of_plomb} - от {now_time} - {list_of_parties})'
        wb2.save(Manifest_name2)
        #send_mail(Manifest_name2, Manifest_name_short_name2)
        object_name = Manifest_name_short_name
        comment = 'Манифест огрузки: Манифест сформирован'
        insert_user_action(object_name, comment)


        flash(f'Манифест {Manifest_name_short_name} сформирован', category='success')
        winsound.PlaySound('Snd\priezjayte-k-nam-esche.wav', winsound.SND_FILENAME)
    else:
        flash(f'ЕСТЬ ОТКАЗНЫЕ (НЕ ОТРАБОТАННЫЕ) ПЛОМБЫ!', category='error')


def manifest_to_xls_GBS(df_manifest_total):
    now = datetime.datetime.now().strftime("%d.%m.%Y")
    now_nextday = (datetime.datetime.now() + datetime.timedelta(days=1)).strftime("%d.%m.%Y")
    now_time = datetime.datetime.now().strftime("%d.%m.%Y %H-%M")
    df = df_manifest_total
    delta = datetime.timedelta(hours=-10, minutes=0)
    event_date = datetime.datetime.now() + delta
    event_date = event_date.strftime("%Y-%m-%d %H:%M:%S")
    #insert_event_API_test(df, event_date)
    party_numb = df_manifest_total['party_numb'].values[0]

    group_Weight_df = df_manifest_total.groupby('parcel_plomb_numb')['parcel_weight'].sum()
    group_Weight_df = group_Weight_df.rename('Вес_мешка', inplace=True)
    df_total = pd.merge(df_manifest_total, group_Weight_df, how='left', left_on='parcel_plomb_numb',
                        right_on='parcel_plomb_numb')
    df_total['№ п.п.'] = np.arange(len(df_total))[::+1] + 1
    df_total['Номер индивидуальной     накладной'] = df_manifest_total['parcel_numb']
    df_total['Трекинг'] = df_total['parcel_numb']
    df_total['Номер накладной'] = df_total['party_numb']
    df_total['Вес посылки'] = df_total['parcel_weight']
    df_total['Мест'] = df_total.Вес_мешка.eq(df_total.Вес_мешка.shift()).astype('str')
    df_total['Вес_мешка'] = np.where(df_total['Мест'] == 'True', '0', df_total['Вес_мешка'])

    df_total['Мест'] = df_total.parcel_plomb_numb.eq(df_total.parcel_plomb_numb.shift()).astype('str')
    df_total['parcel_plomb_numb'] = np.where(df_total['Мест'] == 'True', '', df_total['parcel_plomb_numb'])

    df_total['Вес мешка'] = df_total['Вес_мешка'].astype(float)

    df_total = df_total.reindex(columns=['№ п.п.', 'Номер индивидуальной     накладной',
                                         'Трекинг', 'Номер накладной', 'Вес посылки',
                                         'parcel_plomb_numb', 'Вес мешка'])

    df_total = df_total.drop_duplicates(subset='Трекинг', keep='first')
    if not df_manifest_total['custom_status_short'].astype(str).str.contains('ИЗЪЯТИЕ').any():

        df_parc_numb = df_manifest_total[['parcel_numb', 'parcel_weight', 'parcel_plomb_numb']]
        print(df_parc_numb)
        con_track = sl.connect('TRACKS.db')
        with con_track:
            data = con_track.execute("select count(*) from sqlite_master where type='table' and name='parcels_search_tracks'")
            for row in data:
                # если таких таблиц нет
                if row[0] == 0:
                    # создаём таблицу
                    with con_track:
                        con_track.execute("""
                                                                CREATE TABLE parcels_search_tracks (
                                                                ID INTEGER PRIMARY KEY AUTOINCREMENT,
                                                                parcel_numb VARCHAR(25) NOT NULL UNIQUE ON CONFLICT REPLACE,
                                                                parcel_weight FLOAT,
                                                                parcel_plomb_numb VARCHAR(25)
                                                                );
                                                                """)
            df_parc_numb.to_sql('parcels_search_tracks', con=con_track, if_exists='replace', index=False)
        with con_track:
            query = """SELECT * FROM parcels_search_tracks                          
                                LEFT JOIN tracks ON tracks.parcel_numb = parcels_search_tracks.parcel_numb
                                """
            df_manifest_total = pd.read_sql(query, con_track)

        quont_of_plomb = len(df_manifest_total.drop_duplicates(subset='parcel_plomb_numb'))
        quont_of_parc = len(df_manifest_total.drop_duplicates(subset='parcel_numb'))
        weight_total = df_manifest_total['parcel_weight'].sum().round(3)
        df_manifest_total = df_manifest_total.loc[:, ~df_manifest_total.columns.duplicated()].copy()
        df_manifest_total['№ п.п.'] = np.arange(len(df_manifest_total))[::+1] + 1
        df_manifest_total = df_manifest_total[['№ п.п.', 'track_numb', 'parcel_numb', 'parcel_weight', 'parcel_plomb_numb']]

        df_manifest_total = df_manifest_total.rename(columns={'№ п.п.': '№ п/п',
                                                              'track_numb': 'Номер индивидуальной накладной',
                                                              'parcel_numb': 'Трекинг',
                                                              'parcel_weight': 'Вес посылки',
                                                              'parcel_plomb_numb': 'Пломба'
                                                              })


        Manifest_name = f'{download_folder}GBS Manifest {party_numb} pacs-{quont_of_plomb} ({weight_total} kg) to {now_nextday}.xlsx'
        Manifest_name_short_name = f'МАНИФЕСТ ОТГРУЗКИ GBS {party_numb} мест {quont_of_plomb} ({weight_total} кг) к {now_nextday}'

        writer = pd.ExcelWriter(Manifest_name, engine='xlsxwriter')
        df_manifest_total.to_excel(writer, sheet_name='Sheet1', index=False)

        writer.save()

        wb = openpyxl.load_workbook(Manifest_name)
        ws = wb.active
        ws.insert_rows(1, 1)
        ws['F2'].value = 'Номер реестра выпуска'
        ws['G2'].value = 'Дата уведомления о выпуске'
        ws['H2'].value = 'Подготовлено к отгрузке'
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length + 5

        len_table = len(ws['A'])
        ws['A1'].value = f'готово к выдаче {quont_of_parc} посылка'
        ws['A1'].font = Font(size=12, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:H1')

        ws[f'B{len_table + 1}'].value = f'ИТОГО: {quont_of_parc} ПОСЫЛКА'
        ws[f'B{len_table + 1}'].font = Font(size=12, bold=True)
        ws[f'B{len_table + 1}'].alignment = Alignment(horizontal='center')
        ws.merge_cells(f'B{len_table + 1}:H{len_table + 1}')

        """ws.column_dimensions['A'].width = 9
        ws.column_dimensions['B'].width = 31
        ws.column_dimensions['C'].width = 31
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 31"""


        wb.save(Manifest_name)

        df_manifest_for_driver = df_total.drop_duplicates(subset='parcel_plomb_numb')
        df_manifest_for_driver = df_manifest_for_driver[df_manifest_for_driver['parcel_plomb_numb'].astype(bool)]
        writer = pd.ExcelWriter('system.xlsx', engine='xlsxwriter')
        df_manifest_for_driver.to_excel(writer, sheet_name='Sheet1', index=False)
        for column in df_manifest_for_driver:
            column_width = max(df_manifest_for_driver[column].astype(str).map(len).max(), len(column))
            col_idx = df_manifest_for_driver.columns.get_loc(column)
            writer.sheets['Sheet1'].set_column(col_idx, col_idx, column_width)
            writer.sheets['Sheet1'].set_column(0, 3, 10)
            writer.sheets['Sheet1'].set_column(1, 3, 20)
            writer.sheets['Sheet1'].set_column(2, 3, 20)
            writer.sheets['Sheet1'].set_column(3, 3, 20)
            writer.sheets['Sheet1'].set_column(4, 3, 30)
            writer.sheets['Sheet1'].set_column(5, 3, 20)

        writer.save()
        wb2 = openpyxl.load_workbook('system.xlsx')
        ws2 = wb2.active
        ws2.insert_rows(1, 2)

        wb = openpyxl.load_workbook('Akt.xlsx')
        ws = wb.active
        ws['B1'].value = f'{now} а/м М246ВР пломба № 52388385'

        head = "A1:G3"  # Заголовок таблицы, в котором есть объединенные ячейки
        for _range in ws.merged_cells.ranges:
            boundaries = range_boundaries(str(_range))
            ws2.merge_cells(start_column=boundaries[0], start_row=boundaries[1],
                            end_column=boundaries[2], end_row=boundaries[3])
        for row_number, row in enumerate(ws[head]):
            for col_number, cell in enumerate(row):
                ws2.cell(row_number + 1, col_number + 1, cell.value)
                if cell.has_style:
                    ws2.cell(row_number + 1, col_number + 1).font = copy(cell.font)
                    ws2.cell(row_number + 1, col_number + 1).fill = copy(cell.fill)
                    ws2.cell(row_number + 1, col_number + 1).border = copy(cell.border)
                    ws2.cell(row_number + 1, col_number + 1).number_format = copy(cell.number_format)
                    ws2.cell(row_number + 1, col_number + 1).protection = copy(cell.protection)
                    ws2.cell(row_number + 1, col_number + 1).alignment = copy(cell.alignment)
                    ws2.cell(row_number + 1, col_number + 1).quotePrefix = copy(cell.quotePrefix)
                    ws2.cell(row_number + 1, col_number + 1).pivotButton = copy(cell.pivotButton)

        len_A = len(ws2['A']) + 1
        ws2[f"E{len_A}"] = 'ИТОГО (пломб / вес):'
        ws2[f"G{len_A}"] = ws2[f"G{len_A}"].number_format = '0.000'
        ws2[f"G{len_A}"] = f"=SUM(G4:G{len_A - 1})"
        ws2[f"F{len_A}"] = quont_of_plomb
        ws2[f"D{len_A + 4}"] = 'Груз сдал_________'
        ws2[f"F{len_A + 4}"] = 'Груз принял_________'

        ws2.sheet_properties.pageSetUpPr.fitToPage = True
        ws2.page_setup.fitToHeight = False
        cm = int(1 / 4)
        ws2.page_margins = PageMargins(left=cm, right=cm, top=cm, bottom=cm)
        Manifest_name2 = f'{download_folder}TTN {party_numb} мест {quont_of_plomb} - {now_time}.xlsx'
        Manifest_name_short_name2 = f'ТТН {party_numb} мест {quont_of_plomb} - {now_time})'
        wb2.save(Manifest_name2)

        #send_mail(Manifest_name, Manifest_name_short_name)
        object_name = Manifest_name_short_name
        comment = 'Манифест огрузки: Манифест GBS сформирован'
        insert_user_action(object_name, comment)
        flash(f'Манифест {Manifest_name_short_name} сформирован', category='success')
        winsound.PlaySound('Snd\priezjayte-k-nam-esche.wav', winsound.SND_FILENAME)
    else:
        flash(f'ЕСТЬ ОТКАЗНЫЕ (НЕ ОТРАБОТАННЫЕ) ПЛОМБЫ!', category='error')
