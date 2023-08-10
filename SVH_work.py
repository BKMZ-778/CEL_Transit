import time
import winsound
import pandas as pd
import numpy as np
from tkinter import ttk
import tkinter as tk
from tkinter import *
from tkinter import filedialog
import tkinter.messagebox as mb
import datetime
import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from copy import copy
from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet.page import PageMargins
import random
import xlsxwriter
from tkinter import Button, Frame, Tk
import pandastable as pt
from pandastable import Table

now = datetime.datetime.now().strftime("%d.%m.%Y")

pd.set_option('display.max_columns', None)
file_name = filedialog.askopenfilename()
df = pd.read_excel(file_name, sheet_name=0, engine='openpyxl')
df['Вес брутто'] = df['Вес брутто'].replace(',', '.', regex=True).astype(float)
df['Статус_ТО'] = df['Статус ТО']
df_true_status_names = df.drop_duplicates()
df['Статус_ТО'] = df['Статус_ТО'].replace(to_replace='Выпуск товаров без уплаты таможенных платежей', value='ВЫПУСК', regex=True)
df['Статус_ТО'] = df['Статус_ТО'].replace(to_replace='Выпуск товаров разрешен, таможенные платежи уплачены', value='ВЫПУСК', regex=True)
for cel in df['Статус_ТО']:
    if cel != 'ВЫПУСК':
        df.loc[df['Статус_ТО'] == cel, 'Статус_ТО'] = 'ИЗЪЯТИЕ'
df['Пломба'] = df['Пломба'].astype(str)

party_numb = df['Общая накладная'].values[0]
df_count_plomb = df['Пломба'].drop_duplicates()
plomb_quont = len(df_count_plomb)
len_done_plomb = 0
df_refuse_plomb = df.loc[df["Статус_ТО"] != 'ВЫПУСК']
df_refuse_plomb1 = df_refuse_plomb['Пломба'].drop_duplicates()
refuse_plomb_quont = len(df_refuse_plomb1)
plomb_list = []
refuse_parcels_list = []
parcels_list = []
options = {'align': 'w',
'cellbackgr': '#F4F4F3',
'cellwidth': 80,
'floatprecision': 2,
'thousandseparator': '',
'font': 'Arial',
'fontsize': 30,
'fontstyle': '',
'grid_color': '#ABB1AD',
'linewidth': 4,
'rowheight': 50,
'rowselectedcolor': '#E4DED4',
'textcolor': 'black'}

def plomb_serch(plomb):
    global plomb_list
    global len_done_plomb
    global dTDa1
    if dTDa1:
        dTDa1.destroy()
    plomb = str(entry_plomb.get())
    df_plomb = df.loc[df['Пломба'] == plomb]
    df_plomb = df_plomb.reindex(
        columns=['Пломба', 'Трек-номер', 'Статус_ТО', 'Вес брутто'])
    df_plomb = df_plomb.sort_values(by=['Трек-номер'])
    df_plomb = df_plomb.reset_index(drop=True)
    #raws = df_plomb.loc[df_plomb['Статус ТО'] != 'ВЫПУСК', 'Статус ТО']
    raws1 = df_plomb[df_plomb["Статус_ТО"] != 'ВЫПУСК'].index
    dTDa1 = tk.Toplevel(width='800', height=750)
    dTDa1.title('Накладные')
    dTDaPT = pt.Table(dTDa1, dataframe=df_plomb,
                      showtoolbar=True, showstatusbar=True,
                      width='800', height='750', maxcellwidth=500)
    dTDaPT.setRowHeight(60)
    dTDaPT.setFont()
    pt.config.apply_options(options, dTDaPT)
    entry_plomb.delete(0, 'end')
    if len(list(raws1)) != 0:
        dTDaPT.setRowColors(rows=list(raws1), clr='red', cols='all')
        dTDaPT.show()
        print('!')
        winsound.PlaySound('Snd_Open_Bag.wav', winsound.SND_FILENAME)
        plomb_list.append(plomb)
        plomb_list = list(set(plomb_list))
        len_done_plomb = len(plomb_list)
        label.config(text=f'обработанно: {len_done_plomb} \nиз {plomb_quont} (откзн:{refuse_plomb_quont})',
                     fg='black', font=('Times', 20))
        print(len_done_plomb)
        entry_plomb.grab_set()
        entry_plomb.focus_set()
    elif df_plomb.empty:
        dTDa1.destroy()
        winsound.PlaySound('Snd_NoPlomb.wav', winsound.SND_FILENAME)
        entry_plomb.focus_set()
    else:
        dTDaPT.show()
        winsound.PlaySound('Snd_All_Issue.wav', winsound.SND_FILENAME)
        print('ВЫПУСК')
        # window.grab_set()
        # window.focus_force()
        plomb_list.append(plomb)
        plomb_list = list(set(plomb_list))
        len_done_plomb = len(plomb_list)
        label.config(text=f'обработанно: {len_done_plomb} \nиз {plomb_quont} (откзн:{refuse_plomb_quont})',
                     fg='black', font=('Times', 20))
        print(len_done_plomb)
        entry_plomb.grab_set()
        entry_plomb.focus_set()

def new_plomb(new_plomb_numb):
    old_plomb_numb = str(entry_plomb_old.get())
    new_plomb_numb = str(entry_NEW_plomb.get())
    print(old_plomb_numb)
    print(new_plomb_numb)
    df['Пломба'] = df['Пломба'].replace(old_plomb_numb, new_plomb_numb)
    print(df.loc[df['Пломба'] == new_plomb_numb])
    #entry_plomb_old.grab_set()
    entry_plomb_old.focus_set()
    entry_NEW_plomb.delete(0, 'end')
    entry_plomb_old.delete(0, 'end')
    window2.bind('<Return>', old_plomb)

def old_plomb(old_plomb_numb):
    if dTDa1:
        dTDa1.destroy()
    old_plomb_numb = str(entry_plomb_old.get())
    if df.loc[df['Пломба'] == old_plomb_numb].empty:
        winsound.PlaySound('Snd_NoPlomb.wav', winsound.SND_FILENAME)
        entry_plomb_old.delete(0, 'end')
        entry_plomb_old.focus_set()
    else:
        entry_NEW_plomb.focus_set()
        window2.bind('<Return>', new_plomb)
def save_df_new_plomb():
    writer = pd.ExcelWriter('New_plomb.xlsx', engine='xlsxwriter')
    df.to_excel(writer, sheet_name='Sheet1', index=False)
    writer.save()

df_total_plomb_to_manifest = pd.DataFrame()
def plomb_to_manifest(plomb_to_manifest_numb):
    if dTDa1:
        dTDa1.destroy()
    global df_total_plomb_to_manifest
    plomb_to_manifest_numb = entry_plomb_manifest.get()
    df_plomb_to_manifest = df.loc[df['Пломба'] == plomb_to_manifest_numb]
    if df_plomb_to_manifest.empty:
        winsound.PlaySound('Snd_NoPlomb.wav', winsound.SND_FILENAME)
        entry_plomb_manifest.delete(0, 'end')
        entry_plomb_manifest.focus_set()
    else:
        df_total_plomb_to_manifest = df_total_plomb_to_manifest.append(df_plomb_to_manifest)
        df_total_plomb_to_manifest = df_total_plomb_to_manifest[df_total_plomb_to_manifest.Статус_ТО == 'ВЫПУСК']
        entry_plomb_manifest.delete(0, 'end')
        entry_plomb_manifest.focus_set()
    print(df_total_plomb_to_manifest)

def save_manifest():
    df_total_plomb_to_manifest['№ п.п.'] = np.arange(len(df_total_plomb_to_manifest))[::+1] + 1
    df_total_plomb_to_manifest['Номер индивидуальной     накладной'] = df_total_plomb_to_manifest['Трек-номер']
    df_total_plomb_to_manifest['Трекинг'] = df_total_plomb_to_manifest['Трек-номер']
    df_total_plomb_to_manifest['Номер накладной'] =df_total_plomb_to_manifest['Общая накладная']
    df_total_plomb_to_manifest['Вес посылки'] = df_total_plomb_to_manifest['Вес брутто']
    group_Weight_df = df_total_plomb_to_manifest.groupby('Пломба')['Вес брутто'].sum()
    group_Weight_df = group_Weight_df.rename('Вес мешка', inplace=True)
    df_total = pd.merge(df_total_plomb_to_manifest, group_Weight_df, how='left', left_on='Пломба', right_on='Пломба')
    df_total = df_total.reindex(columns=['№ п.п.', 'Номер индивидуальной     накладной',
                                         'Трекинг', 'Номер накладной', 'Вес посылки',
                                         'Пломба', 'Вес мешка'])

    df_total = df_total.drop_duplicates(subset='Трекинг', keep='first')
    writer = pd.ExcelWriter('system.xlsx', engine='xlsxwriter')
    df_total.to_excel(writer, sheet_name='Sheet1', index=False)
    for column in df_total:
        column_width = max(df_total[column].astype(str).map(len).max(), len(column))
        col_idx = df_total.columns.get_loc(column)
        writer.sheets['Sheet1'].set_column(col_idx, col_idx, column_width)
        writer.sheets['Sheet1'].set_column(0, 3, 10)
        writer.sheets['Sheet1'].set_column(1, 3, 20)
        writer.sheets['Sheet1'].set_column(2, 3, 20)
        writer.sheets['Sheet1'].set_column(3, 3, 20)
        writer.sheets['Sheet1'].set_column(4, 3, 30)
        writer.sheets['Sheet1'].set_column(5, 3, 20)

    writer.save()
    wb2 = openpyxl.load_workbook('system.xlsx')
    ws2 = wb2.active
    ws2.insert_rows(1, 2)

    wb = openpyxl.load_workbook('Akt.xlsx')
    ws = wb.active
    ws['B1'].value = f'{now} а/м М246ВР пломба № 52388385'

    head = "A1:G3"  # Заголовок таблицы, в котором есть объединенные ячейки
    for _range in ws.merged_cells.ranges:
        boundaries = range_boundaries(str(_range))
        ws2.merge_cells(start_column=boundaries[0], start_row=boundaries[1],
                        end_column=boundaries[2], end_row=boundaries[3])
    for row_number, row in enumerate(ws[head]):
        for col_number, cell in enumerate(row):
            ws2.cell(row_number + 1, col_number + 1, cell.value)
            if cell.has_style:
                ws2.cell(row_number + 1, col_number + 1).font = copy(cell.font)
                ws2.cell(row_number + 1, col_number + 1).fill = copy(cell.fill)
                ws2.cell(row_number + 1, col_number + 1).border = copy(cell.border)
                ws2.cell(row_number + 1, col_number + 1).number_format = copy(cell.number_format)
                ws2.cell(row_number + 1, col_number + 1).protection = copy(cell.protection)
                ws2.cell(row_number + 1, col_number + 1).alignment = copy(cell.alignment)
                ws2.cell(row_number + 1, col_number + 1).quotePrefix = copy(cell.quotePrefix)
                ws2.cell(row_number + 1, col_number + 1).pivotButton = copy(cell.pivotButton)

    wb2.save(f'Manifest {now} ({party_numb}).xlsx')
    winsound.PlaySound('priezjayte-k-nam-esche.wav', winsound.SND_FILENAME)
    msg = f"Манифест {now} ({party_numb}) сформирован"
    mb.showinfo("МАНИФЕСТ", msg)

def search_parcel(parcel):
    global dTDa2
    if dTDa1:
        dTDa1.destroy()
    global refuse_parcels_list
    global parcels_list
    global parcel_plomb
    global len_parcels_list
    global df_plomb
    try:
        parcel = str(entry_parcel.get())

        #parcel_plomb = df.loc[df['Трек-номер'] == parcel, 'Пломба'].values[0]
        print(parcel)
        print(parcel_plomb)
        if parcel_plomb != df.loc[df['Трек-номер'] == parcel, 'Пломба'].values[0] and parcel_plomb != 0 and len_parcels_list != len(df_plomb):
            winsound.PlaySound('sovpadenie.wav', winsound.SND_FILENAME)
            res = mb.askquestion('Накладная из другой пломбы', 'Пломба не закончена, хотите переключиться на другую?')
            if res == 'yes':
                dTDa2 = None
                parcel_plomb = 0
                df_plomb = 0
                refuse_parcels_list = []
                parcels_list = []
                search_parcel_continue()
            else:
                entry_parcel.delete(0, 'end')
                entry_parcel.focus_set()
        else:
            search_parcel_continue()
    except Exception as e:
        print(e)
        search_parcel_continue()
        pass

def search_parcel_continue():
    global dTDa2
    if dTDa1:
        dTDa1.destroy()
    global refuse_parcels_list
    global parcels_list
    global parcel_plomb
    global len_parcels_list
    global df_plomb
    try:
        parcel = str(entry_parcel.get())
        parcel_plomb = df.loc[df['Трек-номер'] == parcel, 'Пломба'].values[0]
        df_plomb = df.loc[df['Пломба'] == parcel_plomb]
        print(len(df_plomb))
        df_refuse_parcel_in_plomb = df_plomb.loc[df_plomb['Статус_ТО'] == 'ИЗЪЯТИЕ']
        df_refuse_parcel_in_plomb = df_refuse_parcel_in_plomb.drop_duplicates(subset='Трек-номер', keep='first')
        quont_refuse_parcels_inplomb = len(df_refuse_parcel_in_plomb)
        if len(list(set(parcels_list))) == 0:
            df_plomb = df_plomb.reindex(
                columns=['Пломба', 'Трек-номер', 'Статус_ТО', 'Вес брутто'])
            df_plomb = df_plomb.sort_values(by=['Трек-номер'])
            df_plomb = df_plomb.reset_index(drop=True)
            raws1 = df_plomb[df_plomb["Статус_ТО"] != 'ВЫПУСК'].index
            print(raws1)
            dTDa2 = tk.Toplevel(width='800', height=750)
            dTDa2.title('Накладные')
            dTDaPT2 = pt.Table(dTDa2, dataframe=df_plomb,
                               showtoolbar=True, showstatusbar=True,
                               width='800', height='750', maxcellwidth=500)
            dTDaPT2.setRowHeight(60)
            dTDaPT2.setFont()
            pt.config.apply_options(options, dTDaPT2)
            dTDaPT2.setRowColors(rows=list(raws1), clr='red', cols='all')
            dTDaPT2.show()


    except Exception as e:
        print(e)
    try:
        if df.loc[df['Трек-номер'] == parcel, 'Статус_ТО'].values[0] == 'ВЫПУСК':
            winsound.PlaySound('Snd_Issue.wav', winsound.SND_FILENAME)
            parcels_list.append(parcel)
            print(parcels_list)
            len_parcels_list = len(list(set(parcels_list)))
            print(len_parcels_list)
            try:
                if refuse_parcels_list is not None:
                    len_refuse_parcels_list = len(list(set(refuse_parcels_list)))
                    label_refuse_parc.config(
                        text=f'Номер пломбы: \n{parcel_plomb}\nнайдено отказных: {len_refuse_parcels_list}\n(из {quont_refuse_parcels_inplomb}'
                             f'\nВсего обработано: {len_parcels_list})',
                        fg='black', font=('Times', 20))
                else:
                    label_refuse_parc.config(
                        text=f'Номер пломбы: \n{parcel_plomb}\nнайдено отказных: {0}\n(из {quont_refuse_parcels_inplomb}'
                             f'\nВсего обработано: {len_parcels_list})',
                        fg='black', font=('Times', 20))
            except EXCEPTION:
                pass
            entry_parcel.delete(0, 'end')
            entry_parcel.focus_set()
        elif df.loc[df['Трек-номер'] == parcel, 'Статус_ТО'].values[0] == 'ИЗЪЯТИЕ':
            winsound.PlaySound('Snd_CancelIssue.wav', winsound.SND_FILENAME)
            try:
                refuse_parcels_list.append(parcel)
                parcels_list.append(parcel)
                len_parcels_list = len(list(set(parcels_list)))
                len_refuse_parcels_list = len(list(set(refuse_parcels_list)))
                label_refuse_parc.config(
                    text=f'Номер пломбы: \n{parcel_plomb}\nнайдено отказных: {len_refuse_parcels_list}\n(из {quont_refuse_parcels_inplomb}'
                         f'\nВсего обработано: {len_parcels_list})',
                    fg='black', font=('Times', 20))
            except EXCEPTION as e:
                print(e)
            if len_refuse_parcels_list == quont_refuse_parcels_inplomb:
                winsound.PlaySound('Snd_All_Parcel_Found.wav', winsound.SND_FILENAME)
                entry_parcel.delete(0, 'end')
                entry_parcel.focus_set()
            else:
                entry_parcel.delete(0, 'end')
                entry_parcel.focus_set()
        if len_parcels_list == len(df_plomb):
            winsound.PlaySound('Snd_MestoEnd.wav', winsound.SND_FILENAME)
            dTDa2.destroy()
            refuse_parcels_list = []
            parcels_list = []
    except Exception as e:
        print(e)
        winsound.PlaySound('Snd_Parcel_Not_Found.wav',
                           winsound.SND_FILENAME)
        entry_parcel.delete(0, 'end')
        entry_parcel.focus_set()

def remains():
    place = entry_place.get()
    print(place)
    if place == "":
        msg = f"Значение Место должно быть заполненно!"
        mb.showinfo("Нет значения Места", msg)
    else:
        keep = ['Выпуск']
        print(df_true_status_names)
        df_remains = df_true_status_names[~df_true_status_names.Статус_ТО.str.contains('|'.join(keep))]
        df_remains['Пломба'] = entry_place.get()
        print(df_remains)
        writer = pd.ExcelWriter(f'Не выпущенные {now} {party_numb}.xlsx', engine='xlsxwriter')
        df_remains.to_excel(writer, sheet_name='Sheet1', index=False)
        for column in df_remains:
            column_width = max(df_remains[column].astype(str).map(len).max(), len(column))
            col_idx = df_remains.columns.get_loc(column)
            writer.sheets['Sheet1'].set_column(col_idx, col_idx, column_width)
            writer.sheets['Sheet1'].set_column(0, 3, 10)
            writer.sheets['Sheet1'].set_column(1, 3, 20)
            writer.sheets['Sheet1'].set_column(2, 3, 20)
            writer.sheets['Sheet1'].set_column(3, 3, 20)
            writer.sheets['Sheet1'].set_column(4, 3, 30)
            writer.sheets['Sheet1'].set_column(5, 3, 20)
        writer.save()
        winsound.PlaySound('priezjayte-k-nam-esche.wav', winsound.SND_FILENAME)
        msg = f"Не выпущенные {now} ({party_numb}) файл сформирован"
        mb.showinfo("Не выпущенные", msg)

dTDa1 = None
dTDa2 = None
parcel_plomb = 0
df_plomb = 0
len_parcels_list = 0
parcels_list = []
parcel = 0
window = tk.Tk()
window.title('ОТБОР МЕШКОВ')
window.geometry("300x250+1200+100")

window1 = tk.Tk()
window1.title('ОТБОР ПОСЫЛКОК')
window1.geometry("350x320+790+440")

window2 = tk.Tk()
window2.title('НОВЫЙ НОМЕР ПЛОМБЫ')
window2.geometry("350x320+1150+440")

window3 = tk.Tk()
window3.title('МАНИФЕСТ')
window3.geometry("400x300+790+100")
"""
window4 = tk.Tk()
window4.title('Довыпуски')
window4.geometry("400x150+800+750")
"""
# 0
entry_plomb_lb = tk.Label(window, text="ПЛОМБА", font='hank 19 bold')
entry_plomb = tk.Entry(window, font='hank 40 bold', width=20)
label = Label(window, text=f'обработанно: {len_done_plomb} \nиз {plomb_quont} (откзн:{refuse_plomb_quont})',
              fg='black', font=('Times', 20))
# 1
entry_parcel_lb = tk.Label(window1, text="Трек номер", font='hank 19 bold')
entry_parcel = tk.Entry(window1, font='hank 40 bold', width=20)
label_refuse_parc = Label(window1, text=f'Номер пломбы: \nнайдено отказных: (из )',
                          fg='black', font=('Times', 20))
# 2
entry_plomb_old_lb = tk.Label(window2, text="СТАРАЯ ПЛОМБА", font='hank 19 bold')
entry_plomb_old = tk.Entry(window2, font='hank 40 bold', width=20)
entry_NEW_plomb_lb = tk.Label(window2, text="НОВАЯ ПЛОМБА", font='hank 19 bold')
entry_NEW_plomb = tk.Entry(window2, font='hank 40 bold', width=20)
button = tk.Button(window2, text="ОК", width=15, height=2, bg="lightgrey", fg="black", font='hank 12 bold', command=save_df_new_plomb)

# 3
entry_plomb_manifest_lb = tk.Label(window3, text="ПЛОМБА в Манифест", font='hank 19 bold')
entry_place_lb = tk.Label(window3, text="Место для отказных", font='hank 15')
entry_plomb_manifest = tk.Entry(window3, font='hank 40 bold', width=20)
entry_place = tk.Entry(window3, font='hank 15 bold', width=12)
button_manifest = tk.Button(window3, text="Сформировать манифест", width=25, height=2, bg="lightgrey", fg="black",
                            font='hank 12 bold', command=save_manifest)
button_remains = tk.Button(window3, text="Выгрузить не выпущенные", width=25, height=2, bg="lightgrey", fg="black",
                            font='hank 12 bold', command=remains)
# 4

window.bind('<Return>', plomb_serch)
window1.bind('<Return>', search_parcel)
window2.bind('<Return>', old_plomb)
window3.bind('<Return>', plomb_to_manifest)

# 0
entry_plomb_lb.pack()
entry_plomb.pack()
label.pack()

# 1
entry_parcel_lb.pack()
entry_parcel.pack()
label_refuse_parc.pack()
# 2
entry_plomb_old_lb.pack()
entry_plomb_old.pack()
entry_NEW_plomb_lb.pack()
entry_NEW_plomb.pack()
button.pack()
# 3
entry_plomb_manifest_lb.pack()
entry_plomb_manifest.pack()
button_manifest.pack()
entry_place_lb.pack()
entry_place.pack()
button_remains.pack()
# 4


window.mainloop()
window2.mainloop()
