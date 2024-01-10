import pandas as pd
import openpyxl
from tkinter import filedialog
import tkinter.messagebox as mb
import datetime
from openpyxl.worksheet.page import PageMargins
import os

#xlApp = client.Dispatch("Excel.Application")

now = datetime.datetime.now().strftime("%d.%m.%Y")
def varian_1():
    curse_RUB_EUR = 104
    curse_CN_EUR = 7.73
    msg = "Выберите файл с Шаблоном партии"
    mb.showinfo("Файл Шаблона", msg)
    file_name = filedialog.askopenfilename()
    df = pd.read_excel(file_name, sheet_name=0, engine='openpyxl')
    df = df[['Номер отправления ИМ', 'ФИО получателя', 'Общая стоимость накладной(посылки)', 'Валюта Объявленной стоимости товара']]
    currency = df['Валюта Объявленной стоимости товара'].values[0]
    print(currency)
    if currency == 'RUB':
        df['Общая стоимость накладной(посылки)'] = df['Общая стоимость накладной(посылки)'].apply(lambda x: x / curse_RUB_EUR)
    elif currency == 'CNY':
        df['Общая стоимость накладной(посылки)'] = df['Общая стоимость накладной(посылки)'].apply(lambda x: x / curse_CN_EUR)
    df = df.loc[df['Общая стоимость накладной(посылки)'] > 1000]
    print(df['Общая стоимость накладной(посылки)'])

    i = 757
    for parcel_numb in df['Номер отправления ИМ']:
        i += 1
        buyer = df.loc[df['Номер отправления ИМ'] == parcel_numb]['ФИО получателя'].values[0]
        cost = df.loc[df['Номер отправления ИМ'] == parcel_numb]['Общая стоимость накладной(посылки)'].values[0]
        fees = 500
        if currency == 'RUB':
            duty = (cost - 1000) * 0.15 * curse_RUB_EUR
        elif currency == 'CNY':
            duty = (cost - 1000) * 0.15 * curse_CN_EUR
        comission = (duty + fees) * 0.05
        sum_all = fees + duty + comission

        wb = openpyxl.load_workbook('СЧЕТ_шаблон.xlsx')
        ws = wb.active
        ws['A14'].value = f'Счёт на оплату № {i} от {now}'
        ws['B16'].value = buyer
        ws['B18'].value = parcel_numb
        ws['F21'].value = fees
        ws['G21'].value = fees
        ws['F22'].value = duty
        ws['G22'].value = duty
        ws['F23'].value = comission
        ws['G23'].value = comission
        ws['G25'].value = sum_all
        ws['G27'].value = sum_all
        ws['A25'].value = f'Всего наименований 3 на {sum_all} (сумма прописью)'

        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToHeight = False
        cm = 1 / 4
        ws.page_margins = PageMargins(left=cm, right=cm, top=cm, bottom=cm)

        wb.save(f'Счёт № {i} от {now}.xlsx')

    msg = f'Счета сформированы'
    mb.showinfo("Информация", msg)

def variant_2():
    msg = "Выберите файл с платниками"
    mb.showinfo("Файл Шаблона", msg)
    file_name = filedialog.askopenfilename()
    only_file_name = os.path.basename(file_name)
    print(only_file_name)
    used_folder = f'{only_file_name}'
    if not os.path.isdir(used_folder):
        os.makedirs(used_folder, exist_ok=True)
    df = pd.read_excel(file_name, sheet_name=0, engine='openpyxl')
    df = df[['Трек-номер', 'Получатель', 'Таможенная пошлина',
             'Таможенные сборы']]
    print(df)
    with open("bill_number.txt") as bill_number_file:
        i = int(bill_number_file.read())
    for parcel_numb in df['Трек-номер']:
        i += 1
        buyer = df.loc[df['Трек-номер'] == parcel_numb]['Получатель'].values[0]
        duty = df.loc[df['Трек-номер'] == parcel_numb]['Таможенная пошлина'].replace(to_replace=',', value='.', regex=True).astype(float).values[0]
        fees = df.loc[df['Трек-номер'] == parcel_numb]['Таможенные сборы'].replace(to_replace=',', value='.', regex=True).astype(float).values[0]
        comission = round(((duty + fees) * 0.05), 2)
        sum_all = round(fees + duty + comission, 2)

        wb = openpyxl.load_workbook('СЧЕТ_шаблон.xlsx')
        ws = wb.active
        ws['A14'].value = f'Счёт на оплату № {i} от {now}'
        ws['B16'].value = buyer
        ws['B18'].value = parcel_numb
        ws['F21'].value = fees
        ws['G21'].value = fees
        ws['F22'].value = duty
        ws['G22'].value = duty
        ws['F23'].value = comission
        ws['G23'].value = comission
        ws['G25'].value = sum_all
        ws['G27'].value = sum_all
        ws['A25'].value = f'Всего наименований 3 на {sum_all} рублей'

        #ws.sheet_properties.pageSetUpPr.fitToPage = True
        #ws.page_setup.fitToHeight = False
        cm = 1 / 4
        ws.page_margins = PageMargins(left=cm, right=cm, top=cm, bottom=cm)

        wb.save(f'{used_folder}/{buyer} {parcel_numb} № {i} от {now}.xlsx')
        #pdfkit.from_file(f'Счёт № {i} от {now}.xlsx', f'Счёт № {i} от {now}.pdf', configuration=config)
        #books = xlApp.Workbooks.Open('Мингалеев Владислав Рустемович CEL7000161380CD № 1015 от 15.10.2023.xlsx')
        #ws = books.Worksheets[0]
        # ws.Visible = 1
        #ws.ExportAsFixedFormat(0, f'{buyer} {parcel_numb} № {i} от {now}.pdf')
    msg = f'Счета сформированы'
    mb.showinfo("Информация", msg)
    with open("bill_number.txt", 'w') as bill_number_file:
        bill_number_file.write(str(i))


variant_2()