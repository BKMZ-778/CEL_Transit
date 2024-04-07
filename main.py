import pandas as pd
import numpy as np
import tkinter as tk
from tkinter import *
from tkinter import filedialog
import tkinter.messagebox as mb
import datetime
import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from copy import copy
from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet.page import PageMargins
import random
import xlsxwriter
import os

now = datetime.datetime.now().strftime("%d.%m.%Y")
#now = (datetime.datetime.now() - timedelta(days=1)).strftime("%d.%m.%Y")  #Yesterday date
now1 = datetime.datetime.now().strftime("%d%m")

df_trigers = pd.read_excel('Triger_dict.xlsx')
print(df_trigers)
triger_dict = df_trigers.set_index('trigger').to_dict()['weight']

triger_list = list(triger_dict.keys())



selection1 = "ШАБЛОН МТК СЭЛ VL"
def get_selection(selection):
    print(selection)
    global selection1
    selection1 = selection
    print(selection1)
    return selection1

def check_party():
    num_party = entry_party.get()
    now = entry_date.get()
    numberAndDate = f'{num_party} от {now}'
    msg = "Выберите файл с реестром"
    mb.showinfo("批包括", msg)
    file_name = filedialog.askopenfilename()
    df = pd.read_excel(file_name, sheet_name=0, header=None, engine='openpyxl',
                       usecols='A, K, L, M, N, T', dtype={'M': str, 'T': str}, skiprows=1)
    df.columns = ['Номер накладной', 'Кол.', 'Наименование', 'цена', 'ссылка', 'вес товара']
    print(df)
    price_err = df['цена'].isnull().any()
    if price_err == True:
        msg = "Есть пустые ячейки в столбце цена, исправьте и начните заново!"
        mb.showerror("Столбец цена: ошибка", msg)
    else:
        pass
    if df['цена'].isin([0]).any():
        msg = "Есть нулевые (0) ячейки в столбце цена, исправьте и начните заново!"
        mb.showerror("Столбец цена: ошибка", msg)
    df['цена'] = df['цена'].replace(to_replace=',,', value='.', regex=True)
    df['цена'] = df['цена'].replace(to_replace='\.\.', value='.', regex=True)
    df['цена'] = df['цена'].replace(to_replace=',', value='.', regex=True)
    df['цена'] = df['цена'].replace(to_replace='，', value='.', regex=True)
    df['цена'] = df['цена'].replace(to_replace='^\.', value='', regex=True)
    df['цена'] = df['цена'].replace(to_replace='^,', value='', regex=True)
    df['цена'] = df['цена'].replace(to_replace=',$', value='', regex=True)
    try:
        df['цена'] = df['цена'].replace(to_replace='\.$', value='', regex=True).astype('float')
    except ValueError as msg:
        mb.showinfo("Столбец цена: присутствует нечисловое значение!",
                    f'{str(msg)}\nПроверьте выставлены ли столбцы по шаблону, \nзатем проверьте сам столбец с ценой!')

    df['цена'] = df['цена'].round(2)

    df['Стоимость'] = df['Кол.'].multiply(df['цена'], axis='index')

    weight_err = df['вес товара'].isnull().any()
    if weight_err == True:
        msg = "Есть пустые ячейки в столбце вес, исправьте и начните заново!"
        mb.showerror("Столбец вес: ошибка", msg)
    else:
        pass
    if df['вес товара'].isin([0]).any():
        msg = "Есть нулевые '0' ячейки в столбце вес, исправьте и начните заново!"
        mb.showerror("Столбец вес: ошибка", msg)
    df['вес товара'] = df['вес товара'].replace(to_replace=',,', value='.', regex=True)
    df['вес товара'] = df['вес товара'].replace(to_replace='\.\.', value='.', regex=True)
    df['вес товара'] = df['вес товара'].replace(to_replace=',', value='.', regex=True)
    df['вес товара'] = df['вес товара'].replace(to_replace='，', value='.', regex=True)
    df['вес товара'] = df['вес товара'].replace(to_replace='^\.', value='', regex=True)
    df['вес товара'] = df['вес товара'].replace(to_replace='^,', value='', regex=True)
    df['вес товара'] = df['вес товара'].replace(to_replace=',$', value='', regex=True)
    try:
        df['вес товара'] = df['вес товара'].replace(to_replace='\.$', value='', regex=True).astype('float')
    except ValueError as msg:
        mb.showinfo("Столбец вес: присутствует нечисловое значение!",
                    f'{str(msg)}\nПроверьте выставлены ли столбцы по шаблону, \nзатем проверьте сам столбец с весом!')

    try:
        cost = df['Стоимость']
        name = df['Наименование']
        weight = df['вес товара']
        parcel_numb = df['Номер накладной']
        good_link = df['ссылка']
        good_quont = df['Кол.']
        warning_df = pd.DataFrame()
        for i in range(0, len(df)):
            print(i)
            good_weight = weight[i] / good_quont[i]
            for trigger in triger_list:
                trigger_weight = triger_dict[trigger]
                if trigger.lower() in name[i].lower():
                    if good_weight >= trigger_weight:
                        df_to_append = pd.DataFrame({'parcel_numb': [parcel_numb[i]], 'name': [name[i]],
                                                     'good_link': [good_link[i]], 'weight': good_weight,
                                                     'trigger': [trigger], 'trigger_weight': [trigger_weight],
                                                     'cost': [cost[i]], 'class': 0})
                        warning_df = pd.concat([warning_df, df_to_append])
                elif weight[i] >= 10:
                    df_to_append = pd.DataFrame({'parcel_numb': [parcel_numb[i]], 'name': [name[i]],
                                                 'good_link': [good_link[i]], 'weight': good_weight,
                                                 'trigger': '', 'trigger_weight': '',
                                                 'cost': [cost[i]], 'class': 1})
                    warning_df = pd.concat([warning_df, df_to_append])
                    break
            if cost[i] > 1000:
                df_to_append = pd.DataFrame({'parcel_numb': [parcel_numb[i]], 'name': [name[i]],
                                             'good_link': [good_link[i]], 'weight': good_weight,
                                             'trigger': '', 'trigger_weight': '',
                                             'cost': [cost[i]], 'class': 2})
                warning_df = pd.concat([warning_df, df_to_append])

        warning_df = warning_df.sort_values(by=['class', 'weight'])
        print(warning_df)
        writer = pd.ExcelWriter(f'WARNING_{num_party}.xlsx', engine='xlsxwriter')
        warning_df.to_excel(writer, sheet_name='Sheet1', index=False)
        writer.save()
        msg = f"Готово!"
        mb.showinfo("Риски", msg)
    except Exception as e:
        msg = f"Не удалось сформировать Warning list, возможно нет рисков, ошибки: {e}"
        mb.showinfo("Триггер", msg)
        pass

def start():
    file_name_sample = f'{Sample_choice_var.get()}.xlsx'
    num_party = entry_party.get()
    now = entry_date.get()
    numberAndDate = f'{num_party} от {now}'
    msg = "Выберите файл с реестром"
    mb.showinfo("批包括", msg)
    file_name = filedialog.askopenfilename()
    print(selection1)
    df = pd.read_excel(file_name, sheet_name=0, header=None, engine='openpyxl',
                       usecols='A, K, L, M, N, T', dtype={'M': str, 'T': str}, skiprows=1)
    df.columns = ['Номер накладной', 'Кол.', 'Наименование', 'цена', 'ссылка', 'вес товара']
    print(df)
    price_err = df['цена'].isnull().any()
    if price_err == True:
        msg = "Есть пустые ячейки в столбце цена, исправьте и начните заново!"
        mb.showerror("Столбец цена: ошибка", msg)
    else:
        pass
    if df['цена'].isin([0]).any():
        msg = "Есть нулевые (0) ячейки в столбце цена, исправьте и начните заново!"
        mb.showerror("Столбец цена: ошибка", msg)
    df['цена'] = df['цена'].replace(to_replace=',,', value='.', regex=True)
    df['цена'] = df['цена'].replace(to_replace='\.\.', value='.', regex=True)
    df['цена'] = df['цена'].replace(to_replace=',', value='.', regex=True)
    df['цена'] = df['цена'].replace(to_replace='，', value='.', regex=True)
    df['цена'] = df['цена'].replace(to_replace='^\.', value='', regex=True)
    df['цена'] = df['цена'].replace(to_replace='^,', value='', regex=True)
    df['цена'] = df['цена'].replace(to_replace=',$', value='', regex=True)
    try:
        df['цена'] = df['цена'].replace(to_replace='\.$', value='', regex=True).astype('float')
    except ValueError as msg:
        mb.showinfo("Столбец цена: присутствует нечисловое значение!", f'{str(msg)}\nПроверьте выставлены ли столбцы по шаблону, \nзатем проверьте сам столбец с ценой!')

    df['цена'] = df['цена'].round(2)
    print(selection1)
    if selection1 == "ШАБЛОН МТТ" or selection1 == "ШАБЛОН Белаонова CaiNiao" or selection1 == "ШАБЛОН РЭ":
        df_cost_err_much = df.loc[df['цена'] >= 900]
        print(df_cost_err_much)
        if not df_cost_err_much.empty:
            msg = df_cost_err_much['цена'].tolist()
            mb.showerror("Столбец цена: присутствуют большие цены, мы уменьшим их", msg)
            df.loc[df["цена"] >= 900, "цена"] = random.randint(20, 40)
        else:
            pass
        try:
            df['цена'] = df['цена'].apply(lambda x: round(x * float(entry_koef.get().replace(',', '.')), 2))
        except ValueError as msg1:
            mb.showinfo("Поле коэффициент уменьшения цены: присутствует нечисловое значение!", f'{str(msg1)}')
    else:
        pass

    df['Стоимость'] = df['Кол.'].multiply(df['цена'], axis='index')

    weight_err = df['вес товара'].isnull().any()
    if weight_err == True:
        msg = "Есть пустые ячейки в столбце вес, исправьте и начните заново!"
        mb.showerror("Столбец вес: ошибка", msg)
    else:
        pass
    if df['вес товара'].isin([0]).any():
        msg = "Есть нулевые '0' ячейки в столбце вес, исправьте и начните заново!"
        mb.showerror("Столбец вес: ошибка", msg)
    df['вес товара'] = df['вес товара'].replace(to_replace=',,', value='.', regex=True)
    df['вес товара'] = df['вес товара'].replace(to_replace='\.\.', value='.', regex=True)
    df['вес товара'] = df['вес товара'].replace(to_replace=',', value='.', regex=True)
    df['вес товара'] = df['вес товара'].replace(to_replace='，', value='.', regex=True)
    df['вес товара'] = df['вес товара'].replace(to_replace='^\.', value='', regex=True)
    df['вес товара'] = df['вес товара'].replace(to_replace='^,', value='', regex=True)
    df['вес товара'] = df['вес товара'].replace(to_replace=',$', value='', regex=True)
    try:
        df['вес товара'] = df['вес товара'].replace(to_replace='\.$', value='', regex=True).astype('float')
    except ValueError as msg:
        mb.showinfo("Столбец вес: присутствует нечисловое значение!",
                    f'{str(msg)}\nПроверьте выставлены ли столбцы по шаблону, \nзатем проверьте сам столбец с весом!')


    df_group = df.groupby('Номер накладной', sort=False).sum()
    df_group = df_group.reindex(columns=['вес товара']).rename(columns={'вес товара': 'вес Накладной'})
    df = pd.merge(df, df_group, how='left', left_on='Номер накладной' , right_on='Номер накладной')

    msg = "Выберите файл с загрузкой по мешкам"
    mb.showinfo("批所有", msg)
    file_name1 = filedialog.askopenfilename()
    print(file_name1)
    df_w = pd.read_excel(file_name1, sheet_name=0, engine='openpyxl', usecols='B,C,D,F,G',
                          dtype={'Stamp 铅封号': str})

    mapping = {df_w.columns[1]: 'Номер накладной', df_w.columns[2]: 'Трек номер',
               df_w.columns[3]: 'вес Накладной по загрузке', df_w.columns[0]: 'Пломба'}

    df_w = df_w.rename(columns=mapping)  # (columns={'Order number 单号': 'Номер накладной',
    # [4]: 'Трек номер',
    # 'Gross weight, kg  毛重 公斤': 'вес Накладной по загрузке',
    # 'Stamp 铅封号': 'Пломба'})
    print(df_w)
    weight_brut = df_w['вес Накладной по загрузке'].iloc[-1]
    weight_net = df_w['вес Накладной по загрузке'].iloc[-2]
    if weight_brut / weight_net > 2:
        weight_net = df_w['вес Накладной по загрузке'].iloc[-3]
    else:
        pass
    print(weight_brut)
    print(weight_net)
    weight_all_chinabag = weight_brut - weight_net
    count_chinabag = len(df_w['Пломба'].unique()) - 1
    Chinabag_weight = round(weight_all_chinabag / count_chinabag, 3)
    msg = f"вес брутто {weight_brut},\n вес нетто {weight_net},\n кол-во упаковок (пломб) {count_chinabag},\n\n Вес единицы упаковки {Chinabag_weight}"
    mb.showinfo("Информация по загрузке", msg)

    df_w = df_w.drop_duplicates(subset='Номер накладной', keep='first')
    if df_w['вес Накладной по загрузке'].isin([0]).any():
        msg = "Есть нулевые '0' ячейки в столбце вес Накладной по загрузке, исправьте и начните заново!"
        mb.showerror("Столбец вес Накладной по загрузке: ошибка", msg)
    if df_w['вес Накладной по загрузке'].isnull().any():
        msg = "Есть пустые ячейки в столбце вес Накладной по загрузке, исправьте и начните заново!"
        mb.showerror("Столбец вес Накладной по загрузке: ошибка", msg)

    test_cmpr_df = df['Номер накладной'].drop_duplicates()
    test_cmpr_df_w = df_w['Номер накладной'].dropna()

    for i in test_cmpr_df:
        test_comp1 = test_cmpr_df_w.isin([i]).any()
        if test_comp1 == False:
            msg = f"Накладная {i} не найдена в загрузке"
            mb.showerror("Ошибка", msg)

    for i in test_cmpr_df_w:
        test_comp2 = test_cmpr_df.isin([i]).any()
        if test_comp2 == False:
            msg = f"Накладная {i} лишняя в загрузке"
            mb.showerror("Ошибка", msg)

    df = pd.merge(df, df_w, how='left', left_on='Номер накладной', right_on='Номер накладной')
    df['Вес нетто'] = np.round(df['вес Накладной по загрузке']/df['вес Накладной']*df['вес товара'],
                       decimals=3)
    df_pc_qt = df['Пломба'].value_counts()
    df = pd.merge(df, df_pc_qt, how='left', left_on='Пломба', right_index=True)
    if not b.get():
        df['Вес брутто'] = df['Вес нетто'] + np.round(Chinabag_weight / df['Пломба_y'], decimals=3)
    else:
        bag_wht = entry_bag_wht.get().replace(',', '.')
        try:
            df['Вес брутто'] = df['Вес нетто'] + np.round(float(bag_wht) / df['Пломба_y'], decimals=3)
        except ValueError as msg:
            mb.showinfo(f'Не можем преобразовать вес мешка: {entry_bag_wht.get()}\n в число', str(msg))
    df_s = df.sort_values(by=['Пломба_x', 'Номер накладной'])
    print(df_s)

    msg = "Выберите Базу кодов ТН ВЭД"
    mb.showinfo("ТНВЭД", msg)
    file_name2 = filedialog.askopenfilename()
    print(file_name2)
    dfTnVed = pd.read_excel(file_name2, sheet_name=0, engine='openpyxl', usecols='A,B', dtype={'Наименование': str})
    dfTnVed['Наименование'] = dfTnVed['Наименование'].str.lower()
    dfTnVed = dfTnVed.drop_duplicates(subset='Наименование', keep='first')
    df_s['Наименование'] = df_s['Наименование'].str.lower()
    df_s = pd.merge(df_s, dfTnVed, how='left', left_on='Наименование', right_on='Наименование')
    df_s['Мест'] = df_s.Пломба_x.eq(df_s.Пломба_x.shift()).astype('str').replace(to_replace=['False', 'True'],
                                                                         value=['1', '0'], regex=True).astype('int')
    df_s['№ П/П'] = np.arange(len(df_s))[::+1]+1
    df_s['Ед.изм.'] = 'ШТ'
    df_s['УПАКОВКА'] = Pac_choice_var.get()
    df_s['ВЕС ФАКТ'] = None
    print(df_s)
    group_Weight_df = df_s.groupby('Пломба_x')['Вес брутто'].sum()
    group_Weight_df = group_Weight_df.rename('Вес брутто места', inplace=True)
    df_s = pd.merge(df_s, group_Weight_df, how='left', left_on='Пломба_x', right_on='Пломба_x')
    df_s = df_s.rename(columns={'Пломба_x': 'Пломба'})

    df_s = df_s.reindex(columns=['№ П/П', 'Наименование', 'Номер накладной', 'Трек номер',
                      'Пломба', 'КОД ТНВЭД', 'Кол.', 'Ед.изм.', 'Мест', 'УПАКОВКА', 'Вес брутто',
                      'Вес нетто', 'Вес брутто места', 'ВЕС ФАКТ', 'цена', 'Стоимость'])

    writer = pd.ExcelWriter('pac_for_spec.xlsx', engine='xlsxwriter')
    df_s.to_excel(writer, sheet_name='Sheet1', index=False)
    writer.save()

    df_s.columns = range(df_s.shape[1])

    df_title = pd.read_excel(file_name_sample, sheet_name='Упаковочный', engine='openpyxl', header=None,
                             skiprows=[9, 10, 11, 12, 13, 14])
    print(df_title)
    df_downtitle = pd.read_excel(file_name_sample, sheet_name='Упаковочный', engine='openpyxl', header=None,
                                 skiprows=[0, 1, 2, 3, 4, 5, 6, 7, 8, 9])
    print(df_downtitle)
    df_downtitle.replace(to_replace='{fullDate}', value=now, regex=True, inplace=True)
    df_pack = pd.concat([df_title, df_s, df_downtitle], axis=0)
    print(df_pack)
    df_pack.replace(to_replace='{numberAndDate}', value=numberAndDate, regex=True, inplace=True)
    writer = pd.ExcelWriter('PAC_CEL.xlsx', engine='xlsxwriter')
    df_pack.to_excel(writer, sheet_name='Sheet1', index=False, header=None)
    writer.save()

    wb = openpyxl.load_workbook('PAC_CEL.xlsx')
    ws = wb.active
    ws.merge_cells('A1:J1')
    ws.merge_cells('A2:A4')
    ws.merge_cells('A5:A8')
    ws.merge_cells('B2:B4')
    ws.merge_cells('B5:B8')
    ws.merge_cells('C2:D3')
    ws.merge_cells('C4:D5')
    ws.merge_cells('C6:D6')
    ws.merge_cells('C7:D8')
    ws.merge_cells('E2:J3')
    ws.merge_cells('E4:J5')
    ws.merge_cells('E6:J6')
    ws.merge_cells('E7:J7')
    ws.merge_cells('E8:J8')
    ws.merge_cells('L2:P2')
    ws.merge_cells('L3:P3')
    ws.merge_cells('L4:P4')
    ws.merge_cells('L5:P5')
    ws.merge_cells('L6:P6')
    ws.merge_cells('L7:P7')
    ws.merge_cells('L8:P8')
    ws.merge_cells('L1:P1')

    ws.column_dimensions['A'].width = 9
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 6
    ws.column_dimensions['H'].width = 6
    ws.column_dimensions['I'].width = 6
    ws.column_dimensions['J'].width = 12
    ws.column_dimensions['K'].width = 15
    ws.column_dimensions['L'].width = 10
    ws.column_dimensions['M'].width = 8
    ws.column_dimensions['N'].width = 8
    ws.column_dimensions['O'].width = 15
    ws.column_dimensions['P'].width = 15

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 85
    ws.row_dimensions[3].height = 30
    ws.row_dimensions[4].height = 12
    ws.row_dimensions[5].height = 30
    ws.row_dimensions[6].height = 15
    ws.row_dimensions[7].height = 30
    ws.row_dimensions[8].height = 15
    ws.row_dimensions[9].height = 30

    thins = Side(border_style="thin", color="00000000")
    for r, row in enumerate(range(1, 10), start=1):
        for c, col in enumerate(range(1, 17), start=1):
            ws.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
            ws.cell(row=row, column=col).border = Border(top=thins, bottom=thins, left=thins, right=thins)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="left", vertical="center", wrapText=True)

    len_A = len(ws['A']) - 3

    ws.insert_rows(len_A + 1, 1)

    ws[f"K{len_A + 1}"] = ws[f"K{len_A + 1}"].number_format = '0.000'
    ws[f"L{len_A + 1}"] = ws[f"L{len_A + 1}"].number_format = '0.000'
    ws[f"P{len_A + 1}"] = ws[f"P{len_A + 1}"].number_format = '0.00'

    ws[f"G{len_A + 1}"] = f"=SUM(G1:G{len_A})"
    ws[f"I{len_A + 1}"] = f"=SUM(I1:I{len_A})"
    ws[f"K{len_A + 1}"] = f"=SUM(K1:K{len_A})"
    ws[f"L{len_A + 1}"] = f"=SUM(L1:L{len_A})"
    ws[f"P{len_A + 1}"] = f"=SUM(P1:P{len_A})"

    ws.print_area = f'A1:P{len_A + 15}'
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = False
    cm = 1 / 4
    ws.page_margins = PageMargins(left=cm, right=cm, top=cm, bottom=cm)
    ws.auto_filter.ref = f"A9:K{len_A - 3}"

    wb.save(f'Упаковочный лист № PAC_ {entry_party.get()} от {now}.xlsx')

    msg = f'Упаковочный лист № PAC_ {entry_party.get()} сформирован'
    mb.showinfo("Информация", msg)

    s = df_s[5].isnull().any()
    if s == True:

        msg = f"Не хватает кодов ТНВЭД, сейчас они будут добавлены в {file_name2}!"
        mb.showinfo("Информация", msg)

        df = pd.read_excel('pac_for_spec.xlsx', sheet_name=0, engine='openpyxl')

        df2 = df[df['КОД ТНВЭД'].isnull()]
        df2 = df2[['Наименование', 'КОД ТНВЭД']].drop_duplicates('Наименование')
        print(df2)
        writer = pd.ExcelWriter('TNVD_NEW.xlsx', engine='openpyxl')
        df2.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
        writer.save()

        wb_new = openpyxl.load_workbook('TNVD_NEW.xlsx')
        ws_new = wb_new.active
        wb_tnvd = openpyxl.load_workbook(file_name2)
        ws_tnvd = wb_tnvd['Лист1']

        len_A = len(ws_tnvd['A'])
        for row_number, row in enumerate(ws_new):
            for col_number, cell in enumerate(row):
                ws_tnvd.cell(row_number + len_A + 1, 1, cell.value)

        ws_ktl = wb_tnvd['Каталог']
        len_ktl = len(ws_ktl['A'])
        len_new = len(ws_new['A'])

        len_A = len(ws_tnvd['A'])
        i = len_A - len_new - 1
        for r, row in enumerate(range(len_A - len_new, len_A + 1), start=1):
            i += 1
            ws_tnvd.cell(row=row,
                         column=4).value = f'=LOOKUP(2,1/SEARCH(Каталог!$A$1:$A${len_ktl},A{i}),Каталог!$A$1:$A${len_ktl})'

        wb_tnvd.save(file_name2)

        msg = f"База кодов {file_name2} обновлена,\n подберите коды и начните заново!"
        mb.showinfo("Информация", msg)

    else:
        pass


def Specf():
    file_name_sample = f'{Sample_choice_var.get()}.xlsx'
    num_party = entry_party.get()
    now = entry_date.get()
    print(num_party)
    numberAndDate = f'{num_party} от {now}'

    df = pd.read_excel('pac_for_spec.xlsx', sheet_name=0, engine='openpyxl', usecols='E,B,F,G,H,I,J,K,L,P')

    df_Specif_gr = df.groupby('КОД ТНВЭД', sort=False).sum()
    df['Наименование'] = df['Наименование'] + ' ' + df['Кол.'].astype(str) + ' ' + 'ШТ'
    df_text = df.groupby(['КОД ТНВЭД'])['Наименование'].apply(lambda x: ', '.join(x)).reset_index()

    df['Размещен в мешках / Кол-во мешков'] = \
    df.drop_duplicates(subset=['КОД ТНВЭД', 'Пломба'], keep='first').groupby(['КОД ТНВЭД'])['Пломба'].transform('count')

    df = df.drop(columns=['Наименование', 'Кол.', 'Мест', 'Вес брутто', 'Вес нетто', 'Стоимость'], axis=1)
    df_Specification = pd.merge(df_Specif_gr, df, how='outer', left_on='КОД ТНВЭД',
                                right_on='КОД ТНВЭД').drop_duplicates(subset='КОД ТНВЭД', keep='first')
    df_Specification = pd.merge(df_Specification, df_text, how='outer', left_on='КОД ТНВЭД',
                                right_on='КОД ТНВЭД').drop_duplicates(subset='КОД ТНВЭД', keep='first')
    writer = pd.ExcelWriter('TEXT.xlsx', engine='xlsxwriter')
    df_Specification.to_excel(writer, sheet_name='Sheet1', index=False)
    writer.save()
    df_Specification['№ П/П'] = np.arange(len(df_Specification))[::+1] + 1
    df_Specification = df_Specification.reindex(
        columns=['№ П/П', 'Наименование', 'Размещен в мешках / Кол-во мешков', 'КОД ТНВЭД', 'Кол.', 'Ед.изм.', 'Мест',
                 'УПАКОВКА', 'Вес брутто', 'Вес нетто', 'Стоимость'])
    len_spec = len(df_Specification)



    writer = pd.ExcelWriter('Specification.xlsx', engine='xlsxwriter')
    df_Specification.to_excel(writer, sheet_name='Sheet1', index=False)
    for column in df_Specification:
        column_width = max(df_Specification[column].astype(str).map(len).max(), len(column))
        col_idx = df_Specification.columns.get_loc(column)
        writer.sheets['Sheet1'].set_column(col_idx, col_idx, column_width)
        writer.sheets['Sheet1'].set_column(0, 3, 10)
        writer.sheets['Sheet1'].set_column(2, 3, 30)
    writer.save()

    wb2 = openpyxl.load_workbook('Specification.xlsx')
    ws2 = wb2.active
    ws2.insert_rows(1, 8)

    wb = openpyxl.load_workbook(file_name_sample)
    ws = wb['СпецификацияT']

    head = "A1:K8"  # Заголовок таблицы, в котором есть объединенные ячейки

    for _range in ws.merged_cells.ranges:
        boundaries = range_boundaries(str(_range))
        ws2.merge_cells(start_column=boundaries[0], start_row=boundaries[1],
                        end_column=boundaries[2], end_row=boundaries[3])

    for row_number, row in enumerate(ws[head]):
        for col_number, cell in enumerate(row):
            ws2.cell(row_number + 1, col_number + 1, cell.value)
            if cell.has_style:
                ws2.cell(row_number + 1, col_number + 1).font = copy(cell.font)
                ws2.cell(row_number + 1, col_number + 1).fill = copy(cell.fill)
                ws2.cell(row_number + 1, col_number + 1).border = copy(cell.border)
                ws2.cell(row_number + 1, col_number + 1).number_format = copy(cell.number_format)
                ws2.cell(row_number + 1, col_number + 1).protection = copy(cell.protection)
                ws2.cell(row_number + 1, col_number + 1).alignment = copy(cell.alignment)
                ws2.cell(row_number + 1, col_number + 1).quotePrefix = copy(cell.quotePrefix)
                ws2.cell(row_number + 1, col_number + 1).pivotButton = copy(cell.pivotButton)

    for r, row in enumerate(range(1, 10), start=1):
        for c, col in enumerate(range(1, 17), start=1):
            ws2.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
    ws2.cell(row=1, column=1).alignment = Alignment(horizontal="left", vertical="center", wrapText=True)

    wb3 = openpyxl.load_workbook(file_name_sample)
    ws3 = wb3['СпецификацияDT']

    len_A = len(ws2['A']) + 1
    for row_number, row in enumerate(ws3[head]):
        for col_number, cell in enumerate(row):
            ws2.cell(row_number + len_A, col_number + 1, cell.value)
            if cell.has_style:
                ws2.cell(row_number + len_A, col_number + 1).font = copy(cell.font)
                ws2.cell(row_number + len_A, col_number + 1).fill = copy(cell.fill)
                ws2.cell(row_number + len_A, col_number + 1).border = copy(cell.border)
                ws2.cell(row_number + len_A, col_number + 1).number_format = copy(cell.number_format)
                ws2.cell(row_number + len_A, col_number + 1).protection = copy(cell.protection)
                ws2.cell(row_number + len_A, col_number + 1).alignment = copy(cell.alignment)
                ws2.cell(row_number + len_A, col_number + 1).quotePrefix = copy(cell.quotePrefix)
                ws2.cell(row_number + len_A, col_number + 1).pivotButton = copy(cell.pivotButton)

    ws2.merge_cells('A1:D1')
    ws2.merge_cells(f'H{len_A + 6}:J{len_A + 11}')
    ws2.merge_cells(f'A{len_A + 6}:F{len_A + 11}')
    ws2.cell(1, 1).value = ws.cell(1, 1).value.replace("{numberAndDate}", numberAndDate)
    ws2.cell(8, 4).value = ws.cell(8, 4).value.replace("{numberAndDate}", numberAndDate)
    ws2.cell(len_A + 4, 10).value = now

    ws2.column_dimensions['A'].width = 9
    ws2.column_dimensions['B'].width = 40
    ws2.column_dimensions['C'].width = 17
    ws2.column_dimensions['D'].width = 17
    ws2.column_dimensions['E'].width = 12
    ws2.column_dimensions['F'].width = 10
    ws2.column_dimensions['G'].width = 6
    ws2.column_dimensions['H'].width = 14
    ws2.column_dimensions['I'].width = 17
    ws2.column_dimensions['J'].width = 12
    ws2.column_dimensions['K'].width = 17

    ws2.row_dimensions[1].height = 16
    ws2.row_dimensions[2].height = 70
    ws2.row_dimensions[3].height = 30
    ws2.row_dimensions[4].height = 12
    ws2.row_dimensions[5].height = 30
    ws2.row_dimensions[6].height = 15
    ws2.row_dimensions[7].height = 30
    ws2.row_dimensions[8].height = 15
    ws2.row_dimensions[9].height = 50

    ws2[f"K{len_A}"] = ws2[f"K{len_A}"].number_format = '0.00'
    ws2[f"K{len_A}"] = f"=SUM(K10:K{len_A-1})"

    ws2.sheet_properties.pageSetUpPr.fitToPage = True
    ws2.page_setup.fitToHeight = False
    cm = 1 / 4
    ws2.page_margins = PageMargins(left=cm, right=cm, top=cm, bottom=cm)
    ws2.auto_filter.ref = f"A9:K{len_A - 4}"

    wb2.save(f'Спецификация № {entry_party.get()} от {now}.xlsx')


    msg = f'Спецификация № {entry_party.get()} сформирована'
    mb.showinfo("Информация", msg)

    wb_cmr = openpyxl.load_workbook(file_name_sample)
    ws_cmr = wb_cmr['смр']
    ws_cmr.cell(row=3, column=32).value = f'471{now1}-{entry_party.get()}'
    ws_cmr.cell(row=23, column=10).value = f'{entry_party.get()}'
    ws_cmr.cell(row=23, column=16).value = f'{now}'

    df = pd.read_excel('pac_for_spec.xlsx', sheet_name=0, engine='openpyxl')

    total_weight = df['Вес брутто'].sum()
    total_price = df['Стоимость'].sum()
    total_places = df['Мест'].sum()
    print(total_weight)

    ws_cmr["AB26"].number_format = '0.000'
    ws_cmr[f"J35"].number_format = '0.00'
    ws_cmr["AB26"].value = total_weight
    ws_cmr[f"J35"].value = total_price
    ws_cmr['B26'].value = f'Общее кол-во мест: {total_places}, (интернет заказы для личного пользования)'
    pfd = wb_cmr['Упаковочный']  # делаем страницу активной
    wb_cmr.remove(pfd)  # и тут же удаляем ее
    pfd = wb_cmr['СпецификацияT']  # делаем страницу активной
    wb_cmr.remove(pfd)  # и тут же удаляем ее
    pfd = wb_cmr['СпецификацияDT']  # делаем страницу активной
    wb_cmr.remove(pfd)  # и тут же удаляем ее
    wb_cmr.save(f'СМР № {entry_party.get()} от {now}.xlsx')
    now_short = entry_date.get().replace('.', '')[:-4]
    msg = f'СМР № 471{now_short}-{entry_party.get()} сформирована'
    mb.showinfo("Информация", msg)

def pac_split():
    now = entry_date.get()
    file_name_pac_split = f'Упаковочный лист № PAC_ {entry_party.get()} от {now}.xlsx'
    wb_pac = openpyxl.load_workbook(file_name_pac_split)
    ws_pac = wb_pac.active
    len_pac_a = len(ws_pac['A'])
    print(len_pac_a)
    print(entry_party.get())
    print(file_name_pac_split)

    if 20000 > len_pac_a > 16000:
        wb_pac.save(f'Упаковочный лист № PAC_ {entry_party.get()}-1 от {now}.xlsx')
        wb_pac_1 = openpyxl.load_workbook(f'Упаковочный лист № PAC_ {entry_party.get()}-1 от {now}.xlsx')
        ws_pac_1 = wb_pac_1.active
        ws_pac_1['A1'].value = f'Упаковочный лист № PAC_ {entry_party.get()}-1 от {now}'
        ws_pac_1['E8'].value = f'Упаковочный лист № PAC_ {entry_party.get()}-1 от {now}'
        ws_pac_1.delete_rows(4000 + 10, len_pac_a - 4 - 4000 - 9)

        leng_p1 = len(ws_pac_1['a'])
        ws_pac_1[f"k{leng_p1 - 3}"] = ws_pac_1[f"k{leng_p1 - 3}"].number_format = '0.000'
        ws_pac_1[f"l{leng_p1 - 3}"] = ws_pac_1[f"l{leng_p1 - 3}"].number_format = '0.000'
        ws_pac_1[f"p{leng_p1 - 3}"] = ws_pac_1[f"p{leng_p1 - 3}"].number_format = '0.00'

        ws_pac_1[f"g{leng_p1 - 3}"] = f"=sum(g10:g{leng_p1 - 4})"
        ws_pac_1[f"i{leng_p1 - 3}"] = f"=sum(i10:i{leng_p1 - 4})"
        ws_pac_1[f"k{leng_p1 - 3}"] = f"=sum(k10:k{leng_p1 - 4})"
        ws_pac_1[f"l{leng_p1 - 3}"] = f"=sum(l10:l{leng_p1 - 4})"
        ws_pac_1[f"p{leng_p1 - 3}"] = f"=sum(p10:p{leng_p1 - 4})"
        ws_pac_1.print_area = f'A1:P{leng_p1 + 15}'
        ws_pac_1.sheet_view.topLeftCell = 'A1'
        ws_pac_1.views.sheetView[0].selection[0].activeCell = 'A1'
        ws_pac_1.views.sheetView[0].selection[0].sqref = 'A1'
        wb_pac_1.save(f'Упаковочный лист № PAC_ {entry_party.get()}-1 от {now}.xlsx')

        wb_pac.save(f'Упаковочный лист № PAC_ {entry_party.get()}-2 от {now}.xlsx')
        wb_pac_2 = openpyxl.load_workbook(f'Упаковочный лист № PAC_ {entry_party.get()}-2 от {now}.xlsx')
        ws_pac_2 = wb_pac_2.active
        ws_pac_2['A1'].value = f'Упаковочный лист № PAC_ {entry_party.get()}-2 от {now}'
        ws_pac_2['E8'].value = f'Упаковочный лист № PAC_ {entry_party.get()}-2 от {now}'
        ws_pac_2.delete_rows(10, 4000)
        ws_pac_2.delete_rows(8000 + 10 - 4000, len_pac_a - 4 - 8000 - 9)

        leng_p2 = len(ws_pac_2['a'])
        ws_pac_2[f"k{leng_p2 - 3}"] = ws_pac_2[f"k{leng_p2 - 3}"].number_format = '0.000'
        ws_pac_2[f"l{leng_p2 - 3}"] = ws_pac_2[f"l{leng_p2 - 3}"].number_format = '0.000'
        ws_pac_2[f"p{leng_p2 - 3}"] = ws_pac_2[f"p{leng_p2 - 3}"].number_format = '0.00'
        ws_pac_2[f"g{leng_p2 - 3}"] = f"=sum(g10:g{leng_p2 - 4})"
        ws_pac_2[f"i{leng_p2 - 3}"] = f"=sum(i10:i{leng_p2 - 4})"
        ws_pac_2[f"k{leng_p2 - 3}"] = f"=sum(k10:k{leng_p2 - 4})"
        ws_pac_2[f"l{leng_p2 - 3}"] = f"=sum(l10:l{leng_p2 - 4})"
        ws_pac_2[f"p{leng_p2 - 3}"] = f"=sum(p10:p{leng_p2 - 4})"
        ws_pac_2.print_area = f'A1:P{leng_p2 + 15}'
        ws_pac_2.sheet_view.topLeftCell = 'A1'
        ws_pac_2.views.sheetView[0].selection[0].activeCell = 'A1'
        ws_pac_2.views.sheetView[0].selection[0].sqref = 'A1'
        wb_pac_2.save(f'Упаковочный лист № PAC_ {entry_party.get()}-2 от {now}.xlsx')

        wb_pac.save(f'Упаковочный лист № PAC_ {entry_party.get()}-3 от {now}.xlsx')
        wb_pac_3 = openpyxl.load_workbook(f'Упаковочный лист № PAC_ {entry_party.get()}-3 от {now}.xlsx')
        ws_pac_3 = wb_pac_3.active
        ws_pac_3['A1'].value = f'Упаковочный лист № PAC_ {entry_party.get()}-3 от {now}'
        ws_pac_3['E8'].value = f'Упаковочный лист № PAC_ {entry_party.get()}-3 от {now}'
        ws_pac_3.delete_rows(10, 8000)
        ws_pac_3.delete_rows(12000 + 10 - 8000, len_pac_a - 4 - 12000 - 9)

        leng_p3 = len(ws_pac_3['a'])
        ws_pac_3[f"k{leng_p3 - 3}"] = ws_pac_3[f"k{leng_p3 - 3}"].number_format = '0.000'
        ws_pac_3[f"l{leng_p3 - 3}"] = ws_pac_3[f"l{leng_p3 - 3}"].number_format = '0.000'
        ws_pac_3[f"p{leng_p3 - 3}"] = ws_pac_3[f"p{leng_p3 - 3}"].number_format = '0.00'
        ws_pac_3[f"g{leng_p3 - 3}"] = f"=sum(g10:g{leng_p3 - 4})"
        ws_pac_3[f"i{leng_p3 - 3}"] = f"=sum(i10:i{leng_p3 - 4})"
        ws_pac_3[f"k{leng_p3 - 3}"] = f"=sum(k10:k{leng_p3 - 4})"
        ws_pac_3[f"l{leng_p3 - 3}"] = f"=sum(l10:l{leng_p3 - 4})"
        ws_pac_3[f"p{leng_p3 - 3}"] = f"=sum(p10:p{leng_p3 - 4})"
        ws_pac_3.print_area = f'A1:P{leng_p3 + 15}'
        ws_pac_3.sheet_view.topLeftCell = 'A1'
        ws_pac_3.views.sheetView[0].selection[0].activeCell = 'A1'
        ws_pac_3.views.sheetView[0].selection[0].sqref = 'A1'
        wb_pac_3.save(f'Упаковочный лист № PAC_ {entry_party.get()}-3 от {now}.xlsx')

        wb_pac.save(f'Упаковочный лист № PAC_ {entry_party.get()}-4 от {now}.xlsx')
        wb_pac_4 = openpyxl.load_workbook(f'Упаковочный лист № PAC_ {entry_party.get()}-4 от {now}.xlsx')
        ws_pac_4 = wb_pac_4.active
        ws_pac_4['A1'].value = f'Упаковочный лист № PAC_ {entry_party.get()}-4 от {now}'
        ws_pac_4['E8'].value = f'Упаковочный лист № PAC_ {entry_party.get()}-4 от {now}'
        ws_pac_4.delete_rows(10, 12000)
        ws_pac_4.delete_rows(16000 + 10 - 12000, len_pac_a - 4 - 16000 - 9)

        leng_p4 = len(ws_pac_4['a'])
        print(leng_p4)
        ws_pac_4[f"k{leng_p4 - 3}"] = ws_pac_4[f"k{leng_p4 - 3}"].number_format = '0.000'
        ws_pac_4[f"l{leng_p4 - 3}"] = ws_pac_4[f"l{leng_p4 - 3}"].number_format = '0.000'
        ws_pac_4[f"p{leng_p4 - 3}"] = ws_pac_4[f"p{leng_p4 - 3}"].number_format = '0.00'
        ws_pac_4[f"g{leng_p4 - 3}"] = f"=sum(g10:g{leng_p4 - 4})"
        ws_pac_4[f"i{leng_p4 - 3}"] = f"=sum(i10:i{leng_p4 - 4})"
        ws_pac_4[f"k{leng_p4 - 3}"] = f"=sum(k10:k{leng_p4 - 4})"
        ws_pac_4[f"l{leng_p4 - 3}"] = f"=sum(l10:l{leng_p4 - 4})"
        ws_pac_4[f"p{leng_p4 - 3}"] = f"=sum(p10:p{leng_p4 - 4})"
        ws_pac_4.print_area = f'A1:P{leng_p4 + 15}'
        ws_pac_4.sheet_view.topLeftCell = 'A1'
        ws_pac_4.views.sheetView[0].selection[0].activeCell = 'A1'
        ws_pac_4.views.sheetView[0].selection[0].sqref = 'A1'
        wb_pac_4.save(f'Упаковочный лист № PAC_ {entry_party.get()}-4 от {now}.xlsx')

        wb_pac.save(f'Упаковочный лист № PAC_ {entry_party.get()}-5 от {now}.xlsx')
        wb_pac_5 = openpyxl.load_workbook(f'Упаковочный лист № PAC_ {entry_party.get()}-5 от {now}.xlsx')
        ws_pac_5 = wb_pac_5.active
        ws_pac_5['A1'].value = f'Упаковочный лист № PAC_ {entry_party.get()}-5 от {now}'
        ws_pac_5['E8'].value = f'Упаковочный лист № PAC_ {entry_party.get()}-5 от {now}'
        ws_pac_5.delete_rows(10, 16000)

        leng_p5 = len(ws_pac_5['a'])
        print(leng_p4)
        ws_pac_5[f"k{leng_p5 - 3}"] = ws_pac_5[f"k{leng_p5 - 3}"].number_format = '0.000'
        ws_pac_5[f"l{leng_p5 - 3}"] = ws_pac_5[f"l{leng_p5 - 3}"].number_format = '0.000'
        ws_pac_5[f"p{leng_p5 - 3}"] = ws_pac_5[f"p{leng_p5 - 3}"].number_format = '0.00'
        ws_pac_5[f"g{leng_p5 - 3}"] = f"=sum(g10:g{leng_p5 - 4})"
        ws_pac_5[f"i{leng_p5 - 3}"] = f"=sum(i10:i{leng_p5 - 4})"
        ws_pac_5[f"k{leng_p5 - 3}"] = f"=sum(k10:k{leng_p5 - 4})"
        ws_pac_5[f"l{leng_p5 - 3}"] = f"=sum(l10:l{leng_p5 - 4})"
        ws_pac_5[f"p{leng_p5 - 3}"] = f"=sum(p10:p{leng_p5 - 4})"
        ws_pac_5.print_area = f'A1:P{leng_p5 + 15}'
        ws_pac_5.sheet_view.topLeftCell = 'A1'
        ws_pac_5.views.sheetView[0].selection[0].activeCell = 'A1'
        ws_pac_5.views.sheetView[0].selection[0].sqref = 'A1'
        wb_pac_5.save(f'Упаковочный лист № PAC_ {entry_party.get()}-5 от {now}.xlsx')

    elif 16000 > len_pac_a > 12000:
        wb_pac.save(f'Упаковочный лист № PAC_ {entry_party.get()}-1 от {now}.xlsx')
        wb_pac_1 = openpyxl.load_workbook(f'Упаковочный лист № PAC_ {entry_party.get()}-1 от {now}.xlsx')
        ws_pac_1 = wb_pac_1.active
        ws_pac_1['A1'].value = f'Упаковочный лист № PAC_ {entry_party.get()}-1 от {now}'
        ws_pac_1['E8'].value = f'Упаковочный лист № PAC_ {entry_party.get()}-1 от {now}'
        ws_pac_1.delete_rows(4000 + 10, len_pac_a - 4 - 4000 - 9)

        leng_p1 = len(ws_pac_1['a'])
        ws_pac_1[f"k{leng_p1 - 3}"] = ws_pac_1[f"k{leng_p1 - 3}"].number_format = '0.000'
        ws_pac_1[f"l{leng_p1 - 3}"] = ws_pac_1[f"l{leng_p1 - 3}"].number_format = '0.000'

        ws_pac_1[f"p{leng_p1 - 3}"] = ws_pac_1[f"p{leng_p1 - 3}"].number_format = '0.00'

        ws_pac_1[f"g{leng_p1 - 3}"] = f"=sum(g10:g{leng_p1 - 4})"
        ws_pac_1[f"i{leng_p1 - 3}"] = f"=sum(i10:i{leng_p1 - 4})"
        ws_pac_1[f"k{leng_p1 - 3}"] = f"=sum(k10:k{leng_p1 - 4})"
        ws_pac_1[f"l{leng_p1 - 3}"] = f"=sum(l10:l{leng_p1 - 4})"
        ws_pac_1[f"p{leng_p1 - 3}"] = f"=sum(p10:p{leng_p1 - 4})"
        ws_pac_1.print_area = f'A1:P{leng_p1 + 15}'
        ws_pac_1.sheet_view.topLeftCell = 'A1'
        ws_pac_1.views.sheetView[0].selection[0].activeCell = 'A1'
        ws_pac_1.views.sheetView[0].selection[0].sqref = 'A1'
        wb_pac_1.save(f'Упаковочный лист № PAC_ {entry_party.get()}-1 от {now}.xlsx')

        wb_pac.save(f'Упаковочный лист № PAC_ {entry_party.get()}-2 от {now}.xlsx')
        wb_pac_2 = openpyxl.load_workbook(f'Упаковочный лист № PAC_ {entry_party.get()}-2 от {now}.xlsx')
        ws_pac_2 = wb_pac_2.active
        ws_pac_2['A1'].value = f'Упаковочный лист № PAC_ {entry_party.get()}-2 от {now}'
        ws_pac_2['E8'].value = f'Упаковочный лист № PAC_ {entry_party.get()}-2 от {now}'
        ws_pac_2.delete_rows(10, 4000)
        ws_pac_2.delete_rows(8000 + 10 - 4000, len_pac_a - 4 - 8000 - 9)

        leng_p2 = len(ws_pac_2['a'])
        ws_pac_2[f"k{leng_p2 - 3}"] = ws_pac_2[f"k{leng_p2 - 3}"].number_format = '0.000'
        ws_pac_2[f"l{leng_p2 - 3}"] = ws_pac_2[f"l{leng_p2 - 3}"].number_format = '0.000'
        ws_pac_2[f"p{leng_p2 - 3}"] = ws_pac_2[f"p{leng_p2 - 3}"].number_format = '0.00'
        ws_pac_2[f"g{leng_p2 - 3}"] = f"=sum(g10:g{leng_p2 - 4})"
        ws_pac_2[f"i{leng_p2 - 3}"] = f"=sum(i10:i{leng_p2 - 4})"
        ws_pac_2[f"k{leng_p2 - 3}"] = f"=sum(k10:k{leng_p2 - 4})"
        ws_pac_2[f"l{leng_p2 - 3}"] = f"=sum(l10:l{leng_p2 - 4})"

        ws_pac_2[f"p{leng_p2 - 3}"] = f"=sum(p10:p{leng_p2 - 4})"
        ws_pac_2.print_area = f'A1:P{leng_p2 + 15}'
        ws_pac_2.sheet_view.topLeftCell = 'A1'
        ws_pac_2.views.sheetView[0].selection[0].activeCell = 'A1'
        ws_pac_2.views.sheetView[0].selection[0].sqref = 'A1'
        wb_pac_2.save(f'Упаковочный лист № PAC_ {entry_party.get()}-2 от {now}.xlsx')

        wb_pac.save(f'Упаковочный лист № PAC_ {entry_party.get()}-3 от {now}.xlsx')
        wb_pac_3 = openpyxl.load_workbook(f'Упаковочный лист № PAC_ {entry_party.get()}-3 от {now}.xlsx')
        ws_pac_3 = wb_pac_3.active
        ws_pac_3['A1'].value = f'Упаковочный лист № PAC_ {entry_party.get()}-3 от {now}'
        ws_pac_3['E8'].value = f'Упаковочный лист № PAC_ {entry_party.get()}-3 от {now}'
        ws_pac_3.delete_rows(10, 8000)
        ws_pac_3.delete_rows(12000 + 10 - 8000, len_pac_a - 4 - 12000 - 9)

        leng_p3 = len(ws_pac_3['a'])
        ws_pac_3[f"k{leng_p3 - 3}"] = ws_pac_3[f"k{leng_p3 - 3}"].number_format = '0.000'
        ws_pac_3[f"l{leng_p3 - 3}"] = ws_pac_3[f"l{leng_p3 - 3}"].number_format = '0.000'
        ws_pac_3[f"p{leng_p3 - 3}"] = ws_pac_3[f"p{leng_p3 - 3}"].number_format = '0.00'
        ws_pac_3[f"g{leng_p3 - 3}"] = f"=sum(g10:g{leng_p3 - 4})"
        ws_pac_3[f"i{leng_p3 - 3}"] = f"=sum(i10:i{leng_p3 - 4})"
        ws_pac_3[f"k{leng_p3 - 3}"] = f"=sum(k10:k{leng_p3 - 4})"
        ws_pac_3[f"l{leng_p3 - 3}"] = f"=sum(l10:l{leng_p3 - 4})"
        ws_pac_3[f"p{leng_p3 - 3}"] = f"=sum(p10:p{leng_p3 - 4})"
        ws_pac_3.print_area = f'A1:P{leng_p3 + 15}'
        ws_pac_3.sheet_view.topLeftCell = 'A1'
        ws_pac_3.views.sheetView[0].selection[0].activeCell = 'A1'
        ws_pac_3.views.sheetView[0].selection[0].sqref = 'A1'
        wb_pac_3.save(f'Упаковочный лист № PAC_ {entry_party.get()}-3 от {now}.xlsx')

        wb_pac.save(f'Упаковочный лист № PAC_ {entry_party.get()}-4 от {now}.xlsx')
        wb_pac_4 = openpyxl.load_workbook(f'Упаковочный лист № PAC_ {entry_party.get()}-4 от {now}.xlsx')
        ws_pac_4 = wb_pac_4.active
        ws_pac_4['A1'].value = f'Упаковочный лист № PAC_ {entry_party.get()}-4 от {now}'
        ws_pac_4['E8'].value = f'Упаковочный лист № PAC_ {entry_party.get()}-4 от {now}'
        ws_pac_4.delete_rows(10, 12000)

        leng_p4 = len(ws_pac_4['a'])
        print(leng_p4)
        ws_pac_4[f"k{leng_p4 - 3}"] = ws_pac_4[f"k{leng_p4 - 3}"].number_format = '0.000'
        ws_pac_4[f"l{leng_p4 - 3}"] = ws_pac_4[f"l{leng_p4 - 3}"].number_format = '0.000'
        ws_pac_4[f"p{leng_p4 - 3}"] = ws_pac_4[f"p{leng_p4 - 3}"].number_format = '0.00'
        ws_pac_4[f"g{leng_p4 - 3}"] = f"=sum(g10:g{leng_p4 - 4})"
        ws_pac_4[f"i{leng_p4 - 3}"] = f"=sum(i10:i{leng_p4 - 4})"
        ws_pac_4[f"k{leng_p4 - 3}"] = f"=sum(k10:k{leng_p4 - 4})"
        ws_pac_4[f"l{leng_p4 - 3}"] = f"=sum(l10:l{leng_p4 - 4})"
        ws_pac_4[f"p{leng_p4 - 3}"] = f"=sum(p10:p{leng_p4 - 4})"
        ws_pac_4.print_area = f'A1:P{leng_p4 + 15}'
        ws_pac_4.sheet_view.topLeftCell = 'A1'
        ws_pac_4.views.sheetView[0].selection[0].activeCell = 'A1'
        ws_pac_4.views.sheetView[0].selection[0].sqref = 'A1'
        wb_pac_4.save(f'Упаковочный лист № PAC_ {entry_party.get()}-4 от {now}.xlsx')

    elif 12000 > len_pac_a > 8000:
        wb_pac.save(f'Упаковочный лист № PAC_ {entry_party.get()}-1 от {now}.xlsx')
        wb_pac_1 = openpyxl.load_workbook(f'Упаковочный лист № PAC_ {entry_party.get()}-1 от {now}.xlsx')
        ws_pac_1 = wb_pac_1.active
        ws_pac_1['A1'].value = f'Упаковочный лист № PAC_ {entry_party.get()}-1 от {now}'
        ws_pac_1['E8'].value = f'Упаковочный лист № PAC_ {entry_party.get()}-1 от {now}'
        ws_pac_1.delete_rows(4000 + 10, len_pac_a - 4 - 4000 - 9)

        leng_p1 = len(ws_pac_1['a'])
        print(leng_p1)
        ws_pac_1[f"k{leng_p1 - 3}"] = ws_pac_1[f"k{leng_p1 - 3}"].number_format = '0.000'
        ws_pac_1[f"l{leng_p1 - 3}"] = ws_pac_1[f"l{leng_p1 - 3}"].number_format = '0.000'
        ws_pac_1[f"p{leng_p1 - 3}"] = ws_pac_1[f"p{leng_p1 - 3}"].number_format = '0.00'

        ws_pac_1[f"g{leng_p1 - 3}"] = f"=sum(g10:g{leng_p1 - 4})"
        ws_pac_1[f"i{leng_p1 - 3}"] = f"=sum(i10:i{leng_p1 - 4})"
        ws_pac_1[f"k{leng_p1 - 3}"] = f"=sum(k10:k{leng_p1 - 4})"
        ws_pac_1[f"l{leng_p1 - 3}"] = f"=sum(l10:l{leng_p1 - 4})"
        ws_pac_1[f"p{leng_p1 - 3}"] = f"=sum(p10:p{leng_p1 - 4})"
        ws_pac_1.print_area = f'A1:P{leng_p1 + 15}'
        ws_pac_1.sheet_view.topLeftCell = 'A1'
        ws_pac_1.views.sheetView[0].selection[0].activeCell = 'A1'
        ws_pac_1.views.sheetView[0].selection[0].sqref = 'A1'
        wb_pac_1.save(f'Упаковочный лист № PAC_ {entry_party.get()}-1 от {now}.xlsx')

        wb_pac.save(f'Упаковочный лист № PAC_ {entry_party.get()}-2 от {now}.xlsx')
        wb_pac_2 = openpyxl.load_workbook(f'Упаковочный лист № PAC_ {entry_party.get()}-2 от {now}.xlsx')
        ws_pac_2 = wb_pac_2.active
        ws_pac_2['A1'].value = f'Упаковочный лист № PAC_ {entry_party.get()}-2 от {now}'
        ws_pac_2['E8'].value = f'Упаковочный лист № PAC_ {entry_party.get()}-2 от {now}'
        ws_pac_2.delete_rows(10, 4000)
        ws_pac_2.delete_rows(8000 + 10 - 4000, len_pac_a - 4 - 8000 - 9)

        leng_p2 = len(ws_pac_2['a'])
        ws_pac_2[f"k{leng_p2 - 3}"] = ws_pac_2[f"k{leng_p2 - 3}"].number_format = '0.000'
        ws_pac_2[f"l{leng_p2 - 3}"] = ws_pac_2[f"l{leng_p2 - 3}"].number_format = '0.000'
        ws_pac_2[f"p{leng_p2 - 3}"] = ws_pac_2[f"p{leng_p2 - 3}"].number_format = '0.00'
        ws_pac_2[f"g{leng_p2 - 3}"] = f"=sum(g10:g{leng_p2 - 4})"
        ws_pac_2[f"i{leng_p2 - 3}"] = f"=sum(i10:i{leng_p2 - 4})"
        ws_pac_2[f"k{leng_p2 - 3}"] = f"=sum(k10:k{leng_p2 - 4})"
        ws_pac_2[f"l{leng_p2 - 3}"] = f"=sum(l10:l{leng_p2 - 4})"
        ws_pac_2[f"p{leng_p2 - 3}"] = f"=sum(p10:p{leng_p2 - 4})"
        ws_pac_2.print_area = f'A1:P{leng_p2 + 15}'
        ws_pac_2.sheet_view.topLeftCell = 'A1'
        ws_pac_2.views.sheetView[0].selection[0].activeCell = 'A1'
        ws_pac_2.views.sheetView[0].selection[0].sqref = 'A1'
        wb_pac_2.save(f'Упаковочный лист № PAC_ {entry_party.get()}-2 от {now}.xlsx')

        wb_pac.save(f'Упаковочный лист № PAC_ {entry_party.get()}-3 от {now}.xlsx')
        wb_pac_3 = openpyxl.load_workbook(f'Упаковочный лист № PAC_ {entry_party.get()}-3 от {now}.xlsx')
        ws_pac_3 = wb_pac_3.active
        ws_pac_3['A1'].value = f'Упаковочный лист № PAC_ {entry_party.get()}-3 от {now}'
        ws_pac_3['E8'].value = f'Упаковочный лист № PAC_ {entry_party.get()}-3 от {now}'
        ws_pac_3.delete_rows(10, 8000)

        leng_p3 = len(ws_pac_3['a'])
        ws_pac_3[f"k{leng_p3 - 3}"] = ws_pac_3[f"k{leng_p3 - 3}"].number_format = '0.000'
        ws_pac_3[f"l{leng_p3 - 3}"] = ws_pac_3[f"l{leng_p3 - 3}"].number_format = '0.000'
        ws_pac_3[f"p{leng_p3 - 3}"] = ws_pac_3[f"p{leng_p3 - 3}"].number_format = '0.00'
        ws_pac_3[f"g{leng_p3 - 3}"] = f"=sum(g10:g{leng_p3 - 4})"
        ws_pac_3[f"i{leng_p3 - 3}"] = f"=sum(i10:i{leng_p3 - 4})"
        ws_pac_3[f"k{leng_p3 - 3}"] = f"=sum(k10:k{leng_p3 - 4})"
        ws_pac_3[f"l{leng_p3 - 3}"] = f"=sum(l10:l{leng_p3 - 4})"
        ws_pac_3[f"p{leng_p3 - 3}"] = f"=sum(p10:p{leng_p3 - 4})"
        ws_pac_3.print_area = f'A1:P{leng_p3 + 15}'
        ws_pac_3.sheet_view.topLeftCell = 'A1'
        ws_pac_3.views.sheetView[0].selection[0].activeCell = 'A1'
        ws_pac_3.views.sheetView[0].selection[0].sqref = 'A1'
        wb_pac_3.save(f'Упаковочный лист № PAC_ {entry_party.get()}-3 от {now}.xlsx')

    elif 8000 > len_pac_a > 4000:
        wb_pac.save(f'Упаковочный лист № PAC_ {entry_party.get()}-1 от {now}.xlsx')
        wb_pac_1 = openpyxl.load_workbook(f'Упаковочный лист № PAC_ {entry_party.get()}-1 от {now}.xlsx')
        ws_pac_1 = wb_pac_1.active
        ws_pac_1['A1'].value = f'Упаковочный лист № PAC_ {entry_party.get()}-1 от {now}'
        ws_pac_1['E8'].value = f'Упаковочный лист № PAC_ {entry_party.get()}-1 от {now}'
        ws_pac_1.delete_rows(4000 + 10, len_pac_a - 4 - 4000 - 9)

        leng_p1 = len(ws_pac_1['a'])
        ws_pac_1[f"k{leng_p1 - 3}"] = ws_pac_1[f"k{leng_p1 - 3}"].number_format = '0.000'
        ws_pac_1[f"l{leng_p1 - 3}"] = ws_pac_1[f"l{leng_p1 - 3}"].number_format = '0.000'
        ws_pac_1[f"p{leng_p1 - 3}"] = ws_pac_1[f"p{leng_p1 - 3}"].number_format = '0.00'

        ws_pac_1[f"g{leng_p1 - 3}"] = f"=sum(g10:g{leng_p1 - 4})"
        ws_pac_1[f"i{leng_p1 - 3}"] = f"=sum(i10:i{leng_p1 - 4})"
        ws_pac_1[f"k{leng_p1 - 3}"] = f"=sum(k10:k{leng_p1 - 4})"
        ws_pac_1[f"l{leng_p1 - 3}"] = f"=sum(l10:l{leng_p1 - 4})"
        ws_pac_1[f"p{leng_p1 - 3}"] = f"=sum(p10:p{leng_p1 - 4})"
        ws_pac_1.print_area = f'A1:P{leng_p1 + 15}'
        ws_pac_1.sheet_view.topLeftCell = 'A1'
        ws_pac_1.views.sheetView[0].selection[0].activeCell = 'A1'
        ws_pac_1.views.sheetView[0].selection[0].sqref = 'A1'
        wb_pac_1.save(f'Упаковочный лист № PAC_ {entry_party.get()}-1 от {now}.xlsx')

        wb_pac.save(f'Упаковочный лист № PAC_ {entry_party.get()}-2 от {now}.xlsx')
        wb_pac_2 = openpyxl.load_workbook(f'Упаковочный лист № PAC_ {entry_party.get()}-2 от {now}.xlsx')
        ws_pac_2 = wb_pac_2.active
        ws_pac_2['A1'].value = f'Упаковочный лист № PAC_ {entry_party.get()}-2 от {now}'
        ws_pac_2['E8'].value = f'Упаковочный лист № PAC_ {entry_party.get()}-2 от {now}'
        ws_pac_2.delete_rows(10, 4000)

        leng_p2 = len(ws_pac_2['a'])
        ws_pac_2[f"k{leng_p2 - 3}"] = ws_pac_2[f"k{leng_p2 - 3}"].number_format = '0.000'
        ws_pac_2[f"l{leng_p2 - 3}"] = ws_pac_2[f"l{leng_p2 - 3}"].number_format = '0.000'
        ws_pac_2[f"p{leng_p2 - 3}"] = ws_pac_2[f"p{leng_p2 - 3}"].number_format = '0.00'
        ws_pac_2[f"g{leng_p2 - 3}"] = f"=sum(g10:g{leng_p2 - 4})"
        ws_pac_2[f"i{leng_p2 - 3}"] = f"=sum(i10:i{leng_p2 - 4})"
        ws_pac_2[f"k{leng_p2 - 3}"] = f"=sum(k10:k{leng_p2 - 4})"
        ws_pac_2[f"l{leng_p2 - 3}"] = f"=sum(l10:l{leng_p2 - 4})"
        ws_pac_2[f"p{leng_p2 - 3}"] = f"=sum(p10:p{leng_p2 - 4})"
        ws_pac_2.print_area = f'A1:P{leng_p2 + 15}'
        ws_pac_2.sheet_view.topLeftCell = 'A1'
        ws_pac_2.views.sheetView[0].selection[0].activeCell = 'A1'
        ws_pac_2.views.sheetView[0].selection[0].sqref = 'A1'
        wb_pac_2.save(f'Упаковочный лист № PAC_ {entry_party.get()}-2 от {now}.xlsx')

    elif 4000 > len_pac_a:
        pass

    else:
        pass

    msg = f'Упаковочный лист разделен!'
    mb.showinfo("Информация", msg)

def load_info():
    msg = "Выберите файл с загрузкой по мешкам"
    mb.showinfo("批所有", msg)
    file_name1 = filedialog.askopenfilename()
    print(file_name1)
    df_w = pd.read_excel(file_name1, sheet_name=0, engine='openpyxl', usecols='B,C,D,F,G',
                         dtype={'Stamp 铅封号': str})

    mapping = {df_w.columns[1]: 'Номер накладной', df_w.columns[2]: 'Трек номер',
               df_w.columns[3]: 'вес Накладной по загрузке', df_w.columns[0]: 'Пломба'}

    df_w = df_w.rename(columns=mapping)  # (columns={'Order number 单号': 'Номер накладной',
    # [4]: 'Трек номер',
    # 'Gross weight, kg  毛重 公斤': 'вес Накладной по загрузке',
    # 'Stamp 铅封号': 'Пломба'})
    print(df_w)
    weight_brut = df_w['вес Накладной по загрузке'].iloc[-1]
    weight_net = df_w['вес Накладной по загрузке'].iloc[-2]
    if weight_brut / weight_net > 2:
        weight_net = df_w['вес Накладной по загрузке'].iloc[-3]
    else:
        pass
    print(weight_brut)
    print(weight_net)
    weight_all_chinabag = weight_brut - weight_net
    count_chinabag = len(df_w['Пломба'].unique()) - 1
    Chinabag_weight = round(weight_all_chinabag / count_chinabag, 4)
    msg = f"вес брутто {weight_brut},\n вес нетто {weight_net},\n кол-во упаковок (пломб) {count_chinabag},\n\n Вес единицы упаковки {Chinabag_weight}"
    mb.showinfo("Информация по загрузке", msg)

window = tk.Tk()
window.title('Формирование транзитных документов')
window.geometry("450x520+500+200")
party_name = tk.Label(window, text="Номер партии", font='hank 9 bold')
entry_bag_wht_lb = tk.Label(window, text="Вес 1 ед. тары", font='hank 9 bold')
entry_date_lb = tk.Label(window, text="Дата", font='hank 9 bold')
entry_pac_var_lb = tk.Label(window, text="Вид упаковки", font='hank 9 bold')
entry_sample_var_lb = tk.Label(window, text="Шаблон", font='hank 9 bold')
entry_koef_lb = tk.Label(window, text="Коэффициент уменьшения цены (работает только для Цайняо, МТТ и РЭ)", font='hank 9 bold')

a = tk.StringVar(value='OZON-218')
b = tk.StringVar()
c = tk.StringVar(value=now)
koef = tk.StringVar(value='0,5')
#pac_var = tk.StringVar(value='мешок п/п')
entry_party = tk.Entry(window, width=20, textvariable=a)
entry_bag_wht = tk.Entry(window, width=20, textvariable=b)
entry_date = tk.Entry(window, width=20, textvariable=c)
entry_koef = tk.Entry(window, width=20, textvariable=koef)
# entry_pac_var = tk.Entry(window, width=20, textvariable=pac_var)

button = tk.Button(text="Упаковочный лист", width=24, height=2, bg="lightgrey", fg="black", command=start)
button.configure(font=('hank', 10))
button_spc = tk.Button(text="Спецификация + СМР", width=24, height=2, bg="lightgrey", fg="black", command=Specf)
button_spc.configure(font=('hank', 10))
button_pac_split = tk.Button(text="Разделить Упаковочный лист", width=24, height=2, bg="lightgrey", fg="black", command=pac_split)
button_pac_split.configure(font=('hank', 10))
#button_ad_tnvd = tk.Button(text="Обновить базу Кодов ТНВЭД", width=14, height=2, bg="lightgrey", fg="black", command=update_tnved)
#button_ad_tnvd.configure(font=('hank', 10))
# button_cmr = tk.Button(text="Ввести СМР", width=14, height=2, bg="lightgrey", fg="black", command=get_party)
# button_cmr.configure(font=('hank', 10))
button_load_info = tk.Button(text="Информация по загрузке", width=24, height=2, bg="lightgrey", fg="black", command=load_info)
button_load_info.configure(font=('hank', 10))
button_check = tk.Button(text="Проверить на риск", width=24, height=2, bg="lightgrey", fg="black", command=check_party)
button_check.configure(font=('hank', 10))


Pac_list = [   #to chose packing
"Коробка",
"Мешок п/п",
"Ящик"
] #etc
Pac_choice_var = StringVar(window)
Pac_choice_var.set(Pac_list[0]) # set default value
w = OptionMenu(window, Pac_choice_var, *Pac_list)

Sample_list = [   #to chose sample
"ШАБЛОН OZON",
"ШАБЛОН МТК CEL",
"ШАБЛОН CaiNiao",
"ШАБЛОН МТТ",
"ШАБЛОН HECNY",
"ШАБЛОН HV",
"ШАБЛОН SUI",
"ШАБЛОН AKB",
"ШАБЛОН EUB"
]
Sample_choice_var = StringVar(window)
Sample_choice_var.set(Sample_list[0]) # set default value
w2 = OptionMenu(window, Sample_choice_var, *Sample_list, command=get_selection)

button_load_info.pack()
button_check.pack()
entry_sample_var_lb.pack()
w2.pack()
entry_pac_var_lb.pack()
w.pack()
party_name.pack()
entry_party.pack()
entry_bag_wht_lb.pack()
entry_bag_wht.pack()
entry_date_lb.pack()
entry_date.pack()
entry_koef_lb.pack()
entry_koef.pack()
# button_cmr.pack()
button.pack()
button_spc.pack()
button_pac_split.pack()
#button_ad_tnvd.pack()
window.mainloop()