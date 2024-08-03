import pandas as pd
from tkinter import filedialog
import tkinter as tk
from tkinter import filedialog
import tkinter.messagebox as mb
import datetime
import numpy as np
import openpyxl
import requests
import xlsxwriter
import re
import random
import xml.etree.ElementTree as ET

from openpyxl.styles import PatternFill

now = datetime.datetime.now().strftime("%d.%m.%Y")


def start():
    msg = "Выберите файл с реестром"
    mb.showinfo("批包括", msg)
    file_name = filedialog.askopenfilename()
    try:
        df = pd.read_excel(file_name, sheet_name=0, header=None, engine='openpyxl',
                   skiprows=1, usecols='A:I, K:T, V, W, AD', converters={7: str, 14: str, 15: str, 18: str})

        df.columns = ['Номер накладной СДЭК', 'Фамилия', 'Имя', 'Отчество',
                      'Адрес получателя', 'Город', 'Область', 'Индекс', 'Телефон',
                      'Количество единиц товара', 'Наименование товара', 'Стоимость ед. товарной позиции',
                      'Ссылка на товар', 'Серия паспорта', 'Номер паспорта', 'Дата выдачи', 'Дата рождения',
                      'Идентификационный налоговый номер', 'Вес брутто (Вес позиции)', 'Клиент', 'Номер отправления ИМ', 'Емейл']
    except:

        df = pd.read_excel(file_name, sheet_name=0, header=None, engine='openpyxl',
                           skiprows=1, usecols='A:I, K:T, V, W', converters={7: str, 14: str, 15: str, 18: str})
        df.columns = ['Номер накладной СДЭК', 'Фамилия', 'Имя', 'Отчество',
                      'Адрес получателя', 'Город', 'Область', 'Индекс', 'Телефон',
                      'Количество единиц товара', 'Наименование товара', 'Стоимость ед. товарной позиции',
                      'Ссылка на товар', 'Серия паспорта', 'Номер паспорта', 'Дата выдачи', 'Дата рождения',
                      'Идентификационный налоговый номер', 'Вес брутто (Вес позиции)', 'Клиент', 'Номер отправления ИМ']
    len_df_income = len(df)
    print(len_df_income)
    df['Фамилия'] = df["Фамилия"].str.replace("\W+", " ", regex=True)
    df['Стоимость ед. товарной позиции'] = df['Стоимость ед. товарной позиции'].replace(to_replace=',,', value='.', regex=True)
    df['Стоимость ед. товарной позиции'] = df['Стоимость ед. товарной позиции'].replace(to_replace='\.\.', value='.', regex=True)
    df['Стоимость ед. товарной позиции'] = df['Стоимость ед. товарной позиции'].replace(to_replace='，', value='.',
                                                                                        regex=True)
    df['Стоимость ед. товарной позиции'] = df['Стоимость ед. товарной позиции'].replace(to_replace=',', value='.', regex=True)
    df['Стоимость ед. товарной позиции'] = df['Стоимость ед. товарной позиции'].replace(to_replace='，', value='.', regex=True)
    df['Стоимость ед. товарной позиции'] = df['Стоимость ед. товарной позиции'].replace(to_replace='^\.', value='', regex=True)
    df['Стоимость ед. товарной позиции'] = df['Стоимость ед. товарной позиции'].replace(to_replace='^,', value='', regex=True)
    df['Стоимость ед. товарной позиции'] = df['Стоимость ед. товарной позиции'].replace(to_replace=',$', value='', regex=True)
    try:
        df['Стоимость ед. товарной позиции'] = df['Стоимость ед. товарной позиции'].replace(to_replace='\.$', value='', regex=True).astype('float')
    except ValueError as msg:
        mb.showinfo("Столбец цена: присутствует нечисловое значение!", f'{str(msg)}\nПроверьте выставлены ли столбцы по шаблону, \nзатем проверьте сам столбец с ценой!')

    s = df['Стоимость ед. товарной позиции'].isnull().any()
    if s == True:
        msg = "Проверьте цены, есть пустые значения! Исправьте и начните заново"
        mb.showinfo("Информация", msg)
    else:
        pass
    #df['Наименование товара'] = df['Наименование товара'].replace(to_replace='&', value='', regex=True)
    df['Стоимость позиции'] = df['Количество единиц товара'].multiply(df['Стоимость ед. товарной позиции'], axis='index')
    df['Вес брутто (Вес позиции)'] = df['Вес брутто (Вес позиции)'].replace(to_replace=',,', value='.', regex=True)
    df['Вес брутто (Вес позиции)'] = df['Вес брутто (Вес позиции)'].replace(to_replace='\.\.', value='.', regex=True)
    df['Вес брутто (Вес позиции)'] = df['Вес брутто (Вес позиции)'].replace(to_replace=',', value='.', regex=True)
    df['Вес брутто (Вес позиции)'] = df['Вес брутто (Вес позиции)'].replace(to_replace='，', value='.', regex=True)
    df['Вес брутто (Вес позиции)'] = df['Вес брутто (Вес позиции)'].replace(to_replace='^\.', value='', regex=True)
    df['Вес брутто (Вес позиции)'] = df['Вес брутто (Вес позиции)'].replace(to_replace='^,', value='', regex=True)
    df['Вес брутто (Вес позиции)'] = df['Вес брутто (Вес позиции)'].replace(to_replace=',$', value='', regex=True)
    try:
        df['Вес брутто (Вес позиции)'] = df['Вес брутто (Вес позиции)'].replace(to_replace='\.$', value='', regex=True).astype('float')
    except ValueError as msg:
        mb.showinfo("Столбец вес: присутствует нечисловое значение!", f'{str(msg)}\nПроверьте выставлены ли столбцы по шаблону, \nзатем проверьте сам столбец с весом!')

    W_0 = df['Вес брутто (Вес позиции)'].isnull().any()
    if W_0 == True:
        msg = "Проверьте вес, есть пустые значения! Исправьте и начните заново"
        mb.showinfo("Информация", msg)
    else:
        pass
    print('weight ok')
    df['Дата выдачи'] = df['Дата выдачи'].replace(to_replace=',', value='.', regex=True)
    df['Дата рождения'] = df['Дата рождения'].replace(to_replace=',', value='.', regex=True)
    try:
        df['Дата выдачи work'] = pd.to_datetime(df['Дата выдачи'], errors='coerce', dayfirst=True).dt.strftime('%d.%m.%Y')
        mask = df['Дата выдачи work'].isnull()
        df.loc[mask, 'Дата выдачи work'] = df['Дата выдачи'].astype(str) + '(error)'
        df['Дата выдачи'] = df['Дата выдачи work']
        df['Дата выдачи'] = df['Дата выдачи'].replace(to_replace='nan(error)', value='')
    except Exception as e:
        msg = f"Проверьте Дату выдачи, есть кривые значения! Исправьте и начните заново\n\nТекст ошибки:\n\n{e}"
        mb.showinfo("Информация", msg)
    try:
        df['Дата рождения work'] = pd.to_datetime(df['Дата рождения'], errors='coerce', dayfirst=True).dt.strftime('%d.%m.%Y')
        mask = df['Дата рождения work'].isnull()
        df.loc[mask, 'Дата рождения work'] = df['Дата рождения'].astype(str) + '(error)'
        df['Дата рождения'] = df['Дата рождения work']
        df['Дата рождения'] = df['Дата рождения'].replace(to_replace='nan(error)', value='')
    except Exception as e:
        msg = f"Проверьте Дату рождения, есть кривые значения!\n\nТекст ошибки:\n\n{e}"
        mb.showinfo("Информация", msg)
    print('dates ok')
    df['Телефон'] = df['Телефон'].astype(str)
    df['Телефон'] = df['Телефон'] = df['Телефон'].replace(to_replace='\D', value='', regex=True)
    i = 0
    for col in df['Телефон']:
        i += 1
        print(i, col)
        try:
            len_phone = len(col)
            if len_phone == 12:
                df.loc[df['Телефон'] == col, 'Телефон'] = col[:-1]
            if len_phone == 10:
                df.loc[df['Телефон'] == col, 'Телефон'] = f'7{col}'
            if len_phone == 9:
                df.loc[df['Телефон'] == col, 'Телефон'] = f'79{col}'
            if col[0] != '7':
                df.loc[df['Телефон'] == col, 'Телефон'] = '7' + col[1:]
            if col[1] != '9':
                df.loc[df['Телефон'] == col, 'Телефон'] = '79' + col[2:]
        except Exception as e:
            msg = f"Проверьте Телефоны!\n\nТекст ошибки:\n\n{e}"
            mb.showinfo("Информация", msg)

    print('phone ok')
    df['Серия паспорта'] = df['Серия паспорта'].astype(str)

    df['Серия паспорта'] = df['Серия паспорта'] = df['Серия паспорта'].replace(to_replace='\W', value='', regex=True)
    df['Серия паспорта'] = df['Серия паспорта'] = df['Серия паспорта'].replace(to_replace='\s', value='', regex=True)


    df['Номер паспорта'] = df['Номер паспорта'].astype(str)
    df['Номер паспорта'] = df['Номер паспорта'] = df['Номер паспорта'].replace(to_replace='\W', value='', regex=True)
    df['Номер паспорта'] = df['Номер паспорта'] = df['Номер паспорта'].replace(to_replace='\s', value='', regex=True)
    i = 0
    for col in df['Номер паспорта']:
        i += 1
        if len(col) == 5:
            df.loc[df['Номер паспорта'] == col, 'Номер паспорта'] = f'0{col}'
        elif len(col) == 4:
            df.loc[df['Номер паспорта'] == col, 'Номер паспорта'] = f'00{col}'
        if any(c.isalpha() for c in col):
            try:
                chars = re.findall(r'[a-zA-Z]+', col)[0]
                if chars != 'nan':
                    print(col)
                    print("alpha")
                    print(chars)
                    df.loc[df['Номер паспорта'] == col, 'Серия паспорта'] = f'{chars}'
                    print(df.loc[df['Номер паспорта'] == col, 'Серия паспорта'])
            except:
                print('error', col)

    print('pass numb ok')
    df_group = df.groupby('Номер накладной СДЭК', sort=False)[['Вес брутто (Вес позиции)', 'Стоимость позиции']].sum()

    df_group = df_group.rename(columns={'Вес брутто (Вес позиции)': 'Общий Вес места (накладной)',
                                        'Стоимость позиции': 'Общая стоимость накладной(посылки)'})

    df = pd.merge(df, df_group, how='left', left_on='Номер накладной СДЭК', right_on='Номер накладной СДЭК')
    print('merge ok')
    msg = "Выберите файл с загрузкой по мешкам"
    mb.showinfo("批所有", msg)
    file_name1 = filedialog.askopenfilename()
    df_w = pd.read_excel(file_name1, sheet_name=0, engine='openpyxl', usecols='B,C,F',
                              dtype={'Stamp 铅封号': str})
    df_w = df_w.rename(columns={'Order number 单号': 'Номер накладной',
                           'Gross weight, kg  毛重 公斤': 'вес Накладной по загрузке',
                           'Stamp 铅封号': 'Номер пломбы'})
    weight_brut = df_w['вес Накладной по загрузке'].iloc[-1]
    weight_net = df_w['вес Накладной по загрузке'].iloc[-2]
    if weight_brut / weight_net > 2:
        weight_net = df_w['вес Накладной по загрузке'].iloc[-3]
    else:
        pass
    weight_all_chinabag = weight_brut - weight_net
    count_chinabag = len(df_w['Номер пломбы'].unique()) - 1
    Chinabag_weight = round(weight_all_chinabag / count_chinabag, 3)
    msg = f"вес брутто {weight_brut},\n вес нетто {weight_net},\n кол-во упаковок (пломб) {count_chinabag},\n\n Вес единицы упаковки {Chinabag_weight}"
    mb.showinfo("Информация по загрузке", msg)
    print(df_w)
    test_cmpr_df = df['Номер накладной СДЭК'].drop_duplicates().dropna()
    test_cmpr_df_w = df_w['Номер накладной'].dropna()

    for i in test_cmpr_df:
        print(f'{i} test1')
        test_comp1 = test_cmpr_df_w.isin([i]).any()
        if test_comp1 == False:
            msg = f"Накладная {i} не найдена в загрузке"
            mb.showerror("Ошибка", msg)

    for i in test_cmpr_df_w:
        print(f'{i} test2')
        test_comp2 = test_cmpr_df.isin([i]).any()
        if test_comp2 == False:
            msg = f"Накладная {i} лишняя в загрузке"
            mb.showerror("Ошибка", msg)

    df = pd.merge(df, df_w, how='left', left_on='Номер накладной СДЭК' , right_on='Номер накладной')
    print('merge df, df_w ok')
    df['Вес нетто'] = np.round(df['вес Накладной по загрузке'] / df['Общий Вес места (накладной)'] * df['Вес брутто (Вес позиции)'],
                               decimals=3)
    df_pc_qt = df['Номер пломбы'].value_counts()
    df = pd.merge(df, df_pc_qt, how='left', left_on='Номер пломбы', right_index=True)

    df['Номер пломбы'] = df['Номер пломбы_x']
    print('merge df, df_pc_qt ok')
    df['Вес брутто (Вес позиции)'] = df['Вес нетто'] + np.round(Chinabag_weight / df['Номер пломбы_y'], decimals=3)
    print(df['Вес брутто (Вес позиции)'])
    df['Общий Вес места (накладной)'] = df.groupby('Номер накладной СДЭК')['Вес брутто (Вес позиции)'].transform('sum')
    df['Орган выдачи'] = 'УФМС России'
    df['Код страны получателя'] = 'RU'
    try:
        df['Отчество'] = df['Отчество'].replace(0, '')
        df['Фамилия'] = df['Фамилия'].replace(0, '')
        df['Имя'] = df['Имя'].replace(0, '')
        df['ФИО получателя'] = df['Фамилия'] + ' ' + df['Имя'] + ' ' + df['Отчество'].astype(str)
        df['ФИО получателя'] = df['ФИО получателя'].replace(to_replace=' nan', value='', regex=True)
    except:
        msg = f"в столбцах ФИО получателя есть числа вместо текста.\n\nЛучше проверить, исправить и начать заново"
        mb.showerror("Ошибка", msg)
    country_dict = {'УФМС Узбекистана': 'UZ', 'УФМС Казахстана': 'KZ',
                    'УФМС Азербайджана': 'AZ', 'УФМС Армении': 'AM',
                    'УФМС Беларуси': 'BY', 'УФМС Киргизии': 'KG'}
    for col in df['Серия паспорта']:
        if any(c.isalpha() for c in col) and col != 'nan':
            fio = df.loc[df['Серия паспорта'] == col, 'ФИО получателя'].values[0]
            print(fio)
            if 'оглы' in str.lower(fio) or 'углы' in str.lower(fio) or 'угли' in str.lower(fio) or 'ugli' in str.lower(fio) or 'ogli' in str.lower(fio) or 'ahmed' in str.lower(fio):
                df.loc[df['Серия паспорта'] == col, 'Орган выдачи'] = f'УФМС Узбекистана'
                df.loc[df['Серия паспорта'] == col, 'Код страны получателя'] = f'UZ'
            if 'кызы' in str.lower(fio) or 'kizi' in str.lower(fio):
                df.loc[df['Серия паспорта'] == col, 'Орган выдачи'] = f'УФМС Казахстан'
                df.loc[df['Серия паспорта'] == col, 'Код страны получателя'] = f'KZ'
            else:
                print(col)
                org_issue, countru_code = random.choice(list(country_dict.items()))
                print(org_issue, countru_code)
                df.loc[df['Серия паспорта'] == col, 'Орган выдачи'] = f'{org_issue}'
                df.loc[df['Серия паспорта'] == col, 'Код страны получателя'] = f'{countru_code}'

    df['Код ТН ВЭД'] = None
    df['Длина коробки, см'] = None
    df['Ширина, см'] = None
    df['Высота коробки, см'] = None
    df['Отправитель по AWB'] = 'CO.,LTD. Hunchun Changda Electronic Commerce Co. Ltd'
    party_numb = entry_party.get()
    if 'OZON' in party_numb:
        df['Валюта Объявленной стоимости товара'] = 'RUB'
    else:
        df['Валюта Объявленной стоимости товара'] = 'CNY'
    df['№ AWB'] = party_numb
    df['Дата AWB'] = now
    df['Страна отправления'] = 'CN'
    df['Торгующая страна'] = 'CN'
    df['Условия поставки'] = 'DAP'

    df['Краткое наименование страны получателя'] = 'Россия'
    df['Код документа (паспорта)'] = '21'
    df['Ссылка на товар'] = df['Ссылка на товар'].str.slice(0, 250)
    df['Город'] = df['Город'].str.slice(0, 34)
    df['Область'] = df['Область'].str.slice(0, 34)
    df['Адрес получателя'] = df['Адрес получателя'].str.replace('Россия, ', '')
    df['Адрес получателя'] = df['Адрес получателя'].str[-49:]
    #df['Адрес получателя'] = df['Адрес получателя'].str.replace('^([^,])+', '', regex=True)
    df['Серия паспорта'] = df['Серия паспорта'].replace('nan', '', regex=True)
    df['Номер паспорта'] = df['Номер паспорта'].replace('nan', '', regex=True)
    df['Длина коробки, см'] = None
    df['пол (1 - женский, 0 - мужской)'] = None
    df['Признак платности (1 - платный, 0 - нет)'] = None

    df['Примечание'] = df['Клиент'].replace(to_replace='OZON空运', value='OZON-AIR', regex=True)
    df['Примечание'] = df['Примечание'].replace(to_replace='空运-轻小件', value='OZON-AIR', regex=True)
    df['Примечание'] = df['Примечание'].replace(to_replace='OZON空运到门', value='OZON-AIR',
                                                              regex=True)
    df['Примечание'] = df['Примечание'].replace(to_replace='OZON陆运', value='OZON-LAND',
                                                              regex=True)
    df['Примечание'] = df['Примечание'].replace(to_replace='陆运-轻小件', value='OZON-LAND',
                                                              regex=True)
    df['Примечание'] = df['Примечание'].replace(to_replace='OZON陆运到门', value='OZON-LAND',
                                                              regex=True)
    df['Примечание'] = df['Примечание'].replace(to_replace='自提', value='JD-AIR',
                                                              regex=True)
    df['Примечание'] = df['Примечание'].replace(to_replace='到门', value='JD-AIR',
                                                              regex=True)

    df = df.reindex(columns=['Номер отправления ИМ', 'Номер пломбы', 'ФИО получателя', 'Фамилия',
                          'Имя', 'Отчество', 'Индекс', 'Область', 'Город', 'Адрес получателя', 'Телефон',
                          'Емейл', 'Серия паспорта', 'Номер паспорта', 'Дата выдачи', 'Орган выдачи', 'Дата рождения',
                             'Идентификационный налоговый номер', 'Ссылка на товар', 'Наименование товара', 'Код ТН ВЭД',
                             'Количество единиц товара', 'Стоимость ед. товарной позиции', 'Стоимость позиции',
                             'Общая стоимость накладной(посылки)', 'Валюта Объявленной стоимости товара', 'Вес брутто (Вес позиции)',
                             'Общий Вес места (накладной)', 'Длина коробки, см', 'Ширина, см', 'Высота коробки, см', 'Отправитель по AWB',
                              '№ AWB', 'Дата AWB', 'Страна отправления', 'Торгующая страна', 'Условия поставки', 'Код страны получателя',
                             'Краткое наименование страны получателя', 'Код документа (паспорта)', 'Номер накладной СДЭК',
                             'пол (1 - женский, 0 - мужской)', 'Признак платности (1 - платный, 0 - нет)', 'Примечание'])
    #if 'OZON' in party_numb:
    #    df = df.drop('Клиент', axis=1)
    print(len_df_income)
    len_df_finish = len(df)
    print(len_df_finish)
    df = df.sort_values(['Номер пломбы', 'Номер накладной СДЭК'], ascending=[True, True])
    writer = pd.ExcelWriter(f'Шаблон_{party_numb}.xlsx', engine='xlsxwriter')
    df.to_excel(writer, sheet_name='Sheet1', index=False)
    writer.close()
    msg = "Шаблон сформирован!"
    mb.showinfo("Информация", msg)


def divide_df():
    party_numb = entry_party.get()
    msg = "Выберите Шаблон для деления"
    mb.showinfo("Шаблон для деления", msg)
    file_name = filedialog.askopenfilename()
    df = pd.read_excel(file_name, sheet_name=0, engine='openpyxl', converters={6: str, 10: str, 12: str,
                                                                               13: str, 17: str})
    """writer = pd.ExcelWriter(f'Шаблон_{party_numb}.xlsx', engine='xlsxwriter')
    df.to_excel(writer, sheet_name='Sheet1', index=False)
    writer.save()"""
    df_LD = df.loc[df['Клиент'] == 'AIR-LD']
    df_LD = df_LD.drop('Клиент', axis=1)
    JD_filter_list = ['到门', '自提']
    df_JD = df.loc[df['Клиент'].isin(JD_filter_list)]
    df_JD = df_JD.drop('Клиент', axis=1)
    others_filter_list = ['AIR-LD', '到门', '自提']
    df_others = df.loc[~df['Клиент'].isin(others_filter_list)]
    df_others = df_others.drop('Клиент', axis=1)

    if df_others.empty:
        pass
    else:
        writer = pd.ExcelWriter(f'Шаблон_{party_numb}-others.xlsx', engine='xlsxwriter')
        df_others.to_excel(writer, sheet_name='Sheet1', index=False)
        writer.close()
    if df_LD.empty:
        pass
    else:
        writer = pd.ExcelWriter(f'Шаблон_{party_numb}-LD.xlsx', engine='xlsxwriter')
        df_LD.to_excel(writer, sheet_name='Sheet1', index=False)
        writer.close()
    if df_JD.empty:
        pass
    else:
        writer = pd.ExcelWriter(f'Шаблон_{party_numb}-JD.xlsx', engine='xlsxwriter')
        df_JD.to_excel(writer, sheet_name='Sheet1', index=False)
        writer.close()

    writer = pd.ExcelWriter(f'Шаблон_{party_numb}- Для СВХ.xlsx', engine='xlsxwriter')
    df.to_excel(writer, sheet_name='Sheet1', index=False)
    writer.close()


    msg = "Шаблон разделен!"
    mb.showinfo("Информация", msg)


def get_currency_rate(char_code_currency="USD"):
    """
    :param char_code_currency: can see here - https://www.cbr.ru/scripts/XML_daily.asp
    :return: float_number
    """
    return float(
    ET.fromstring(requests.get('https://www.cbr.ru/scripts/XML_daily.asp').text).find(
    './Valute[CharCode="EUR"]/Value').text.replace(',', '.')
    )


def commerce_products_search():
    filename = filedialog.askopenfilename()
    df = pd.read_excel(filename, sheet_name=0, engine='openpyxl', header=None, usecols='T, U, AO', skiprows=1, converters={1: str})
    df.columns = ['description', 'tnvedCode', 'track']
    df = df.reindex(columns=['track', 'description', 'tnvedCode'])
    print(df)
    df['tnvedCode9'] = df['tnvedCode'].str[:9]
    df['tnvedCode6'] = df['tnvedCode'].str[:6]
    df['tnvedCode4'] = df['tnvedCode'].str[:4]

    df_warnings = pd.read_excel('TNVEDнедляличного.xlsx', sheet_name=0, engine='openpyxl', usecols='A, B, C, D', converters={'triger_tnved': str,
                                                                                                                       'exclud': str})

    df_merged_10 = pd.merge(df, df_warnings, how='left', left_on='tnvedCode', right_on='triger_tnved')
    df_merged_9 = pd.merge(df_merged_10, df_warnings, how='left', left_on='tnvedCode9', right_on='triger_tnved')
    df_merged_6 = pd.merge(df_merged_9, df_warnings, how='left', left_on='tnvedCode6', right_on='triger_tnved')
    df_merged_4 = pd.merge(df_merged_6, df_warnings, how='left', left_on='tnvedCode4', right_on='triger_tnved')

    df_merged_4.columns = ['track', 'description', 'tnvedCode', 'tnvedCode9', 'tnvedCode6', 'tnvedCode4', 'Триггер_код', 'Описание группы/запрет'
                           , 'Постановление', 'Искл.', 'По 9 знакам', 'Описание группы/запрет_9', 'Постановление_9', 'Искл._9',
                           'По 6 знакам', 'Описание группы/запрет_6', 'Постановление_6', 'Искл._6'
                           , 'По 4 знакам', 'Описание группы/запрет_4', 'Постановление_4', 'Искл._4']
    df = df_merged_4
    df = df.drop('Искл.', axis=1)
    df = df.drop('Искл._6', axis=1)
    df = df.drop('Искл._9', axis=1)
    #df['Искл._4'] = df['Искл._4']

    df['Постановление'] = df['Постановление'].fillna(df['Постановление_9'])
    df['Постановление'] = df['Постановление'].fillna(df['Постановление_6'])
    df['Постановление'] = df['Постановление'].fillna(df['Постановление_4'])

    df['Триггер_код'] = df['Триггер_код'].fillna(df['По 9 знакам'])
    df['Триггер_код'] = df['Триггер_код'].fillna(df['По 6 знакам'])
    df['Триггер_код'] = df['Триггер_код'].fillna(df['По 4 знакам'])

    df['Описание группы/запрет'] = df['Описание группы/запрет'].fillna(df['Описание группы/запрет_9'])
    df['Описание группы/запрет'] = df['Описание группы/запрет'].fillna(df['Описание группы/запрет_6'])
    df['Описание группы/запрет'] = df['Описание группы/запрет'].fillna(df['Описание группы/запрет_4'])

    df = df.drop('tnvedCode9', axis=1)

    df = df.reindex(columns=['description', 'tnvedCode', 'tnvedCode6', 'tnvedCode4', 'Триггер_код', 'Описание группы/запрет'
                           , 'Постановление', 'Искл.', 'track'])

    writer = pd.ExcelWriter(f'TNVED_commerce.xlsx', engine='openpyxl')
    df.to_excel(writer, sheet_name='Sheet1', index=False)
    writer.close()

    df_trigger_words_export = pd.read_excel("trigger_words_commerce.xlsx", header=None)
    list_of_trigger_words = df_trigger_words_export[0].to_list()
    wb = openpyxl.load_workbook('TNVED_commerce.xlsx')
    ws = wb.active
    orangeFill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
    redFill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    pinkFill = PatternFill(start_color='FFC0CB', end_color='FFC0CB', fill_type='solid')

    i = 0
    for row in ws['B']:
        i += 1
        tnved = str(row.value)
        for triger_word in list_of_trigger_words:
            if triger_word in ws[f'A{i}'].value.lower():
                ws[f'A{i}'].fill = orangeFill
        if ws[f'E{i}'].value is not None and i != 1:
            print(ws[f'E{i}'].value)
            ws[f'A{i}'].fill = redFill
            ws[f'B{i}'].fill = redFill
            ws[f'E{i}'].fill = redFill


    wb.save('TNVED_commerce.xlsx')
    msg = "ГОТОВО!"
    mb.showinfo("ИНФО", msg)


window = tk.Tk()
window.title('Формирование шаблона OZON/CEL')
window.geometry("500x250+400+400")
name = tk.Label(window, text="Номер CMR")

a = tk.StringVar(value='OZON-4')
entry_party = tk.Entry(window,  width=20, textvariable=a)


button = tk.Button(text="Делаем шаблон!", width=20, height=2, bg="lightgrey", fg="black", command=start)
button.configure(font=('hank', 10))
button_divide = tk.Button(text="Разделить", width=20, height=2, bg="lightgrey", fg="black", command=divide_df)
button_divide.configure(font=('hank', 10))
button3 = tk.Button(text="Выделить не для личного пользования", width=30, height=2, bg="lightgrey", fg="black", command=commerce_products_search)
button3.configure(font=('hank', 10))
name.pack()
entry_party.pack()
button.pack()
button_divide.pack()
button3.pack()
window.mainloop()
